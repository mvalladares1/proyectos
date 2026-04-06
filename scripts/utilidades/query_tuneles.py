import xmlrpc.client
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from collections import defaultdict

# Conexión a Odoo
url = "https://riofuturo.server98c6e.oerpondemand.net"
db = "riofuturo-master"
username = "mvalladares@riofuturo.cl"
password = "c0766224bec30cac071ffe43a858c9ccbd521ddd"

print("Conectando a Odoo...")
common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common", context=None)
uid = common.authenticate(db, username, password, {})
if not uid:
    raise Exception("Error de autenticación")
print(f"Autenticado. UID: {uid}")

models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object", context=None)

def search_read(model, domain, fields, limit=0):
    return models.execute_kw(db, uid, password, model, 'search_read', [domain], {'fields': fields, 'limit': limit})

# Buscar los centros de trabajo (workcenters) que coincidan con los túneles
print("\nBuscando centros de trabajo de túneles estáticos...")
tunnels_names = [
    "[1.1] PROCESO CONGELADO TÚNEL ESTÁTICO 1",
    "[1.2] PROCESO CONGELADO TÚNEL ESTÁTICO 2",
    "[1.3] PROCESO CONGELADO TÚNEL ESTÁTICO 3",
]

workcenters = search_read('mrp.workcenter', [('name', 'in', tunnels_names)], ['id', 'name'])
print(f"Centros de trabajo encontrados: {len(workcenters)}")
for wc in workcenters:
    print(f"  - {wc['name']} (ID: {wc['id']})")

if not workcenters:
    # Intentar búsqueda parcial
    print("\nNo se encontraron con nombre exacto. Buscando con 'TÚNEL ESTÁTICO'...")
    workcenters = search_read('mrp.workcenter', [('name', 'ilike', 'TÚNEL ESTÁTICO')], ['id', 'name'])
    print(f"Encontrados: {len(workcenters)}")
    for wc in workcenters:
        print(f"  - {wc['name']} (ID: {wc['id']})")
    
    if not workcenters:
        print("\nBuscando con 'TUNEL ESTATICO' (sin tildes)...")
        workcenters = search_read('mrp.workcenter', [('name', 'ilike', 'TUNEL ESTATICO')], ['id', 'name'])
        print(f"Encontrados: {len(workcenters)}")
        for wc in workcenters:
            print(f"  - {wc['name']} (ID: {wc['id']})")

if not workcenters:
    print("\nListando todos los centros de trabajo para diagnosticar...")
    all_wc = search_read('mrp.workcenter', [], ['id', 'name'])
    for wc in all_wc:
        print(f"  - {wc['name']} (ID: {wc['id']})")
    raise Exception("No se encontraron los centros de trabajo de los túneles")

wc_ids = [wc['id'] for wc in workcenters]
wc_names = {wc['id']: wc['name'] for wc in workcenters}

# Buscar órdenes de producción asociadas a estos centros de trabajo
print("\nBuscando órdenes de producción...")
# Las work orders (mrp.workorder) tienen el workcenter_id
workorders = search_read('mrp.workorder', [
    ('workcenter_id', 'in', wc_ids),
    ('state', '=', 'done'),
], ['id', 'production_id', 'workcenter_id', 'qty_produced'])

print(f"Work orders encontradas: {len(workorders)}")

if not workorders:
    # Intentar con todos los estados
    print("Buscando en todos los estados...")
    workorders = search_read('mrp.workorder', [
        ('workcenter_id', 'in', wc_ids),
    ], ['id', 'production_id', 'workcenter_id', 'qty_produced', 'state'])
    print(f"Work orders encontradas (todos estados): {len(workorders)}")
    if workorders:
        states = set(wo['state'] for wo in workorders)
        print(f"  Estados: {states}")

production_ids = list(set(wo['production_id'][0] for wo in workorders if wo['production_id']))
print(f"Órdenes de producción únicas: {len(production_ids)}")

# Obtener info de las órdenes de producción
print("\nObteniendo detalle de órdenes de producción...")
kg_por_fruta_tunel = defaultdict(lambda: defaultdict(float))
kg_por_fruta_total = defaultdict(float)

batch_size = 100
for i in range(0, len(production_ids), batch_size):
    batch = production_ids[i:i+batch_size]
    productions = search_read('mrp.production', [('id', 'in', batch)], [
        'id', 'product_id', 'product_qty', 'qty_produced', 'state', 'product_uom_id'
    ])
    
    # Mapear production_id -> workcenter
    prod_wc = {}
    for wo in workorders:
        if wo['production_id']:
            pid = wo['production_id'][0]
            if pid in batch:
                prod_wc[pid] = wo['workcenter_id'][0]
    
    for prod in productions:
        pid = prod['id']
        product_name = prod['product_id'][1] if prod['product_id'] else 'Desconocido'
        qty = prod.get('qty_produced', 0) or prod.get('product_qty', 0)
        wc_id = prod_wc.get(pid)
        tunnel_name = wc_names.get(wc_id, 'Desconocido') if wc_id else 'Desconocido'
        
        kg_por_fruta_tunel[product_name][tunnel_name] += qty
        kg_por_fruta_total[product_name] += qty

    print(f"  Procesados {min(i+batch_size, len(production_ids))}/{len(production_ids)}")

# Crear Excel
print("\nGenerando Excel...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "KG por Tipo de Fruta"

# Estilos
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
total_font = Font(bold=True, size=11)
total_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Headers
tunnel_cols = sorted(set(wc_names.values()))
headers = ["Tipo de Fruta"] + tunnel_cols + ["TOTAL KG"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Datos
row = 2
for fruta in sorted(kg_por_fruta_total.keys(), key=lambda x: kg_por_fruta_total[x], reverse=True):
    ws.cell(row=row, column=1, value=fruta).border = thin_border
    total_row = 0
    for col_idx, tunnel in enumerate(tunnel_cols, 2):
        val = kg_por_fruta_tunel[fruta].get(tunnel, 0)
        cell = ws.cell(row=row, column=col_idx, value=round(val, 2))
        cell.number_format = '#,##0.00'
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right')
        total_row += val
    
    total_cell = ws.cell(row=row, column=len(headers), value=round(total_row, 2))
    total_cell.number_format = '#,##0.00'
    total_cell.font = total_font
    total_cell.border = thin_border
    total_cell.alignment = Alignment(horizontal='right')
    row += 1

# Fila de totales
ws.cell(row=row, column=1, value="TOTAL").font = total_font
ws.cell(row=row, column=1).fill = total_fill
ws.cell(row=row, column=1).border = thin_border
for col_idx, tunnel in enumerate(tunnel_cols, 2):
    total = sum(kg_por_fruta_tunel[f].get(tunnel, 0) for f in kg_por_fruta_total)
    cell = ws.cell(row=row, column=col_idx, value=round(total, 2))
    cell.number_format = '#,##0.00'
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right')

grand_total = sum(kg_por_fruta_total.values())
gt_cell = ws.cell(row=row, column=len(headers), value=round(grand_total, 2))
gt_cell.number_format = '#,##0.00'
gt_cell.font = Font(bold=True, size=12)
gt_cell.fill = total_fill
gt_cell.border = thin_border
gt_cell.alignment = Alignment(horizontal='right')

# Ajustar anchos
for col in range(1, len(headers) + 1):
    max_len = max(len(str(ws.cell(row=r, column=col).value or '')) for r in range(1, row + 1))
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(max_len + 4, 15)

output_file = "KG_Tuneles_Estaticos.xlsx"
wb.save(output_file)
print(f"\n¡Excel generado exitosamente: {output_file}")
print(f"Total de tipos de fruta: {len(kg_por_fruta_total)}")
print(f"Total general KG: {grand_total:,.2f}")
