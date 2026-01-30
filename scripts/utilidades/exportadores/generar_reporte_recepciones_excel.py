"""
Script para generar reporte Excel de recepciones con detalles de defectos.
Especialmente enfocado en mora y frambuesa como solicitó el usuario.
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from shared.odoo_client import OdooClient
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configurar credenciales
USERNAME = "mvalladares@riofuturo.cl"
PASSWORD = "c0766224bec30cac071ffe43a858c9ccbd521ddd"

# Rango de fechas a analizar (últimos 30 días por defecto)
FECHA_DESDE = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
FECHA_HASTA = datetime.now().strftime("%Y-%m-%d")

print("=" * 100)
print("GENERACIÓN DE REPORTE EXCEL DE RECEPCIONES")
print("=" * 100)
print(f"Período: {FECHA_DESDE} hasta {FECHA_HASTA}")
print("=" * 100)

# Conectar a Odoo
print("Conectando a Odoo...")
odoo = OdooClient(username=USERNAME, password=PASSWORD)
print("✅ Conectado exitosamente")

# ============ ANÁLISIS DE CAMPOS DISPONIBLES ============
print("\n" + "=" * 100)
print("ANÁLISIS DE CAMPOS DISPONIBLES EN LOS MODELOS")
print("=" * 100)

# 1. Analizar stock.picking
print("\n🔍 Analizando campos en stock.picking...")
campos_picking = odoo.execute('stock.picking', 'fields_get', [], {'attributes': ['string', 'type']})

print("Campos relacionados con 'guia' o 'despacho':")
campos_guia = {k: v for k, v in campos_picking.items() if 'guia' in k.lower() or 'despacho' in k.lower()}
for campo, info in campos_guia.items():
    print(f"  - {campo}: {info.get('string', 'N/A')}")

print("\nCampos relacionados con 'categoria':")
campos_categoria = {k: v for k, v in campos_picking.items() if 'categor' in k.lower()}
for campo, info in campos_categoria.items():
    print(f"  - {campo}: {info.get('string', 'N/A')}")

# 2. Analizar product.template
print("\n🔍 Analizando campos en product.template...")
campos_template = odoo.execute('product.template', 'fields_get', [], {'attributes': ['string', 'type']})

print("Campos x_studio disponibles:")
campos_x_studio = {k: v for k, v in campos_template.items() if k.startswith('x_studio')}
for campo, info in list(campos_x_studio.items())[:20]:  # Mostrar solo los primeros 20
    print(f"  - {campo}: {info.get('string', 'N/A')}")

# 3. Analizar quality.check
print("\n🔍 Analizando campos en quality.check...")
campos_quality = odoo.execute('quality.check', 'fields_get', [], {'attributes': ['string', 'type']})

print("Campos relacionados con defectos:")
campos_defectos = {k: v for k, v in campos_quality.items() if any(x in k.lower() for x in ['defect', 'hongos', 'inmadura', 'crumble'])}
for campo, info in campos_defectos.items():
    print(f"  - {campo}: {info.get('string', 'N/A')}")

print("\n" + "=" * 100)
print("DETERMINANDO CAMPOS A USAR")
print("=" * 100)

# Determinar qué campo usar para guía de despacho en stock.picking
campo_guia = None
for opcion in ['x_studio_gua_de_despacho', 'x_studio_gua_despacho', 'x_studio_guia_despacho', 'x_guia_despacho']:
    if opcion in campos_picking:
        campo_guia = opcion
        break
print(f"✅ Campo para guía de despacho: {campo_guia}")

# Determinar campo para categoría de producto en stock.picking
campo_categoria_prod = None
for opcion in ['x_studio_categora_de_producto', 'x_studio_categoria_de_producto', 'x_categoria_producto']:
    if opcion in campos_picking:
        campo_categoria_prod = opcion
        break
print(f"✅ Campo para categoría de producto: {campo_categoria_prod}")

# Determinar campos en product.template
campos_template_usar = {}
for campo_nombre, opciones in {
    'variedad': ['x_studio_categora_variedad', 'x_studio_variedad', 'x_variedad'],
    'calibre': ['x_studio_calibre', 'x_calibre'],
    'tipo_categoria': ['x_studio_tipos_de_categoria', 'x_tipos_categoria', 'x_studio_tipo_categoria'],
    'manejo': ['x_studio_categora_tipo_de_manejo', 'x_studio_manejo_del_producto', 'x_manejo_producto', 'x_studio_manejo']
}.items():
    for opcion in opciones:
        if opcion in campos_template:
            campos_template_usar[campo_nombre] = opcion
            break
    if campo_nombre not in campos_template_usar:
        campos_template_usar[campo_nombre] = None

print(f"✅ Campos product.template:")
for k, v in campos_template_usar.items():
    print(f"   - {k}: {v}")

# Determinar campos en quality.check
campos_quality_usar = {}
for campo_nombre, opciones in {
    'tipo_fruta_qc': ['x_studio_tipo_de_fruta'],  # Nuevo campo encontrado en quality.check
    'pallet': ['x_studio_n_de_palet_o_paquete', 'x_studio_n_palet', 'x_n_pallet', 'x_studio_n_pallet'],
    'clasificacion': ['x_studio_calific_final', 'x_studio_calificacin_final', 'x_studio_clasificacion', 'x_clasificacion'],
    'total_defectos': ['x_studio_total_def_calidad', 'x_studio_total_de_defectos_', 'x_total_defectos'],
    'temperatura': ['x_studio_temperatura', 'x_temperatura'],
    'hongos': ['x_studio_hongos', 'x_hongos'],
    'inmadura': ['x_studio_inmadura', 'x_inmadura'],
    'sobremadura': ['x_studio_sobremadura', 'x_studio_sobre_madura', 'x_sobremadura'],
    'deshidratado': ['x_studio_deshidratado', 'x_deshidratado'],
    'crumble': ['x_studio_crumble', 'x_crumble'],
    'dano_mecanico': ['x_studio_dao_mecanico', 'x_studio_dano_mecanico', 'x_dano_mecanico'],
    'dano_insecto': ['x_studio_presencia_de_insectos', 'x_studio_totdaoinsecto', 'x_studio_dano_de_insecto', 'x_dano_insecto'],
    'deformes': ['x_studio_frutos_deformes', 'x_studio_deformes', 'x_deformes'],
    'fruta_verde': ['x_studio_fruta_verde', 'x_fruta_verde'],
    'herida_partida': ['x_studio_heridapartidamolida', 'x_studio_heridapartiduramolida', 'x_studio_heridapartida', 'x_herida_partida']
}.items():
    for opcion in opciones:
        if opcion in campos_quality:
            campos_quality_usar[campo_nombre] = opcion
            break
    if campo_nombre not in campos_quality_usar:
        campos_quality_usar[campo_nombre] = None

print(f"✅ Campos quality.check:")
for k, v in campos_quality_usar.items():
    print(f"   - {k}: {v}")

print("=" * 100)

# IDs de picking types por origen (según debug_recepciones_kg.py)
ORIGEN_PICKING_MAP = {
    "RFP": 1,
    "VILKUN": 217,
    "SAN JOSE": 164
}

# 1. Obtener recepciones
print("\n📦 Obteniendo recepciones...")
todas_recepciones = []

# Construir dominio dinámicamente
for origen, picking_type_id in ORIGEN_PICKING_MAP.items():
    domain = [
        ('picking_type_id', '=', picking_type_id),
        ('scheduled_date', '>=', FECHA_DESDE),
        ('scheduled_date', '<=', FECHA_HASTA),
        ('state', '=', 'done')
    ]
    
    # Agregar filtro de categoría si el campo existe
    if campo_categoria_prod:
        domain.append((campo_categoria_prod, '=', 'MP'))
    
    # Construir lista de campos a leer
    campos_leer = ['id', 'name', 'scheduled_date', 'partner_id']
    if campo_guia:
        campos_leer.append(campo_guia)
    
    recepciones = odoo.search_read(
        'stock.picking',
        domain,
        campos_leer,
        limit=5000
    )
    
    print(f"  {origen}: {len(recepciones)} recepciones")
    
    for rec in recepciones:
        rec['origen'] = origen
    
    todas_recepciones.extend(recepciones)

print(f"\n✅ Total de recepciones: {len(todas_recepciones)}")

if not todas_recepciones:
    print("❌ No se encontraron recepciones en el período especificado")
    sys.exit(0)

# 2. Obtener movimientos de stock (productos recepcionados)
print("\n📦 Obteniendo movimientos de stock...")
picking_ids = [r['id'] for r in todas_recepciones]

movimientos = odoo.search_read(
    'stock.move',
    [
        ('picking_id', 'in', picking_ids),
        ('state', '=', 'done')
    ],
    ['id', 'product_id', 'quantity_done', 'picking_id'],
    limit=50000
)

print(f"  Total movimientos: {len(movimientos)}")

# 3. Obtener información de productos
print("\n🍓 Obteniendo información de productos...")
product_ids = list(set([m['product_id'][0] for m in movimientos if m.get('product_id')]))

productos = odoo.search_read(
    'product.product',
    [('id', 'in', product_ids)],
    ['id', 'name', 'product_tmpl_id', 'categ_id'],
    limit=10000
)

# Crear mapeo de producto_id -> info
productos_map = {p['id']: p for p in productos}

# Obtener templates para más información
template_ids = list(set([p['product_tmpl_id'][0] for p in productos if p.get('product_tmpl_id')]))

# Construir lista de campos a leer dinámicamente
template_campos_leer = ['id']
for campo in campos_template_usar.values():
    if campo:
        template_campos_leer.append(campo)

templates = odoo.search_read(
    'product.template',
    [('id', 'in', template_ids)],
    template_campos_leer,
    limit=10000
)

templates_map = {t['id']: t for t in templates}

print(f"  Total productos: {len(productos)}")
print(f"  Total templates: {len(templates)}")

# 4. Obtener líneas de análisis de calidad (quality.check)
print("\n🔬 Obteniendo análisis de calidad...")

# Construir lista de campos a leer dinámicamente
quality_campos_leer = ['id', 'picking_id', 'create_date']
for campo in campos_quality_usar.values():
    if campo:
        quality_campos_leer.append(campo)

# Buscar quality.check relacionados con estos pickings
quality_checks = odoo.search_read(
    'quality.check',
    [('picking_id', 'in', picking_ids)],
    quality_campos_leer,
    limit=50000
)

print(f"  Total quality checks: {len(quality_checks)}")

# Organizar quality checks por picking_id
quality_by_picking = {}
for qc in quality_checks:
    picking_id = qc.get('picking_id')
    if picking_id:
        picking_id = picking_id[0] if isinstance(picking_id, (list, tuple)) else picking_id
        if picking_id not in quality_by_picking:
            quality_by_picking[picking_id] = []
        quality_by_picking[picking_id].append(qc)

# 5. Construir datos para el Excel
print("\n📊 Construyendo datos para Excel...")
datos_excel = []

for recepcion in todas_recepciones:
    picking_id = recepcion['id']
    albaran = recepcion.get('name', '')
    fecha = recepcion.get('scheduled_date', '')
    if fecha:
        fecha = datetime.strptime(fecha, '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y')
    
    proveedor = recepcion.get('partner_id')
    proveedor = proveedor[1] if isinstance(proveedor, (list, tuple)) else str(proveedor) if proveedor else ''
    
    guia_despacho = recepcion.get(campo_guia, '') if campo_guia else ''
    origen = recepcion.get('origen', '')
    
    # Obtener movimientos para esta recepción
    movs_recepcion = [m for m in movimientos if m.get('picking_id') and m['picking_id'][0] == picking_id]
    
    # Obtener quality checks para esta recepción
    qcs_recepcion = quality_by_picking.get(picking_id, [])
    
    # Para cada pallet en los quality checks
    for qc in qcs_recepcion:
        # Usar campos dinámicos
        n_pallet = qc.get(campos_quality_usar.get('pallet'), '') if campos_quality_usar.get('pallet') else ''
        calificacion = qc.get(campos_quality_usar.get('clasificacion'), '') if campos_quality_usar.get('clasificacion') else ''
        temperatura = qc.get(campos_quality_usar.get('temperatura'), 0) or 0 if campos_quality_usar.get('temperatura') else 0
        total_defectos_pct = qc.get(campos_quality_usar.get('total_defectos'), 0) or 0 if campos_quality_usar.get('total_defectos') else 0
        
        # IMPORTANTE: Obtener tipo de fruta desde quality.check si está disponible
        tipo_fruta_from_qc = qc.get(campos_quality_usar.get('tipo_fruta_qc'), '') if campos_quality_usar.get('tipo_fruta_qc') else ''
        
        # Defectos en gramos (usar campos dinámicos)
        hongos = qc.get(campos_quality_usar.get('hongos'), 0) or 0 if campos_quality_usar.get('hongos') else 0
        inmadura = qc.get(campos_quality_usar.get('inmadura'), 0) or 0 if campos_quality_usar.get('inmadura') else 0
        sobremadura = qc.get(campos_quality_usar.get('sobremadura'), 0) or 0 if campos_quality_usar.get('sobremadura') else 0
        deshidratado = qc.get(campos_quality_usar.get('deshidratado'), 0) or 0 if campos_quality_usar.get('deshidratado') else 0
        crumble = qc.get(campos_quality_usar.get('crumble'), 0) or 0 if campos_quality_usar.get('crumble') else 0
        dano_mecanico = qc.get(campos_quality_usar.get('dano_mecanico'), 0) or 0 if campos_quality_usar.get('dano_mecanico') else 0
        dano_insecto = qc.get(campos_quality_usar.get('dano_insecto'), 0) or 0 if campos_quality_usar.get('dano_insecto') else 0
        deformes = qc.get(campos_quality_usar.get('deformes'), 0) or 0 if campos_quality_usar.get('deformes') else 0
        fruta_verde = qc.get(campos_quality_usar.get('fruta_verde'), 0) or 0 if campos_quality_usar.get('fruta_verde') else 0
        herida_partida = qc.get(campos_quality_usar.get('herida_partida'), 0) or 0 if campos_quality_usar.get('herida_partida') else 0
        
        # Para cada producto en los movimientos
        for mov in movs_recepcion:
            qty = mov.get('quantity_done', 0) or 0
            if qty == 0:
                continue
                
            prod_id = mov.get('product_id')
            if not prod_id:
                continue
            prod_id = prod_id[0] if isinstance(prod_id, (list, tuple)) else prod_id
            
            prod_info = productos_map.get(prod_id, {})
            producto_nombre = prod_info.get('name', '')
            
            # Info del template
            tmpl_id = prod_info.get('product_tmpl_id')
            tmpl_id = tmpl_id[0] if isinstance(tmpl_id, (list, tuple)) else tmpl_id
            tmpl_info = templates_map.get(tmpl_id, {})
            
            # Usar campos dinámicos para extraer info del template
            campo_variedad = campos_template_usar.get('variedad')
            variedad = tmpl_info.get(campo_variedad, '') if campo_variedad else ''
            if isinstance(variedad, (list, tuple)) and len(variedad) > 1:
                variedad = variedad[1]
            elif variedad:
                variedad = str(variedad)
            else:
                variedad = ''
            
            campo_calibre = campos_template_usar.get('calibre')
            calibre = tmpl_info.get(campo_calibre, '') if campo_calibre else ''
            if isinstance(calibre, (list, tuple)) and len(calibre) > 1:
                calibre = calibre[1]
            elif calibre:
                calibre = str(calibre)
            else:
                calibre = ''
            
            # Usar tipo de fruta desde quality.check si está disponible, sino desde template
            # Esto es crítico para identificar mora y frambuesa
            campo_tipo = campos_template_usar.get('tipo_categoria')
            if tipo_fruta_from_qc:
                tipo_fruta = tipo_fruta_from_qc
            elif campo_tipo:
                tipo_fruta_tmpl = tmpl_info.get(campo_tipo, '') if campo_tipo else ''
                if isinstance(tipo_fruta_tmpl, (list, tuple)) and len(tipo_fruta_tmpl) > 1:
                    tipo_fruta = tipo_fruta_tmpl[1]
                elif tipo_fruta_tmpl:
                    tipo_fruta = str(tipo_fruta_tmpl)
                else:
                    tipo_fruta = ''
            else:
                tipo_fruta = ''
            
            campo_manejo = campos_template_usar.get('manejo')
            manejo = tmpl_info.get(campo_manejo, '') if campo_manejo else ''
            if isinstance(manejo, (list, tuple)) and len(manejo) > 1:
                manejo = manejo[1]
            elif manejo:
                manejo = str(manejo)
            else:
                manejo = ''
            
            # Determinar si es Mora o Frambuesa (especialmente importante)
            tipo_fruta_lower = tipo_fruta.lower()
            es_mora = 'mora' in tipo_fruta_lower
            es_frambuesa = 'frambuesa' in tipo_fruta_lower
            
            datos_excel.append({
                'Fecha': fecha,
                'Proveedor': proveedor,
                'Guía Despacho': guia_despacho,
                'Origen': origen,
                'Albarán': albaran,
                'Producto': producto_nombre,
                'Variedad': variedad,
                'Calibre': calibre,
                'Tipo Fruta': tipo_fruta,
                'Manejo': manejo,
                'Kg': qty,
                'N° Pallet': n_pallet,
                'Calificación': calificacion,
                'Temperatura °C': temperatura,
                '% Defectos': total_defectos_pct,
                'Hongos (g)': hongos,
                'Inmadura (g)': inmadura,
                'Sobremadura (g)': sobremadura,
                'Deshidratado (g)': deshidratado,
                'Crumble (g)': crumble,
                'Daño Mecánico (g)': dano_mecanico,
                'Daño Insecto (g)': dano_insecto,
                'Deformes (g)': deformes,
                'Fruta Verde (g)': fruta_verde,
                'Herida/Partida (g)': herida_partida,
            })

print(f"  Total filas para Excel: {len(datos_excel)}")

# 6. Crear Excel con formato
print("\n📄 Creando archivo Excel...")

df = pd.DataFrame(datos_excel)

# Generar nombre de archivo con timestamp
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
output_file = f"reporte_recepciones_{timestamp}.xlsx"

# Usar openpyxl para mejor formato
wb = Workbook()
ws = wb.active
ws.title = "Recepciones"

# Estilo de encabezado
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Bordes
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Agregar encabezados
headers = df.columns.tolist()
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Agregar datos
for row_num, row_data in enumerate(df.values, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = thin_border
        
        # Alineación especial para números
        if col_num > 10:  # Columnas numéricas
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.alignment = Alignment(horizontal="left")

# Ajustar ancho de columnas
for col_num, header in enumerate(headers, 1):
    column_letter = ws.cell(row=1, column=col_num).column_letter
    
    # Anchos específicos por tipo de columna
    if 'Fecha' in header:
        ws.column_dimensions[column_letter].width = 12
    elif 'Proveedor' in header or 'Producto' in header:
        ws.column_dimensions[column_letter].width = 30
    elif '(g)' in header or '°C' in header or '%' in header:
        ws.column_dimensions[column_letter].width = 12
    else:
        ws.column_dimensions[column_letter].width = 15

# Congelar primera fila
ws.freeze_panes = "A2"

# Guardar
wb.save(output_file)

print(f"\n✅ REPORTE GENERADO EXITOSAMENTE")
print("=" * 100)
print(f"Archivo: {output_file}")
print(f"Total registros: {len(datos_excel)}")
print(f"Columnas: {len(headers)}")
print("=" * 100)

# Mostrar resumen por tipo de fruta
print("\n📊 RESUMEN POR TIPO DE FRUTA:")
print("-" * 100)
if not df.empty:
    resumen = df.groupby('Tipo Fruta').agg({
        'Kg': 'sum',
        'Albarán': 'count'
    }).reset_index()
    resumen.columns = ['Tipo Fruta', 'Kg Total', 'N° Recepciones']
    print(resumen.to_string(index=False))
    
    # Enfoque especial en Mora y Frambuesa
    print("\n🍓 ENFOQUE EN MORA Y FRAMBUESA:")
    print("-" * 100)
    mora_frambuesa = df[df['Tipo Fruta'].str.lower().str.contains('mora|frambuesa', na=False)]
    if not mora_frambuesa.empty:
        print(f"Total registros Mora/Frambuesa: {len(mora_frambuesa)}")
        print(f"Total Kg: {mora_frambuesa['Kg'].sum():,.2f}")
        print(f"Defectos promedio %: {mora_frambuesa['% Defectos'].mean():.2f}%")
    else:
        print("No se encontraron registros de Mora o Frambuesa")

print("\n" + "=" * 100)
