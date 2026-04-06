import xmlrpc.client
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from collections import defaultdict

# Conexión a Odoo
url = "https://riofuturo.server98c6e.oerpondemand.net"
db = "riofuturo-master"
username = "mvalladares@riofuturo.cl"
password = "c0766224bec30cac071ffe43a858c9ccbd521ddd"

print("Conectando a Odoo...")
common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common", context=None)
uid = common.authenticate(db, username, password, {})
models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object", context=None)

def search_read(model, domain, fields, limit=0):
    return models.execute_kw(db, uid, password, model, 'search_read', [domain], {'fields': fields, 'limit': limit})

# 1. Buscar los 3 productos de túneles estáticos (sin VLK)
print("Buscando productos de túneles estáticos 1, 2 y 3...")
tunnel_products = search_read('product.product', [
    ('name', 'in', [
        'PROCESO CONGELADO TÚNEL ESTÁTICO 1',
        'PROCESO CONGELADO TÚNEL ESTÁTICO 2',
        'PROCESO CONGELADO TÚNEL ESTÁTICO 3',
    ])
], ['id', 'name'])

if not tunnel_products:
    tunnel_products = search_read('product.product', [
        ('name', 'ilike', 'PROCESO CONGELADO TÚNEL ESTÁTICO'),
        ('name', 'not ilike', 'VLK'),
    ], ['id', 'name'])

# Filtrar solo los 3 túneles estáticos (sin VLK)
tunnel_products = [p for p in tunnel_products if 'VLK' not in p['name'].upper()]
print(f"Túneles encontrados: {len(tunnel_products)}")
for p in tunnel_products:
    print(f"  - {p['name']} (ID: {p['id']})")

product_ids = [p['id'] for p in tunnel_products]
product_names = {p['id']: p['name'] for p in tunnel_products}

# Nombres cortos para los túneles
short_names = {}
for p in tunnel_products:
    name = p['name']
    if 'ESTÁTICO 1' in name or 'ESTATICO 1' in name:
        short_names[p['id']] = '[1.1] Túnel Estático 1'
    elif 'ESTÁTICO 2' in name or 'ESTATICO 2' in name:
        short_names[p['id']] = '[1.2] Túnel Estático 2'
    elif 'ESTÁTICO 3' in name or 'ESTATICO 3' in name:
        short_names[p['id']] = '[1.3] Túnel Estático 3'
    else:
        short_names[p['id']] = name

# 2. Buscar producciones terminadas
print("\nBuscando producciones terminadas...")
productions = search_read('mrp.production', [
    ('product_id', 'in', product_ids),
    ('state', '=', 'done'),
], ['id', 'name', 'product_id', 'move_raw_ids', 'qty_produced', 'date_finished'])
print(f"Producciones encontradas: {len(productions)}")

# 3. Mapear move_ids a túnel
all_raw_ids = []
move_to_tunnel = {}
for prod in productions:
    tunnel_id = prod['product_id'][0]
    tunnel_short = short_names.get(tunnel_id, prod['product_id'][1])
    for mid in prod['move_raw_ids']:
        all_raw_ids.append(mid)
        move_to_tunnel[mid] = tunnel_short

# 4. Leer movimientos de materia prima
print(f"\nLeyendo {len(all_raw_ids)} movimientos de materia prima...")

# Items a excluir (no son fruta)
EXCLUDE_PATTERNS = ['ELECTRICIDAD', 'PROVISIÓN', 'PROVISION', 'CAJA EXPORT', 'PROCESO CONGELADO']

kg_por_fruta_tunel = defaultdict(lambda: defaultdict(float))
kg_por_fruta_total = defaultdict(float)

batch_size = 200
for i in range(0, len(all_raw_ids), batch_size):
    batch = all_raw_ids[i:i+batch_size]
    moves = search_read('stock.move', [('id', 'in', batch)], [
        'id', 'product_id', 'product_uom_qty', 'quantity_done', 'product_uom'
    ])
    
    for move in moves:
        product_name = move['product_id'][1] if move['product_id'] else 'Desconocido'
        
        # Excluir items que no son fruta
        if any(pat in product_name.upper() for pat in EXCLUDE_PATTERNS):
            continue
        
        qty = move.get('quantity_done', 0) or move.get('product_uom_qty', 0)
        if qty <= 0:
            continue
            
        tunnel = move_to_tunnel.get(move['id'], 'Desconocido')
        
        kg_por_fruta_tunel[product_name][tunnel] += qty
        kg_por_fruta_total[product_name] += qty
    
    print(f"  Procesados {min(i+batch_size, len(all_raw_ids))}/{len(all_raw_ids)}")

# 5. Crear Excel
print("\nGenerando Excel...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "KG por Tipo de Fruta"

# Estilos
header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
data_font = Font(size=10, name="Calibri")
total_font = Font(bold=True, size=11, name="Calibri")
total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
grand_total_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
grand_total_font = Font(bold=True, size=12, name="Calibri", color="FFFFFF")
alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)

# Título
ws.merge_cells('A1:E1')
title_cell = ws.cell(row=1, column=1, value="KG Totales por Tipo de Fruta - Túneles Estáticos 1, 2 y 3")
title_cell.font = Font(bold=True, size=14, name="Calibri", color="1F4E79")
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

# Subtítulo
ws.merge_cells('A2:E2')
subtitle = ws.cell(row=2, column=1, value=f"Total de producciones procesadas: {len(productions)}")
subtitle.font = Font(size=10, italic=True, name="Calibri", color="666666")
subtitle.alignment = Alignment(horizontal='center')
ws.row_dimensions[2].height = 20

# Headers - Row 3
tunnel_cols = sorted(set(
    t for frutas in kg_por_fruta_tunel.values() for t in frutas.keys()
))
headers = ["Tipo de Fruta"] + tunnel_cols + ["TOTAL KG"]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws.row_dimensions[3].height = 35

# Datos - desde Row 4
row = 4
sorted_frutas = sorted(kg_por_fruta_total.keys(), key=lambda x: kg_por_fruta_total[x], reverse=True)

for idx, fruta in enumerate(sorted_frutas):
    fill = alt_fill if idx % 2 == 0 else PatternFill(fill_type=None)
    
    cell = ws.cell(row=row, column=1, value=fruta)
    cell.font = data_font
    cell.border = thin_border
    cell.fill = fill
    
    total_row = 0
    for col_idx, tunnel in enumerate(tunnel_cols, 2):
        val = kg_por_fruta_tunel[fruta].get(tunnel, 0)
        cell = ws.cell(row=row, column=col_idx, value=round(val, 2))
        cell.number_format = '#,##0.00'
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right')
        cell.fill = fill
        total_row += val
    
    total_cell = ws.cell(row=row, column=len(headers), value=round(total_row, 2))
    total_cell.number_format = '#,##0.00'
    total_cell.font = Font(bold=True, size=10, name="Calibri")
    total_cell.border = thin_border
    total_cell.alignment = Alignment(horizontal='right')
    total_cell.fill = fill
    row += 1

# Fila de totales
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=row, column=col)
    cell.fill = total_fill
    cell.font = total_font
    cell.border = thin_border

ws.cell(row=row, column=1, value="TOTAL GENERAL")

for col_idx, tunnel in enumerate(tunnel_cols, 2):
    total = sum(kg_por_fruta_tunel[f].get(tunnel, 0) for f in kg_por_fruta_total)
    cell = ws.cell(row=row, column=col_idx, value=round(total, 2))
    cell.number_format = '#,##0.00'
    cell.alignment = Alignment(horizontal='right')

grand_total = sum(kg_por_fruta_total.values())
gt_cell = ws.cell(row=row, column=len(headers), value=round(grand_total, 2))
gt_cell.number_format = '#,##0.00'
gt_cell.font = grand_total_font
gt_cell.fill = grand_total_fill
gt_cell.alignment = Alignment(horizontal='right')

# Ajustar anchos de columna
ws.column_dimensions['A'].width = 45
for col in range(2, len(headers) + 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

# Freeze panes
ws.freeze_panes = 'B4'

# Auto filter
ws.auto_filter.ref = f"A3:{openpyxl.utils.get_column_letter(len(headers))}{row}"

output_file = "KG_Tuneles_Estaticos.xlsx"
wb.save(output_file)

print(f"\n{'='*60}")
print(f"  Excel generado: {output_file}")
print(f"  Tipos de fruta: {len(kg_por_fruta_total)}")
print(f"  Producciones:   {len(productions)}")
print(f"  Total KG:       {grand_total:,.2f}")
print(f"{'='*60}")

# Resumen por túnel
print("\nResumen por túnel:")
for tunnel in tunnel_cols:
    total_tunnel = sum(kg_por_fruta_tunel[f].get(tunnel, 0) for f in kg_por_fruta_total)
    print(f"  {tunnel}: {total_tunnel:,.2f} kg")
