import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data = {
    31322: [(289, "Comercial Zeniz Organico Limitada")],
    30306: [(3262, "AGRICOLA COX LTDA"), (3978, "AGRÍCOLA SANTA ELENA LIMITADA"), (3261, "AGRICOLA COX LTDA"), (3257, "AGRICOLA COX LTDA"), (140, "CARLOS ALBERTO KLEIN KOCH"), (4903, "AGRICOLA Y FORESTAL SAN ALEJANDRO LTDA"), (4904, "AGRICOLA Y FORESTAL SAN ALEJANDRO LTDA")],
    34968: [(556, "AGRÍCOLA TRES ROBLES"), (4914, "AGRICOLA Y FORESTAL SAN ALEJANDRO LTDA"), (196, "AGRÍCOLA ITAMALAL SPA")],
    31448: [(282, "Comercial Zeniz Organico Limitada"), (276, "Comercial Zeniz Organico Limitada"), (289, "Comercial Zeniz Organico Limitada")],
    33469: [(3977, "AGRÍCOLA SANTA ELENA LIMITADA"), (3978, "AGRÍCOLA SANTA ELENA LIMITADA"), (4903, "AGRICOLA Y FORESTAL SAN ALEJANDRO LTDA")],
    30733: [(3258, "AGRICOLA COX LTDA"), (3257, "AGRICOLA COX LTDA"), (3254, "AGRICOLA COX LTDA")],
    34415: [(309, "Comercial Zeniz Organico Limitada"), (314, "Comercial Zeniz Organico Limitada")],
    11907: [(700, "AGRICOLA LA CORTINA DE COLBUN LIMITADA")],
    21507: [(3952, "AGRÍCOLA SANTA ELENA LIMITADA"), (3948, "AGRÍCOLA SANTA ELENA LIMITADA"), (3946, "AGRÍCOLA SANTA ELENA LIMITADA")],
    22248: [(4887, "AGRICOLA Y FORESTAL SAN ALEJANDRO LTDA"), (4673, "SOCIEDAD AGRICOLA DEL HUERTO LTDA"), (3952, "AGRÍCOLA SANTA ELENA LIMITADA"), (3188, "AGRICOLA COX LTDA")],
    24438: [(3235, "AGRICOLA COX LTDA"), (3186, "AGRICOLA COX LTDA"), (4895, "AGRICOLA Y FORESTAL SAN ALEJANDRO LTDA"), (88, "BLUE FOX LIMITADA")],
    18133: [(4882, "AGRICOLA Y FORESTAL SAN ALEJANDRO LTDA"), (2271, "AGRICOLA ASTURIAS LIMITADA"), (3175, "AGRICOLA COX LTDA")],
    25912: [(106, "BLUE FOX LIMITADA"), (107, "BLUE FOX LIMITADA")],
    25532: [(84, "BLUE FOX LIMITADA"), (86, "BLUE FOX LIMITADA"), (708, "AGRICOLA LA CORTINA DE COLBUN LIMITADA")],
    30933: [(21110, "VICTOR OSVALDO FROHLICH MOHR"), (4649, "SOC AGRICOLA RIO CHEPU LTDA"), (4648, "SOC AGRICOLA RIO CHEPU LTDA"), (139, "CARLOS ALBERTO KLEIN KOCH")],
    33112: [(3266, "AGRICOLA COX LTDA"), (4910, "AGRICOLA Y FORESTAL SAN ALEJANDRO LTDA"), (3265, "AGRICOLA COX LTDA")],
    23126: [(3189, "AGRICOLA COX LTDA"), (164, "JACQUELINE MOURGUET BESOAIN"), (3192, "AGRICOLA COX LTDA")],
    31834: [(4909, "AGRICOLA Y FORESTAL SAN ALEJANDRO LTDA"), (3265, "AGRICOLA COX LTDA")],
    30074: [(3973, "AGRÍCOLA SANTA ELENA LIMITADA"), (3974, "AGRÍCOLA SANTA ELENA LIMITADA"), (276, "Comercial Zeniz Organico Limitada")],
    36134: [(2133, "VICTOR OSVALDO FROHLICH MOHR"), (3268, "AGRICOLA COX LTDA"), (736, "AGRICOLA LA CORTINA DE COLBUN LIMITADA")],
    36044: [(2133, "VICTOR OSVALDO FROHLICH MOHR"), (3268, "AGRICOLA COX LTDA"), (736, "AGRICOLA LA CORTINA DE COLBUN LIMITADA")],
    24887: [(132, "CARLOS ALBERTO KLEIN KOCH"), (3231, "AGRICOLA COX LTDA"), (4892, "AGRICOLA Y FORESTAL SAN ALEJANDRO LTDA")],
    24231: [(3232, "AGRICOLA COX LTDA"), (3228, "AGRICOLA COX LTDA"), (3195, "AGRICOLA COX LTDA")],
    33185: [(4912, "AGRICOLA Y FORESTAL SAN ALEJANDRO LTDA")],
    36135: [(4912, "AGRICOLA Y FORESTAL SAN ALEJANDRO LTDA")],
    22636: [(3198, "AGRICOLA COX LTDA"), (89, "BLUE FOX LIMITADA"), (85, "BLUE FOX LIMITADA")],
    36148: [(559, "AGRÍCOLA TRES ROBLES"), (3992, "AGRÍCOLA SANTA ELENA LIMITADA")],
    36147: [(559, "AGRÍCOLA TRES ROBLES"), (3992, "AGRÍCOLA SANTA ELENA LIMITADA")],
    28113: [(720, "AGRICOLA LA CORTINA DE COLBUN LIMITADA"), (102, "BLUE FOX LIMITADA")],
    26939: [(246, "AGRICOLA Y FORESTAL EL PORVENIR SPA"), (238, "Comercial Zeniz Organico Limitada")],
    28318: [(3252, "AGRICOLA COX LTDA"), (3967, "AGRÍCOLA SANTA ELENA LIMITADA"), (3965, "AGRÍCOLA SANTA ELENA LIMITADA")],
    24672: [(132, "CARLOS ALBERTO KLEIN KOCH"), (3231, "AGRICOLA COX LTDA"), (539, "AGRÍCOLA TRES ROBLES")],
    25995: [(156, "AGRÍCOLA ITAMALAL SPA"), (2101, "VICTOR OSVALDO FROHLICH MOHR"), (156, "AGRÍCOLA ITAMALAL SPA")],
    21819: [(2083, "VICTOR OSVALDO FROHLICH MOHR"), (4889, "AGRICOLA Y FORESTAL SAN ALEJANDRO LTDA"), (3955, "AGRÍCOLA SANTA ELENA LIMITADA")],
    25360: [(3959, "AGRÍCOLA SANTA ELENA LIMITADA"), (3232, "AGRICOLA COX LTDA")],
    30787: [(3258, "AGRICOLA COX LTDA"), (3256, "AGRICOLA COX LTDA"), (264, "Comercial Zeniz Organico Limitada")],
    49147: [(2073, "VICTOR OSVALDO FROHLICH MOHR"), (3918, "AGRÍCOLA SANTA ELENA LIMITADA")],
    18401: [(3936, "AGRÍCOLA SANTA ELENA LIMITADA"), (4878, "AGRICOLA Y FORESTAL SAN ALEJANDRO LTDA")],
    19052: [(3936, "AGRÍCOLA SANTA ELENA LIMITADA"), (4878, "AGRICOLA Y FORESTAL SAN ALEJANDRO LTDA")],
    30513: [(3975, "AGRÍCOLA SANTA ELENA LIMITADA"), (3254, "AGRICOLA COX LTDA")],
    30120: [(732, "AGRICOLA LA CORTINA DE COLBUN LIMITADA"), (3252, "AGRICOLA COX LTDA"), (171, "AGRÍCOLA ITAMALAL SPA")],
    35001: [(556, "AGRÍCOLA TRES ROBLES"), (4914, "AGRICOLA Y FORESTAL SAN ALEJANDRO LTDA")],
    35188: [(314, "Comercial Zeniz Organico Limitada"), (286, "Comercial Zeniz Organico Limitada")],
    35621: [(304, "Comercial Zeniz Organico Limitada"), (690, "AGRICOLA LOVENGREEN REUS LIMITADA"), (264, "Comercial Zeniz Organico Limitada")],
    25763: [(106, "BLUE FOX LIMITADA"), (720, "AGRICOLA LA CORTINA DE COLBUN LIMITADA"), (97, "BLUE FOX LIMITADA")],
    27775: [(3249, "AGRICOLA COX LTDA"), (3247, "AGRICOLA COX LTDA"), (3246, "AGRICOLA COX LTDA")],
}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Trazabilidad Packs"

# Estilos
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
pack_font = Font(bold=True, size=11, color="1e3a5f")
pack_fill = PatternFill(start_color="dce6f1", end_color="dce6f1", fill_type="solid")
thin_border = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)

# Headers
headers = ["Pallet (PACK)", "Guía Despacho", "Productor"]
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

row = 2
for pack, guias in data.items():
    start_row = row
    for guia_num, productor in guias:
        ws.cell(row=row, column=1, value=pack).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=1).font = pack_font
        ws.cell(row=row, column=2, value=guia_num).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=productor).border = thin_border
        row += 1
    # Merge pack column if multiple rows
    if row - start_row > 1:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

# Auto-width
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 50

# Freeze header
ws.freeze_panes = "A2"

output = r"c:\Users\HP\Desktop\dash\proyectos\Trazabilidad_Packs.xlsx"
wb.save(output)
print(f"Excel generado: {output}")
print(f"Total packs: {len(data)}")
print(f"Total guías: {sum(len(v) for v in data.values())}")
