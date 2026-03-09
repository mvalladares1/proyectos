"""
Reporte Diario de Producción por Sala y Turno.
Genera/actualiza un Excel con la siguiente info por Sala/Turno:
  - Nombre Cliente
  - # Contenedor (PO asociada)
  - Dotación del proceso
  - Kilos turno (kg PT producidos)
  - Kilos / Horas Hombre
  - Kilos / Horas Efectivas de Proceso
  - Detenciones (total de tiempo)
  - Kilos aprobados para el cliente

Uso:
  python scripts/reporte_diario_produccion.py                    # Hoy
  python scripts/reporte_diario_produccion.py 2026-03-09         # Fecha específica
  python scripts/reporte_diario_produccion.py 2026-03-01 2026-03-09  # Rango
"""
import sys
import os
from pathlib import Path
from datetime import datetime, date
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Setup path para imports del proyecto
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from shared.odoo_client import OdooClient


# ==================================================
# FUNCIONES AUXILIARES
# ==================================================

def clasificar_turno(dt, dt_fin=None):
    """Clasifica en turno Día o Tarde basado en el punto medio del proceso.
    Día: L-J 8:00-17:30, V 8:00-16:30, S 8:00-13:00
    Tarde: el resto
    """
    if dt is None:
        return "Día"
    if dt_fin is not None and dt_fin > dt:
        medio = dt + (dt_fin - dt) / 2
    else:
        medio = dt
    hora = medio.hour + medio.minute / 60.0
    dow = medio.weekday()  # 0=Lun, 4=Vie, 5=Sáb
    if dow <= 3:
        return "Día" if hora < 17.5 else "Tarde"
    elif dow == 4:
        return "Día" if hora < 16.5 else "Tarde"
    elif dow == 5:
        return "Día" if hora < 13 else "Tarde"
    return "Día"


# ==================================================
# CONFIGURACIÓN
# ==================================================
OUTPUT_DIR = PROJECT_ROOT / "output"
OUTPUT_DIR.mkdir(exist_ok=True)
EXCEL_FILE = OUTPUT_DIR / "reporte_diario_produccion.xlsx"

# Estilos
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SALA_FONT = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
SALA_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
DATA_FONT = Font(size=10, name="Calibri")
TOTAL_FONT = Font(bold=True, size=11, name="Calibri")
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)


def get_odoo():
    """Conecta a Odoo usando credenciales del .env o defaults."""
    return OdooClient(
        username="mvalladares@riofuturo.cl",
        password="c0766224bec30cac071ffe43a858c9ccbd521ddd"
    )


def fetch_mos(odoo, fecha_inicio, fecha_fin):
    """Obtiene las MOs terminadas del período."""
    domain = [
        '|',
        '&', ['x_studio_inicio_de_proceso', '>=', f'{fecha_inicio} 00:00:00'],
             ['x_studio_inicio_de_proceso', '<=', f'{fecha_fin} 23:59:59'],
        '&', ['x_studio_termino_de_proceso', '>=', f'{fecha_inicio} 00:00:00'],
             ['x_studio_termino_de_proceso', '<=', f'{fecha_fin} 23:59:59'],
        ['state', '!=', 'cancel']
    ]

    fields = [
        'id', 'name', 'product_id', 'state',
        'x_studio_sala_de_proceso',
        'x_studio_clientes',
        'x_studio_po_cliente_1',
        'x_studio_dotacin',
        'x_studio_hh_efectiva',
        'x_studio_hh',
        'x_studio_horas_detencion_totales',
        'x_studio_inicio_de_proceso',
        'x_studio_termino_de_proceso',
        'move_finished_ids', 'move_byproduct_ids',
    ]

    print(f"  Buscando MOs del {fecha_inicio} al {fecha_fin}...")
    mos = odoo.search_read('mrp.production', domain, fields, limit=5000,
                           order='x_studio_inicio_de_proceso asc')
    print(f"  → {len(mos)} MOs encontradas")
    return mos or []


def fetch_kg_pt(odoo, mos):
    """Calcula kg PT por MO desde move_finished_ids (excluyendo merma)."""
    # Recopilar move IDs
    all_finished = []
    for mo in mos:
        all_finished.extend(mo.get('move_finished_ids', []))
        all_finished.extend(mo.get('move_byproduct_ids', []))
    all_finished = list(set(all_finished))

    if not all_finished:
        return {}

    # Obtener move lines
    move_lines = odoo.search_read(
        'stock.move.line',
        [['move_id', 'in', all_finished]],
        ['move_id', 'product_id', 'qty_done'],
        limit=50000
    )

    # Obtener info de productos para detectar merma
    product_ids = set()
    for ml in move_lines or []:
        prod = ml.get('product_id')
        if prod:
            pid = prod[0] if isinstance(prod, (list, tuple)) else prod
            product_ids.add(pid)

    # product.product → product.template (para categoría)
    categ_map = {}
    if product_ids:
        products = odoo.search_read('product.product', [['id', 'in', list(product_ids)]],
                                    ['id', 'product_tmpl_id'], limit=20000)
        tmpl_ids = set()
        prod_to_tmpl = {}
        for p in products:
            tmpl = p.get('product_tmpl_id')
            if tmpl:
                tmpl_id = tmpl[0] if isinstance(tmpl, (list, tuple)) else tmpl
                tmpl_ids.add(tmpl_id)
                prod_to_tmpl[p['id']] = tmpl_id

        if tmpl_ids:
            templates = odoo.read('product.template', list(tmpl_ids), ['id', 'categ_id'])
            tmpl_categ = {}
            for t in templates:
                categ = t.get('categ_id')
                categ_name = categ[1] if isinstance(categ, (list, tuple)) and len(categ) > 1 else ''
                tmpl_categ[t['id']] = categ_name

            for pid, tid in prod_to_tmpl.items():
                categ_map[pid] = tmpl_categ.get(tid, '')

    # Mapear move_id → mo_id
    move_to_mo = {}
    for mo in mos:
        mo_id = mo.get('id')
        for mid in mo.get('move_finished_ids', []) + mo.get('move_byproduct_ids', []):
            move_to_mo[mid] = mo_id

    # Acumular kg_pt por MO (excluyendo merma y proceso intermedio)
    result = defaultdict(float)
    for ml in move_lines or []:
        move_info = ml.get('move_id')
        if not move_info:
            continue
        move_id = move_info[0] if isinstance(move_info, (list, tuple)) else move_info
        mo_id = move_to_mo.get(move_id)
        if mo_id is None:
            continue

        prod = ml.get('product_id')
        pid = prod[0] if isinstance(prod, (list, tuple)) else prod
        prod_name = prod[1] if isinstance(prod, (list, tuple)) and len(prod) > 1 else ''
        categ_name = categ_map.get(pid, '')

        # Excluir merma
        if 'MERMA' in categ_name.upper():
            continue
        # Excluir proceso intermedio
        if prod_name and 'PROCESO' in prod_name.upper():
            if prod_name.startswith('[3]') or prod_name.startswith('[1.') or \
               prod_name.startswith('[2.') or prod_name.startswith('[4]'):
                continue
            if 'PROCESOS' in categ_name.upper():
                continue

        result[mo_id] += ml.get('qty_done', 0) or 0

    return dict(result)


def fetch_quality_approval(odoo, mo_ids):
    """Obtiene kg aprobados por MO desde quality.check."""
    if not mo_ids:
        return {}

    batch_size = 200
    all_checks = []
    for i in range(0, len(mo_ids), batch_size):
        batch = mo_ids[i:i + batch_size]
        checks = odoo.search_read(
            'quality.check',
            [['production_id', 'in', batch],
             ['x_studio_aprovado', '=', 'Aprobado']],
            ['production_id', 'x_studio_aprovado', 'product_id', 'measure'],
            limit=10000
        )
        all_checks.extend(checks or [])

    # Acumular por MO
    result = defaultdict(int)
    for c in all_checks:
        prod = c.get('production_id')
        if prod:
            mo_id = prod[0] if isinstance(prod, (list, tuple)) else prod
            result[mo_id] += 1

    return dict(result)


def process_data(mos, kg_pt_map, quality_map):
    """Agrupa MOs por Sala → Turno → Fecha y calcula métricas."""
    rows = []

    for mo in mos:
        mo_id = mo.get('id')

        # Sala
        sala_raw = mo.get('x_studio_sala_de_proceso')
        sala = sala_raw[1] if isinstance(sala_raw, (list, tuple)) and len(sala_raw) > 1 else str(sala_raw or 'Sin Sala')

        # Turno (derivado del horario de proceso)
        dt_inicio = None
        dt_fin = None
        try:
            raw_inicio = mo.get('x_studio_inicio_de_proceso')
            if raw_inicio:
                dt_inicio = datetime.strptime(str(raw_inicio), '%Y-%m-%d %H:%M:%S') if isinstance(raw_inicio, str) else raw_inicio
            raw_fin = mo.get('x_studio_termino_de_proceso')
            if raw_fin:
                dt_fin = datetime.strptime(str(raw_fin), '%Y-%m-%d %H:%M:%S') if isinstance(raw_fin, str) else raw_fin
        except Exception:
            pass
        turno = clasificar_turno(dt_inicio, dt_fin)

        # Cliente
        cliente_raw = mo.get('x_studio_clientes')
        cliente = cliente_raw[1] if isinstance(cliente_raw, (list, tuple)) and len(cliente_raw) > 1 else str(cliente_raw or '')

        # Contenedor
        contenedor = mo.get('x_studio_po_cliente_1') or ''

        # Dotación
        dotacion = mo.get('x_studio_dotacin') or 0
        if not isinstance(dotacion, (int, float)):
            dotacion = 0

        # KG PT
        kg_pt = round(kg_pt_map.get(mo_id, 0), 2)

        # Duración
        inicio = mo.get('x_studio_inicio_de_proceso')
        fin = mo.get('x_studio_termino_de_proceso')
        duracion_horas = 0
        fecha_proceso = ''
        if inicio:
            try:
                d0 = datetime.strptime(str(inicio), "%Y-%m-%d %H:%M:%S") if isinstance(inicio, str) else inicio
                fecha_proceso = d0.strftime('%Y-%m-%d')
            except Exception:
                fecha_proceso = str(inicio)[:10]
        if inicio and fin:
            try:
                d0 = datetime.strptime(str(inicio), "%Y-%m-%d %H:%M:%S") if isinstance(inicio, str) else inicio
                d1 = datetime.strptime(str(fin), "%Y-%m-%d %H:%M:%S") if isinstance(fin, str) else fin
                duracion_horas = round((d1 - d0).total_seconds() / 3600, 2)
            except Exception:
                pass

        # HH
        hh = mo.get('x_studio_hh') or 0
        if not isinstance(hh, (int, float)):
            hh = 0
        hh_efectiva = mo.get('x_studio_hh_efectiva') or 0
        if not isinstance(hh_efectiva, (int, float)):
            hh_efectiva = 0

        # Detenciones
        detenciones = mo.get('x_studio_horas_detencion_totales') or 0
        if not isinstance(detenciones, (int, float)):
            detenciones = 0

        # KG/HH
        kg_hh = round(kg_pt / hh, 2) if hh > 0 else 0

        # KG/Hora Efectiva de Proceso
        horas_efectivas = max(duracion_horas - detenciones, 0)
        kg_hora_ef = round(kg_pt / horas_efectivas, 2) if horas_efectivas > 0 else 0

        # Quality
        quality_count = quality_map.get(mo_id, 0)

        rows.append({
            'fecha': fecha_proceso,
            'sala': sala,
            'turno': turno,
            'mo_name': mo.get('name', ''),
            'cliente': cliente,
            'contenedor': contenedor,
            'dotacion': dotacion,
            'kg_pt': kg_pt,
            'duracion_horas': duracion_horas,
            'hh': hh,
            'hh_efectiva': hh_efectiva,
            'kg_hh': kg_hh,
            'kg_hora_efectiva': kg_hora_ef,
            'detenciones': detenciones,
            'quality_aprobados': quality_count,
        })

    return rows


def write_excel(rows, fecha_inicio, fecha_fin, save_to_file=True):
    """Escribe datos al Excel. Retorna bytes del workbook.
    Si save_to_file=True, también guarda al archivo acumulado.
    """
    import io

    # Workbook para descarga (siempre limpio, solo el día)
    wb = openpyxl.Workbook()
    ws = wb.active

    # Nombre de hoja
    if fecha_inicio == fecha_fin:
        sheet_name = fecha_inicio
    else:
        sheet_name = f"{fecha_inicio}_a_{fecha_fin}"
    sheet_name = sheet_name[:31]
    ws.title = sheet_name

    # Título
    ws.merge_cells('A1:N1')
    title_cell = ws.cell(row=1, column=1,
                         value=f"📊 Reporte Producción: {fecha_inicio}" + (f" a {fecha_fin}" if fecha_inicio != fecha_fin else ""))
    title_cell.font = Font(bold=True, size=14, name="Calibri", color="1F4E79")
    title_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    # Info
    ws.merge_cells('A2:N2')
    info = ws.cell(row=2, column=1,
                   value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Total MOs: {len(rows)}")
    info.font = Font(size=10, italic=True, color="666666")
    info.alignment = Alignment(horizontal='center')

    # Agrupar por sala → turno
    salas = defaultdict(list)
    for r in rows:
        salas[r['sala']].append(r)

    # Ordenar salas
    salas_ordered = sorted(salas.keys())

    HEADERS = [
        "Fecha", "OF", "Cliente", "# Contenedor", "Dotación",
        "Kilos Turno (PT)", "Duración (hrs)", "HH Total", "HH Efectiva",
        "KG/HH", "KG/Hora Ef. Proceso", "Detenciones (hrs)",
        "Checks Aprobados", "Turno"
    ]
    COL_WIDTHS = [12, 14, 25, 18, 12, 16, 14, 12, 14, 12, 18, 16, 16, 12]

    row_idx = 4  # Start row

    for sala in salas_ordered:
        sala_rows = salas[sala]

        # SALA HEADER
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(HEADERS))
        sala_cell = ws.cell(row=row_idx, column=1, value=f"🏭 {sala}")
        sala_cell.font = SALA_FONT
        sala_cell.fill = SALA_FILL
        sala_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row_idx].height = 28
        for c in range(2, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=c).fill = SALA_FILL
        row_idx += 1

        # Agrupar por turno dentro de sala
        turnos = defaultdict(list)
        for r in sala_rows:
            turnos[r['turno']].append(r)

        for turno in sorted(turnos.keys()):
            turno_rows = turnos[turno]

            # Turno sub-header
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(HEADERS))
            turno_cell = ws.cell(row=row_idx, column=1, value=f"   ⏰ Turno: {turno}")
            turno_cell.font = Font(bold=True, size=11, name="Calibri", color="2E75B6")
            turno_cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
            for c in range(2, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=c).fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
            row_idx += 1

            # Column headers
            for col, h in enumerate(HEADERS, 1):
                c = ws.cell(row=row_idx, column=col, value=h)
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
                c.alignment = Alignment(horizontal='center', wrap_text=True)
                c.border = THIN_BORDER
            ws.row_dimensions[row_idx].height = 24
            row_idx += 1

            # Data
            turno_rows.sort(key=lambda r: (r['fecha'], r['mo_name']))
            sum_kg = 0
            sum_hh = 0
            sum_hh_ef = 0
            sum_det = 0
            sum_duracion = 0

            for i, r in enumerate(turno_rows):
                fill = ALT_FILL if i % 2 == 0 else PatternFill(fill_type=None)
                vals = [
                    r['fecha'],
                    r['mo_name'],
                    r['cliente'],
                    r['contenedor'],
                    r['dotacion'],
                    r['kg_pt'],
                    r['duracion_horas'],
                    r['hh'],
                    r['hh_efectiva'],
                    r['kg_hh'],
                    r['kg_hora_efectiva'],
                    r['detenciones'],
                    r['quality_aprobados'],
                    r['turno'],
                ]
                for col, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.font = DATA_FONT
                    cell.border = THIN_BORDER
                    cell.fill = fill
                    # Formato numérico
                    if col in (5, 6, 7, 8, 9, 10, 11, 12):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')
                    elif col == 13:
                        cell.alignment = Alignment(horizontal='center')

                sum_kg += r['kg_pt']
                sum_hh += r['hh']
                sum_hh_ef += r['hh_efectiva']
                sum_det += r['detenciones']
                sum_duracion += r['duracion_horas']
                row_idx += 1

            # Totales del turno
            ws.cell(row=row_idx, column=1, value="").border = THIN_BORDER
            for c in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=c).fill = TOTAL_FILL
                ws.cell(row=row_idx, column=c).font = TOTAL_FONT
                ws.cell(row=row_idx, column=c).border = THIN_BORDER

            ws.cell(row=row_idx, column=3, value=f"TOTAL {turno} ({len(turno_rows)} OFs)")
            ws.cell(row=row_idx, column=6, value=round(sum_kg, 2)).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=7, value=round(sum_duracion, 2)).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=8, value=round(sum_hh, 2)).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=9, value=round(sum_hh_ef, 2)).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=10, value=round(sum_kg / sum_hh, 2) if sum_hh > 0 else 0).number_format = '#,##0.00'
            horas_ef_t = max(sum_duracion - sum_det, 0)
            ws.cell(row=row_idx, column=11, value=round(sum_kg / horas_ef_t, 2) if horas_ef_t > 0 else 0).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=12, value=round(sum_det, 2)).number_format = '#,##0.00'
            row_idx += 1

        # Espacio entre salas
        row_idx += 1

    # Ajustar anchos
    for col, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Congelar paneles
    ws.freeze_panes = "A4"

    # Generar bytes para descarga
    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()

    # También guardar al archivo acumulado si se pide
    if save_to_file:
        if EXCEL_FILE.exists():
            wb_acum = openpyxl.load_workbook(str(EXCEL_FILE))
        else:
            wb_acum = openpyxl.Workbook()
            wb_acum.remove(wb_acum.active)

        if sheet_name in wb_acum.sheetnames:
            del wb_acum[sheet_name]

        # Copiar la hoja al workbook acumulado
        ws_acum = wb_acum.create_sheet(title=sheet_name)
        from copy import copy
        for row in ws.iter_rows():
            for cell in row:
                new_cell = ws_acum.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.number_format = cell.number_format

        # Copiar anchos y alturas
        for col_letter, dim in ws.column_dimensions.items():
            ws_acum.column_dimensions[col_letter].width = dim.width
        for row_num, dim in ws.row_dimensions.items():
            ws_acum.row_dimensions[row_num].height = dim.height

        # Copiar merges
        for merge in ws.merged_cells.ranges:
            ws_acum.merge_cells(str(merge))

        ws_acum.freeze_panes = "A4"
        wb_acum.save(str(EXCEL_FILE))
        print(f"\n✅ Excel guardado: {EXCEL_FILE}")
        print(f"   Hoja: '{sheet_name}' con {len(rows)} registros")

    return excel_bytes


def main():
    args = sys.argv[1:]

    if len(args) == 0:
        fecha_inicio = date.today().strftime('%Y-%m-%d')
        fecha_fin = fecha_inicio
    elif len(args) == 1:
        fecha_inicio = args[0]
        fecha_fin = fecha_inicio
    elif len(args) >= 2:
        fecha_inicio = args[0]
        fecha_fin = args[1]
    else:
        print("Uso: python scripts/reporte_diario_produccion.py [fecha_inicio] [fecha_fin]")
        sys.exit(1)

    print(f"📊 Reporte Diario de Producción")
    print(f"   Período: {fecha_inicio} → {fecha_fin}")
    print(f"   Output:  {EXCEL_FILE}")
    print(f"{'='*60}")

    # 1. Conectar
    print("\n🔌 Conectando a Odoo...")
    odoo = get_odoo()
    print("   ✅ Conectado")

    # 2. Obtener MOs
    print("\n📦 Obteniendo MOs...")
    mos = fetch_mos(odoo, fecha_inicio, fecha_fin)
    if not mos:
        print("   ⚠️  No se encontraron MOs en este período.")
        sys.exit(0)

    # 3. Calcular KG PT
    print("\n⚖️  Calculando KG de producto terminado...")
    kg_pt_map = fetch_kg_pt(odoo, mos)
    print(f"   → {len(kg_pt_map)} MOs con producción")

    # 4. Quality checks
    print("\n✅ Obteniendo checks de calidad aprobados...")
    mo_ids = [mo['id'] for mo in mos]
    quality_map = fetch_quality_approval(odoo, mo_ids)
    print(f"   → {sum(quality_map.values())} checks aprobados en {len(quality_map)} MOs")

    # 5. Procesar
    print("\n🔧 Procesando datos...")
    rows = process_data(mos, kg_pt_map, quality_map)

    # 6. Escribir Excel
    print("\n📝 Escribiendo Excel...")
    write_excel(rows, fecha_inicio, fecha_fin)

    print(f"\n{'='*60}")
    print("✅ ¡Listo! Abre el archivo para revisar.")


if __name__ == '__main__':
    main()
