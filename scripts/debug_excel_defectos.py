"""
Script de debug para probar la generación del Excel de defectos localmente.
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Usar credenciales estáticas existentes
USERNAME = "mvalladares@riofuturo.cl"
PASSWORD = "c0766224bec30cac071ffe43a858c9ccbd521ddd"

print("=" * 80)
print("DEBUG: Generación de Excel de Defectos")
print("=" * 80)

try:
    from backend.services.recepcion_defectos_service import generar_reporte_defectos_excel
    
    print("\n1. Llamando a generar_reporte_defectos_excel...")
    
    xlsx_bytes = generar_reporte_defectos_excel(
        username=USERNAME,
        password=PASSWORD,
        fecha_inicio="2025-11-17",
        fecha_fin="2026-02-05",
        origenes=None,  # Todos los orígenes
        solo_hechas=True
    )
    
    print(f"\n2. Excel generado: {len(xlsx_bytes)} bytes")
    
    # Verificar que es un archivo Excel válido
    if xlsx_bytes[:4] == b'PK\x03\x04':
        print("   ✅ El archivo comienza con la firma ZIP correcta (PK)")
    else:
        print(f"   ❌ El archivo NO comienza con PK, comienza con: {xlsx_bytes[:20]}")
    
    # Guardar el archivo
    output_file = "debug_excel_defectos.xlsx"
    with open(output_file, 'wb') as f:
        f.write(xlsx_bytes)
    
    print(f"\n3. Archivo guardado como: {output_file}")
    
    # Intentar abrir con openpyxl para verificar
    from openpyxl import load_workbook
    wb = load_workbook(output_file)
    ws = wb.active
    
    print(f"\n4. Excel abierto correctamente con openpyxl")
    print(f"   - Hoja activa: {ws.title}")
    print(f"   - Filas: {ws.max_row}")
    print(f"   - Columnas: {ws.max_column}")
    
    if ws.max_row > 1:
        print(f"\n   Primeras columnas:")
        for col in range(1, min(6, ws.max_column + 1)):
            print(f"     - {ws.cell(1, col).value}")
    
    print("\n✅ TEST COMPLETADO EXITOSAMENTE")
    
except Exception as e:
    import traceback
    print(f"\n❌ ERROR: {type(e).__name__}: {e}")
    print("\nTraceback:")
    traceback.print_exc()
