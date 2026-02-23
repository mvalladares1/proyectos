import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta

wb = openpyxl.Workbook()

# ===================== ESTILOS =====================
HEADER = {
    "font": Font(bold=True, color="FFFFFF", size=11, name="Calibri"),
    "fill": PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid"),
    "align": Alignment(horizontal="center", vertical="center", wrap_text=True),
}
WORKER_HEADER = {
    "font": Font(bold=True, size=11, color="1e3a5f", name="Calibri"),
    "fill": PatternFill(start_color="d6e4f0", end_color="d6e4f0", fill_type="solid"),
}
NORMAL = Font(size=10, name="Calibri")
NORMAL_BOLD = Font(size=10, name="Calibri", bold=True)
GREEN_FONT = Font(bold=True, size=10, color="006100", name="Calibri")
GREEN_FILL = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")
RED_FONT = Font(bold=True, size=10, color="9c0006", name="Calibri")
RED_FILL = PatternFill(start_color="fce4ec", end_color="fce4ec", fill_type="solid")
YELLOW_FONT = Font(bold=True, size=10, color="856404", name="Calibri")
YELLOW_FILL = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
GRAY_FONT = Font(size=10, color="808080", name="Calibri", italic=True)
TOTAL_FONT = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
TOTAL_FILL = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
SUBTOTAL_FONT = Font(bold=True, size=10, color="1e3a5f", name="Calibri")
SUBTOTAL_FILL = PatternFill(start_color="e8f0fe", end_color="e8f0fe", fill_type="solid")
BORDER = Border(
    left=Side(style='thin', color='C0C0C0'),
    right=Side(style='thin', color='C0C0C0'),
    top=Side(style='thin', color='C0C0C0'),
    bottom=Side(style='thin', color='C0C0C0'),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

DIAS_ES = {0: "Lunes", 1: "Martes", 2: "Miércoles", 3: "Jueves", 4: "Viernes", 5: "Sábado", 6: "Domingo"}

def parse_fecha(s):
    d, m, y = s.split("/")
    return date(int(y), int(m), int(d))

def fmt_fecha(d):
    return d.strftime("%d/%m/%Y")

def dia_nombre(d):
    return DIAS_ES[d.weekday()]

# ===================== DATOS =====================
# Período: 1 al 20 de febrero 2026
# Doble turno = 8 horas extra
# Solo se registran novedades; el resto = asistencia normal

# Calcular días hábiles del período (Lunes a Sábado)
PERIODO_INICIO = date(2026, 2, 1)
PERIODO_FIN = date(2026, 2, 20)
dias_habiles = []
d = PERIODO_INICIO
while d <= PERIODO_FIN:
    if d.weekday() < 6:  # Lunes(0) a Sábado(5)
        dias_habiles.append(d)
    d += timedelta(days=1)
TOTAL_DIAS_HABILES = len(dias_habiles)

trabajadores = [
    {
        "nombre": "Belén Montes",
        "novedades": [
            {"fecha": date(2026, 2, 16), "tipo": "Turno Día", "horas": 0, "obs": "Asistencia normal"},
            {"fecha": date(2026, 2, 17), "tipo": "Turno Día", "horas": 0, "obs": "Asistencia normal"},
            {"fecha": date(2026, 2, 18), "tipo": "Doble Turno", "horas": 8, "obs": "Doble turno"},
            {"fecha": date(2026, 2, 19), "tipo": "Turno Día", "horas": 0, "obs": "Asistencia normal"},
            {"fecha": date(2026, 2, 20), "tipo": "Turno Día", "horas": 0, "obs": "Asistencia normal"},
        ]
    },
    {
        "nombre": "Jordana Álvarez",
        "novedades": [
            {"fecha": date(2026, 2, 4), "tipo": "Doble Turno", "horas": 8, "obs": "Doble turno"},
            {"fecha": date(2026, 2, 6), "tipo": "Doble Turno", "horas": 8, "obs": "Doble turno"},
            {"fecha": date(2026, 2, 17), "tipo": "Doble Turno", "horas": 8, "obs": "Doble turno"},
            {"fecha": date(2026, 2, 18), "tipo": "Doble Turno", "horas": 8, "obs": "Doble turno"},
        ]
    },
    {
        "nombre": "Javiera Venegas",
        "novedades": [
            {"fecha": date(2026, 2, 16), "tipo": "Falta", "horas": 0, "obs": "No asistió"},
            {"fecha": date(2026, 2, 17), "tipo": "Licencia", "horas": 0, "obs": "Licencia médica"},
            {"fecha": date(2026, 2, 18), "tipo": "Licencia", "horas": 0, "obs": "Licencia médica"},
            {"fecha": date(2026, 2, 19), "tipo": "Turno Día", "horas": 0, "obs": "Asistió normalmente"},
            {"fecha": date(2026, 2, 20), "tipo": "Turno Día", "horas": 0, "obs": "Asistió normalmente"},
        ]
    },
    {
        "nombre": "Juan Pablo",
        "novedades": []  # Asistencia completa L-S, sin novedades
    },
    {
        "nombre": "Juliana Delgad",
        "novedades": [
            {"fecha": date(2026, 2, 5), "tipo": "Atraso", "horas": 0, "obs": "Ingresó a las 08:40, 40 minutos tarde"},
            {"fecha": date(2026, 2, 6), "tipo": "Administrativo", "horas": 0, "obs": "Día administrativo"},
        ]
    },
    {
        "nombre": "Ignacia Aguilar",
        "novedades": [
            {"fecha": date(2026, 2, 1), "tipo": "Horas Extra", "horas": 4.5, "obs": "4 horas y media extra (domingo)"},
            {"fecha": date(2026, 2, 3), "tipo": "Doble Turno", "horas": 8, "obs": "Doble turno"},
            {"fecha": date(2026, 2, 17), "tipo": "Horas Extra", "horas": 4.5, "obs": "4 horas y media extra"},
        ]
    },
    {
        "nombre": "Emilio",
        "novedades": [
            {"fecha": date(2026, 2, 8), "tipo": "Horas Extra", "horas": 3, "obs": "3 horas extra (domingo)"},
        ]
    },
    {
        "nombre": "Alejandro Godoy",
        "novedades": [
            {"fecha": date(2026, 2, 7), "tipo": "Horas Extra", "horas": 2, "obs": "2 horas extra"},
            {"fecha": date(2026, 2, 9), "tipo": "Horas Extra", "horas": 1.5, "obs": "1 hora y media extra"},
            {"fecha": date(2026, 2, 10), "tipo": "Horas Extra", "horas": 1, "obs": "1 hora extra"},
            {"fecha": date(2026, 2, 12), "tipo": "Horas Extra", "horas": 1.5, "obs": "1 hora y media extra"},
            {"fecha": date(2026, 2, 14), "tipo": "Horas Extra", "horas": 1, "obs": "1 hora extra (viernes)"},
            {"fecha": date(2026, 2, 20), "tipo": "Salida Anticipada", "horas": -4.5, "obs": "Se retiró a las 12:00 (salía a las 16:30)"},
        ]
    },
    {
        "nombre": "Francisca Barriga",
        "novedades": [
            {"fecha": date(2026, 2, 16), "tipo": "Doble Turno", "horas": 8, "obs": "Doble turno"},
        ]
    },
    {
        "nombre": "Nicole Montes",
        "novedades": [
            {"fecha": date(2026, 2, 3), "tipo": "Horas Extra", "horas": 0.5, "obs": "30 minutos extra"},
            {"fecha": date(2026, 2, 5), "tipo": "Horas Extra", "horas": 1.5, "obs": "1 hora y media extra"},
            {"fecha": date(2026, 2, 7), "tipo": "Horas Extra", "horas": 0.87, "obs": "52 minutos extra"},
            {"fecha": date(2026, 2, 10), "tipo": "Horas Extra", "horas": 3, "obs": "3 horas extra"},
            {"fecha": date(2026, 2, 12), "tipo": "Horas Extra", "horas": 1.5, "obs": "1 hora y media extra"},
            {"fecha": date(2026, 2, 13), "tipo": "Horas Extra", "horas": 2, "obs": "2 horas extra"},
            {"fecha": date(2026, 2, 14), "tipo": "Horas Extra", "horas": 0.5, "obs": "30 minutos extra"},
        ]
    },
]


def apply_cell(ws, row, col, value, font=NORMAL, fill=None, alignment=CENTER):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.alignment = alignment
    c.border = BORDER
    if fill:
        c.fill = fill
    return c


# ===================== HOJA 1: DETALLE POR TRABAJADOR =====================
ws1 = wb.active
ws1.title = "Detalle"

headers = ["Nº", "Trabajador", "Fecha", "Día", "Tipo Novedad", "Horas Extra", "Observación"]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.font = HEADER["font"]
    c.fill = HEADER["fill"]
    c.alignment = HEADER["align"]
    c.border = BORDER
ws1.row_dimensions[1].height = 35

row = 2
num = 1
for t in trabajadores:
    nombre = t["nombre"]
    novedades = t["novedades"]
    
    # Fila de encabezado del trabajador
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws1.cell(row=row, column=1, value=f"  {nombre}")
    c.font = WORKER_HEADER["font"]
    c.fill = WORKER_HEADER["fill"]
    c.alignment = Alignment(horizontal="left", vertical="center")
    for cc in range(1, 8):
        ws1.cell(row=row, column=cc).border = BORDER
        ws1.cell(row=row, column=cc).fill = WORKER_HEADER["fill"]
    ws1.row_dimensions[row].height = 28
    row += 1
    
    if not novedades:
        # Asistencia completa sin novedades
        apply_cell(ws1, row, 1, num, NORMAL)
        apply_cell(ws1, row, 2, nombre, NORMAL, alignment=LEFT)
        apply_cell(ws1, row, 3, "01/02 – 20/02", NORMAL)
        apply_cell(ws1, row, 4, "—", GRAY_FONT)
        apply_cell(ws1, row, 5, "Asistencia Completa", GREEN_FONT, GREEN_FILL)
        apply_cell(ws1, row, 6, "0", GRAY_FONT)
        apply_cell(ws1, row, 7, f"Sin novedades, asistencia completa los {TOTAL_DIAS_HABILES} días hábiles del período", NORMAL, alignment=LEFT)
        num += 1
        row += 1
    else:
        for nov in novedades:
            fecha = nov["fecha"]
            tipo = nov["tipo"]
            horas = nov["horas"]
            obs = nov["obs"]
            
            apply_cell(ws1, row, 1, num, NORMAL)
            apply_cell(ws1, row, 2, nombre, NORMAL, alignment=LEFT)
            apply_cell(ws1, row, 3, fmt_fecha(fecha), NORMAL)
            apply_cell(ws1, row, 4, dia_nombre(fecha), NORMAL)
            
            # Tipo con color
            if tipo in ("Doble Turno", "Horas Extra"):
                apply_cell(ws1, row, 5, tipo, GREEN_FONT, GREEN_FILL)
                apply_cell(ws1, row, 6, horas, GREEN_FONT, GREEN_FILL)
            elif tipo == "Falta":
                apply_cell(ws1, row, 5, tipo, RED_FONT, RED_FILL)
                apply_cell(ws1, row, 6, "—", RED_FONT, RED_FILL)
            elif tipo == "Licencia":
                apply_cell(ws1, row, 5, tipo, YELLOW_FONT, YELLOW_FILL)
                apply_cell(ws1, row, 6, "—", YELLOW_FONT, YELLOW_FILL)
            elif tipo == "Salida Anticipada":
                apply_cell(ws1, row, 5, tipo, RED_FONT, RED_FILL)
                apply_cell(ws1, row, 6, f"{abs(horas)}h menos", RED_FONT, RED_FILL)
            elif tipo == "Atraso":
                apply_cell(ws1, row, 5, tipo, RED_FONT, RED_FILL)
                apply_cell(ws1, row, 6, "—", RED_FONT, RED_FILL)
            elif tipo == "Administrativo":
                apply_cell(ws1, row, 5, tipo, YELLOW_FONT, YELLOW_FILL)
                apply_cell(ws1, row, 6, "—", YELLOW_FONT, YELLOW_FILL)
            else:
                apply_cell(ws1, row, 5, tipo, NORMAL)
                apply_cell(ws1, row, 6, horas if horas > 0 else "—", NORMAL)
            
            apply_cell(ws1, row, 7, obs, NORMAL, alignment=LEFT)
            
            num += 1
            row += 1
        
        # Fila indicando asistencia completa en los demás días
        dias_con_novedad = len(novedades)
        dias_normales = TOTAL_DIAS_HABILES - dias_con_novedad
        if dias_normales > 0:
            apply_cell(ws1, row, 1, "", GRAY_FONT)
            apply_cell(ws1, row, 2, "", GRAY_FONT)
            apply_cell(ws1, row, 3, "Resto del período", GRAY_FONT)
            apply_cell(ws1, row, 4, f"{dias_normales} días", GRAY_FONT)
            apply_cell(ws1, row, 5, "Asistencia Completa", GREEN_FONT, GREEN_FILL)
            apply_cell(ws1, row, 6, "—", GRAY_FONT)
            apply_cell(ws1, row, 7, f"Asistencia normal los {dias_normales} días hábiles restantes", GRAY_FONT, alignment=LEFT)
            row += 1
    
    # Subtotal por trabajador
    total_horas = sum(n["horas"] for n in novedades)
    dobles = sum(1 for n in novedades if n["tipo"] == "Doble Turno")
    faltas = sum(1 for n in novedades if n["tipo"] == "Falta")
    licencias = sum(1 for n in novedades if n["tipo"] == "Licencia")
    n_salidas = sum(1 for n in novedades if n["tipo"] == "Salida Anticipada")
    salidas = sum(n["horas"] for n in novedades if n["tipo"] == "Salida Anticipada")
    atrasos = sum(1 for n in novedades if n["tipo"] == "Atraso")
    administrativos = sum(1 for n in novedades if n["tipo"] == "Administrativo")
    
    resumen_parts = []
    if dobles:
        resumen_parts.append(f"{dobles} doble(s) turno")
    extras_otros = [n for n in novedades if n["tipo"] == "Horas Extra"]
    if extras_otros:
        h_extras = sum(n["horas"] for n in extras_otros)
        resumen_parts.append(f"{h_extras}h extras adicionales")
    if n_salidas:
        resumen_parts.append(f"{n_salidas} salida(s) anticipada(s) ({abs(salidas)}h)")
    if faltas:
        resumen_parts.append(f"{faltas} falta(s)")
    if licencias:
        resumen_parts.append(f"{licencias} licencia(s)")
    if atrasos:
        resumen_parts.append(f"{atrasos} atraso(s)")
    if administrativos:
        resumen_parts.append(f"{administrativos} día(s) administrativo(s)")
    if not resumen_parts:
        resumen_parts.append("Sin novedades")
    
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws1.cell(row=row, column=1, value=f"  Subtotal {nombre}: {' | '.join(resumen_parts)}")
    c.font = SUBTOTAL_FONT
    c.fill = SUBTOTAL_FILL
    c.alignment = LEFT
    for cc in range(1, 8):
        ws1.cell(row=row, column=cc).border = BORDER
        ws1.cell(row=row, column=cc).fill = SUBTOTAL_FILL
    apply_cell(ws1, row, 6, total_horas if total_horas > 0 else 0, SUBTOTAL_FONT, SUBTOTAL_FILL)
    apply_cell(ws1, row, 7, "", NORMAL, SUBTOTAL_FILL)
    row += 1
    
    # Espacio
    row += 1

# Anchos hoja 1
ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 22
ws1.column_dimensions['F'].width = 14
ws1.column_dimensions['G'].width = 45
ws1.freeze_panes = "A2"


# ===================== HOJA 2: RESUMEN =====================
ws2 = wb.create_sheet("Resumen")

headers2 = ["Nº", "Trabajador", "Dobles Turno", "Horas Extra\n(adicionales)", "Total\nHoras Extra", "Faltas", "Licencias", "Estado"]
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = HEADER["font"]
    c.fill = HEADER["fill"]
    c.alignment = HEADER["align"]
    c.border = BORDER
ws2.row_dimensions[1].height = 40

# Título
ws2.merge_cells("A1:A1")

row2 = 2
total_general_horas = 0

for idx, t in enumerate(trabajadores, 1):
    nombre = t["nombre"]
    novedades = t["novedades"]
    
    dobles = sum(1 for n in novedades if n["tipo"] == "Doble Turno")
    h_dobles = dobles * 8
    extras = sum(n["horas"] for n in novedades if n["tipo"] == "Horas Extra")
    salidas = sum(n["horas"] for n in novedades if n["tipo"] == "Salida Anticipada")
    total_h = h_dobles + extras + salidas
    faltas = sum(1 for n in novedades if n["tipo"] == "Falta")
    licencias = sum(1 for n in novedades if n["tipo"] == "Licencia")
    n_salidas = sum(1 for n in novedades if n["tipo"] == "Salida Anticipada")
    atrasos = sum(1 for n in novedades if n["tipo"] == "Atraso")
    administrativos = sum(1 for n in novedades if n["tipo"] == "Administrativo")
    
    # Estado
    if total_h > 0 and faltas == 0 and licencias == 0:
        estado = "✓ Con horas extra"
        est_font, est_fill = GREEN_FONT, GREEN_FILL
    elif faltas > 0:
        estado = "✗ Con inasistencias"
        est_font, est_fill = RED_FONT, RED_FILL
    elif licencias > 0 and total_h == 0:
        estado = "⚠ Con licencia"
        est_font, est_fill = YELLOW_FONT, YELLOW_FILL
    else:
        estado = "— Normal"
        est_font, est_fill = NORMAL, None
    
    apply_cell(ws2, row2, 1, idx, NORMAL)
    apply_cell(ws2, row2, 2, nombre, NORMAL_BOLD, alignment=LEFT)
    apply_cell(ws2, row2, 3, dobles if dobles > 0 else "—", GREEN_FONT if dobles > 0 else GRAY_FONT, GREEN_FILL if dobles > 0 else None)
    apply_cell(ws2, row2, 4, extras if extras > 0 else "—", GREEN_FONT if extras > 0 else GRAY_FONT, GREEN_FILL if extras > 0 else None)
    apply_cell(ws2, row2, 5, total_h if total_h > 0 else 0, Font(bold=True, size=12, color="006100", name="Calibri") if total_h > 0 else GRAY_FONT, GREEN_FILL if total_h > 0 else None)
    apply_cell(ws2, row2, 6, faltas if faltas > 0 else "—", RED_FONT if faltas > 0 else GRAY_FONT, RED_FILL if faltas > 0 else None)
    apply_cell(ws2, row2, 7, licencias if licencias > 0 else "—", YELLOW_FONT if licencias > 0 else GRAY_FONT, YELLOW_FILL if licencias > 0 else None)
    apply_cell(ws2, row2, 8, estado, est_font, est_fill, LEFT)
    
    ws2.row_dimensions[row2].height = 25
    total_general_horas += total_h
    row2 += 1

# Fila TOTAL
for cc in range(1, 9):
    ws2.cell(row=row2, column=cc).fill = TOTAL_FILL
    ws2.cell(row=row2, column=cc).font = TOTAL_FONT
    ws2.cell(row=row2, column=cc).alignment = CENTER
    ws2.cell(row=row2, column=cc).border = BORDER
ws2.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=2)
ws2.cell(row=row2, column=1).value = "TOTAL GENERAL"
ws2.cell(row=row2, column=1).alignment = Alignment(horizontal="right", vertical="center")
ws2.cell(row=row2, column=5).value = total_general_horas
ws2.row_dimensions[row2].height = 30

row2 += 2
# Leyenda
leyenda = [
    ("Verde", "Horas extra / Doble turno", GREEN_FILL),
    ("Rojo", "Falta / Inasistencia", RED_FILL),
    ("Amarillo", "Licencia médica", YELLOW_FILL),
]
ws2.cell(row=row2, column=2, value="LEYENDA:").font = Font(bold=True, size=10, name="Calibri")
row2 += 1
for color, desc, fill in leyenda:
    ws2.cell(row=row2, column=2).fill = fill
    ws2.cell(row=row2, column=2).border = BORDER
    ws2.cell(row=row2, column=3, value=desc).font = NORMAL
    row2 += 1

row2 += 1
ws2.cell(row=row2, column=2, value="Período: 01/02/2026 al 20/02/2026").font = Font(italic=True, size=10, name="Calibri", color="666666")

# Anchos hoja 2
ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 16
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 12
ws2.column_dimensions['H'].width = 24
ws2.freeze_panes = "A2"

output = r"c:\Users\HP\Desktop\Horas_Extra_Febrero_2026.xlsx"
wb.save(output)
print(f"Excel generado: {output}")
print(f"Trabajadores: {len(trabajadores)}")
print(f"Total horas extra: {total_general_horas}")
