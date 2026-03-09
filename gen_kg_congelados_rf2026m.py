import xmlrpc.client
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from collections import defaultdict

# Conexión a Odoo
url = "https://riofuturo.server98c6e.oerpondemand.net"
db = "riofuturo-master"
username = "mvalladares@riofuturo.cl"
password = "c0766224bec30cac071ffe43a858c9ccbd521ddd"

print("Conectando a Odoo...")
common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common", context=None)
uid = common.authenticate(db, username, password, {})
models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object", context=None)

def search_read(model, domain, fields, limit=0):
    return models.execute_kw(db, uid, password, model, 'search_read', [domain], {'fields': fields, 'limit': limit})

# 1. Buscar TODOS los paquetes RF2026M
print("Buscando paquetes RF2026M...")
all_packages = search_read('stock.quant.package', [('name', '=like', 'RF2026M%')], ['id', 'name'])
print(f"Total paquetes RF2026M: {len(all_packages)}")
pkg_ids = [p['id'] for p in all_packages]
pkg_names = {p['id']: p['name'] for p in all_packages}

# 2. Buscar todas las move lines con result_package_id en esos paquetes
print("Buscando move lines de recepciones...")
all_move_lines = []
batch_size = 100
for i in range(0, len(pkg_ids), batch_size):
    batch = pkg_ids[i:i+batch_size]
    mls = search_read('stock.move.line', [
        ('result_package_id', 'in', batch),
        ('state', '=', 'done'),
    ], ['id', 'product_id', 'qty_done', 'picking_id', 'result_package_id', 'lot_id'])
    all_move_lines.extend(mls)
    print(f"  Lote {i+1}-{min(i+batch_size, len(pkg_ids))}: {len(mls)} move lines")

print(f"Total move lines: {len(all_move_lines)}")

# 3. Obtener info de los pickings (recepciones)
pick_ids = list(set(ml['picking_id'][0] for ml in all_move_lines if ml.get('picking_id')))
print(f"Pickings únicos: {len(pick_ids)}")

pick_info = {}
for i in range(0, len(pick_ids), batch_size):
    batch = pick_ids[i:i+batch_size]
    picks = search_read('stock.picking', [('id', 'in', batch)], ['id', 'name', 'picking_type_id', 'date_done', 'origin'])
    for p in picks:
        pick_info[p['id']] = p

# 4. Agregar datos por tipo de fruta
kg_por_fruta = defaultdict(float)
kg_por_recepcion = defaultdict(float)
detalle = []  # Para hoja de detalle

for ml in all_move_lines:
    product = ml['product_id'][1] if ml['product_id'] else 'Desconocido'
    qty = ml['qty_done'] or 0
    pkg_name = pkg_names.get(ml['result_package_id'][0], '') if ml['result_package_id'] else ''
    pick_id = ml['picking_id'][0] if ml['picking_id'] else None
    pick = pick_info.get(pick_id, {})
    pick_name = pick.get('name', '')
    pick_date = pick.get('date_done', '')
    
    kg_por_fruta[product] += qty
    kg_por_recepcion[pick_name] += qty
    detalle.append({
        'paquete': pkg_name,
        'recepcion': pick_name,
        'fecha': str(pick_date)[:10] if pick_date else '',
        'producto': product,
        'kg': qty,
    })

total_kg = sum(kg_por_fruta.values())

# 5. Crear Excel
print("\nGenerando Excel...")
wb = openpyxl.Workbook()

# Estilos
header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
data_font = Font(size=10, name="Calibri")
total_font = Font(bold=True, size=11, name="Calibri")
total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
grand_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
grand_font = Font(bold=True, size=12, name="Calibri", color="FFFFFF")
alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
thin_border = Border(
    left=Side(style='thin', color='B4C6E7'), right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'), bottom=Side(style='thin', color='B4C6E7')
)

# ==================== HOJA 1: Resumen por Tipo de Fruta ====================
ws1 = wb.active
ws1.title = "KG por Tipo de Fruta"

ws1.merge_cells('A1:C1')
t = ws1.cell(row=1, column=1, value="KG Totales por Tipo de Fruta - Pallets RF2026M (Congelados)")
t.font = Font(bold=True, size=14, name="Calibri", color="1F4E79")
t.alignment = Alignment(horizontal='center')
ws1.row_dimensions[1].height = 30

ws1.merge_cells('A2:C2')
ws1.cell(row=2, column=1, value=f"Pallets: {len(all_packages)} | Recepciones: {len(pick_ids)} | Total: {total_kg:,.2f} kg")
ws1['A2'].font = Font(size=10, italic=True, color="666666")
ws1['A2'].alignment = Alignment(horizontal='center')

headers1 = ["Tipo de Fruta", "KG Totales", "% del Total"]
for col, h in enumerate(headers1, 1):
    c = ws1.cell(row=3, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal='center'); c.border = thin_border

row = 4
for idx, (fruta, kg) in enumerate(sorted(kg_por_fruta.items(), key=lambda x: x[1], reverse=True)):
    fill = alt_fill if idx % 2 == 0 else PatternFill(fill_type=None)
    c1 = ws1.cell(row=row, column=1, value=fruta); c1.font = data_font; c1.border = thin_border; c1.fill = fill
    c2 = ws1.cell(row=row, column=2, value=round(kg, 2)); c2.number_format = '#,##0.00'; c2.font = data_font; c2.border = thin_border; c2.alignment = Alignment(horizontal='right'); c2.fill = fill
    c3 = ws1.cell(row=row, column=3, value=round(kg/total_kg*100, 1) if total_kg else 0); c3.number_format = '0.0"%"'; c3.font = data_font; c3.border = thin_border; c3.alignment = Alignment(horizontal='right'); c3.fill = fill
    row += 1

# Total
for col in range(1, 4):
    ws1.cell(row=row, column=col).fill = total_fill; ws1.cell(row=row, column=col).font = total_font; ws1.cell(row=row, column=col).border = thin_border
ws1.cell(row=row, column=1, value="TOTAL CONGELADOS")
c = ws1.cell(row=row, column=2, value=round(total_kg, 2)); c.number_format = '#,##0.00'; c.font = grand_font; c.fill = grand_fill; c.alignment = Alignment(horizontal='right')
ws1.cell(row=row, column=3, value=100.0).number_format = '0.0"%"'

ws1.column_dimensions['A'].width = 45
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 14
ws1.freeze_panes = 'A4'
ws1.auto_filter.ref = f"A3:C{row}"

# ==================== HOJA 2: Resumen por Recepción ====================
ws2 = wb.create_sheet("KG por Recepción")

headers2 = ["Recepción", "Fecha", "KG Totales"]
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal='center'); c.border = thin_border

row2 = 2
recep_data = []
for pick_name, kg in sorted(kg_por_recepcion.items(), key=lambda x: x[1], reverse=True):
    pi = [v for v in pick_info.values() if v['name'] == pick_name]
    fecha = str(pi[0].get('date_done', ''))[:10] if pi else ''
    recep_data.append((pick_name, fecha, kg))

for idx, (pick_name, fecha, kg) in enumerate(recep_data):
    fill = alt_fill if idx % 2 == 0 else PatternFill(fill_type=None)
    ws2.cell(row=row2, column=1, value=pick_name).font = data_font; ws2.cell(row=row2, column=1).border = thin_border; ws2.cell(row=row2, column=1).fill = fill
    ws2.cell(row=row2, column=2, value=fecha).font = data_font; ws2.cell(row=row2, column=2).border = thin_border; ws2.cell(row=row2, column=2).fill = fill
    c = ws2.cell(row=row2, column=3, value=round(kg, 2)); c.number_format = '#,##0.00'; c.font = data_font; c.border = thin_border; c.alignment = Alignment(horizontal='right'); c.fill = fill
    row2 += 1

for col in range(1, 4):
    ws2.cell(row=row2, column=col).fill = total_fill; ws2.cell(row=row2, column=col).font = total_font; ws2.cell(row=row2, column=col).border = thin_border
ws2.cell(row=row2, column=1, value="TOTAL")
c = ws2.cell(row=row2, column=3, value=round(total_kg, 2)); c.number_format = '#,##0.00'; c.font = grand_font; c.fill = grand_fill; c.alignment = Alignment(horizontal='right')

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 18
ws2.freeze_panes = 'A2'

# ==================== HOJA 3: Detalle por Pallet ====================
ws3 = wb.create_sheet("Detalle por Pallet")

headers3 = ["Pallet", "Recepción", "Fecha", "Producto", "KG"]
for col, h in enumerate(headers3, 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal='center'); c.border = thin_border

detalle_sorted = sorted(detalle, key=lambda x: x['paquete'])
row3 = 2
for idx, d in enumerate(detalle_sorted):
    fill = alt_fill if idx % 2 == 0 else PatternFill(fill_type=None)
    for col, key in enumerate(['paquete', 'recepcion', 'fecha', 'producto', 'kg'], 1):
        c = ws3.cell(row=row3, column=col, value=d[key])
        c.font = data_font; c.border = thin_border; c.fill = fill
        if key == 'kg':
            c.number_format = '#,##0.00'
            c.alignment = Alignment(horizontal='right')
    row3 += 1

for col in range(1, 6):
    ws3.cell(row=row3, column=col).fill = total_fill; ws3.cell(row=row3, column=col).font = total_font; ws3.cell(row=row3, column=col).border = thin_border
ws3.cell(row=row3, column=1, value="TOTAL")
c = ws3.cell(row=row3, column=5, value=round(total_kg, 2)); c.number_format = '#,##0.00'; c.font = grand_font; c.fill = grand_fill; c.alignment = Alignment(horizontal='right')

ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 35
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 45
ws3.column_dimensions['E'].width = 15
ws3.freeze_panes = 'A2'
ws3.auto_filter.ref = f"A1:E{row3}"

# Guardar
output_file = "KG_Congelados_RF2026M.xlsx"
wb.save(output_file)

print(f"\n{'='*60}")
print(f"  Excel generado: {output_file}")
print(f"  Pallets RF2026M:     {len(all_packages)}")
print(f"  Recepciones:         {len(pick_ids)}")
print(f"  Tipos de fruta:      {len(kg_por_fruta)}")
print(f"  TOTAL KG CONGELADOS: {total_kg:,.2f}")
print(f"{'='*60}")

print("\nDesglose por tipo de fruta:")
for fruta, kg in sorted(kg_por_fruta.items(), key=lambda x: x[1], reverse=True):
    print(f"  {fruta}: {kg:,.2f} kg ({kg/total_kg*100:.1f}%)")
