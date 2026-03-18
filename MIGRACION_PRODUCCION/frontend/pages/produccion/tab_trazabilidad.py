"""
Tab Trazabilidad de Pallets - trazado automatico completo.

Soporta:
  - Busqueda por Pallet o por Lote
  - Trazabilidad hacia atras (-> Recepcion/Origen)
  - Trazabilidad hacia adelante (-> Destino/Consumo/Despacho)
"""
import streamlit as st
import io
from datetime import datetime
from typing import Dict, List, Optional
import httpx
from .shared import API_URL

TREE_KEY = "traz_tree_v2"


# =============================================================
# API
# =============================================================

def _full_trace(package_name: str, username: str, password: str,
                direction: str = "backward") -> Dict:
    """Llama al endpoint que traza un pallet en la direccion indicada."""
    r = httpx.get(
        f"{API_URL}/api/v1/etiquetas/full_trace",
        params={
            "package_name": package_name,
            "username": username,
            "password": password,
            "max_levels": 12,
            "direction": direction,
        },
        timeout=120.0,
    )
    r.raise_for_status()
    return r.json()


def _trace_lot(lot_name: str, username: str, password: str,
               direction: str = "backward") -> Dict:
    """Llama al endpoint que traza un lote completo."""
    r = httpx.get(
        f"{API_URL}/api/v1/etiquetas/trace_lot",
        params={
            "lot_name": lot_name,
            "username": username,
            "password": password,
            "max_levels": 12,
            "direction": direction,
        },
        timeout=180.0,
    )
    r.raise_for_status()
    return r.json()


# =============================================================
# Helpers
# =============================================================

def _short(name: str) -> str:
    """PACK0048229 -> 048229"""
    if not name:
        return ""
    n = name.strip()
    if n.upper().startswith("PACK"):
        return n[4:].strip() or n
    return n


def _build_children_map(nodes: List[Dict]) -> Dict[int, List[Dict]]:
    """Agrupa nodos por parent_node_id -> hijos."""
    children: Dict[int, List[Dict]] = {}
    for n in nodes:
        pid = n.get("parent_node_id")
        if pid is not None:
            children.setdefault(pid, []).append(n)
    return children


def _build_chain(node: Dict, idx: Dict[int, Dict]) -> List[str]:
    """Construye cadena de nombres desde root hasta el nodo."""
    chain = [_short(node["pkg_name"])]
    cur = node
    while cur.get("parent_node_id") and cur["parent_node_id"] in idx:
        cur = idx[cur["parent_node_id"]]
        chain.append(_short(cur["pkg_name"]))
    chain.reverse()
    return chain


# =============================================================
# Render principal
# =============================================================

def render(username: str, password: str):
    st.subheader("\U0001f50d Trazabilidad de Pallets")
    st.caption(
        "Ingresa un pallet o un lote y el sistema traza automaticamente toda la cadena. "
        "Puedes trazar hacia atras (hasta recepcion) o hacia adelante (hasta destino/despacho)."
    )

    # Estado
    if TREE_KEY not in st.session_state:
        st.session_state[TREE_KEY] = None

    # Limpiar estado antiguo (v1)
    for old_key in ["traz_tree", "traz_acumulado"]:
        if old_key in st.session_state:
            del st.session_state[old_key]

    data = st.session_state[TREE_KEY]

    # -- Opciones de busqueda --
    col_mode, col_dir = st.columns(2)
    with col_mode:
        modo = st.radio(
            "Buscar por",
            ["Pallet", "Lote"],
            horizontal=True,
            key="traz_modo",
        )
    with col_dir:
        direccion = st.radio(
            "Direccion",
            ["\u2b05\ufe0f Hacia atras (\u2192 Recepcion)", "\u27a1\ufe0f Hacia adelante (\u2192 Destino)"],
            horizontal=True,
            key="traz_dir",
        )

    dir_value = "backward" if "atras" in direccion else "forward"

    # -- Input --
    col_in, col_btn, col_rst = st.columns([3, 1.2, 0.8])
    with col_in:
        if modo == "Pallet":
            search_term = st.text_input(
                "Pallet",
                placeholder="Ej: PACK0012345",
                key="traz_input_v2",
                label_visibility="collapsed",
            )
        else:
            search_term = st.text_input(
                "Lote",
                placeholder="Ej: LOT-2025-001 o nombre de lote",
                key="traz_input_lote",
                label_visibility="collapsed",
            )
    with col_btn:
        trazar = st.button("\U0001f680 Trazar", type="primary", use_container_width=True)
    with col_rst:
        if data and st.button("\U0001f504 Limpiar", use_container_width=True):
            st.session_state[TREE_KEY] = None
            st.rerun()

    # -- Trazar --
    if trazar:
        if not search_term or not search_term.strip():
            st.warning(f"Ingresa un nombre de {'pallet' if modo == 'Pallet' else 'lote'}.")
            return

        nombre = search_term.strip().upper()
        dir_label = "hacia recepcion" if dir_value == "backward" else "hacia destino"

        try:
            if modo == "Pallet":
                with st.spinner(f"Trazando **{nombre}** {dir_label}..."):
                    result = _full_trace(nombre, username, password, direction=dir_value)
            else:
                with st.spinner(f"Trazando lote **{nombre}** {dir_label}..."):
                    result = _trace_lot(nombre, username, password, direction=dir_value)

            st.session_state[TREE_KEY] = result
            data = result
        except httpx.HTTPStatusError as e:
            if e.response.status_code == 404:
                if modo == "Pallet":
                    st.error(f"No se encontro el pallet **{nombre}** en Odoo.")
                else:
                    st.error(f"No se encontraron pallets para el lote **{nombre}**.")
            else:
                st.error(f"Error del servidor: {e.response.text}")
            return
        except Exception as e:
            st.error(f"Error de conexion: {e}")
            return

    if data is None:
        st.info(
            "Ingresa un pallet o lote y presiona **Trazar** para ver la cadena completa "
            "de trazabilidad."
        )
        return

    # Resultado
    _render_result(data)


# =============================================================
# Render resultado
# =============================================================

def _render_result(data: Dict):
    """Muestra el resultado del trazado completo."""
    nodes = data.get("nodes", [])
    root_name = data.get("root_name", "?")
    num_levels = data.get("levels", 0)
    leaf_count = data.get("leaf_count", 0)
    rec_count = data.get("reception_count", 0)
    dispatch_count = data.get("dispatch_count", 0)
    direction = data.get("direction", "backward")
    lot_name = data.get("lot_name")
    pallets_found = data.get("pallets_found")

    if not nodes:
        st.warning("No se encontraron datos.")
        return

    st.divider()

    # -- Info de direccion --
    if direction == "forward":
        st.markdown("**Direccion:** \u27a1\ufe0f Hacia adelante (Origen \u2192 Destino/Despacho)")
    else:
        st.markdown("**Direccion:** \u2b05\ufe0f Hacia atras (Destino \u2192 Recepcion/Origen)")

    # -- Metricas --
    if lot_name:
        c0, c1, c2, c3, c4 = st.columns(5)
        c0.metric("Lote", lot_name)
        c1.metric("Pallets encontrados", pallets_found or 0)
        c2.metric("Niveles", num_levels)
        if direction == "forward":
            c3.metric("Despachos", dispatch_count)
        else:
            c3.metric("Recepciones", rec_count)
        c4.metric("Total nodos", len(nodes))
    else:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Pallet", root_name)
        c2.metric("Niveles", num_levels)
        if direction == "forward":
            c3.metric("Despachos", dispatch_count)
        else:
            c3.metric("Recepciones", rec_count)
        c4.metric("Total pallets", len(nodes))

    # -- Status messages --
    if direction == "backward":
        if rec_count > 0 and rec_count == leaf_count:
            st.success("\u2705 Trazabilidad completa \u2014 todos los caminos llegan a recepcion.")
        elif rec_count > 0:
            st.info(f"\u2139\ufe0f {rec_count} de {leaf_count} hojas son recepciones.")
    else:
        if dispatch_count > 0:
            st.success(f"\u2705 {dispatch_count} camino(s) llegan a despacho.")
        elif leaf_count > 0:
            st.info(f"\u2139\ufe0f {leaf_count} hojas encontradas (sin despacho asociado).")

    # -- Arbol visual --
    st.markdown("### \U0001f333 Arbol de Trazabilidad")

    idx = {n["node_id"]: n for n in nodes}
    children_map = _build_children_map(nodes)
    root = next((n for n in nodes if n.get("parent_node_id") is None), None)

    # Set para rastrear recepciones/despachos ya mostrados (guia+productor / cliente+picking)
    seen_set: set = set()

    if root:
        _render_tree_node(root, idx, children_map, depth=0, direction=direction, seen=seen_set)

    # -- Tablas de resultados --
    if direction == "backward":
        _render_recepciones_table(nodes, idx)
    else:
        _render_despachos_table(nodes, idx)

    # -- Excel --
    st.divider()
    excel = _generar_excel(data)
    dir_suffix = "adelante" if direction == "forward" else "atras"
    base_name = root_name.replace(":", "").replace(" ", "_")
    fname = f"Trazabilidad_{base_name}_{dir_suffix}_{datetime.now():%Y%m%d_%H%M}.xlsx"
    st.download_button(
        "\u2b07\ufe0f Descargar Excel",
        data=excel,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )


def _render_recepciones_table(nodes: List[Dict], idx: Dict):
    """Tabla de recepciones encontradas (backward), deduplicada por guia+productor."""
    receptions = [n for n in nodes if n.get("reception_info")]
    if not receptions:
        return
    st.markdown("### \U0001f4cb Recepciones Encontradas")
    seen: Dict[tuple, dict] = {}  # (guia, productor) -> row dict
    ordered_keys: List[tuple] = []
    for n in receptions:
        rec = n["reception_info"]
        guia = rec.get("guia_despacho", "") or "\u2014"
        prov = rec.get("proveedor", "") or "\u2014"
        key = (guia, prov)
        chain = _build_chain(n, idx)
        if key in seen:
            # Agregar el pallet a la fila existente
            prev = seen[key]
            prev["Pallet Origen"] += f", {n['pkg_name']}"
        else:
            seen[key] = {
                "Pallet Origen": n["pkg_name"],
                "Cadena": " \u2192 ".join(chain),
                "Guia Despacho": guia,
                "Productor": prov,
                "Picking": rec.get("picking_name", "") or "\u2014",
                "Fecha": rec.get("fecha", "") or "\u2014",
            }
            ordered_keys.append(key)
    rows = [seen[k] for k in ordered_keys]
    st.dataframe(rows, use_container_width=True, hide_index=True)


def _render_despachos_table(nodes: List[Dict], idx: Dict):
    """Tabla de despachos encontrados (forward), deduplicada por cliente+picking."""
    dispatches = [n for n in nodes if n.get("dispatch_info")]
    if dispatches:
        st.markdown("### \U0001f69a Despachos Encontrados")
        seen: Dict[tuple, dict] = {}
        ordered_keys: List[tuple] = []
        for n in dispatches:
            disp = n["dispatch_info"]
            cliente = disp.get("cliente", "") or "\u2014"
            picking = disp.get("picking_name", "") or "\u2014"
            key = (cliente, picking)
            chain = _build_chain(n, idx)
            if key in seen:
                seen[key]["Pallet Destino"] += f", {n['pkg_name']}"
            else:
                seen[key] = {
                    "Pallet Destino": n["pkg_name"],
                    "Cadena": " \u2192 ".join(chain),
                    "Cliente": cliente,
                    "Picking": picking,
                    "Origen": disp.get("origin", "") or "\u2014",
                    "Fecha": disp.get("fecha", "") or "\u2014",
                }
                ordered_keys.append(key)
        rows = [seen[k] for k in ordered_keys]
        st.dataframe(rows, use_container_width=True, hide_index=True)

    # Mostrar tambien hojas sin despacho como destinos finales
    leaves_no_dispatch = [n for n in nodes if n.get("is_leaf") and not n.get("dispatch_info") and n.get("level", 0) > 0]
    if leaves_no_dispatch:
        st.markdown("### \U0001f4e6 Destinos Finales (sin despacho)")
        rows = []
        for n in leaves_no_dispatch:
            chain = _build_chain(n, idx)
            rows.append({
                "Pallet": n["pkg_name"],
                "Cadena": " \u2192 ".join(chain),
                "OP": n.get("mo_name", "") or "\u2014",
                "Producto": n.get("product_name", "") or "\u2014",
                "Cantidad (kg)": f"{n.get('qty', 0):.1f}" if n.get("qty") else "\u2014",
            })
        st.dataframe(rows, use_container_width=True, hide_index=True)


# =============================================================
# Arbol visual recursivo
# =============================================================

def _render_tree_node(node: Dict, idx: Dict, children_map: Dict,
                      depth: int, direction: str = "backward",
                      seen: set = None):
    """Renderiza un nodo y sus hijos recursivamente como arbol con indentacion.
    Deduplica recepciones/despachos con misma guia+productor o cliente+picking."""
    if seen is None:
        seen = set()
    pad = "&nbsp;&nbsp;&nbsp;&nbsp;" * depth
    nid = node["node_id"]
    name = node["pkg_name"]
    rec = node.get("reception_info")
    disp = node.get("dispatch_info")
    mo = node.get("mo_name")
    kids = children_map.get(nid, [])

    if depth == 0:
        # Raiz
        if name.startswith("LOTE:"):
            st.markdown(f"\U0001f3f7\ufe0f **{name}**")
        elif direction == "forward":
            label = "(origen)" if rec else "(punto de partida)"
            extra = ""
            if rec:
                guia = rec.get("guia_despacho", "") or ""
                prov = rec.get("proveedor", "") or ""
                if guia or prov:
                    extra = f" \u2014 Guia: **{guia}** \u00b7 Productor: **{prov}**"
            st.markdown(f"\U0001f4e6 **`{name}`** {label}{extra}")
        else:
            st.markdown(f"\U0001f4e6 **`{name}`** (destino final)")
    elif rec and direction == "backward":
        guia = rec.get("guia_despacho", "") or "\u2014"
        prov = rec.get("proveedor", "") or "\u2014"
        fecha = rec.get("fecha", "") or ""
        dedup_key = (guia, prov)
        if dedup_key in seen:
            # Ya se mostro esta guia+productor, solo poner referencia corta
            st.markdown(
                f"{pad}\U0001f7e2 `{name}` \u2014 *(Guia {guia} \u00b7 {prov} \u2014 ver arriba)*"
            )
        else:
            seen.add(dedup_key)
            st.markdown(
                f"{pad}\U0001f7e2 **`{name}`** \u2014 Guia: **{guia}** \u00b7 Productor: **{prov}** \u00b7 {fecha}"
            )
    elif disp and direction == "forward":
        cliente = disp.get("cliente", "") or "\u2014"
        fecha = disp.get("fecha", "") or ""
        picking = disp.get("picking_name", "") or ""
        dedup_key = (cliente, picking)
        if dedup_key in seen:
            st.markdown(
                f"{pad}\U0001f535 `{name}` \u2014 *(Cliente {cliente} \u00b7 {picking} \u2014 ver arriba)*"
            )
        else:
            seen.add(dedup_key)
            st.markdown(
                f"{pad}\U0001f535 **`{name}`** \u2014 Cliente: **{cliente}** \u00b7 {picking} \u00b7 {fecha}"
            )
    elif node.get("is_leaf"):
        # Hoja sin recepcion/despacho
        sin_label = "sin origen" if direction == "backward" else "sin destino"
        st.markdown(f"{pad}\U0001f7e1 `{name}` \u2014 {sin_label} conocido")
    else:
        # Nodo intermedio
        mo_label = f" *(OP: {mo})*" if mo else ""
        st.markdown(f"{pad}\U0001f539 **`{name}`**{mo_label}")

    # Hijos
    for child in kids:
        _render_tree_node(child, idx, children_map, depth + 1, direction=direction, seen=seen)


# =============================================================
# Excel -- 3 hojas
# =============================================================

def _generar_excel(data: Dict) -> bytes:
    """
    Excel con 3 hojas:
      1) Trazabilidad -- una fila por recepcion/despacho con cadena completa
      2) Arbol Completo -- todos los nodos
      3) Arbol Visual -- indentado
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    nodes = data.get("nodes", [])
    root_name = data.get("root_name", "?")
    direction = data.get("direction", "backward")
    idx = {n["node_id"]: n for n in nodes}

    wb = Workbook()

    # Estilos
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al = Alignment(vertical="center", wrap_text=True)
    rec_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    root_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    disp_fill = PatternFill(start_color="BDE0FE", end_color="BDE0FE", fill_type="solid")
    hdr2_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

    def _write_header(ws, headers, fill=None):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = fill or hdr_fill
            cell.font = hdr_font
            cell.alignment = center
            cell.border = border

    def _set_widths(ws, widths):
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

    # ===========================================================
    # HOJA 1 -- Trazabilidad
    # ===========================================================
    ws1 = wb.active
    ws1.title = "Trazabilidad"

    if direction == "forward":
        headers1 = [
            "Pack (Origen)", "Pallet Destino",
            "Cadena Completa", "Cliente", "Picking Despacho", "Fecha Despacho",
        ]
        _write_header(ws1, headers1)
        _set_widths(ws1, [20, 20, 55, 35, 20, 14])

        dispatches = [n for n in nodes if n.get("dispatch_info")]
        if not dispatches:
            dispatches = [n for n in nodes if n.get("is_leaf") and n.get("level", 0) > 0]

        row = 2
        for n in dispatches:
            disp = n.get("dispatch_info") or {}
            chain = _build_chain(n, idx)
            chain_str = " \u2192 ".join(chain)
            ws1.cell(row=row, column=1, value=root_name)
            ws1.cell(row=row, column=2, value=n["pkg_name"])
            ws1.cell(row=row, column=3, value=chain_str)
            ws1.cell(row=row, column=4, value=disp.get("cliente", ""))
            ws1.cell(row=row, column=5, value=disp.get("picking_name", ""))
            ws1.cell(row=row, column=6, value=disp.get("fecha", ""))
            for c in range(1, 7):
                cell = ws1.cell(row=row, column=c)
                cell.border = border
                cell.alignment = left_al
                cell.fill = disp_fill if disp else rec_fill
            row += 1
    else:
        headers1 = [
            "Pack (Destino)", "Pallet Origen (Recepcion)",
            "Cadena Completa", "Guia Despacho", "Productor",
            "Picking", "Fecha Recepcion",
        ]
        _write_header(ws1, headers1)
        _set_widths(ws1, [20, 20, 55, 20, 35, 20, 14])

        receptions = [n for n in nodes if n.get("reception_info")]
        # Deduplicar por (guia, productor)
        seen_excel: dict = {}  # (guia, prov) -> {pallets: [], first_node}
        excel_order: list = []
        for n in receptions:
            rec = n["reception_info"]
            guia = rec.get("guia_despacho", "")
            prov = rec.get("proveedor", "")
            key = (guia, prov)
            if key in seen_excel:
                seen_excel[key]["pallets"].append(n["pkg_name"])
            else:
                seen_excel[key] = {"pallets": [n["pkg_name"]], "node": n, "rec": rec}
                excel_order.append(key)

        row = 2
        for key in excel_order:
            entry = seen_excel[key]
            rec = entry["rec"]
            n = entry["node"]
            pallets_str = ", ".join(entry["pallets"])
            chain = _build_chain(n, idx)
            chain_str = " \u2192 ".join(chain)
            ws1.cell(row=row, column=1, value=root_name)
            ws1.cell(row=row, column=2, value=pallets_str)
            ws1.cell(row=row, column=3, value=chain_str)
            ws1.cell(row=row, column=4, value=rec.get("guia_despacho", ""))
            ws1.cell(row=row, column=5, value=rec.get("proveedor", ""))
            ws1.cell(row=row, column=6, value=rec.get("picking_name", ""))
            ws1.cell(row=row, column=7, value=rec.get("fecha", ""))
            for c in range(1, 8):
                cell = ws1.cell(row=row, column=c)
                cell.border = border
                cell.alignment = left_al
                cell.fill = rec_fill
            row += 1

        if not receptions:
            ws1.cell(row=2, column=1, value=root_name)
            ws1.cell(row=2, column=2, value="(sin recepciones encontradas)")
            for c in range(1, 8):
                ws1.cell(row=2, column=c).border = border

    ws1.freeze_panes = "A2"

    # ===========================================================
    # HOJA 2 -- Arbol Completo
    # ===========================================================
    ws2 = wb.create_sheet("Arbol Completo")

    headers2 = [
        "Nivel", "Pallet", "Padre", "OP",
        "Producto", "Cantidad (kg)", "Lote", "Tipo",
        "Guia Despacho", "Productor/Cliente", "Picking", "Fecha",
    ]
    _write_header(ws2, headers2, fill=hdr2_fill)
    _set_widths(ws2, [8, 20, 20, 22, 45, 14, 20, 14, 20, 35, 20, 14])

    sorted_nodes = sorted(nodes, key=lambda n: (n.get("level", 0), n.get("pkg_name", "")))

    row2 = 2
    for n in sorted_nodes:
        level = n.get("level", 0)
        rec = n.get("reception_info")
        disp = n.get("dispatch_info")
        is_leaf = n.get("is_leaf", False)

        if level == 0:
            tipo = "Origen" if direction == "forward" else "Destino"
        elif rec:
            tipo = "Recepcion"
        elif disp:
            tipo = "Despacho"
        elif is_leaf:
            tipo = "Hoja"
        else:
            tipo = "Intermedio"

        parent = idx.get(n.get("parent_node_id"))
        parent_name = parent["pkg_name"] if parent else ""

        ws2.cell(row=row2, column=1, value=level)
        ws2.cell(row=row2, column=2, value=n.get("pkg_name", "")).font = Font(bold=True)
        ws2.cell(row=row2, column=3, value=parent_name)
        ws2.cell(row=row2, column=4, value=n.get("mo_name", "") or "")
        ws2.cell(row=row2, column=5, value=n.get("product_name", "") or "")
        qty = n.get("qty")
        ws2.cell(row=row2, column=6, value=f"{qty:.1f}" if qty else "")
        ws2.cell(row=row2, column=7, value=n.get("lot_name", "") or "")
        ws2.cell(row=row2, column=8, value=tipo)

        if rec:
            ws2.cell(row=row2, column=9, value=rec.get("guia_despacho", ""))
            ws2.cell(row=row2, column=10, value=rec.get("proveedor", ""))
            ws2.cell(row=row2, column=11, value=rec.get("picking_name", ""))
            ws2.cell(row=row2, column=12, value=rec.get("fecha", ""))
        elif disp:
            ws2.cell(row=row2, column=9, value="")
            ws2.cell(row=row2, column=10, value=disp.get("cliente", ""))
            ws2.cell(row=row2, column=11, value=disp.get("picking_name", ""))
            ws2.cell(row=row2, column=12, value=disp.get("fecha", ""))
        else:
            for c in range(9, 13):
                ws2.cell(row=row2, column=c, value="")

        for c in range(1, 13):
            cell = ws2.cell(row=row2, column=c)
            cell.border = border
            cell.alignment = center if c in (1, 6, 8) else left_al

        if level == 0:
            for c in range(1, 13):
                ws2.cell(row=row2, column=c).fill = root_fill
        elif rec:
            for c in range(1, 13):
                ws2.cell(row=row2, column=c).fill = rec_fill
        elif disp:
            for c in range(1, 13):
                ws2.cell(row=row2, column=c).fill = disp_fill

        row2 += 1

    ws2.freeze_panes = "A2"

    # ===========================================================
    # HOJA 3 -- Arbol Visual
    # ===========================================================
    ws3 = wb.create_sheet("Arbol Visual")

    tree_hdr_fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    indent_font = Font(name="Consolas", size=11)
    indent_bold = Font(name="Consolas", size=11, bold=True)
    rec_font = Font(name="Consolas", size=11, color="006100", bold=True)
    disp_font = Font(name="Consolas", size=11, color="1F4E79", bold=True)
    leaf_font = Font(name="Consolas", size=11, color="BF8F00")
    root_font = Font(name="Consolas", size=12, bold=True, color="1F4E79")

    headers3 = ["Arbol de Pallets", "OP / Proceso", "Guia/Cliente", "Productor/Destino", "Producto", "Cantidad (kg)"]
    _write_header(ws3, headers3, fill=tree_hdr_fill)
    _set_widths(ws3, [65, 24, 18, 35, 45, 14])

    children_map = _build_children_map(nodes)
    root_node = next((n for n in nodes if n.get("parent_node_id") is None), None)

    tree_row = [2]  # mutable counter

    def _write_tree_row(node, depth):
        r = tree_row[0]
        tree_row[0] += 1

        rec_info = node.get("reception_info")
        disp_info = node.get("dispatch_info")
        is_leaf = node.get("is_leaf", False)
        nid = node["node_id"]
        name = node["pkg_name"]
        mo = node.get("mo_name") or ""
        kids = children_map.get(nid, [])

        if depth == 0:
            if name.startswith("LOTE:"):
                prefix = "\U0001f3f7 "
                label = f"{prefix}{name}"
            else:
                prefix = "\u25c6 "
                dir_label = "ORIGEN" if direction == "forward" else "DESTINO"
                label = f"{prefix}{name}  ({dir_label})"
            font_choice = root_font
            fill_choice = root_fill
        elif rec_info and direction == "backward":
            prefix = "    " * depth + "\u2514\u2500 \u2705 "
            label = f"{prefix}{name}  [RECEPCION]"
            font_choice = rec_font
            fill_choice = rec_fill
        elif disp_info and direction == "forward":
            prefix = "    " * depth + "\u2514\u2500 \U0001f69a "
            label = f"{prefix}{name}  [DESPACHO]"
            font_choice = disp_font
            fill_choice = disp_fill
        elif is_leaf:
            prefix = "    " * depth + "\u2514\u2500 \u26a0 "
            sin = "SIN ORIGEN" if direction == "backward" else "SIN DESTINO"
            label = f"{prefix}{name}  [{sin}]"
            font_choice = leaf_font
            fill_choice = None
        else:
            connector = "\u251c\u2500 "
            prefix = "    " * depth + connector
            label = f"{prefix}{name}"
            font_choice = indent_bold
            fill_choice = None

        cell_tree = ws3.cell(row=r, column=1, value=label)
        cell_tree.font = font_choice
        cell_tree.alignment = Alignment(vertical="center")
        cell_tree.border = border
        if fill_choice:
            cell_tree.fill = fill_choice

        ws3.cell(row=r, column=2, value=mo).border = border
        ws3.cell(row=r, column=2).alignment = left_al

        if rec_info:
            ws3.cell(row=r, column=3, value=rec_info.get("guia_despacho", "")).border = border
            ws3.cell(row=r, column=4, value=rec_info.get("proveedor", "")).border = border
        elif disp_info:
            ws3.cell(row=r, column=3, value=disp_info.get("cliente", "")).border = border
            ws3.cell(row=r, column=4, value=disp_info.get("picking_name", "")).border = border
        else:
            ws3.cell(row=r, column=3).border = border
            ws3.cell(row=r, column=4).border = border
        ws3.cell(row=r, column=3).alignment = center
        ws3.cell(row=r, column=4).alignment = left_al

        ws3.cell(row=r, column=5, value=node.get("product_name", "") or "").border = border
        ws3.cell(row=r, column=5).alignment = left_al

        qty = node.get("qty")
        ws3.cell(row=r, column=6, value=f"{qty:.1f}" if qty else "").border = border
        ws3.cell(row=r, column=6).alignment = center

        for child in kids:
            _write_tree_row(child, depth + 1)

    if root_node:
        _write_tree_row(root_node, 0)

    ws3.freeze_panes = "A2"

    # Guardar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
