"""
Tab Trazabilidad Interactiva de Pallets – paso a paso.

Flujo:
  1) El usuario ingresa un pallet destino final
  2) El sistema busca los paquetes origen que lo conformaron (candidatos)
  3) El usuario selecciona cuáles seguir trazando hacia atrás
  4) Se repite hasta llegar a paquetes de recepción (sin orígenes)
  5) Se genera Excel con la cadena completa de trazabilidad

Cada paquete destino se compone de paquetes origen.
Esos paquetes origen, en un proceso anterior, fueron paquetes destino
compuestos por otros paquetes origen, y así sucesivamente hasta
llegar a la recepción.
"""
import streamlit as st
import pandas as pd
import io
from datetime import datetime
from typing import Dict, List, Optional
import httpx
from .shared import API_URL

TREE_KEY = "traz_tree"
OLD_SESSION_KEY = "traz_acumulado"  # clave antigua, limpiar si existe


# ═════════════════════════════════════════════════════════════
# Helpers API
# ═════════════════════════════════════════════════════════════

def _api_get(url: str, params: dict, timeout: float = 60.0):
    """GET al backend con timeout."""
    r = httpx.get(url, params=params, timeout=timeout)
    r.raise_for_status()
    return r.json()


def _find_package(name: str, username: str, password: str) -> Optional[Dict]:
    """Busca un package por nombre exacto/similar. Devuelve {id, name} o None."""
    try:
        res = _api_get(
            f"{API_URL}/api/v1/etiquetas/find_package",
            {"package_name": name, "username": username, "password": password},
        )
        if not res:
            return None
        # preferir match exacto
        for r in res:
            if (r.get("name") or "").upper() == name.upper():
                return r
        return res[0]
    except Exception:
        return None


def _get_candidates(pkg_id: int, username: str, password: str,
                    product_name=None, manejo=None, variedad=None) -> Dict:
    """
    Obtiene candidatos previos (paquetes origen) para un paquete destino.
    Retorna dict con: candidates, has_mo, mo_name, is_reception, reception.
    """
    params: dict = {"package_id": pkg_id, "username": username, "password": password}
    if product_name:
        params["product_name"] = product_name
    if manejo:
        params["manejo"] = manejo
    if variedad:
        params["variedad"] = variedad
    try:
        result = _api_get(f"{API_URL}/api/v1/etiquetas/prev_candidates", params)
        if isinstance(result, list):
            return {"candidates": result, "has_mo": False, "is_reception": False, "reception": None, "mo_name": None}
        return result or {"candidates": [], "has_mo": False, "is_reception": False, "reception": None, "mo_name": None}
    except Exception:
        return {"candidates": [], "has_mo": False, "is_reception": False, "reception": None, "mo_name": None}


def _get_package_info(pkg_id: int, username: str, password: str) -> Dict:
    """Obtiene info de etiqueta (producto, manejo, variedad, etc.)."""
    try:
        return _api_get(
            f"{API_URL}/api/v1/etiquetas/info_etiqueta/{pkg_id}",
            {"username": username, "password": password},
            timeout=30.0,
        ) or {}
    except Exception:
        return {}


def _new_node(pkg_id: int, pkg_name: str, **extra) -> Dict:
    """Crea un nodo de paquete en el árbol de trazabilidad."""
    node = {
        "node_id": None,       # ID único dentro del árbol (int auto-incremental)
        "pkg_id": pkg_id,
        "pkg_name": pkg_name,
        "candidates": None,    # None = no cargado, [] = sin orígenes
        "is_leaf": False,
        "parent_node_id": None, # node_id del padre en el árbol
        "parent_pkg_id": None,  # ID del pkt padre (informativo)
        "parent_pkg_name": None,
        "mo_name": None,        # OP que usó este pallet como insumo
        "has_mo": False,
        "is_reception": False,
        "reception_info": None, # {picking_name, guia_despacho, proveedor, fecha}
    }
    node.update(extra)
    return node


def _get_ancestor_ids(tree: Dict, node: Dict) -> set:
    """Devuelve el set de pkg_ids ancestros de un nodo específico (para detección de ciclos)."""
    idx: Dict[int, Dict] = {}
    for lvl in tree.get("levels", []):
        for p in lvl:
            nid = p.get("node_id")
            if nid is not None:
                idx[nid] = p

    ancestors = set()
    cur = node
    while cur.get("parent_node_id") and cur["parent_node_id"] in idx:
        parent = idx[cur["parent_node_id"]]
        ancestors.add(parent["pkg_id"])
        cur = parent
    return ancestors


def _short(name: str) -> str:
    """Acorta PACK0048229 → 048229 para cadenas compactas."""
    if not name:
        return ""
    n = name.strip()
    if n.upper().startswith("PACK"):
        rest = n[4:].strip()
        return rest if rest else n
    return n


def _build_chain_to_root(tree: Dict, leaf_node: Dict) -> List[str]:
    """
    Dada una hoja, recorre parent_node_id hacia arriba
    y devuelve la cadena [root, ..., leaf] con nombres cortos.
    """
    # Build index: node_id → node
    idx: Dict[int, Dict] = {}
    for lvl in tree.get("levels", []):
        for p in lvl:
            nid = p.get("node_id")
            if nid is not None:
                idx[nid] = p

    chain = [_short(leaf_node["pkg_name"])]
    cur = leaf_node
    while cur.get("parent_node_id") and cur["parent_node_id"] in idx:
        parent = idx[cur["parent_node_id"]]
        chain.append(_short(parent["pkg_name"]))
        cur = parent
    chain.reverse()
    return chain


def _render_leaf(p: Dict, indent: int = 1):
    """Renderiza un paquete hoja (recepción o sin orígenes)."""
    pad = "&nbsp;&nbsp;&nbsp;&nbsp;" * indent
    rec = p.get("reception_info")
    if rec:
        guia = rec.get("guia_despacho", "") or "—"
        prov = rec.get("proveedor", "") or "—"
        pick = rec.get("picking_name", "") or ""
        fecha = rec.get("fecha", "") or ""
        st.markdown(
            f"{pad}\U0001F7E2 **`{p['pkg_name']}`** — **RECEPCIÓN** "
            f"| Picking: **{pick}** | Guía: **{guia}** | Proveedor: **{prov}** | Fecha: {fecha}"
        )
    elif p.get("is_reception"):
        st.markdown(f"{pad}\U0001F7E2 **`{p['pkg_name']}`** — recepción (sin detalle adicional)")
    else:
        st.markdown(f"{pad}\U0001F7E1 **`{p['pkg_name']}`** — sin orígenes conocidos")


# ═════════════════════════════════════════════════════════════
# Render principal
# ═════════════════════════════════════════════════════════════

def render(username: str, password: str):
    st.subheader("\U0001F50D Trazabilidad Interactiva de Pallets")
    st.caption(
        "Traza pallets paso a paso hacia atrás. En cada nivel el sistema "
        "muestra qué pallets fueron consumidos por cada destino, con la OP que los unió. "
        "Selecciona cuáles seguir trazando hasta llegar a las recepciones."
    )

    # limpiar estado antiguo
    if OLD_SESSION_KEY in st.session_state:
        del st.session_state[OLD_SESSION_KEY]

    # Inicializar árbol
    if TREE_KEY not in st.session_state:
        st.session_state[TREE_KEY] = None

    tree = st.session_state[TREE_KEY]

    # ── Migrar árboles legacy sin node_id ────────────────────
    if tree is not None and "next_node_id" not in tree:
        nid = 1
        pkg_to_node: Dict[int, int] = {}
        for lvl in tree.get("levels", []):
            for p in lvl:
                p["node_id"] = nid
                pkg_to_node[p["pkg_id"]] = nid
                nid += 1
        tree["next_node_id"] = nid
        for lvl in tree.get("levels", []):
            for p in lvl:
                ppid = p.get("parent_pkg_id")
                if ppid and ppid in pkg_to_node:
                    p["parent_node_id"] = pkg_to_node[ppid]
        st.session_state[TREE_KEY] = tree

    # ── Fase 1: Ingresar pallet raíz ────────────────────────
    if tree is None:
        _render_input_root(username, password)
        return

    # ── Header ───────────────────────────────────────────────
    filters = tree.get("filters", {})
    col_hdr, col_reset = st.columns([4, 1])
    with col_hdr:
        st.markdown(f"### \U0001F3AF Pallet destino: `{tree['root_name']}`")
        parts = []
        if filters.get("product_name"):
            parts.append(f"**Producto:** {filters['product_name']}")
        if filters.get("manejo"):
            parts.append(f"**Manejo:** {filters['manejo']}")
        if filters.get("variedad"):
            parts.append(f"**Variedad:** {filters['variedad']}")
        if parts:
            st.caption(" · ".join(parts))
    with col_reset:
        if st.button("\U0001F504 Nueva", key="traz_reset"):
            st.session_state[TREE_KEY] = None
            st.rerun()

    st.divider()

    # ── Fase 2: Navegar niveles del árbol ────────────────────
    levels = tree["levels"]
    for level_idx, level_pkgs in enumerate(levels):
        is_frontier = level_idx == len(levels) - 1
        _render_level(level_idx, level_pkgs, is_frontier, tree, username, password)

    # ── Fase 3: Resumen y exportar Excel ─────────────────────
    st.divider()
    _render_summary_and_export(tree)


# ═════════════════════════════════════════════════════════════
# Fase 1 – Input pallet raíz
# ═════════════════════════════════════════════════════════════

def _render_input_root(username: str, password: str):
    """Muestra input para el pallet destino raíz."""
    col_in, col_btn = st.columns([3, 1])
    with col_in:
        pallet_name = st.text_input(
            "Pallet destino inicial",
            placeholder="Ej: PACK0012345",
            key="traz_root_input",
        )
    with col_btn:
        st.markdown("<br>", unsafe_allow_html=True)
        iniciar = st.button("\U0001F680 Iniciar Trazabilidad", type="primary")

    if not iniciar:
        st.info(
            "Ingresa un pallet destino para iniciar la trazabilidad paso a paso.\n\n"
            "El sistema buscará los paquetes origen que lo conformaron "
            "(a través de las órdenes de producción MRP). "
            "En cada nivel verás qué pallets fueron consumidos por cuál pallet destino, "
            "con la OP correspondiente, hasta llegar a las recepciones "
            "con su guía de despacho y proveedor."
        )
        return

    if not pallet_name or not pallet_name.strip():
        st.warning("Ingresa un nombre de pallet.")
        return

    nombre = pallet_name.strip().upper()

    with st.spinner(f"Buscando **{nombre}** en Odoo..."):
        pkg = _find_package(nombre, username, password)

    if not pkg:
        st.error(f"No se encontró el paquete **{nombre}** en Odoo.")
        return

    with st.spinner("Obteniendo información del producto..."):
        info = _get_package_info(pkg["id"], username, password)

    tree = {
        "root_name": nombre,
        "root_id": pkg["id"],
        "next_node_id": 2,
        "filters": {
            "product_name": info.get("product_name") or info.get("nombre_producto"),
            "manejo": info.get("manejo"),
            "variedad": info.get("variedad"),
        },
        "levels": [
            [_new_node(pkg["id"], nombre, node_id=1)]
        ],
    }
    st.session_state[TREE_KEY] = tree
    st.rerun()


# ═════════════════════════════════════════════════════════════
# Fase 2 – Renderizar un nivel del árbol
# ═════════════════════════════════════════════════════════════

def _render_level(level_idx: int, level_pkgs: List[Dict], is_frontier: bool,
                  tree: Dict, username: str, password: str):
    """Renderiza un nivel del árbol de paquetes con relaciones claras."""
    filters = tree.get("filters", {})
    levels = tree["levels"]

    # ── Título del nivel ─────────────────────────────────────
    if level_idx == 0:
        label = "\U0001F3AF **Nivel 0** — Pallet Destino Final"
        st.markdown(label)
        p = level_pkgs[0]
        st.markdown(f"&nbsp;&nbsp;&nbsp;&nbsp;\U0001F4E6 **`{p['pkg_name']}`**")
        # Si es frontera (solo nivel 0) → mostrar botón para buscar orígenes
        if is_frontier:
            _render_frontier_controls(level_pkgs, level_idx, tree, username, password)
        return

    # Agrupar nodos de este nivel por su padre (parent_node_id)
    by_parent: Dict[int, List[Dict]] = {}
    orphans: List[Dict] = []
    for p in level_pkgs:
        pid = p.get("parent_node_id")
        if pid:
            by_parent.setdefault(pid, []).append(p)
        else:
            orphans.append(p)

    # Índice de nodos del nivel anterior por node_id
    parent_idx: Dict[int, Dict] = {}
    if level_idx > 0:
        for pp in levels[level_idx - 1]:
            nid = pp.get("node_id")
            if nid is not None:
                parent_idx[nid] = pp

    st.markdown(f"### \U0001F4E6 Nivel {level_idx}")

    # ── Render agrupado por padre ────────────────────────────
    for parent_id, children in by_parent.items():
        parent_node = parent_idx.get(parent_id)
        parent_name = parent_node["pkg_name"] if parent_node else f"ID {parent_id}"
        mo_name = parent_node.get("mo_name") if parent_node else None
        mo_label = f" — OP: **{mo_name}**" if mo_name else ""

        st.markdown(
            f"&nbsp;&nbsp;\U0001F53B **`{parent_name}`** consumió {len(children)} pallet(s){mo_label}:"
        )

        if not is_frontier:
            # Nivel ya confirmado → resumen compacto
            for ch in children:
                is_leaf = ch.get("is_leaf")
                if is_leaf:
                    _render_leaf(ch, indent=2)
                else:
                    st.markdown(f"&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;\U0001F539 `{ch['pkg_name']}`")
        else:
            # FRONTERA → solo mostrar estado de cada paquete del grupo
            for ch in children:
                if ch.get("is_leaf"):
                    _render_leaf(ch, indent=2)
                elif ch.get("candidates") is None:
                    st.markdown(f"&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;\U0001F535 `{ch['pkg_name']}` — pendiente")
                else:
                    st.markdown(f"&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;\U0001F539 `{ch['pkg_name']}`")

    # Huérfanos (no deberían existir pero por robustez)
    if orphans:
        if not is_frontier:
            for p in orphans:
                if p.get("is_leaf"):
                    _render_leaf(p, indent=1)
                else:
                    st.markdown(f"&nbsp;&nbsp;&nbsp;&nbsp;\U0001F539 `{p['pkg_name']}`")
        else:
            for p in orphans:
                if p.get("is_leaf"):
                    _render_leaf(p, indent=1)
                elif p.get("candidates") is None:
                    st.markdown(f"&nbsp;&nbsp;&nbsp;&nbsp;\U0001F535 `{p['pkg_name']}` — pendiente")
                else:
                    st.markdown(f"&nbsp;&nbsp;&nbsp;&nbsp;\U0001F539 `{p['pkg_name']}`")

    # ── FRONTERA: botón único para cargar y seleccionar ──────
    if is_frontier:
        _render_frontier_controls(level_pkgs, level_idx, tree, username, password)


def _render_frontier_controls(level_pkgs: List[Dict], level_idx: int,
                              tree: Dict, username: str, password: str):
    """Renderiza controles de la frontera: botón cargar + selección de candidatos.
    Se llama UNA VEZ por nivel para evitar keys duplicados."""
    filters = tree.get("filters", {})
    levels = tree["levels"]

    any_unloaded = any(
        p.get("candidates") is None and not p.get("is_leaf")
        for p in level_pkgs
    )
    all_leaves = all(
        p.get("is_leaf") or (isinstance(p.get("candidates"), list) and len(p["candidates"]) == 0)
        for p in level_pkgs
    )

    # Si todos son hojas → nada que hacer (ya se renderizaron arriba)
    if all_leaves:
        return

    # ── Botón para cargar candidatos ─────────────────────────
    if any_unloaded:
        unloaded_names = [
            p["pkg_name"] for p in level_pkgs
            if p.get("candidates") is None and not p.get("is_leaf")
        ]
        btn_label = f"\U0001F50D Buscar orígenes de nivel {level_idx} ({len(unloaded_names)} pendientes)"
        if st.button(btn_label, key=f"load_all_{level_idx}", type="primary"):
            with st.spinner("Buscando orígenes por OP en Odoo..."):
                _cand_cache: Dict[int, Dict] = {}  # cache por pkg_id (evita llamadas duplicadas)
                for p in level_pkgs:
                    if p.get("candidates") is None and not p.get("is_leaf"):
                        pid = p["pkg_id"]
                        if pid in _cand_cache:
                            result = _cand_cache[pid]
                        else:
                            result = _get_candidates(
                                pid, username, password,
                                product_name=filters.get("product_name"),
                                manejo=filters.get("manejo"),
                                variedad=filters.get("variedad"),
                            )
                            _cand_cache[pid] = result
                        cands = result.get("candidates", [])
                        # Solo excluir ancestros directos para evitar ciclos
                        ancestors = _get_ancestor_ids(tree, p)
                        ancestors.add(p["pkg_id"])
                        cands = [c for c in cands if c.get("package_id") not in ancestors]
                        p["candidates"] = cands
                        p["has_mo"] = result.get("has_mo", False)
                        p["mo_name"] = result.get("mo_name")
                        p["is_reception"] = result.get("is_reception", False)
                        p["reception_info"] = result.get("reception")
                        if not cands:
                            p["is_leaf"] = True
            st.session_state[TREE_KEY] = tree
            st.rerun()
        return

    # ── Candidatos ya cargados → re-check hojas ─────────────
    all_leaves = all(
        p.get("is_leaf") or (isinstance(p.get("candidates"), list) and len(p["candidates"]) == 0)
        for p in level_pkgs
    )
    if all_leaves:
        for p in level_pkgs:
            p["is_leaf"] = True
        st.session_state[TREE_KEY] = tree
        return

    # ── Mostrar candidatos agrupados por paquete padre ───────
    all_selected: List[Dict] = []

    for p_idx, p in enumerate(level_pkgs):
        if p.get("is_leaf") or not p.get("candidates"):
            if p.get("is_leaf"):
                _render_leaf(p, indent=2)
            continue

        cands = p["candidates"]
        mo_label = f" (OP: **{p.get('mo_name')}**)" if p.get("mo_name") else ""
        st.markdown(
            f"&nbsp;&nbsp;&nbsp;&nbsp;\U0001F53B Orígenes de **`{p['pkg_name']}`**{mo_label} "
            f"— {len(cands)} candidato(s):"
        )

        for c_idx, c in enumerate(cands):
            c_name = c.get("package_name") or str(c.get("package_id"))
            c_id = c.get("package_id")
            c_qty = c.get("qty_total", 0)
            c_prod = c.get("product_name", "")
            c_lot = c.get("lot_name", "")
            c_manejo = c.get("x_manejo") or ""

            cols = st.columns([0.4, 2, 2.5, 1, 1, 1])
            with cols[0]:
                key = f"chk_{level_idx}_{p.get('node_id', p['pkg_id'])}_{c_idx}"
                checked = st.checkbox("", key=key, value=True, label_visibility="collapsed")
            with cols[1]:
                st.markdown(f"**{c_name}**")
            with cols[2]:
                st.caption(c_prod[:50] if c_prod else "—")
            with cols[3]:
                st.caption(f"{c_qty:.1f} kg")
            with cols[4]:
                st.caption(c_manejo or "—")
            with cols[5]:
                st.caption(c_lot or "—")

            if checked and c_id:
                # Cada par padre-candidato es independiente (permite duplicados cross-padre)
                all_selected.append({
                    **c,
                    "_parent_node_id": p.get("node_id"),
                    "_parent_pkg_id": p["pkg_id"],
                    "_parent_pkg_name": p["pkg_name"],
                    "_parent_mo_name": p.get("mo_name"),
                })

    # ── Botón confirmar selección ────────────────────────────
    st.markdown("---")
    if not all_selected:
        st.warning("Selecciona al menos un paquete origen para continuar la trazabilidad.")
    else:
        if st.button(
            f"\u2705 Confirmar y avanzar ({len(all_selected)} paquetes seleccionados)",
            key=f"confirm_{level_idx}",
            type="primary",
        ):
            new_level = []
            for s in all_selected:
                s_name = s.get("package_name") or str(s.get("package_id"))
                s_id = s.get("package_id")
                nid = tree.get("next_node_id", 1)
                tree["next_node_id"] = nid + 1
                new_level.append(_new_node(
                    s_id, s_name,
                    node_id=nid,
                    qty=s.get("qty_total", 0),
                    product_name=s.get("product_name", ""),
                    lot_name=s.get("lot_name", ""),
                    parent_node_id=s.get("_parent_node_id"),
                    parent_pkg_id=s.get("_parent_pkg_id"),
                    parent_pkg_name=s.get("_parent_pkg_name"),
                ))
            levels.append(new_level)
            st.session_state[TREE_KEY] = tree
            st.rerun()


# ═════════════════════════════════════════════════════════════
# Fase 3 – Resumen y exportar Excel
# ═════════════════════════════════════════════════════════════

def _render_summary_and_export(tree: Dict):
    """Métricas de resumen y botón de descarga Excel."""
    levels = tree["levels"]
    total_pkgs = sum(len(lvl) for lvl in levels)

    leaf_count = 0
    reception_count = 0
    pending = 0
    for lvl in levels:
        for p in lvl:
            if p.get("is_leaf") or (isinstance(p.get("candidates"), list) and len(p.get("candidates", [])) == 0):
                leaf_count += 1
                if p.get("reception_info"):
                    reception_count += 1
            if p.get("candidates") is None and not p.get("is_leaf"):
                pending += 1

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Niveles", len(levels))
    c2.metric("Paquetes", total_pkgs)
    c3.metric("Hojas", leaf_count)
    c4.metric("Recepciones", reception_count)
    c5.metric("Pendientes", pending)

    if pending == 0 and leaf_count > 0:
        if reception_count == leaf_count:
            st.success(
                "\u2705 Trazabilidad completa — todos los caminos llegan a recepción "
                "con guía de despacho y proveedor."
            )
        elif reception_count > 0:
            st.info(
                f"\u2139\uFE0F {reception_count} de {leaf_count} hojas son recepciones. "
                f"Las restantes no tienen origen conocido."
            )

    # Árbol visual compacto
    if len(levels) > 1:
        with st.expander("\U0001F333 Ver árbol de trazabilidad completo", expanded=False):
            _render_tree_text(tree)

    # Excel siempre disponible (puede estar parcial)
    excel = _generar_excel_arbol(tree)
    fname = f"Trazabilidad_{tree['root_name']}_{datetime.now():%Y%m%d_%H%M}.xlsx"
    st.download_button(
        "\u2B07\uFE0F Descargar Excel de Trazabilidad",
        data=excel,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )


def _render_tree_text(tree: Dict):
    """Muestra un resumen textual tipo árbol con cadenas y recepciones."""
    levels = tree.get("levels", [])
    if not levels:
        return

    # Recopilar las hojas y sus cadenas
    leaves = []
    for lvl in levels:
        for p in lvl:
            if p.get("is_leaf") or (isinstance(p.get("candidates"), list) and len(p.get("candidates", [])) == 0):
                leaves.append(p)

    if not leaves:
        st.caption("No hay hojas aún. Sigue trazando niveles.")
        return

    root_name = tree.get("root_name", "?")
    lines = [f"**`{root_name}`** (destino final)"]

    # Construir cadenas de cada hoja al root
    for leaf in leaves:
        chain = _build_chain_to_root(tree, leaf)
        chain_str = " → ".join(chain)
        rec = leaf.get("reception_info")
        if rec:
            guia = rec.get("guia_despacho", "—")
            prov = rec.get("proveedor", "—")
            lines.append(f"&nbsp;&nbsp;&nbsp;&nbsp;↳ {chain_str} | **Guía:** {guia} | **Proveedor:** {prov}")
        else:
            lines.append(f"&nbsp;&nbsp;&nbsp;&nbsp;↳ {chain_str} | *(sin recepción)*")

    st.markdown("\n\n".join(lines))


# ═════════════════════════════════════════════════════════════
# Generación Excel – 3 hojas
# ═════════════════════════════════════════════════════════════

def _generar_excel_arbol(tree: Dict) -> bytes:
    """
    Genera Excel con 3 hojas:
      1) Trazabilidad    – formato igual al TrazabilidadPalletService:
                           Pack | Pallets Consumidos | Pallet Origen | Cadena | Guía Despacho | Productor
      2) Árbol Detallado – todos los paquetes nivel por nivel con parent, OP, tipo, recepción
      3) Detalle Candidatos – candidatos por nivel con marca selección
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Estilos ──────────────────────────────────────────────
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    hdr2_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    hdr3_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al = Alignment(vertical="center", wrap_text=True)
    root_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    leaf_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    sel_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    def _write_header(ws, headers, fill=None):
        f = fill or hdr_fill
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = f
            cell.font = hdr_font
            cell.alignment = center
            cell.border = border

    def _set_widths(ws, widths):
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

    levels = tree.get("levels", [])
    filters = tree.get("filters", {})
    root_name = tree.get("root_name", "?")

    # ══════════════════════════════════════════════════════════
    # HOJA 1 – Trazabilidad (formato referencia)
    #   Pack | Pallets Consumidos | Pallet Origen | Cadena de Trazabilidad | Guía Despacho | Productor
    # ══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Trazabilidad"

    headers1 = [
        "Pack (Destino)", "Pallets Consumidos",
        "Pallet Origen", "Cadena de Trazabilidad",
        "Guía Despacho", "Productor",
    ]
    _write_header(ws1, headers1)
    _set_widths(ws1, [22, 45, 18, 55, 22, 35])

    # Recopilar hojas y construir cadenas
    leaves = []
    for lvl in levels:
        for p in lvl:
            if p.get("is_leaf") or (isinstance(p.get("candidates"), list) and len(p.get("candidates", [])) == 0):
                leaves.append(p)

    # Pallets consumidos directos (nivel 1 si existe)
    consumidos_str = ""
    if len(levels) > 1:
        consumidos_str = " - ".join(_short(p["pkg_name"]) for p in levels[1])

    row = 2
    if leaves:
        for leaf in leaves:
            chain = _build_chain_to_root(tree, leaf)
            chain_str = " → ".join(chain)
            rec = leaf.get("reception_info")
            guia = ""
            productor = ""
            if rec:
                guia = rec.get("guia_despacho", "") or ""
                productor = rec.get("proveedor", "") or ""

            ws1.cell(row=row, column=1, value=root_name)
            ws1.cell(row=row, column=2, value=consumidos_str)
            ws1.cell(row=row, column=3, value=_short(leaf["pkg_name"]))
            ws1.cell(row=row, column=4, value=chain_str)
            ws1.cell(row=row, column=5, value=guia)
            ws1.cell(row=row, column=6, value=productor)

            for c in range(1, 7):
                cell = ws1.cell(row=row, column=c)
                cell.border = border
                cell.alignment = left_al
            # Color recepciones
            if rec:
                for c in range(1, 7):
                    ws1.cell(row=row, column=c).fill = leaf_fill

            row += 1
    else:
        # Sin hojas aún — poner al menos el root
        ws1.cell(row=row, column=1, value=root_name)
        ws1.cell(row=row, column=2, value="(trazabilidad pendiente)")
        for c in range(1, 7):
            ws1.cell(row=row, column=c).border = border
            ws1.cell(row=row, column=c).alignment = left_al

    ws1.freeze_panes = "A2"

    # ══════════════════════════════════════════════════════════
    # HOJA 2 – Árbol Detallado (todos los nodos nivel por nivel)
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Árbol Detallado")

    headers2 = [
        "Nivel", "Paquete", "Padre (consumido por)", "OP",
        "Producto", "Cantidad (kg)", "Lote", "Tipo",
        "Guía Despacho", "Productor", "Picking Recepción", "Fecha Recepción",
    ]
    _write_header(ws2, headers2, fill=hdr2_fill)
    _set_widths(ws2, [8, 22, 22, 18, 45, 14, 22, 14, 22, 35, 18, 14])

    row2 = 2
    for level_idx, level_pkgs in enumerate(levels):
        for p in level_pkgs:
            is_leaf = p.get("is_leaf") or (
                isinstance(p.get("candidates"), list) and len(p.get("candidates", [])) == 0
            )
            if level_idx == 0:
                tipo = "Destino"
            elif is_leaf and p.get("reception_info"):
                tipo = "Recepción"
            elif is_leaf:
                tipo = "Hoja (sin origen)"
            else:
                tipo = "Intermedio"

            rec = p.get("reception_info")
            guia = rec.get("guia_despacho", "") if rec else ""
            prov = rec.get("proveedor", "") if rec else ""
            pick_name = rec.get("picking_name", "") if rec else ""
            fecha_rec = rec.get("fecha", "") if rec else ""
            parent_name = p.get("parent_pkg_name", "") or ""
            mo_name = p.get("mo_name", "") or ""

            ws2.cell(row=row2, column=1, value=level_idx)
            ws2.cell(row=row2, column=2, value=p.get("pkg_name", "")).font = Font(bold=True)
            ws2.cell(row=row2, column=3, value=parent_name)
            ws2.cell(row=row2, column=4, value=mo_name)
            prod_name = p.get("product_name", "") or (filters.get("product_name") if level_idx == 0 else "")
            ws2.cell(row=row2, column=5, value=prod_name or "")
            qty = p.get("qty")
            ws2.cell(row=row2, column=6, value=f"{qty:.1f}" if qty else "")
            ws2.cell(row=row2, column=7, value=p.get("lot_name", ""))
            ws2.cell(row=row2, column=8, value=tipo)
            ws2.cell(row=row2, column=9, value=guia)
            ws2.cell(row=row2, column=10, value=prov)
            ws2.cell(row=row2, column=11, value=pick_name)
            ws2.cell(row=row2, column=12, value=fecha_rec)

            for c in range(1, 13):
                cell = ws2.cell(row=row2, column=c)
                cell.border = border
                cell.alignment = center if c in (1, 6, 8) else left_al

            if level_idx == 0:
                for c in range(1, 13):
                    ws2.cell(row=row2, column=c).fill = root_fill
            elif is_leaf and rec:
                for c in range(1, 13):
                    ws2.cell(row=row2, column=c).fill = leaf_fill

            row2 += 1

    ws2.freeze_panes = "A2"

    # ══════════════════════════════════════════════════════════
    # HOJA 3 – Detalle Candidatos
    # ══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Detalle Candidatos")

    headers3 = ["Nivel", "Paquete Destino", "Candidato Origen", "Producto",
                 "Cantidad (kg)", "Lote", "Seleccionado"]
    _write_header(ws3, headers3, fill=hdr3_fill)
    _set_widths(ws3, [8, 22, 22, 45, 14, 22, 14])

    row3 = 2
    for level_idx, level_pkgs in enumerate(levels):
        # Nombres del nivel siguiente para saber cuáles se seleccionaron
        next_names: set = set()
        if level_idx + 1 < len(levels):
            next_names = {p["pkg_name"] for p in levels[level_idx + 1]}

        for p in level_pkgs:
            cands = p.get("candidates") or []
            for c in cands:
                c_name = c.get("package_name") or str(c.get("package_id"))
                selected = "Sí" if c_name in next_names else "No"

                ws3.cell(row=row3, column=1, value=level_idx)
                ws3.cell(row=row3, column=2, value=p.get("pkg_name", ""))
                ws3.cell(row=row3, column=3, value=c_name)
                ws3.cell(row=row3, column=4, value=c.get("product_name", ""))
                qty = c.get("qty_total", 0)
                ws3.cell(row=row3, column=5, value=f"{qty:.1f}" if qty else "")
                ws3.cell(row=row3, column=6, value=c.get("lot_name", ""))
                ws3.cell(row=row3, column=7, value=selected)

                for col in range(1, 8):
                    cell = ws3.cell(row=row3, column=col)
                    cell.border = border
                    cell.alignment = center if col in (1, 5, 7) else left_al

                if selected == "Sí":
                    for col in range(1, 8):
                        ws3.cell(row=row3, column=col).fill = sel_fill

                row3 += 1

    ws3.freeze_panes = "A2"

    # ── Guardar ──────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
