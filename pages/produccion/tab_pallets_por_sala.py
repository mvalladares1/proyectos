"""
Tab: Pallets por Sala
Muestra una vista detallada de pallets de producto terminado filtrados por sala de proceso,
producto, fecha, manejo y tipo de fruta. Solo temporada 2025 y 2026.
Incluye enlace con datos de Control de Calidad desde Excel.
"""
import streamlit as st
import pandas as pd
import requests
import re
import time
from datetime import datetime, timedelta, date
from typing import Dict, List, Optional

from .shared import API_URL, fmt_numero


def _throttle_rerun(key: str = "ps_last_rerun", min_interval: float = 2.0) -> bool:
    """Evita reruns infinitos limitando frecuencia."""
    now = time.time()
    last = st.session_state.get(key, 0)
    if now - last < min_interval:
        return False
    st.session_state[key] = now
    return True


@st.cache_data(ttl=3600, show_spinner=False)
def _obtener_productos_pt(username: str, password: str) -> List[Dict]:
    """Obtiene la lista de productos terminados de Odoo (cacheada 1 hora)."""
    try:
        resp = requests.get(
            f"{API_URL}/api/v1/produccion/productos_pt",
            params={"username": username, "password": password},
            timeout=15
        )
        if resp.status_code == 200:
            return resp.json().get('productos', [])
    except Exception:
        pass
    return []


# --- Constantes del módulo ---
SALA_MAP_INTERNAL = {
    "Sala 1 - Línea Retail": "Sala 1 - Linea Retail",
    "Sala 2 - Línea Granel": "Sala 2 - Linea Granel",
    "Sala 3 - Línea Granel": "Sala 3 - Linea Granel",
    "Sala 3 - Línea Retail": "Sala 3 - Linea Retail",
    "Sala 4 - Línea Retail": "Sala 4 - Linea Retail",
    "Sala 4 - Línea Chocolate": "Sala 4 - Linea Chocolate",
    "Sala - Vilkun": "Sala - Vilkun"
}

GRADO_NOMBRES = {
    '1': 'IQF AA', '2': 'IQF A', '3': 'PSP', '4': 'W&B',
    '5': 'Block', '6': 'Jugo', '7': 'IQF Retail'
}

GRADO_COLORES = {
    'IQF AA': '#FFB302', 'IQF A': '#2196F3', 'PSP': '#9C27B0',
    'W&B': '#795548', 'Block': '#00BCD4', 'Jugo': '#FF5722', 'IQF Retail': '#4CAF50'
}

TEMPORADAS_VALIDAS = ["2025", "2026"]

# Columnas de calidad del Excel (fila 2 = headers, mapeadas en orden)
QUALITY_COLS = [
    'fecha_qc', 'semana', 'turno', 'sala_qc', 'codigo_qc', 'descripcion_producto',
    'familia_calidad', 'n_pallet', 'n_lote', 'muestra', 'hora_monitoreo',
    'peso_bolsa', 'peso_caja', 'temperatura', 'sellado_bolsa', 'sellado_caja',
    'pct_iqf', 'inmadura', 'pct_inmadura', 'partidas', 'pct_partidas',
    'dano_insecto', 'pct_dano_insecto', 'sobremadura', 'pct_sobremadura',
    'deforme', 'pct_deforme', 'desc_extra_1', 'valor_extra_1', 'pct_extra_1',
    'desc_extra_2', 'valor_extra_2', 'pct_extra_2',
    'hongos', 'pct_hongos', 'block', 'pct_block',
    'caliz_und', 'larvas_und',
    'fuera_calibre_menor', 'pct_fuera_calibre_menor',
    'calibre_menor_25', 'pct_menor_25', 'calibre_mayor_12', 'pct_mayor_12',
    'calibre_menor_14', 'pct_menor_14', 'calibre_mayor_14', 'pct_mayor_14',
    'calibre_menor_10', 'pct_menor_10', 'calibre_mayor_13', 'pct_mayor_13',
    'hoja_un', 'pedicelo_caja', 'larvas_caja', 'vidrios', 'madera', 'plastico',
    'brix', 'sabor', 'olor', 'defectos_totales', 'pct_defectos',
    'observaciones', 'pct_iqf2', 'analista',
    'mat_extr_endo', 'mat_extr_exo', 'filtro_producto', 'filtro_lote', 'filtro_pallet'
]

# Columnas principales para tabla interactiva de QC
QUALITY_DISPLAY_COLS = {
    'hora_monitoreo': 'Hora',
    'peso_caja': 'Peso Caja',
    'temperatura': 'Temp',
    'sellado_bolsa': 'Sell.B',
    'sellado_caja': 'Sell.C',
    'pct_iqf': '% IQF',
    'pct_inmadura': '% Inmad.',
    'pct_partidas': '% Partid.',
    'pct_dano_insecto': '% D.Insecto',
    'pct_sobremadura': '% Sobrem.',
    'pct_deforme': '% Deforme',
    'pct_hongos': '% Hongos',
    'pct_defectos': '% Defectos',
    'brix': 'Brix',
    'observaciones': 'Obs.',
}


def _normalize(text: str) -> str:
    """Normaliza texto removiendo acentos para comparaciones."""
    if not text:
        return ""
    text = text.lower()
    for old, new in [("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u")]:
        text = text.replace(old, new)
    return text


def _extraer_numero_pallet(pallet_name: str) -> str:
    """Extrae el número puro del nombre de pallet.
    PACK0013712 -> 13712, PACK0003567 -> 3567, 03567 -> 3567"""
    if not pallet_name:
        return ""
    num = re.sub(r'^PACK0*', '', str(pallet_name), flags=re.IGNORECASE)
    num = num.strip().lstrip('0') or '0'
    return num


def _fmt_pct(val):
    """Formatea un valor de porcentaje del Excel."""
    if val is None or val == '' or val == 0:
        return ''
    try:
        v = float(val)
        return round(v * 100, 2) if v <= 1 else round(v, 2)
    except (ValueError, TypeError):
        return val


def _parsear_excel_calidad(uploaded_file) -> Dict[str, List[Dict]]:
    """
    Parsea un Excel de control de calidad.
    Retorna: { "13712": [registro1, ...], "13713": [...] }
    Lee hojas con "base de datos" o "b.d." en el nombre.
    Headers en fila 2, datos desde fila 3.
    """
    import openpyxl

    calidad_por_pallet: Dict[str, List[Dict]] = {}

    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    except Exception as e:
        st.error(f"Error al leer Excel de calidad: {e}")
        return {}

    hojas_calidad = [
        n for n in wb.sheetnames
        if 'base de datos' in n.lower() or 'b.d.' in n.lower()
    ]

    for nombre_hoja in hojas_calidad:
        ws = wb[nombre_hoja]
        planta_origen = "RIO FUTURO" if "r.f." in nombre_hoja.lower() else "VILKUN"

        for row_idx in range(3, ws.max_row + 1):
            pallet_val = ws.cell(row=row_idx, column=8).value
            if not pallet_val:
                continue

            pallet_num = str(int(pallet_val)) if isinstance(pallet_val, (int, float)) else str(pallet_val).strip().lstrip('0') or '0'
            if not pallet_num or pallet_num == '0':
                continue

            registro = {'planta_qc': planta_origen, 'hoja_origen': nombre_hoja}
            for col_idx in range(1, min(ws.max_column + 1, len(QUALITY_COLS) + 1)):
                col_name = QUALITY_COLS[col_idx - 1] if col_idx - 1 < len(QUALITY_COLS) else f'col_{col_idx}'
                registro[col_name] = ws.cell(row=row_idx, column=col_idx).value

            calidad_por_pallet.setdefault(pallet_num, []).append(registro)

    return calidad_por_pallet


def _cargar_calidad_servidor() -> Dict[str, List[Dict]]:
    """
    Carga automáticamente TODOS los Excel de calidad desde data/calidad/.
    Busca en múltiples rutas posibles (local y container).
    Cachea el resultado en session_state para no re-parsear cada refresh.
    """
    import os
    import glob

    # Si ya está cacheado y no se pidió recarga, retornar
    if st.session_state.get('_qc_server_loaded') and st.session_state.get('ps_calidad_dict'):
        return st.session_state['ps_calidad_dict']

    # Buscar la carpeta data/calidad en rutas posibles
    posibles_rutas = [
        os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'data', 'calidad'),
        '/app/data/calidad',
        os.path.join(os.getcwd(), 'data', 'calidad'),
    ]

    carpeta_calidad = None
    for ruta in posibles_rutas:
        if os.path.isdir(ruta):
            carpeta_calidad = ruta
            break

    if not carpeta_calidad:
        return {}

    archivos_excel = glob.glob(os.path.join(carpeta_calidad, '*.xlsx')) + \
                     glob.glob(os.path.join(carpeta_calidad, '*.xls'))

    if not archivos_excel:
        return {}

    calidad_total: Dict[str, List[Dict]] = {}
    archivos_ok = []
    archivos_error = []

    for archivo in sorted(archivos_excel):
        nombre = os.path.basename(archivo)
        try:
            resultado = _parsear_excel_calidad(archivo)
            for pallet_num, registros in resultado.items():
                calidad_total.setdefault(pallet_num, []).extend(registros)
            archivos_ok.append(nombre)
        except Exception as e:
            archivos_error.append(f"{nombre}: {e}")

    # Guardar en session_state
    st.session_state['ps_calidad_dict'] = calidad_total
    st.session_state['_qc_server_loaded'] = True
    st.session_state['_qc_archivos_cargados'] = archivos_ok
    st.session_state['_qc_archivos_error'] = archivos_error

    return calidad_total


def _render_calidad_pallet(registros_calidad: List[Dict], pallet_name: str):
    """Renderiza los datos de calidad de un pallet de forma visual e interactiva."""
    n = len(registros_calidad)

    # Calcular promedios
    pcts_iqf = [float(r.get('pct_iqf', 0) or 0) for r in registros_calidad if r.get('pct_iqf')]
    pcts_defectos = [float(r.get('pct_defectos', 0) or 0) for r in registros_calidad if r.get('pct_defectos')]
    temps = []
    for r in registros_calidad:
        t = r.get('temperatura')
        if t and str(t).strip() not in ('', 'N/A'):
            try:
                temps.append(float(t))
            except (ValueError, TypeError):
                pass

    avg_iqf = sum(pcts_iqf) / len(pcts_iqf) * 100 if pcts_iqf else None
    avg_defectos = sum(pcts_defectos) / len(pcts_defectos) * 100 if pcts_defectos else None
    avg_temp = sum(temps) / len(temps) if temps else None

    # Determinar estado de calidad
    if avg_defectos is not None:
        if avg_defectos <= 4:
            estado_color, estado_icon, estado_text = "#22c55e", "✅", "APROBADO"
        elif avg_defectos <= 7:
            estado_color, estado_icon, estado_text = "#f59e0b", "⚠️", "OBSERVADO"
        else:
            estado_color, estado_icon, estado_text = "#ef4444", "❌", "RECHAZADO"
    else:
        estado_color, estado_icon, estado_text = "#64748b", "❓", "SIN EVALUAR"

    # Header con estado
    glow = f"0 0 15px {estado_color}33" if estado_text != "SIN EVALUAR" else "none"
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,{estado_color}18 0%,{estado_color}08 100%);
        border:1px solid {estado_color}44; border-left:4px solid {estado_color};
        border-radius:10px; padding:0.8rem 1.2rem; margin-bottom:0.8rem;
        box-shadow:{glow};">
        <div style="display:flex; justify-content:space-between; align-items:center;">
            <div style="display:flex; align-items:center; gap:10px;">
                <span style="font-size:1.3rem;">{estado_icon}</span>
                <div>
                    <div style="font-weight:700; color:{estado_color}; font-size:0.95rem; letter-spacing:0.03em;">
                        {estado_text}
                    </div>
                    <div style="font-size:0.7rem; color:#64748b;">Control de Calidad</div>
                </div>
            </div>
            <span class="stat-chip" style="background:{estado_color}1a; color:{estado_color};">{n} monitoreo{'s' if n > 1 else ''}</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # KPIs de calidad - mini gauges
    qc_metrics = []
    iqf_val = f"{avg_iqf:.1f}%" if avg_iqf is not None else "—"
    iqf_color = "#22c55e" if avg_iqf and avg_iqf >= 90 else ("#f59e0b" if avg_iqf and avg_iqf >= 80 else "#ef4444")
    qc_metrics.append(("% IQF Prom.", iqf_val, iqf_color, "🎯"))

    def_val = f"{avg_defectos:.1f}%" if avg_defectos is not None else "—"
    def_color = "#22c55e" if avg_defectos is not None and avg_defectos <= 4 else ("#f59e0b" if avg_defectos is not None and avg_defectos <= 7 else "#ef4444")
    qc_metrics.append(("% Defectos", def_val, def_color, "⚠️"))

    temp_val = f"{avg_temp:.1f}&deg;C" if avg_temp is not None else "—"
    temp_color = "#22c55e" if avg_temp is not None and avg_temp <= -5 else ("#f59e0b" if avg_temp is not None and avg_temp <= 0 else "#ef4444")
    qc_metrics.append(("Temperatura", temp_val, temp_color, "🌡️"))

    sellados_b = [str(r.get('sellado_bolsa', '')).upper() for r in registros_calidad if r.get('sellado_bolsa')]
    sellados_c = [str(r.get('sellado_caja', '')).upper() for r in registros_calidad if r.get('sellado_caja')]
    todo_ok = all(s == 'C' for s in sellados_b + sellados_c) if (sellados_b or sellados_c) else None
    sell_icon = "✅" if todo_ok else ("❌" if todo_ok is not None else "—")
    sell_color = "#22c55e" if todo_ok else ("#ef4444" if todo_ok is not None else "#64748b")
    qc_metrics.append(("Sellado", sell_icon, sell_color, "📦"))

    kpi_cols = st.columns(4)
    for i, (label, value, color, icon) in enumerate(qc_metrics):
        with kpi_cols[i]:
            st.markdown(f"""<div style="background:linear-gradient(145deg,#0f172a,#1e293b); border:1px solid {color}33;
                border-radius:10px; padding:0.6rem; text-align:center; position:relative; overflow:hidden;">
                <div style="position:absolute; top:0; left:0; right:0; height:2px; background:{color};"></div>
                <div style="font-size:0.9rem; margin-bottom:2px;">{icon}</div>
                <div style="font-size:1.15rem; font-weight:800; color:{color};">{value}</div>
                <div style="font-size:0.6rem; color:#64748b; text-transform:uppercase; letter-spacing:0.05em; font-weight:600;">{label}</div>
            </div>""", unsafe_allow_html=True)

    # Tabla de detalle por monitoreo
    rows = []
    for r in registros_calidad:
        row = {}
        for col_key, col_label in QUALITY_DISPLAY_COLS.items():
            val = r.get(col_key, '')
            if val is None:
                val = ''
            if col_key.startswith('pct_') and val != '':
                try:
                    val = f"{float(val) * 100:.1f}%" if float(val) <= 1 else f"{float(val):.1f}%"
                except (ValueError, TypeError):
                    pass
            elif col_key == 'temperatura' and val != '':
                try:
                    val = f"{float(val):.1f}"
                except (ValueError, TypeError):
                    pass
            elif col_key == 'hora_monitoreo':
                val = str(val)[:8] if val else ''
            row[col_label] = val
        rows.append(row)

    df_qc = pd.DataFrame(rows)
    st.dataframe(df_qc, use_container_width=True, hide_index=True, height=min(200, 35 * len(rows) + 38))

    # Desglose de defectos (barras visuales)
    defectos = {}
    defecto_keys = [
        ('pct_inmadura', 'Inmadura'), ('pct_partidas', 'Partidas'),
        ('pct_dano_insecto', 'D. Insecto'), ('pct_sobremadura', 'Sobremadura'),
        ('pct_deforme', 'Deforme'), ('pct_hongos', 'Hongos'), ('pct_block', 'Block'),
    ]
    for key, label in defecto_keys:
        vals = [float(r.get(key, 0) or 0) for r in registros_calidad if r.get(key)]
        if vals:
            avg = sum(vals) / len(vals) * 100
            if avg > 0:
                defectos[label] = round(avg, 2)

    if defectos:
        max_val = max(defectos.values()) if defectos else 1
        bars_html = ""
        for label, val in sorted(defectos.items(), key=lambda x: x[1], reverse=True):
            width_pct = min(val / max_val * 100, 100) if max_val > 0 else 0
            bar_color = "#ef4444" if val > 4 else ("#f59e0b" if val > 2 else "#22c55e")
            bars_html += (
                '<div style="display:flex;align-items:center;margin:4px 0;gap:8px;">'
                f'<span style="width:100px;font-size:0.72rem;color:#94a3b8;text-align:right;font-weight:500;">{label}</span>'
                '<div style="flex:1;background:#0f172a;border-radius:6px;height:20px;overflow:hidden;border:1px solid #1e293b;">'
                f'<div style="width:{width_pct}%;background:linear-gradient(90deg,{bar_color},{bar_color}bb);height:100%;border-radius:5px;min-width:2px;'
                f'transition:width 0.6s ease;"></div>'
                '</div>'
                f'<span style="width:50px;font-size:0.72rem;color:{bar_color};font-weight:700;">{val:.1f}%</span>'
                '</div>'
            )

        html_block = (
            '<div style="background:linear-gradient(145deg,#0f172a,#1a1e2e);border:1px solid #1e293b;border-radius:10px;padding:0.9rem;margin-top:0.5rem;">'
            '<div style="font-size:0.68rem;color:#64748b;margin-bottom:8px;text-transform:uppercase;font-weight:700;letter-spacing:0.06em;">'
            '📊 Desglose de Defectos (Promedio)</div>'
            f'{bars_html}'
            '</div>'
        )
        st.markdown(html_block, unsafe_allow_html=True)


def _generar_excel_con_calidad(df_display, calidad_dict):
    """Genera un Excel con 3 hojas: Pallets, Calidad (detalle QC), Resumen QC."""
    import io
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # --- HOJA 1: PALLETS ---
        df_display.to_excel(writer, index=False, sheet_name='Pallets')

        # --- HOJA 2: CALIDAD (registros QC enlazados) ---
        qc_rows = []
        for _, row in df_display.iterrows():
            pallet_name = row.get('Pallet', '')
            pallet_num = _extraer_numero_pallet(pallet_name)
            registros = calidad_dict.get(pallet_num, [])
            for r in registros:
                qc_rows.append({
                    'Pallet': pallet_name,
                    'Producto': row.get('Producto', ''),
                    'Grado': row.get('Grado', ''),
                    'Kilos': row.get('Kilos', 0),
                    'OF': row.get('OF', ''),
                    'Sala': row.get('Sala', ''),
                    'Fecha QC': str(r.get('fecha_qc', ''))[:10] if r.get('fecha_qc') else '',
                    'Hora': str(r.get('hora_monitoreo', ''))[:8] if r.get('hora_monitoreo') else '',
                    'Turno': r.get('turno', ''),
                    'Peso Caja': r.get('peso_caja', ''),
                    'Temperatura': r.get('temperatura', ''),
                    'Sellado Bolsa': r.get('sellado_bolsa', ''),
                    'Sellado Caja': r.get('sellado_caja', ''),
                    '% IQF': _fmt_pct(r.get('pct_iqf')),
                    '% Inmadura': _fmt_pct(r.get('pct_inmadura')),
                    '% Partidas': _fmt_pct(r.get('pct_partidas')),
                    '% D. Insecto': _fmt_pct(r.get('pct_dano_insecto')),
                    '% Sobremadura': _fmt_pct(r.get('pct_sobremadura')),
                    '% Deforme': _fmt_pct(r.get('pct_deforme')),
                    '% Hongos': _fmt_pct(r.get('pct_hongos')),
                    '% Block': _fmt_pct(r.get('pct_block')),
                    'Defectos Tot.': r.get('defectos_totales', ''),
                    '% Defectos': _fmt_pct(r.get('pct_defectos')),
                    'Brix': r.get('brix', ''),
                    'Observaciones': r.get('observaciones', ''),
                    'Analista': r.get('analista', ''),
                })

        if qc_rows:
            pd.DataFrame(qc_rows).to_excel(writer, index=False, sheet_name='Calidad')

        # --- HOJA 3: RESUMEN QC POR PALLET ---
        resumen_rows = []
        for _, row in df_display.iterrows():
            pallet_name = row.get('Pallet', '')
            pallet_num = _extraer_numero_pallet(pallet_name)
            registros = calidad_dict.get(pallet_num, [])
            base = {
                'Pallet': pallet_name,
                'Producto': row.get('Producto', ''),
                'Grado': row.get('Grado', ''),
                'Kilos': row.get('Kilos', 0),
                'Sala': row.get('Sala', ''),
            }
            if registros:
                pi = [float(r.get('pct_iqf', 0) or 0) for r in registros if r.get('pct_iqf')]
                pd2 = [float(r.get('pct_defectos', 0) or 0) for r in registros if r.get('pct_defectos')]
                tl = []
                for r in registros:
                    t = r.get('temperatura')
                    if t and str(t).strip() not in ('', 'N/A'):
                        try:
                            tl.append(float(t))
                        except (ValueError, TypeError):
                            pass
                ai = sum(pi) / len(pi) * 100 if pi else None
                ad = sum(pd2) / len(pd2) * 100 if pd2 else None
                at = sum(tl) / len(tl) if tl else None
                estado = "APROBADO" if ad is not None and ad <= 4 else ("OBSERVADO" if ad is not None and ad <= 7 else ("RECHAZADO" if ad is not None else "SIN EVALUAR"))
                base.update({
                    'N Monitoreos': len(registros),
                    '% IQF Prom.': round(ai, 1) if ai is not None else '',
                    '% Defectos Prom.': round(ad, 1) if ad is not None else '',
                    'Temp. Prom.': round(at, 1) if at is not None else '',
                    'Estado': estado,
                })
            else:
                base.update({'N Monitoreos': 0, '% IQF Prom.': '', '% Defectos Prom.': '', 'Temp. Prom.': '', 'Estado': 'SIN DATOS QC'})
            resumen_rows.append(base)

        pd.DataFrame(resumen_rows).to_excel(writer, index=False, sheet_name='Resumen QC')

        # --- ESTILOS ---
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'),
        )
        fills = {
            'Pallets': PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid"),
            'Calidad': PatternFill(start_color="1a5f3a", end_color="1a5f3a", fill_type="solid"),
            'Resumen QC': PatternFill(start_color="5f1a3a", end_color="5f1a3a", fill_type="solid"),
        }

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            fill = fills.get(sheet_name, fills['Pallets'])
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"
            for ci in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=ci)
                cell.font = header_font
                cell.fill = fill
                cell.alignment = header_align
                cell.border = thin_border
            for ri in range(2, ws.max_row + 1):
                for ci in range(1, ws.max_column + 1):
                    ws.cell(row=ri, column=ci).border = thin_border
            for ci in range(1, ws.max_column + 1):
                cl = get_column_letter(ci)
                mx = len(str(ws.cell(row=1, column=ci).value or ""))
                for ri in range(2, min(ws.max_row + 1, 200)):
                    v = ws.cell(row=ri, column=ci).value
                    if v:
                        mx = max(mx, len(str(v)))
                ws.column_dimensions[cl].width = min(mx + 3, 45)

        # Colorear estados en Resumen QC
        if 'Resumen QC' in writer.sheets:
            ws_r = writer.sheets['Resumen QC']
            estado_col = None
            for ci in range(1, ws_r.max_column + 1):
                if ws_r.cell(row=1, column=ci).value == 'Estado':
                    estado_col = ci
                    break
            if estado_col:
                state_fills = {
                    'APROBADO': (PatternFill(start_color="C6EFCE", fill_type="solid"), Font(bold=True, color="006100")),
                    'OBSERVADO': (PatternFill(start_color="FFEB9C", fill_type="solid"), Font(bold=True, color="9C6500")),
                    'RECHAZADO': (PatternFill(start_color="FFC7CE", fill_type="solid"), Font(bold=True, color="9C0006")),
                }
                gray = (PatternFill(start_color="D9D9D9", fill_type="solid"), Font(bold=False, color="333333"))
                for ri in range(2, ws_r.max_row + 1):
                    cell = ws_r.cell(row=ri, column=estado_col)
                    val = str(cell.value or '').upper()
                    f, fo = gray
                    for key, (sf, sfo) in state_fills.items():
                        if key in val:
                            f, fo = sf, sfo
                            break
                    cell.fill = f
                    cell.font = fo

    return buffer.getvalue()


def _filtrar_detalle(detalle_raw, filtros):
    """Aplica filtros dinámicos sobre la lista de detalle en memoria."""
    resultado = list(detalle_raw)

    # Filtrar por Planta
    if filtros.get("planta") and filtros["planta"] != "Todas":
        planta_val = filtros["planta"].upper()
        resultado = [d for d in resultado if d.get('planta', '').upper() == planta_val]

    # Filtrar por Fruta (usando primeros 2 dígitos del código de producto)
    if filtros.get("tipo_fruta") and filtros["tipo_fruta"] != "Todas":
        fruta_lower = _normalize(filtros["tipo_fruta"])
        FRUTA_CODE_MAP = {
            'arandano': ['31', '41'],
            'frambuesa': ['32', '42'],
            'frutilla': ['33', '43'],
            'mora': ['34', '44'],
        }
        codes_esperados = FRUTA_CODE_MAP.get(fruta_lower, [])
        filtered = []
        for d in resultado:
            code = d.get('codigo_producto', '')
            if codes_esperados and len(code) >= 2 and code[:2] in codes_esperados:
                filtered.append(d)
            elif fruta_lower in _normalize(d.get('producto', '')):
                filtered.append(d)
        resultado = filtered

    # Filtrar por Sala de Proceso (normalizado, tolerante a acentos y variaciones)
    if filtros.get("sala") and filtros["sala"] != "Todas":
        target_key = SALA_MAP_INTERNAL.get(filtros["sala"], filtros["sala"])
        target_norm = _normalize(target_key)
        # Extraer solo la parte clave para matching parcial: "sala 1", "sala 2", etc.
        target_parts = target_norm.split(" - ")
        filtered = []
        for d in resultado:
            sala_dato = _normalize(d.get('sala', ''))
            if not sala_dato:
                continue
            # Match exacto normalizado
            if sala_dato == target_norm:
                filtered.append(d)
            # Match parcial: ambas partes del target deben estar presentes
            elif all(part.strip() in sala_dato for part in target_parts):
                filtered.append(d)
            # Match solo por sala base (ej: "sala 1")
            elif target_parts[0].strip() in sala_dato and (len(target_parts) < 2 or target_parts[1].strip() in sala_dato):
                filtered.append(d)
        resultado = filtered

    # Filtrar por Manejo (usando dígito 4 del código de producto)
    if filtros.get("tipo_manejo") and filtros["tipo_manejo"] != "Todos":
        manejo = filtros["tipo_manejo"]
        filtered = []
        for d in resultado:
            code = d.get('codigo_producto', '')
            if len(code) >= 4:
                m_digit = code[3]
                if manejo == "Orgánico" and m_digit == '2':
                    filtered.append(d)
                elif manejo == "Convencional" and m_digit == '1':
                    filtered.append(d)
        resultado = filtered

    # Filtrar por producto seleccionado (selectbox)
    if filtros.get("producto_seleccionado") and filtros["producto_seleccionado"] != "Todos":
        code_sel = filtros["producto_seleccionado"].split(" — ")[0].strip()
        resultado = [d for d in resultado if d.get('codigo_producto', '') == code_sel]

    # Filtrar por texto libre (pallet, OF, lote, producto, código)
    if filtros.get("producto_texto"):
        texto = _normalize(filtros["producto_texto"].strip())
        resultado = [
            d for d in resultado
            if texto in _normalize(d.get('producto', ''))
            or texto in d.get('codigo_producto', '').lower()
            or texto in _normalize(d.get('pallet', ''))
            or texto in _normalize(d.get('orden_fabricacion', ''))
            or texto in _normalize(d.get('lote', ''))
        ]

    # Filtrar por Temporada 2025/2026 basado en fecha
    resultado_temp = []
    for d in resultado:
        fecha = d.get('fecha', '')
        if fecha:
            # Extraer año: soporta "2025-01-15", "2025-01-15 10:00:00", etc.
            year = str(fecha)[:4]
            if year in TEMPORADAS_VALIDAS:
                resultado_temp.append(d)
        else:
            # Si no tiene fecha, incluirlo (no descartar datos válidos)
            resultado_temp.append(d)
    resultado = resultado_temp

    return resultado


def render(username: str, password: str):
    """Renderiza el tab de Pallets por Sala."""

    st.markdown("""
    <div class="ps-hero">
        <h2>🏭 Pallets por Sala de Proceso</h2>
        <div class="subtitle">Detalle de pallets de producto terminado por sala, producto y grado &mdash; Temporada 2025 / 2026</div>
    </div>
    """, unsafe_allow_html=True)

    # === FILTROS ===
    with st.container():
        col1, col2 = st.columns(2)

        with col1:
            fecha_inicio = st.date_input(
                "Fecha Inicio",
                value=datetime(2025, 1, 1),
                key="ps_fecha_inicio"
            )

            tipo_fruta_opciones = ["Todas", "Arándano", "Frambuesa", "Frutilla", "Mora"]
            tipo_fruta = st.selectbox(
                "Tipo de Fruta",
                options=tipo_fruta_opciones,
                index=0,
                key="ps_tipo_fruta"
            )

        with col2:
            fecha_fin = st.date_input(
                "Fecha Fin",
                value=date.today(),
                key="ps_fecha_fin"
            )

            tipo_manejo_opciones = ["Todos", "Orgánico", "Convencional"]
            tipo_manejo = st.selectbox(
                "Tipo de Manejo",
                options=tipo_manejo_opciones,
                index=0,
                key="ps_tipo_manejo"
            )

        col3, col4 = st.columns(2)

        with col3:
            sala_opciones = ["Todas"] + list(SALA_MAP_INTERNAL.keys())
            sala_seleccionada = st.selectbox(
                "🏢 Sala de Proceso",
                options=sala_opciones,
                index=0,
                key="ps_sala",
                help="Sala donde se procesó el pallet"
            )

        with col4:
            planta_opciones = ["Todas", "VILKUN", "RIO FUTURO"]
            planta_seleccionada = st.selectbox(
                "🏭 Planta / Operación",
                options=planta_opciones,
                index=0,
                key="ps_planta"
            )

        # Filtro de productos (selectbox con datos de Odoo)
        opciones_productos = ["Todos"]
        try:
            productos_odoo = _obtener_productos_pt(username, password)
            if productos_odoo:
                opciones_productos += [f"{p['code']} — {p['name']}" for p in productos_odoo]
        except Exception:
            # Fallback: usar datos del cache si existen
            cached_data = st.session_state.get("ps_pallets_data")
            if cached_data and cached_data.get('detalle'):
                productos_unicos = {}
                for d in cached_data['detalle']:
                    code = d.get('codigo_producto', '')
                    name = d.get('producto', '')
                    if code and code not in productos_unicos:
                        productos_unicos[code] = f"{code} — {name}"
                opciones_productos = ["Todos"] + sorted(productos_unicos.values())

        producto_seleccionado = st.selectbox(
            "📦 Filtrar por Producto",
            options=opciones_productos,
            index=0,
            key="ps_producto_sel",
        )

        # Búsqueda libre adicional (pallet, OF, lote)
        producto_texto = st.text_input(
            "🔎 Buscar pallet, OF o lote",
            key="ps_producto_texto",
            placeholder="Ej: PACK0013713, WPF/00123..."
        )

        consultar = st.button("🔍 Consultar Pallets", use_container_width=True, type="primary", key="ps_btn_consultar")

    # === CSS PREMIUM ===
    st.markdown("""
    <style>
        @keyframes shimmer { 0%{background-position:200% 0} 100%{background-position:-200% 0} }
        @keyframes pulse-glow { 0%,100%{box-shadow:0 0 8px rgba(99,102,241,0.15)} 50%{box-shadow:0 0 20px rgba(99,102,241,0.35)} }
        @keyframes float-in { 0%{opacity:0;transform:translateY(12px)} 100%{opacity:1;transform:translateY(0)} }
        @keyframes count-up { 0%{opacity:0;transform:scale(0.8)} 100%{opacity:1;transform:scale(1)} }
        @keyframes gradient-shift { 0%{background-position:0% 50%} 50%{background-position:100% 50%} 100%{background-position:0% 50%} }

        .ps-hero {
            background: linear-gradient(135deg, #0f172a 0%, #1e1b4b 40%, #312e81 70%, #1e1b4b 100%);
            background-size: 300% 300%;
            animation: gradient-shift 8s ease infinite;
            border-radius: 16px;
            padding: 1.5rem 2rem;
            margin-bottom: 1rem;
            border: 1px solid rgba(99,102,241,0.2);
            position: relative;
            overflow: hidden;
        }
        .ps-hero::before {
            content: '';
            position: absolute;
            top: -50%; right: -50%;
            width: 100%; height: 100%;
            background: radial-gradient(circle, rgba(99,102,241,0.08) 0%, transparent 70%);
            pointer-events: none;
        }
        .ps-hero h2 {
            margin: 0 0 0.3rem 0;
            font-size: 1.6rem;
            font-weight: 800;
            color: #e0e7ff;
            letter-spacing: -0.02em;
        }
        .ps-hero .subtitle {
            color: #818cf8;
            font-size: 0.85rem;
            font-weight: 500;
            letter-spacing: 0.02em;
        }

        .kpi-card {
            background: linear-gradient(145deg, rgba(30,41,59,0.95) 0%, rgba(15,23,42,0.98) 100%);
            border: 1px solid rgba(99,102,241,0.15);
            border-radius: 14px;
            padding: 1rem;
            text-align: center;
            position: relative;
            overflow: hidden;
            animation: float-in 0.5s ease-out both;
            transition: all 0.3s cubic-bezier(0.4,0,0.2,1);
        }
        .kpi-card:hover {
            border-color: rgba(99,102,241,0.4);
            transform: translateY(-2px);
            box-shadow: 0 8px 25px rgba(0,0,0,0.3), 0 0 15px rgba(99,102,241,0.1);
        }
        .kpi-card .icon-ring {
            width: 40px; height: 40px;
            border-radius: 12px;
            display: flex; align-items: center; justify-content: center;
            margin: 0 auto 0.5rem;
            font-size: 1.1rem;
        }
        .kpi-card .value {
            font-size: 1.5rem;
            font-weight: 800;
            color: #f1f5f9;
            line-height: 1.1;
            animation: count-up 0.6s ease-out both;
        }
        .kpi-card .label {
            font-size: 0.65rem;
            color: #64748b;
            text-transform: uppercase;
            letter-spacing: 0.08em;
            margin-top: 0.25rem;
            font-weight: 600;
        }
        .kpi-card .accent-bar {
            position: absolute;
            top: 0; left: 0; right: 0;
            height: 3px;
            border-radius: 14px 14px 0 0;
        }

        .qc-banner {
            background: linear-gradient(135deg, rgba(16,185,129,0.08) 0%, rgba(6,78,59,0.12) 100%);
            border: 1px solid rgba(16,185,129,0.2);
            border-radius: 12px;
            padding: 0.8rem 1.2rem;
            display: flex;
            align-items: center;
            gap: 12px;
            margin-bottom: 0.8rem;
        }
        .qc-banner .qc-icon {
            width: 38px; height: 38px;
            background: linear-gradient(135deg, #10b981, #059669);
            border-radius: 10px;
            display: flex; align-items: center; justify-content: center;
            font-size: 1.1rem;
            flex-shrink: 0;
        }
        .qc-banner .qc-text { color: #a7f3d0; font-size: 0.85rem; font-weight: 500; }
        .qc-banner .qc-text strong { color: #ecfdf5; }
        .qc-banner .qc-files { color: #6ee7b7; font-size: 0.72rem; margin-top: 2px; }

        .sala-header {
            display: flex;
            align-items: center;
            gap: 12px;
            padding: 0.4rem 0;
        }
        .sala-badge {
            display: inline-flex;
            align-items: center;
            gap: 5px;
            background: rgba(99,102,241,0.12);
            color: #a5b4fc;
            padding: 3px 10px;
            border-radius: 20px;
            font-size: 0.72rem;
            font-weight: 600;
        }

        .grade-pill {
            display: inline-block;
            padding: 0.35rem 0.6rem;
            border-radius: 10px;
            text-align: center;
            min-width: 80px;
            transition: transform 0.2s;
        }
        .grade-pill:hover { transform: scale(1.05); }
        .grade-pill .g-name { font-size: 0.65rem; font-weight: 700; color: rgba(255,255,255,0.95); text-transform: uppercase; letter-spacing: 0.03em; }
        .grade-pill .g-val { font-size: 1.15rem; font-weight: 800; color: #fff; line-height: 1.2; }
        .grade-pill .g-unit { font-size: 0.55rem; color: rgba(255,255,255,0.7); }

        .coverage-bar-outer {
            background: #1e293b;
            border-radius: 8px;
            height: 8px;
            overflow: hidden;
            margin: 6px 0;
        }
        .coverage-bar-inner {
            height: 100%;
            border-radius: 8px;
            background: linear-gradient(90deg, #6366f1, #8b5cf6);
            transition: width 0.8s ease;
        }

        .export-card {
            background: linear-gradient(145deg, rgba(30,41,59,0.9), rgba(15,23,42,0.95));
            border: 1px solid #334155;
            border-radius: 14px;
            padding: 1.2rem;
            text-align: center;
        }
        .export-card .ex-title {
            font-size: 0.7rem;
            color: #64748b;
            text-transform: uppercase;
            letter-spacing: 0.06em;
            margin-bottom: 0.6rem;
            font-weight: 600;
        }

        .section-divider {
            border: none;
            height: 1px;
            background: linear-gradient(90deg, transparent, rgba(99,102,241,0.3), transparent);
            margin: 1.5rem 0;
        }

        .stat-chip {
            display: inline-flex;
            align-items: center;
            gap: 4px;
            background: rgba(99,102,241,0.1);
            color: #c7d2fe;
            padding: 2px 8px;
            border-radius: 6px;
            font-size: 0.7rem;
            font-weight: 600;
        }
    </style>
    """, unsafe_allow_html=True)

    # === CARGA AUTOMÁTICA DE CALIDAD (silenciosa, cacheada en session_state) ===
    calidad_dict = _cargar_calidad_servidor()

    # === CONSULTA ===
    if consultar or st.session_state.get("ps_pallets_data"):
        if consultar:
            fecha_inicio_str = fecha_inicio.strftime("%Y-%m-%d")
            fecha_fin_str = fecha_fin.strftime("%Y-%m-%d")

            # Todos los filtros de fruta, manejo y sala se aplican en memoria
            # para evitar problemas de acentos/encoding con Odoo ilike

            with st.spinner("⏳ Consultando pallets por sala..."):
                try:
                    params = {
                        "username": username,
                        "password": password,
                        "fecha_inicio": fecha_inicio_str,
                        "fecha_fin": fecha_fin_str,
                        "tipo_operacion": planta_seleccionada
                    }

                    response = requests.get(
                        f"{API_URL}/api/v1/produccion/clasificacion",
                        params=params,
                        timeout=120
                    )

                    if response.status_code == 200:
                        data = response.json()
                        st.session_state.ps_pallets_data = data
                        st.session_state.ps_data_loaded = True
                        count = len(data.get('detalle', []))
                        if count > 0:
                            st.toast(f"✅ {count} pallets cargados")
                            if _throttle_rerun("ps_last_rerun"):
                                st.rerun()
                        else:
                            st.info("ℹ️ No se encontraron pallets para los filtros seleccionados.")
                    else:
                        st.error(f"❌ Error al obtener datos: {response.status_code}")
                        return

                except Exception as e:
                    st.error(f"❌ Error en la consulta: {str(e)}")
                    return

        # --- Procesar datos ---
        data = st.session_state.get("ps_pallets_data")
        if not data:
            st.info("👆 Consulta los datos primero.")
            return

        detalle_raw = data.get('detalle', [])

        # Aplicar filtros en memoria
        filtros = {
            "planta": planta_seleccionada,
            "tipo_fruta": tipo_fruta,
            "sala": sala_seleccionada,
            "tipo_manejo": tipo_manejo,
            "producto_seleccionado": producto_seleccionado,
            "producto_texto": producto_texto,
        }
        detalle_filtrado = _filtrar_detalle(detalle_raw, filtros)

        if not detalle_filtrado:
            st.warning("⚠️ No hay pallets con los filtros seleccionados (solo temporada 2025/2026, producto terminado).")
            return

        # Convertir a DataFrame
        df = pd.DataFrame(detalle_filtrado)

        # Banner de calidad auto-cargada (solo si hay datos QC)
        if calidad_dict:
            archivos_ok = st.session_state.get('_qc_archivos_cargados', [])
            archivos_error = st.session_state.get('_qc_archivos_error', [])
            if archivos_ok:
                total_pallets_qc_loaded = len(calidad_dict)
                total_registros_qc_loaded = sum(len(v) for v in calidad_dict.values())
                archivos_txt = " &bull; ".join(archivos_ok)
                st.markdown(f"""
                <div class="qc-banner">
                    <div class="qc-icon">🔬</div>
                    <div>
                        <div class="qc-text">Control de calidad activo: <strong>{total_registros_qc_loaded}</strong> monitoreos enlazados a <strong>{total_pallets_qc_loaded}</strong> pallets</div>
                        <div class="qc-files">Archivos: {archivos_txt}</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
            if archivos_error:
                for err in archivos_error:
                    st.warning(f"⚠️ Error en archivo: {err}")

        # --- KPIs GENERALES ---
        total_pallets = df['pallet'].nunique()
        total_kg = df['kg'].sum()
        total_productos = df['producto'].nunique()
        salas_activas = df['sala'].nunique()

        # Contar pallets con calidad
        pallets_con_qc = 0
        if calidad_dict:
            for pname in df['pallet'].unique():
                if _extraer_numero_pallet(pname) in calidad_dict:
                    pallets_con_qc += 1

        st.markdown('<hr class="section-divider">', unsafe_allow_html=True)

        kpi_data = [
            ("📦", fmt_numero(total_pallets), "Pallets Únicos", "#6366f1", "linear-gradient(135deg,#6366f1,#818cf8)"),
            ("⚖️", fmt_numero(total_kg), "Total Kilogramos", "#10b981", "linear-gradient(135deg,#10b981,#34d399)"),
            ("🍇", fmt_numero(total_productos), "Productos", "#f59e0b", "linear-gradient(135deg,#f59e0b,#fbbf24)"),
            ("🏢", fmt_numero(salas_activas), "Salas Activas", "#8b5cf6", "linear-gradient(135deg,#8b5cf6,#a78bfa)"),
        ]
        if calidad_dict:
            pct_qc = round(pallets_con_qc / total_pallets * 100) if total_pallets > 0 else 0
            qc_color = "#22c55e" if pct_qc >= 80 else ("#f59e0b" if pct_qc >= 50 else "#ef4444")
            qc_grad = f"linear-gradient(135deg,{qc_color},{qc_color}cc)"
            kpi_data.append(("🔬", f"{pallets_con_qc}/{total_pallets}", f"Con Calidad ({pct_qc}%)", qc_color, qc_grad))

        kpi_cols = st.columns(len(kpi_data))
        for i, (icon, value, label, color, gradient) in enumerate(kpi_data):
            with kpi_cols[i]:
                delay = f"{i * 0.1}s"
                st.markdown(f"""
                <div class="kpi-card" style="animation-delay:{delay};">
                    <div class="accent-bar" style="background:{gradient};"></div>
                    <div class="icon-ring" style="background:{color}22;">{icon}</div>
                    <div class="value">{value}</div>
                    <div class="label">{label}</div>
                </div>""", unsafe_allow_html=True)

        # --- COBERTURA QC ---
        if calidad_dict and total_pallets > 0:
            pct_coverage = round(pallets_con_qc / total_pallets * 100)
            bar_color = "#22c55e" if pct_coverage >= 80 else ("#f59e0b" if pct_coverage >= 50 else "#ef4444")
            st.markdown(f"""
            <div style="background:rgba(30,41,59,0.6); border:1px solid #334155; border-radius:10px; padding:0.6rem 1rem; margin-top:0.8rem;">
                <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:4px;">
                    <span style="font-size:0.72rem; color:#94a3b8; font-weight:600; text-transform:uppercase; letter-spacing:0.05em;">Cobertura de Calidad</span>
                    <span style="font-size:0.8rem; color:{bar_color}; font-weight:700;">{pct_coverage}%</span>
                </div>
                <div class="coverage-bar-outer">
                    <div class="coverage-bar-inner" style="width:{pct_coverage}%; background:linear-gradient(90deg,{bar_color},{bar_color}bb);"></div>
                </div>
            </div>""", unsafe_allow_html=True)

        st.markdown('<hr class="section-divider">', unsafe_allow_html=True)

        # --- RESUMEN POR SALA ---
        st.markdown("""
        <div style="display:flex; align-items:center; gap:10px; margin-bottom:0.5rem;">
            <span style="font-size:1.2rem;">🏢</span>
            <span style="font-size:1.1rem; font-weight:700; color:#e0e7ff;">Resumen por Sala de Proceso</span>
            <span class="stat-chip">%d salas</span>
        </div>""" % salas_activas, unsafe_allow_html=True)

        # Agrupar pallets por sala
        salas_unicas = sorted(df['sala'].unique(), key=lambda x: x if x else "ZZZ")

        for sala in salas_unicas:
            sala_label = sala if sala else "Sin Sala"
            df_sala = df[df['sala'] == sala]

            kg_sala = df_sala['kg'].sum()
            pallets_sala = df_sala['pallet'].nunique()
            productos_sala = df_sala['producto'].nunique()

            # Contar QC en esta sala
            qc_badge = ""
            if calidad_dict:
                sala_con_qc = sum(1 for p in df_sala['pallet'].unique() if _extraer_numero_pallet(p) in calidad_dict)
                if sala_con_qc > 0:
                    qc_pct_sala = round(sala_con_qc / pallets_sala * 100) if pallets_sala > 0 else 0
                    qc_badge = f" | 🔬 {sala_con_qc}/{pallets_sala} QC ({qc_pct_sala}%)"

            # Desglose por grado
            grado_resumen = df_sala.groupby('grado')['kg'].sum().sort_values(ascending=False)

            # Porcentaje de kg respecto al total
            pct_kg_total = round(kg_sala / total_kg * 100, 1) if total_kg > 0 else 0

            with st.expander(f"🏢 **{sala_label}** — {fmt_numero(kg_sala)} kg ({pct_kg_total}%) | {pallets_sala} pallets | {productos_sala} prod.{qc_badge}", expanded=False):

                # Grade pills
                gcols = st.columns(min(len(grado_resumen), 7))
                for i, (grado, kg_g) in enumerate(grado_resumen.items()):
                    if i >= 7:
                        break
                    color = GRADO_COLORES.get(grado, '#64748b')
                    with gcols[i]:
                        st.markdown(f"""
                        <div class="grade-pill" style="background:linear-gradient(145deg,{color},{color}cc);">
                            <div class="g-name">{grado}</div>
                            <div class="g-val">{fmt_numero(kg_g)}</div>
                            <div class="g-unit">kg</div>
                        </div>""", unsafe_allow_html=True)

                st.markdown("")

                # Resumen por producto
                resumen_producto = (
                    df_sala.groupby(['producto', 'codigo_producto', 'grado'])
                    .agg(kg_total=('kg', 'sum'), pallets=('pallet', 'nunique'))
                    .reset_index()
                    .sort_values('kg_total', ascending=False)
                )
                resumen_producto.columns = ['Producto', 'Código', 'Grado', 'Kilos', 'Pallets']
                resumen_producto['Kilos'] = resumen_producto['Kilos'].apply(lambda x: round(x, 2))

                st.caption("📊 Resumen por producto")
                st.dataframe(
                    resumen_producto,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Kilos": st.column_config.NumberColumn(format="%.2f"),
                        "Pallets": st.column_config.NumberColumn(format="%d"),
                    }
                )

                # Detalle pallet por pallet
                df_sala_detail = df_sala[['pallet', 'producto', 'codigo_producto', 'grado', 'kg', 'orden_fabricacion', 'fecha']].copy()
                df_sala_detail['kg'] = df_sala_detail['kg'].apply(lambda x: round(x, 2))
                df_sala_detail.columns = ['Pallet', 'Producto', 'Código', 'Grado', 'Kilos', 'OF', 'Inicio Proceso']
                df_sala_detail = df_sala_detail.sort_values(['Producto', 'Grado', 'Pallet'], ascending=[True, True, True])

                # Agregar indicador QC a la tabla de detalle
                if calidad_dict:
                    df_sala_detail['QC'] = df_sala_detail['Pallet'].apply(
                        lambda p: f"🔬 {len(calidad_dict.get(_extraer_numero_pallet(p), []))}" if _extraer_numero_pallet(p) in calidad_dict else "—"
                    )

                st.caption(f"📋 Detalle: {len(df_sala_detail)} pallets")
                st.dataframe(
                    df_sala_detail,
                    use_container_width=True,
                    hide_index=True,
                    height=min(400, 35 * len(df_sala_detail) + 38),
                    column_config={
                        "Kilos": st.column_config.NumberColumn(format="%.2f"),
                    }
                )

                # === QC INLINE: expander por cada pallet con datos de calidad ===
                if calidad_dict:
                    pallets_con_datos = [
                        p for p in df_sala_detail['Pallet'].unique()
                        if _extraer_numero_pallet(p) in calidad_dict
                    ]
                    for pallet_name in pallets_con_datos:
                        pallet_num = _extraer_numero_pallet(pallet_name)
                        registros = calidad_dict.get(pallet_num, [])
                        if not registros:
                            continue
                        # Resumen rápido para el título del expander
                        pi = [float(r.get('pct_iqf', 0) or 0) for r in registros if r.get('pct_iqf')]
                        pd2 = [float(r.get('pct_defectos', 0) or 0) for r in registros if r.get('pct_defectos')]
                        avg_i = sum(pi) / len(pi) * 100 if pi else 0
                        avg_d = sum(pd2) / len(pd2) * 100 if pd2 else 0
                        icon = "✅" if avg_d <= 4 else ("⚠️" if avg_d <= 7 else "❌")
                        iqf_str = f"IQF {avg_i:.0f}%" if pi else ""
                        def_str = f"Def {avg_d:.1f}%" if pd2 else ""
                        resumen = " | ".join(filter(None, [iqf_str, def_str]))
                        with st.expander(f"🔬 **{pallet_name}** — {icon} {resumen}", expanded=False):
                            _render_calidad_pallet(registros, pallet_name)

        # --- TABLA DETALLADA COMPLETA ---
        st.markdown('<hr class="section-divider">', unsafe_allow_html=True)
        st.markdown("""
        <div style="display:flex; align-items:center; gap:10px; margin-bottom:0.5rem;">
            <span style="font-size:1.2rem;">📋</span>
            <span style="font-size:1.1rem; font-weight:700; color:#e0e7ff;">Detalle Completo de Pallets</span>
            <span class="stat-chip">%d registros</span>
        </div>""" % len(df), unsafe_allow_html=True)

        # Preparar DataFrame de display
        SALA_REVERSE = {v: k for k, v in SALA_MAP_INTERNAL.items()}

        df_display = df[[
            'pallet', 'producto', 'codigo_producto', 'grado', 'kg',
            'orden_fabricacion', 'planta', 'sala', 'fecha'
        ]].copy()
        df_display['sala'] = df_display['sala'].map(lambda x: SALA_REVERSE.get(x, x))
        df_display['kg'] = df_display['kg'].apply(lambda x: round(x, 2))
        df_display.columns = [
            'Pallet', 'Producto', 'Código', 'Grado', 'Kilos',
            'OF', 'Planta', 'Sala', 'Inicio Proceso'
        ]
        df_display = df_display.sort_values(['Sala', 'Producto', 'Grado'], ascending=[True, True, True])

        # Agregar columna QC a la tabla general
        if calidad_dict:
            df_display['QC'] = df_display['Pallet'].apply(
                lambda p: f"🔬 {len(calidad_dict.get(_extraer_numero_pallet(p), []))}" if _extraer_numero_pallet(p) in calidad_dict else "—"
            )

        st.dataframe(
            df_display,
            use_container_width=True,
            hide_index=True,
            height=600,
            column_config={
                "Kilos": st.column_config.NumberColumn(format="%.2f"),
            }
        )

        # --- EXPORTAR ---
        st.markdown('<hr class="section-divider">', unsafe_allow_html=True)
        st.markdown("""
        <div style="display:flex; align-items:center; gap:10px; margin-bottom:0.8rem;">
            <span style="font-size:1.2rem;">💾</span>
            <span style="font-size:1.1rem; font-weight:700; color:#e0e7ff;">Exportar Datos</span>
        </div>""", unsafe_allow_html=True)

        col_exp1, col_exp2 = st.columns(2)

        with col_exp1:
            st.markdown("""<div class="export-card"><div class="ex-title">Formato CSV</div></div>""", unsafe_allow_html=True)
            csv_data = df_display.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="📥 Descargar CSV",
                data=csv_data,
                file_name=f"pallets_por_sala_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True
            )

        with col_exp2:
            qc_label = "con Calidad" if calidad_dict else ""
            st.markdown(f"""<div class="export-card"><div class="ex-title">Formato Excel {qc_label}</div></div>""", unsafe_allow_html=True)
            try:
                excel_bytes = _generar_excel_con_calidad(df_display, calidad_dict)
                label_excel = "📊 Descargar Excel + Calidad" if calidad_dict else "📊 Descargar Excel"
                st.download_button(
                    label=label_excel,
                    data=excel_bytes,
                    file_name=f"pallets_sala_calidad_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as e:
                st.error(f"Error al preparar Excel: {e}")
    else:
        st.markdown("""
        <div style="background:linear-gradient(135deg,rgba(99,102,241,0.06),rgba(139,92,246,0.06));
            border:1px solid rgba(99,102,241,0.15); border-radius:14px; padding:2rem; text-align:center; margin:1rem 0;">
            <div style="font-size:2.5rem; margin-bottom:0.5rem;">📊</div>
            <div style="font-size:1rem; color:#a5b4fc; font-weight:600;">Selecciona los filtros y consultá</div>
            <div style="font-size:0.8rem; color:#64748b; margin-top:0.3rem;">Usá el botón <strong style="color:#818cf8;">Consultar Pallets</strong> para cargar los datos</div>
        </div>""", unsafe_allow_html=True)
