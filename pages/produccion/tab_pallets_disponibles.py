"""
Tab Pallets Disponibles: Muestra pallets con stock que NO están en ninguna fabricación.
Excluye ubicaciones de stock final y cámaras de congelado.
"""
import streamlit as st
import httpx
import pandas as pd
from typing import Dict, List, Any
from streamlit_echarts import st_echarts
from .shared import API_URL


def fetch_pallets_disponibles(username: str, password: str, 
                               planta: str = None,
                               producto_id: int = None,
                               proveedor_id: int = None,
                               fecha_desde: str = None,
                               fecha_hasta: str = None,
                               pallet_codigo: str = None) -> Dict[str, Any]:
    """Obtiene pallets disponibles del backend."""
    params = {
        "username": username,
        "password": password,
    }
    if planta and planta != "Todas":
        params["planta"] = planta
    if producto_id:
        params["producto_id"] = producto_id
    if proveedor_id:
        params["proveedor_id"] = proveedor_id
    if fecha_desde:
        params["fecha_desde"] = fecha_desde
    if fecha_hasta:
        params["fecha_hasta"] = fecha_hasta
    if pallet_codigo:
        params["pallet_codigo"] = pallet_codigo
    
    response = httpx.get(f"{API_URL}/api/v1/produccion/pallets-disponibles",
                         params=params, timeout=120.0)
    response.raise_for_status()
    return response.json()


def fetch_productos_2026(username: str, password: str) -> List[Dict]:
    """Obtiene productos del año 2026."""
    params = {
        "username": username,
        "password": password,
    }
    response = httpx.get(f"{API_URL}/api/v1/produccion/pallets-disponibles/productos-2026",
                         params=params, timeout=60.0)
    response.raise_for_status()
    return response.json().get('productos', [])


def fetch_proveedores_compras(username: str, password: str) -> List[Dict]:
    """Obtiene proveedores del módulo de compras."""
    params = {
        "username": username,
        "password": password,
    }
    response = httpx.get(f"{API_URL}/api/v1/produccion/pallets-disponibles/proveedores",
                         params=params, timeout=60.0)
    response.raise_for_status()
    return response.json().get('proveedores', [])


def render(username: str = None, password: str = None):
    """Renderiza el tab de Pallets Disponibles."""
    
    if not username or not password:
        username = st.session_state.get("username", "")
        password = st.session_state.get("password", "")
    
    # === HEADER ===
    st.markdown("""
    <div style="background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%); 
                padding: 20px; border-radius: 12px; border-left: 4px solid #e94560;
                margin-bottom: 20px;">
        <h2 style="color: #e94560; margin: 0;">📦 Pallets Disponibles</h2>
        <p style="color: #aaa; margin: 5px 0 0 0;">
            Pallets con stock que <b>NO están asignados</b> a ninguna orden de fabricación
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # === FILTROS ===
    col1, col2, col3 = st.columns([2, 2, 2])
    with col1:
        planta_sel = st.selectbox(
            "🏭 Planta",
            ["Todas", "RIO FUTURO", "VILKUN"],
            key="pallets_disp_planta"
        )
    with col2:
        # Cargar productos 2026 si no están en session
        if 'pallets_disp_productos_2026' not in st.session_state:
            try:
                prods_2026 = fetch_productos_2026(username, password)
                st.session_state['pallets_disp_productos_2026'] = prods_2026
            except Exception as e:
                st.session_state['pallets_disp_productos_2026'] = []
        
        productos_2026 = st.session_state.get('pallets_disp_productos_2026', [])
        opciones_producto = ["Todos"] + [p['nombre'] for p in productos_2026]
        producto_sel = st.selectbox(
            "📦 Producto",
            opciones_producto,
            key="pallets_disp_producto"
        )
    with col3:
        # Cargar proveedores si no están en session
        if 'pallets_disp_proveedores' not in st.session_state:
            try:
                provs = fetch_proveedores_compras(username, password)
                st.session_state['pallets_disp_proveedores'] = provs
            except Exception as e:
                st.session_state['pallets_disp_proveedores'] = []
        
        proveedores = st.session_state.get('pallets_disp_proveedores', [])
        opciones_proveedor = ["Todos"] + [p['nombre'] for p in proveedores]
        proveedor_sel = st.selectbox(
            "🥑 Productor (Frescos)",
            opciones_proveedor,
            key="pallets_disp_proveedor"
        )
    
    # Fila 2: fechas y búsqueda de pallet
    from datetime import date, timedelta
    col_f1, col_f2, col_f3, col_btn = st.columns([2, 2, 2, 1])
    with col_f1:
        fecha_desde = st.date_input(
            "📅 Fecha Desde",
            value=date.today() - timedelta(days=30),
            key="pallets_disp_fecha_desde"
        )
    with col_f2:
        fecha_hasta = st.date_input(
            "📅 Fecha Hasta",
            value=date.today(),
            key="pallets_disp_fecha_hasta"
        )
    with col_f3:
        pallet_buscar = st.text_input(
            "🔎 Buscar Pallet",
            placeholder="PACK0012345",
            key="pallets_disp_buscar_codigo",
            help="Busca un pallet específico por su código"
        )
    with col_btn:
        st.markdown("<br>", unsafe_allow_html=True)
        btn_buscar = st.button("🔍 Buscar", type="primary", 
                                use_container_width=True, key="pallets_disp_buscar")
    
    st.markdown("---")
    
    # === CARGAR DATOS ===
    if btn_buscar:
        st.cache_data.clear()
        try:
            # Obtener IDs de producto y proveedor seleccionados
            productos_2026 = st.session_state.get('pallets_disp_productos_2026', [])
            producto_id = None
            if producto_sel and producto_sel != "Todos":
                for p in productos_2026:
                    if p['nombre'] == producto_sel:
                        producto_id = p['id']
                        break
            
            proveedores = st.session_state.get('pallets_disp_proveedores', [])
            proveedor_id = None
            if proveedor_sel and proveedor_sel != "Todos":
                for prov in proveedores:
                    if prov['nombre'] == proveedor_sel:
                        proveedor_id = prov['id']
                        break
            
            # Preparar filtros de fecha
            f_desde = fecha_desde.isoformat() if fecha_desde else None
            f_hasta = fecha_hasta.isoformat() if fecha_hasta else None
            p_codigo = pallet_buscar.strip().upper() if pallet_buscar and pallet_buscar.strip() else None
            
            with st.spinner("Buscando pallets disponibles..."):
                data = fetch_pallets_disponibles(
                    username, password, planta_sel, producto_id, proveedor_id,
                    f_desde, f_hasta, p_codigo
                )
                st.session_state['pallets_disp_data'] = data
                st.session_state['pallets_disp_loaded'] = True
        except Exception as e:
            st.error(f"Error al cargar datos: {str(e)}")
            return
    
    if not st.session_state.get('pallets_disp_loaded', False):
        st.info("👆 Selecciona la planta y presiona **'🔍 Buscar Pallets'** para ver los pallets disponibles")
        return
    
    data = st.session_state.get('pallets_disp_data', {})
    pallets = data.get('pallets', [])
    
    # Recalcular estadísticas con los pallets filtrados
    stats = {
        'total_pallets': len(pallets),
        'total_kg': sum(p.get('cantidad_kg', 0) for p in pallets),
        'congelados': len([p for p in pallets if p.get('tipo') == 'Congelado']),
        'frescos': len([p for p in pallets if p.get('tipo') == 'Fresco']),
    }
    
    if not pallets:
        st.warning("No se encontraron pallets disponibles con los filtros seleccionados")
        return
    
    # === KPIs ===
    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.metric("📦 Total Pallets", f"{stats.get('total_pallets', 0):,}")
    with k2:
        st.metric("⚖️ KG Totales", f"{stats.get('total_kg', 0):,.0f}")
    with k3:
        st.metric("❄️ Congelados", f"{stats.get('congelados', 0):,}")
    with k4:
        st.metric("🌿 Frescos", f"{stats.get('frescos', 0):,}")
    
    st.markdown("---")
    
    # === GRÁFICOS ===
    col_g1, col_g2 = st.columns(2)
    
    with col_g1:
        # Gráfico tipo (Congelado vs Fresco)
        render_grafico_tipo(stats)
    
    with col_g2:
        # Gráfico por planta (recalculado con filtro)
        por_planta_filtrado = {}
        for p in pallets:
            pl = p.get('planta', 'Otro')
            if pl not in por_planta_filtrado:
                por_planta_filtrado[pl] = {'cantidad': 0, 'kg': 0}
            por_planta_filtrado[pl]['cantidad'] += 1
            por_planta_filtrado[pl]['kg'] += p.get('cantidad_kg', 0)
        render_grafico_planta(por_planta_filtrado)
    
    st.markdown("---")
    
    # === GRÁFICO POR UBICACIÓN ===
    render_grafico_ubicacion(pallets)
    
    st.markdown("---")
    
    # === BUSCAR PAQUETE EN ODOO ===
    st.markdown("### 🔗 Buscar Paquete en Odoo")
    
    ODOO_BASE = "https://riofuturo.server98c6e.oerpondemand.net"
    
    col_odoo1, col_odoo2 = st.columns([3, 1])
    with col_odoo1:
        buscar_paquete = st.text_input(
            "📦 Ingresa el nombre del pallet/paquete",
            placeholder="Ej: PACK0012345",
            key="pallets_buscar_odoo"
        )
    with col_odoo2:
        st.markdown("<br>", unsafe_allow_html=True)
        if buscar_paquete:
            # Buscar el pallet en los datos cargados
            pallet_encontrado = None
            for p in pallets:
                if buscar_paquete.upper() in p.get('pallet', '').upper():
                    pallet_encontrado = p
                    break
            
            if pallet_encontrado:
                pallet_id = pallet_encontrado.get('pallet_id', 0)
                odoo_url = f"{ODOO_BASE}/web#id={pallet_id}&model=stock.quant.package&view_type=form"
                st.link_button("🔗 Abrir en Odoo", odoo_url, use_container_width=True)
            else:
                st.warning("No encontrado")
    
    st.markdown("---")
    
    # === TABLA DE PALLETS ===
    st.markdown("### 📋 Detalle de Pallets Disponibles")
    
    # Filtros de tabla
    col_f1, col_f2 = st.columns(2)
    with col_f1:
        tipo_filtro = st.selectbox("Filtrar por tipo", ["Todos", "Congelado", "Fresco"],
                                    key="pallets_tipo_filtro")
    with col_f2:
        buscar_texto = st.text_input("🔎 Buscar pallet/producto/lote", 
                                      key="pallets_buscar_texto")
    
    # Aplicar filtros
    pallets_filtrados = pallets
    if tipo_filtro != "Todos":
        pallets_filtrados = [p for p in pallets_filtrados if p['tipo'] == tipo_filtro]
    if buscar_texto:
        texto = buscar_texto.upper()
        pallets_filtrados = [p for p in pallets_filtrados 
                             if texto in p.get('pallet', '').upper()
                             or texto in p.get('producto', '').upper()
                             or texto in p.get('lote', '').upper()
                             or texto in p.get('ubicacion', '').upper()]
    
    # Crear DataFrame con link a Odoo
    if pallets_filtrados:
        df_data = []
        for p in pallets_filtrados:
            pid = p.get('pallet_id', 0)
            odoo_link = f"{ODOO_BASE}/web#id={pid}&model=stock.quant.package&view_type=form"
            df_data.append({
                'Pallet': p.get('pallet', ''),
                'Lote': p.get('lote', ''),
                'Producto': p.get('producto', ''),
                'KG': p.get('cantidad_kg', 0),
                'Ubicación': p.get('ubicacion', ''),
                'Tipo': p.get('tipo', ''),
                'Planta': p.get('planta', ''),
                'Fecha Creación': p.get('fecha_creacion', ''),
                'Fecha Ingreso': p.get('fecha_ingreso', ''),
                'Productor': p.get('proveedor', ''),
                'Ver en Odoo': odoo_link
            })
        
        df = pd.DataFrame(df_data)
        df['KG'] = df['KG'].apply(lambda x: f"{x:,.1f}")
        
        st.dataframe(
            df,
            use_container_width=True,
            hide_index=True,
            height=500,
            column_config={
                "Pallet": st.column_config.TextColumn("📦 Pallet", width="medium"),
                "Lote": st.column_config.TextColumn("🏷️ Lote", width="medium"),
                "Producto": st.column_config.TextColumn("📋 Producto", width="large"),
                "KG": st.column_config.TextColumn("⚖️ KG", width="small"),
                "Ubicación": st.column_config.TextColumn("📍 Ubicación", width="medium"),
                "Tipo": st.column_config.TextColumn("🔄 Tipo", width="small"),
                "Planta": st.column_config.TextColumn("🏭 Planta", width="small"),
                "Fecha Creación": st.column_config.TextColumn("📅 Creación", width="small"),
                "Fecha Ingreso": st.column_config.TextColumn("📅 Ingreso", width="small"),
                "Productor": st.column_config.TextColumn("🥑 Productor", width="medium"),
                "Ver en Odoo": st.column_config.LinkColumn("🔗 Odoo", width="small", display_text="Abrir"),
            }
        )

        # === DESCARGA EXCEL ===
        from datetime import date as _date
        try:
            excel_bytes = generar_excel_pallets(pallets_filtrados)
            nombre_archivo = f"pallets_disponibles_{_date.today().isoformat()}.xlsx"
            st.download_button(
                label="📥 Descargar Excel",
                data=excel_bytes,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Descarga el Excel con todos los filtros activos: variedad, formato, cajas estimadas, antigüedad, MP/Congelado y más.",
                use_container_width=False,
                type="secondary",
            )
        except Exception as exc_xl:
            st.warning(f"No se pudo generar el Excel: {exc_xl}")

        st.caption(f"Mostrando **{len(pallets_filtrados)}** pallets de {len(pallets)} totales")
    else:
        st.info("No hay pallets que coincidan con los filtros")


def generar_excel_pallets(pallets_data: List[Dict]) -> bytes:
    """Genera un Excel enriquecido con autofilter, filas alternas y todas las columnas de análisis."""
    import io
    import re
    from datetime import date, datetime
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pallets Disponibles"

    today = date.today()

    def extraer_info_producto(nombre: str):
        """Extrae variedad, peso unitario y formato/envase del nombre del producto."""
        # Patrón: [CODIGO] VARIEDAD PESO kg en FORMATO
        m = re.search(r'\[[^\]]+\]\s+(.+?)\s+([\d]+[,\.][\d]+)\s*kg\s+en\s+(\w+)', nombre or '', re.IGNORECASE)
        if m:
            variedad = m.group(1).strip()
            try:
                peso = float(m.group(2).replace(',', '.'))
            except Exception:
                peso = None
            formato = m.group(3).strip().capitalize()
            return variedad, peso, formato
        # Fallback: extraer todo lo que está después del código entre corchetes
        m2 = re.match(r'\[[^\]]+\]\s+(.+)', nombre or '')
        if m2:
            return m2.group(1).strip(), None, None
        return nombre or '', None, None

    def calc_antiguedad(p: Dict):
        fecha_ref = p.get('fecha_ingreso') or p.get('fecha_creacion') or ''
        if fecha_ref:
            try:
                dt = datetime.strptime(str(fecha_ref)[:10], '%Y-%m-%d').date()
                return (today - dt).days
            except Exception:
                pass
        return ''

    headers = [
        'Pallet', 'Lote', 'Producto', 'Variedad', 'Formato',
        'Peso Unitario (kg)', 'Cajas/Bandejas (est.)',
        'Planta', 'Ubicación', 'MP o Congelado', 'Productor',
        'KG Totales', 'Fecha Ingreso', 'Fecha Creación', 'Antigüedad (días)',
    ]

    # Estilos cabecera
    hdr_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    hdr_font = Font(bold=True, color="E94560", size=11, name="Calibri")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
    # Filas alternas
    row_fill_a = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
    row_fill_b = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    row_font = Font(color="DDEEFF", size=10, name="Calibri")
    congelado_font = Font(color="4FC3F7", size=10, name="Calibri")
    mp_font = Font(color="81C784", size=10, name="Calibri")

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center_align
    ws.row_dimensions[1].height = 22

    for r_idx, p in enumerate(pallets_data, 2):
        variedad, peso_unit, formato = extraer_info_producto(p.get('producto', ''))
        kg_total = p.get('cantidad_kg', 0)
        cajas_est = round(kg_total / peso_unit) if peso_unit and peso_unit > 0 else ''
        tipo_raw = p.get('tipo', 'Fresco')
        mp_o_congelado = 'Congelado' if tipo_raw == 'Congelado' else 'Materia Prima'
        antiguedad = calc_antiguedad(p)

        row_data = [
            p.get('pallet', ''),
            p.get('lote', ''),
            p.get('producto', ''),
            variedad,
            formato or '',
            peso_unit if peso_unit is not None else '',
            cajas_est,
            p.get('planta', ''),
            p.get('ubicacion', ''),
            mp_o_congelado,
            p.get('proveedor', ''),
            kg_total,
            p.get('fecha_ingreso', '') or '',
            p.get('fecha_creacion', '') or '',
            antiguedad,
        ]
        ws.append(row_data)

        fill = row_fill_a if r_idx % 2 == 0 else row_fill_b
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.fill = fill
            # Colorear columna MP o Congelado (índice 10)
            if col_idx == 10:
                cell.font = congelado_font if mp_o_congelado == 'Congelado' else mp_font
            else:
                cell.font = row_font
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Autofilter en toda la tabla
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    # Congelar fila de cabecera
    ws.freeze_panes = 'A2'

    # Anchos de columna
    col_widths = [18, 20, 55, 32, 12, 18, 20, 14, 32, 16, 32, 14, 15, 15, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def render_grafico_tipo(stats: Dict):
    """Gráfico de dona: Congelado vs Fresco."""
    st.markdown("#### ❄️🌿 Tipo de Pallet")
    
    congelados = stats.get('congelados', 0)
    frescos = stats.get('frescos', 0)
    
    if congelados == 0 and frescos == 0:
        st.info("Sin datos")
        return
    
    options = {
        "tooltip": {
            "trigger": "item",
            "formatter": "{b}: {c} pallets ({d}%)"
        },
        "legend": {
            "bottom": "0%",
            "textStyle": {"color": "#ddd"}
        },
        "series": [{
            "type": "pie",
            "radius": ["40%", "70%"],
            "center": ["50%", "45%"],
            "avoidLabelOverlap": True,
            "itemStyle": {"borderRadius": 8, "borderColor": "#1a1a2e", "borderWidth": 2},
            "label": {
                "show": True,
                "formatter": "{b}\n{c} pallets",
                "color": "#ddd"
            },
            "data": [
                {"value": congelados, "name": "❄️ Congelado", 
                 "itemStyle": {"color": "#4fc3f7"}},
                {"value": frescos, "name": "🌿 Fresco", 
                 "itemStyle": {"color": "#81c784"}},
            ]
        }]
    }
    
    st_echarts(options=options, height="280px")


def render_grafico_planta(por_planta: Dict):
    """Gráfico de barras: Pallets por planta."""
    st.markdown("#### 🏭 Pallets por Planta")
    
    if not por_planta:
        st.info("Sin datos")
        return
    
    plantas = list(por_planta.keys())
    cantidades = [v['cantidad'] if isinstance(v, dict) else len(v) for v in por_planta.values()]
    kg_totales = [v['kg'] if isinstance(v, dict) else sum(p.get('cantidad_kg', 0) for p in v) for v in por_planta.values()]
    
    colores = {
        'RIO FUTURO': '#4ecdc4',
        'VILKUN': '#ff6b6b',
    }
    
    colors = [colores.get(p, '#ffd93d') for p in plantas]
    
    options = {
        "tooltip": {
            "trigger": "axis",
            "axisPointer": {"type": "shadow"}
        },
        "grid": {"left": "5%", "right": "5%", "bottom": "15%", "top": "10%", "containLabel": True},
        "xAxis": {
            "type": "category",
            "data": plantas,
            "axisLabel": {"color": "#ddd", "fontSize": 12}
        },
        "yAxis": {
            "type": "value",
            "name": "Cantidad",
            "nameTextStyle": {"color": "#aaa"},
            "axisLabel": {"color": "#ccc"},
            "splitLine": {"lineStyle": {"color": "#333"}}
        },
        "series": [
            {
                "name": "Pallets",
                "type": "bar",
                "data": [{"value": c, "itemStyle": {"color": colors[i]}} 
                         for i, c in enumerate(cantidades)],
                "label": {
                    "show": True,
                    "position": "top",
                    "formatter": "{c} pallets",
                    "color": "#fff",
                    "fontSize": 12
                },
                "barWidth": "50%"
            }
        ]
    }
    
    st_echarts(options=options, height="280px")


def render_grafico_ubicacion(pallets: List[Dict]):
    """Gráfico de barras horizontales: Top ubicaciones con más pallets."""
    st.markdown("### 📍 Pallets por Ubicación")
    st.caption("Top ubicaciones con más pallets disponibles")
    
    # Agrupar por ubicación
    ubicaciones = {}
    for p in pallets:
        ub = p.get('ubicacion', 'Sin ubicación')
        if ub not in ubicaciones:
            ubicaciones[ub] = {'pallets': 0, 'kg': 0}
        ubicaciones[ub]['pallets'] += 1
        ubicaciones[ub]['kg'] += p.get('cantidad_kg', 0)
    
    # Ordenar por cantidad de pallets y tomar top 15
    top_ub = sorted(ubicaciones.items(), key=lambda x: x[1]['pallets'], reverse=True)[:15]
    top_ub.reverse()  # Para que el más grande quede arriba en barras horizontales
    
    nombres = [u[0][-35:] for u in top_ub]  # Truncar nombres largos
    cantidades = [u[1]['pallets'] for u in top_ub]
    kg_list = [round(u[1]['kg'], 0) for u in top_ub]
    
    options = {
        "tooltip": {
            "trigger": "axis",
            "axisPointer": {"type": "shadow"},
        },
        "grid": {"left": "30%", "right": "10%", "bottom": "5%", "top": "5%", "containLabel": False},
        "xAxis": {
            "type": "value",
            "name": "Pallets",
            "nameLocation": "middle",
            "nameGap": 30,
            "nameTextStyle": {"color": "#aaa"},
            "axisLabel": {"color": "#ccc"},
            "splitLine": {"lineStyle": {"color": "#333"}}
        },
        "yAxis": {
            "type": "category",
            "data": nombres,
            "axisLabel": {"color": "#ddd", "fontSize": 10},
        },
        "series": [{
            "name": "Pallets",
            "type": "bar",
            "data": cantidades,
            "label": {
                "show": True,
                "position": "right",
                "formatter": "{c}",
                "color": "#fff",
                "fontSize": 11
            },
            "itemStyle": {
                "color": {
                    "type": "linear",
                    "x": 0, "y": 0, "x2": 1, "y2": 0,
                    "colorStops": [
                        {"offset": 0, "color": "#e94560"},
                        {"offset": 1, "color": "#f39c12"}
                    ]
                },
                "borderRadius": [0, 4, 4, 0]
            }
        }]
    }
    
    st_echarts(options=options, height=f"{max(300, len(top_ub) * 30)}px")
