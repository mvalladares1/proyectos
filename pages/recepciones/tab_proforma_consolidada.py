"""
Tab de Proforma Consolidada de Fletes
Genera documento consolidado por transportista con detalle de OCs, kms, kilos, costos
"""

import streamlit as st
import pandas as pd
import xmlrpc.client
from datetime import datetime, timedelta
import io
from typing import Dict, List, Optional
import requests
import json
import base64
from .email_templates import get_proforma_email_template


URL = 'https://riofuturo.server98c6e.oerpondemand.net'
DB = 'riofuturo-master'
API_LOGISTICA_RUTAS = 'https://riofuturoprocesos.com/api/logistica/rutas'
API_LOGISTICA_COSTES = 'https://riofuturoprocesos.com/api/logistica/db/coste-rutas'


def formato_numero_chileno(numero: float, decimales: int = 0) -> str:
    """Formatea número con separador de miles chileno (punto) y decimal (coma)"""
    if decimales == 0:
        formato = f"{numero:,.0f}"
    else:
        formato = f"{numero:,.{decimales}f}"
    # Reemplazar separadores: coma por punto (miles) y punto por coma (decimal)
    return formato.replace(',', 'X').replace('.', ',').replace('X', '.')


def get_odoo_connection(username, password):
    """Conexión a Odoo"""
    try:
        common = xmlrpc.client.ServerProxy(f'{URL}/xmlrpc/2/common')
        uid = common.authenticate(DB, username, password, {})
        models = xmlrpc.client.ServerProxy(f'{URL}/xmlrpc/2/object')
        return models, uid
    except Exception as e:
        st.error(f"Error de conexión a Odoo: {e}")
        return None, None


@st.cache_data(ttl=300)
def obtener_rutas_logistica():
    """Obtener rutas del sistema de logística"""
    try:
        response = requests.get(API_LOGISTICA_RUTAS, timeout=10)
        if response.status_code == 200:
            data = response.json()
            # Manejar si es lista o dict con 'data'
            if isinstance(data, list):
                return data
            elif isinstance(data, dict) and 'data' in data:
                return data['data']
            return []
        return []
    except Exception as e:
        st.warning(f"No se pudo conectar al sistema de logística: {e}")
        return []


def obtener_kg_de_oc_mp(numero_ruta: str, rutas_logistica: List[Dict]) -> Optional[float]:
    """Obtiene los kg de la OC de MP asociada a una ruta de transporte
    
    Flujo:
    OC Flete (OC11476) -> Número Ruta (RT00259) -> OC MP (OC11427) -> Kg reales
    """
    try:
        # Buscar la ruta por su número (campo 'name' contiene el número de ruta como RT00XXX)
        for ruta in rutas_logistica:
            ruta_name = ruta.get('name', '')
            
            if numero_ruta and ruta_name == numero_ruta:
                # Extraer kg de la carga asociada
                total_qnt = ruta.get('total_qnt', 0)
                if total_qnt and total_qnt > 0:
                    return float(total_qnt)
                
                # Intentar obtener de las cargas individuales
                cargas = ruta.get('loads', [])
                if cargas:
                    total_kg = sum(float(c.get('quantity', 0)) for c in cargas if isinstance(cargas, list))
                    if total_kg > 0:
                        return total_kg
        
        return None
    except Exception as e:
        return None


def obtener_nombre_ruta_real(ruta_data: Optional[Dict]) -> str:
    """Extrae el nombre real de la ruta desde los datos de logística"""
    if not ruta_data:
        return 'Sin ruta'
    
    try:
        # Intentar obtener de routes field
        routes_field = ruta_data.get('routes', False)
        if routes_field and isinstance(routes_field, str) and routes_field.startswith('['):
            routes_data = json.loads(routes_field)
            if isinstance(routes_data, list) and len(routes_data) > 0:
                route_info = routes_data[0]
                route_name = route_info.get('route_name', '')
                if route_name:
                    return route_name
        
        # Fallback: construir desde origen/destino
        origen = ruta_data.get('origin', '') or ruta_data.get('origen', '')
        destino = ruta_data.get('destination', '') or ruta_data.get('destino', '')
        
        if origen and destino:
            return f"{origen} - {destino}"
        
        # Último fallback: nombre de ruta
        ruta_name = ruta_data.get('ruta_name', '') or ruta_data.get('name', '')
        if ruta_name:
            return ruta_name
        
        return 'Sin ruta'
    except:
        return 'Sin ruta'


def buscar_ruta_en_logistica(oc_name: str, rutas_logistica: List[Dict]) -> Optional[Dict]:
    """Buscar ruta en sistema de logística por nombre de OC
    
    Busca por múltiples campos para mayor flexibilidad:
    - purchase_order_name (campo principal)
    - po (campo alternativo)
    """
    for ruta in rutas_logistica:
        # Intentar con purchase_order_name
        po_name = ruta.get('purchase_order_name', '')
        # Intentar con po (campo alternativo)
        po_alt = ruta.get('po', '')
        
        if po_name == oc_name or po_alt == oc_name:
            return ruta
    return None


@st.cache_data(ttl=60)
def obtener_ocs_transportes(_models, _uid, username, password, fecha_desde, fecha_hasta):
    """Obtener OCs de TRANSPORTES en el rango de fechas"""
    try:
        # Construir dominio de búsqueda
        domain = [
            ('x_studio_categora_de_producto', '=', 'SERVICIOS'),
            ('x_studio_selection_field_yUNPd', 'ilike', 'TRANSPORTES'),
            ('state', 'in', ['draft', 'sent', 'to approve', 'purchase', 'done']),  # Todos los estados
            ('date_order', '>=', fecha_desde),
            ('date_order', '<=', fecha_hasta)
        ]
        
        ocs = _models.execute_kw(
            DB, _uid, password,
            'purchase.order', 'search_read',
            [domain],
            {'fields': ['id', 'name', 'partner_id', 'date_order', 'amount_untaxed', 'state'], 
             'order': 'date_order desc'}
        )
        
        # Obtener líneas y costo calculado para cada OC
        for oc in ocs:
            lineas = _models.execute_kw(
                DB, _uid, password,
                'purchase.order.line', 'search_read',
                [[('order_id', '=', oc['id'])]],
                {'fields': ['product_id', 'name', 'product_qty', 'price_subtotal']}
            )
            
            oc['costo_lineas'] = sum(linea.get('price_subtotal', 0) for linea in lineas)
            oc['num_viajes'] = sum(linea.get('product_qty', 0) for linea in lineas)
        
        return ocs
    except Exception as e:
        st.error(f"Error obteniendo OCs: {e}")
        return []


def generar_pdf_proforma(datos_consolidados: Dict, fecha_desde: str, fecha_hasta: str) -> bytes:
    """Genera archivo PDF con la proforma consolidada por transportista"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f4788'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=6,
        alignment=TA_CENTER
    )
    
    transportista_style = ParagraphStyle(
        'Transportista',
        parent=styles['Heading2'],
        fontSize=13,
        textColor=colors.HexColor('#2c5aa0'),
        spaceAfter=8,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    story = []
    
    for idx, (transportista, data) in enumerate(datos_consolidados.items()):
        # Si no es el primero, agregar salto de página
        if idx > 0:
            story.append(PageBreak())
        
        # Header principal
        story.append(Paragraph("PROFORMA CONSOLIDADA DE FLETES", title_style))
        story.append(Paragraph(f"Período: {fecha_desde} al {fecha_hasta}", subtitle_style))
        story.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Nombre del transportista
        story.append(Paragraph(f"Transportista: {transportista}", transportista_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Tabla de datos
        table_data = [
            ['OC', 'Fecha', 'Ruta', 'Kms', 'Kilos', 'Costo', '$/km', 'Tipo Camión']
        ]
        
        for oc_data in data['ocs']:
            table_data.append([
                oc_data['oc_name'],
                oc_data['fecha'],
                oc_data['ruta'][:25] if oc_data['ruta'] else 'Sin ruta',  # Truncar si es muy largo
                formato_numero_chileno(oc_data['kms'], 0) if oc_data['kms'] else '0',
                formato_numero_chileno(oc_data['kilos'], 1) if oc_data['kilos'] else '0',
                f"${formato_numero_chileno(oc_data['costo'], 0)}",
                f"${formato_numero_chileno(oc_data['costo_por_km'], 0)}" if oc_data['costo_por_km'] else '$0',
                oc_data['tipo_camion'][:15] if oc_data['tipo_camion'] else 'N/A'
            ])
        
        # Fila de totales
        promedio_km = data['totales']['costo'] / data['totales']['kms'] if data['totales']['kms'] > 0 else 0
        table_data.append([
            'TOTALES',
            '',
            '',
            formato_numero_chileno(data['totales']['kms'], 0),
            formato_numero_chileno(data['totales']['kilos'], 1),
            f"${formato_numero_chileno(data['totales']['costo'], 0)}",
            f"${formato_numero_chileno(promedio_km, 0)}",
            ''
        ])
        
        # Crear tabla
        table = Table(table_data, colWidths=[0.8*inch, 0.8*inch, 1.5*inch, 0.6*inch, 0.6*inch, 0.9*inch, 0.7*inch, 1.1*inch])
        
        # Estilo de tabla
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a90e2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Datos
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            
            # Totales (última fila)
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d3d3d3')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 9),
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#4a90e2')),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.grey),
        ]))
        
        story.append(table)
        
        # Info adicional
        story.append(Spacer(1, 0.3*inch))
        info_text = f"Total de OCs: {len(data['ocs'])} | Total Kms: {formato_numero_chileno(data['totales']['kms'], 0)} | Total Kilos: {formato_numero_chileno(data['totales']['kilos'], 1)}"
        story.append(Paragraph(info_text, subtitle_style))
    
    # Construir PDF
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def generar_excel_proforma(datos_consolidados: Dict, fecha_desde: str, fecha_hasta: str) -> bytes:
    """Genera archivo Excel con la proforma consolidada por transportista"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = Workbook()
    
    for transportista, data in datos_consolidados.items():
        # Crear hoja por transportista
        ws = wb.create_sheet(title=transportista[:30])  # Max 31 chars
        
        # Header
        ws['A1'] = 'PROFORMA CONSOLIDADA DE FLETES'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A2'] = f'Transportista: {transportista}'
        ws['A2'].font = Font(bold=True, size=12)
        ws.merge_cells('A2:H2')
        
        ws['A3'] = f'Período: {fecha_desde} - {fecha_hasta}'
        ws.merge_cells('A3:H3')
        
        ws['A4'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws.merge_cells('A4:H4')
        
        # Headers de tabla
        headers = ['OC', 'Fecha', 'Ruta', 'Kms', 'Kilos', 'Costo', '$/km', 'Tipo Camión']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=6, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        row = 7
        for oc_data in data['ocs']:
            ws.cell(row=row, column=1, value=oc_data['oc_name'])
            ws.cell(row=row, column=2, value=oc_data['fecha'])
            ws.cell(row=row, column=3, value=oc_data['ruta'] or 'Sin ruta')
            ws.cell(row=row, column=4, value=oc_data['kms'] or 0)
            ws.cell(row=row, column=5, value=oc_data['kilos'] or 0)
            ws.cell(row=row, column=6, value=oc_data['costo'])
            ws.cell(row=row, column=7, value=oc_data['costo_por_km'] or 0)
            ws.cell(row=row, column=8, value=oc_data['tipo_camion'] or 'N/A')
            row += 1
        
        # Totales
        row += 1
        ws.cell(row=row, column=1, value='TOTALES')
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=4, value=data['totales']['kms'])
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=5, value=data['totales']['kilos'])
        ws.cell(row=row, column=5).font = Font(bold=True)
        ws.cell(row=row, column=6, value=data['totales']['costo'])
        ws.cell(row=row, column=6).font = Font(bold=True)
        
        # Promedio $/km
        if data['totales']['kms'] > 0:
            promedio_km = data['totales']['costo'] / data['totales']['kms']
            ws.cell(row=row, column=7, value=promedio_km)
            ws.cell(row=row, column=7).font = Font(bold=True)
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 20
    
    # Eliminar hoja por defecto
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Guardar en bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generar_pdf_individual_transportista(transportista: str, data: Dict, fecha_desde: str, fecha_hasta: str) -> bytes:
    """Genera PDF individual para un transportista (estilo similar a proformas de MP)"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                          topMargin=0.5*inch, bottomMargin=0.7*inch,
                          leftMargin=0.6*inch, rightMargin=0.6*inch)
    
    styles = getSampleStyleSheet()
    color_azul = colors.HexColor('#1B4F72')
    color_azul_claro = colors.HexColor('#2E86AB')
    
    title_style = ParagraphStyle('CustomTitle',
                                parent=styles['Heading1'],
                                fontSize=18,
                                alignment=TA_CENTER,
                                textColor=color_azul,
                                spaceAfter=20)
    
    elements = []
    
    # Título
    elements.append(Paragraph("PROFORMA DE FLETES", title_style))
    elements.append(Spacer(1, 12))
    
    # Información del documento
    fecha_envio = datetime.now().strftime("%d-%m-%Y")
    
    info_data = [
        ["Transportista:", transportista[:50], "", "Fecha Envío:", fecha_envio],
        ["Período:", f"{fecha_desde} al {fecha_hasta}", "", "Moneda:", "CLP"],
        ["Total OCs:", str(len(data['ocs'])), "", "", ""],
    ]
    
    info_table = Table(info_data, colWidths=[1.2*inch, 3*inch, 0.5*inch, 1.2*inch, 1.3*inch])
    info_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (3, 0), (3, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    # Tabla de OCs
    table_data = [["OC", "Fecha", "Ruta", "Kms", "Kilos", "Costo", "$/km", "Tipo Camión"]]
    
    for oc_data in data['ocs']:
        table_data.append([
            oc_data['oc_name'],
            oc_data['fecha'],
            oc_data['ruta'][:30] if oc_data['ruta'] else 'Sin ruta',
            formato_numero_chileno(oc_data['kms'], 0),
            formato_numero_chileno(oc_data['kilos'], 1),
            f"${formato_numero_chileno(oc_data['costo'], 0)}",
            f"${formato_numero_chileno(oc_data['costo_por_km'], 0)}",
            oc_data['tipo_camion'][:15] if oc_data['tipo_camion'] else 'N/A'
        ])
    
    # Línea en blanco
    table_data.append(["", "", "", "", "", "", "", ""])
    
    # Totales
    promedio_km = data['totales']['costo'] / data['totales']['kms'] if data['totales']['kms'] > 0 else 0
    table_data.append([
        "", "", "TOTAL:",
        formato_numero_chileno(data['totales']['kms'], 0),
        formato_numero_chileno(data['totales']['kilos'], 1),
        f"${formato_numero_chileno(data['totales']['costo'], 0)} *",
        f"${formato_numero_chileno(promedio_km, 0)}",
        ""
    ])
    
    main_table = Table(table_data, colWidths=[0.9*inch, 0.8*inch, 2.2*inch, 0.6*inch, 0.7*inch, 1.0*inch, 0.7*inch, 1.1*inch])
    main_table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), color_azul),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        
        # Grid
        ('GRID', (0, 0), (-1, len(data['ocs'])), 0.5, colors.grey),
        ('LINEBELOW', (0, len(data['ocs'])), (-1, len(data['ocs'])), 1, colors.black),
        
        # Padding
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        
        # Totales
        ('FONTNAME', (2, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (2, -1), (-1, -1), 9),
        ('BACKGROUND', (2, -1), (-1, -1), colors.HexColor('#E8F4F8')),
    ]))
    elements.append(main_table)
    
    # Nota
    elements.append(Spacer(1, 10))
    nota_style = ParagraphStyle('Nota',
                               parent=styles['Normal'],
                               fontSize=8,
                               textColor=color_azul,
                               alignment=TA_LEFT)
    elements.append(Paragraph("<b>* Este es el monto total en CLP a facturar por servicios de transporte</b>", nota_style))
    
    # Footer
    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle('Footer',
                                 parent=styles['Normal'],
                                 fontSize=7,
                                 textColor=color_azul,
                                 alignment=TA_CENTER)
    elements.append(Paragraph(f"Rio Futuro Procesos SPA | Año {datetime.now().year}", footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def generar_zip_proformas_transportistas(datos_consolidados: Dict, fecha_desde: str, fecha_hasta: str) -> bytes:
    """Genera ZIP con PDFs organizados por carpeta de transportista"""
    import zipfile
    
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for transportista, data in datos_consolidados.items():
            # Generar PDF individual
            pdf_bytes = generar_pdf_individual_transportista(transportista, data, fecha_desde, fecha_hasta)
            
            # Crear carpeta por transportista (sanitizar nombre)
            carpeta_nombre = transportista.replace('/', '_').replace('\\', '_').replace(' ', '_')[:50]
            nombre_pdf = f"Proforma_Fletes_{fecha_desde}_{fecha_hasta}.pdf"
            
            # Agregar al ZIP en carpeta del transportista
            ruta_en_zip = f"{carpeta_nombre}/{nombre_pdf}"
            zip_file.writestr(ruta_en_zip, pdf_bytes)
    
    zip_buffer.seek(0)
    return zip_buffer.getvalue()


def get_email_template_transportista(transportista: str, data: Dict, fecha_desde: str, fecha_hasta: str) -> Dict[str, str]:
    """Genera template de email para envío a transportista (estilo similar a proformas MP)"""
    
    total_kms = data['totales']['kms']
    total_kilos = data['totales']['kilos']
    total_costo = data['totales']['costo']
    cant_ocs = len(data['ocs'])
    
    # Agrupar OCs por ruta para el desglose
    from collections import defaultdict
    rutas_agrupadas = defaultdict(lambda: {"ocs": [], "kms": 0, "kilos": 0, "costo": 0})
    
    for oc_data in data['ocs']:
        ruta = oc_data['ruta'] if oc_data['ruta'] and oc_data['ruta'] != 'Sin ruta' else 'Sin ruta'
        rutas_agrupadas[ruta]["ocs"].append(oc_data['oc_name'])
        rutas_agrupadas[ruta]["kms"] += oc_data['kms']
        rutas_agrupadas[ruta]["kilos"] += oc_data['kilos']
        rutas_agrupadas[ruta]["costo"] += oc_data['costo']
    
    # Generar HTML de rutas o de OCs individuales (si no hay rutas con datos)
    rutas_html = ""
    tiene_datos_ruta = any(datos['kms'] > 0 for datos in rutas_agrupadas.values())
    
    if tiene_datos_ruta:
        # Mostrar agrupado por ruta con datos completos
        for idx, (ruta, datos_ruta) in enumerate(list(rutas_agrupadas.items())[:10]):
            if datos_ruta['kms'] == 0 and datos_ruta['kilos'] == 0:
                continue  # Saltar rutas sin datos
            
            kms_fmt = formato_numero_chileno(datos_ruta['kms'], 0)
            kilos_fmt = formato_numero_chileno(datos_ruta['kilos'], 1)
            costo_fmt = f"${formato_numero_chileno(datos_ruta['costo'], 0)}"
            ocs_str = ', '.join(datos_ruta['ocs'][:3])
            if len(datos_ruta['ocs']) > 3:
                ocs_str += f" (+{len(datos_ruta['ocs']) - 3} más)"
            
            rutas_html += f'''<li style="margin-bottom: 10px;">
                <strong>{ruta}:</strong> {kms_fmt} km, {kilos_fmt} kg - {costo_fmt}<br>
                <span style="color: #666; font-size: 12px;">OCs: {ocs_str}</span>
            </li>\n'''
    else:
        # Mostrar OCs individuales sin agrupar (cuando no hay datos de ruta)
        for oc_data in data['ocs'][:10]:
            costo_fmt = f"${formato_numero_chileno(oc_data['costo'], 0)}"
            fecha_oc = oc_data['fecha']
            
            rutas_html += f'''<li style="margin-bottom: 10px;">
                <strong>{oc_data['oc_name']}</strong> ({fecha_oc}): {costo_fmt}
            </li>\n'''
    
    if len(data['ocs']) > 10:
        rutas_html += f'<li style="color: #666;"><em>...y {len(data["ocs"]) - 10} OCs más</em></li>'
    elif len(rutas_agrupadas) > 10 and tiene_datos_ruta:
        rutas_html += f'<li style="color: #666;"><em>...y {len(rutas_agrupadas) - 10} rutas más</em></li>'
    
    # Formatear totales
    total_costo_fmt = f"${formato_numero_chileno(total_costo, 0)}"
    
    # Generar sección de totales condicional
    totales_extras = ""
    if total_kms > 0:
        total_kms_fmt = formato_numero_chileno(total_kms, 0)
        totales_extras += f"                    • Kilómetros: {total_kms_fmt} km<br>\n"
    if total_kilos > 0:
        total_kilos_fmt = formato_numero_chileno(total_kilos, 1)
        totales_extras += f"                    • Kilos transportados: {total_kilos_fmt} kg<br>\n"
    
    asunto = f"Proforma de Fletes {fecha_desde} al {fecha_hasta} - Rio Futuro"
    
    cuerpo_html = f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
        <div style="background-color: #1B4F72; padding: 20px; text-align: center;">
            <h2 style="color: #FFFFFF; margin: 0; font-size: 24px;">Proforma de Servicios de Transporte</h2>
        </div>
        
        <div style="padding: 30px; background-color: #f9f9f9;">
            <p style="color: #333; font-size: 15px;">Estimado(a) <strong>{transportista}</strong>,</p>
            
            <p style="color: #555; line-height: 1.6;">
                Adjunto encontrará la proforma correspondiente a <strong style="color: #1B4F72;">{cant_ocs} OC(s) de transporte</strong> 
                del período <strong>{fecha_desde} al {fecha_hasta}</strong>, Detalle:
            </p>
            
            <div style="background-color: #E8F4F8; border-left: 4px solid #2E86AB; padding: 15px; margin: 20px 0;">
                <p style="margin: 0 0 10px 0; color: #333; font-size: 14px;"><strong>{'Ordenes de Compra:' if not tiene_datos_ruta else 'Resumen de Servicios:'}</strong></p>
                <ul style="color: #555; margin: 10px 0; padding-left: 20px; list-style-type: disc;">
                    {rutas_html}
                </ul>
                <hr style="border: none; border-top: 1px solid #ccc; margin: 15px 0;">
                <p style="margin: 0; color: #333; font-size: 13px;">
                    {f'<strong>Totales:</strong><br>{totales_extras}' if totales_extras else ''}
                </p>
                <p style="margin: 10px 0 0 0; color: #1B4F72; font-size: 18px; font-weight: bold;">
                    Total a Facturar: {total_costo_fmt} CLP
                </p>
            </div>
            
            <p style="color: #555; line-height: 1.6;">
                Por favor revise el documento adjunto con el detalle completo y no dude en contactarnos si tiene alguna consulta.
            </p>
            
            <p style="color: #333; margin-top: 30px;">Saludos cordiales,</p>
            <p style="color: #1B4F72; font-weight: bold; font-size: 16px; margin: 5px 0;">Rio Futuro Procesos</p>
        </div>
        
        <div style="background-color: #1B4F72; padding: 15px; text-align: center;">
            <p style="font-size: 11px; color: #FFFFFF; margin: 0;">
                Este correo fue enviado automáticamente desde el sistema de gestión de Rio Futuro.
            </p>
        </div>
    </div>
    """
    
    return {
        "subject": asunto,
        "body_html": cuerpo_html
    }


def render(username: str, password: str):
    """Renderiza el tab de Proforma Consolidada"""
    st.markdown("## 📄 Proforma Consolidada de Fletes")
    st.markdown("Genere documentos consolidados por transportista con el detalle de movimientos del período")
    
    models, uid = get_odoo_connection(username, password)
    if not models or not uid:
        return
    
    # Filtros
    col1, col2, col3 = st.columns([2, 2, 1])
    
    with col1:
        fecha_desde = st.date_input(
            "Fecha Desde",
            value=datetime.now().replace(day=1),
            key="proforma_fecha_desde"
        )
    
    with col2:
        fecha_hasta = st.date_input(
            "Fecha Hasta",
            value=datetime.now(),
            key="proforma_fecha_hasta"
        )
    
    with col3:
        st.markdown("<br>", unsafe_allow_html=True)  # Espaciado
        cargar_datos = st.button("🔄 Cargar Datos", type="primary", use_container_width=True)
    
    # Convertir a string
    fecha_desde_str = fecha_desde.strftime('%Y-%m-%d')
    fecha_hasta_str = fecha_hasta.strftime('%Y-%m-%d')
    
    # Inicializar session state para datos cargados
    if 'proforma_ocs_cargadas' not in st.session_state:
        st.session_state.proforma_ocs_cargadas = None
    if 'proforma_rutas_logistica' not in st.session_state:
        st.session_state.proforma_rutas_logistica = None
    if 'proforma_fecha_carga' not in st.session_state:
        st.session_state.proforma_fecha_carga = None
    
    # Solo cargar datos si se presiona el botón
    rango_actual = f"{fecha_desde_str}_{fecha_hasta_str}"
    if cargar_datos:
        # Obtener datos
        with st.spinner("Cargando OCs de transportes y datos de logística..."):
            st.session_state.proforma_ocs_cargadas = obtener_ocs_transportes(models, uid, username, password, fecha_desde_str, fecha_hasta_str)
            st.session_state.proforma_rutas_logistica = obtener_rutas_logistica()
            st.session_state.proforma_fecha_carga = rango_actual
    
    # Verificar si hay datos cargados
    if st.session_state.proforma_ocs_cargadas is None:
        st.info("👆 Selecciona el rango de fechas y presiona **Cargar Datos** para comenzar")
        return
    
    # Usar datos desde session_state
    ocs = st.session_state.proforma_ocs_cargadas
    rutas_logistica = st.session_state.proforma_rutas_logistica
    
    # Debug info
    if rutas_logistica:
        rutas_con_oc = [r for r in rutas_logistica if r.get('purchase_order_name') or r.get('po')]
        st.info(f"ℹ️ Sistema de logística: {len(rutas_logistica)} rutas totales, {len(rutas_con_oc)} con OC asignada")
    
    if not ocs:
        st.info("No hay OCs de TRANSPORTES confirmadas en el período seleccionado")
        return
    
    st.success(f"✅ {len(ocs)} OCs encontradas")
    
    # Obtener lista de transportistas únicos
    transportistas_unicos = sorted(list(set([
        oc['partner_id'][1] if oc.get('partner_id') and isinstance(oc['partner_id'], (list, tuple)) else 'N/A'
        for oc in ocs
    ])))
    
    # Filtro de transportistas
    st.markdown("### 🚚 Filtrar por Transportista")
    transportistas_seleccionados = st.multiselect(
        "Seleccione los transportistas a incluir (deje vacío para todos)",
        options=transportistas_unicos,
        default=None,
        key="filtro_transportistas"
    )
    
    # Filtrar OCs si hay transportistas seleccionados
    if transportistas_seleccionados:
        ocs = [oc for oc in ocs if (
            oc.get('partner_id') and 
            isinstance(oc['partner_id'], (list, tuple)) and 
            oc['partner_id'][1] in transportistas_seleccionados
        )]
        st.info(f"📋 Mostrando {len(ocs)} OCs de {len(transportistas_seleccionados)} transportista(s) seleccionado(s)")
    
    if not ocs:
        st.warning("No hay OCs para los transportistas seleccionados")
        return
    
    # Enriquecer con datos de logística
    datos_tabla = []
    for oc in ocs:
        transportista = oc['partner_id'][1] if oc.get('partner_id') and isinstance(oc['partner_id'], (list, tuple)) else 'N/A'
        
        ruta_info = buscar_ruta_en_logistica(oc['name'], rutas_logistica)
        
        kms = ruta_info.get('total_distance_km', 0) if ruta_info else 0
        costo = oc['costo_lineas']
        costo_por_km = (costo / kms) if kms > 0 else 0
        
        # Obtener número de ruta (campo 'name' de la API de logística, ej: RT00469)
        numero_ruta = ruta_info.get('name', '') if ruta_info else ''
        
        kilos = 0
        if ruta_info:
            # Intentar obtener kg de la OC de MP asociada a la ruta
            kilos_mp = obtener_kg_de_oc_mp(numero_ruta, rutas_logistica)
            if kilos_mp is not None and kilos_mp > 0:
                kilos = kilos_mp
            else:
                # Fallback al total_qnt del sistema de logística
                kilos = ruta_info.get('total_qnt', 0)
        
        # Obtener nombre de ruta real de la app logística y tipo de camión
        nombre_ruta_logistica = 'Sin nombre'
        tipo_camion = 'N/A'
        if ruta_info:
            # Usar la función para obtener nombre de ruta real desde la app logística
            nombre_ruta_logistica = obtener_nombre_ruta_real(ruta_info)
            
            # Obtener tipo de camión
            routes_field = ruta_info.get('routes', False)
            if routes_field and isinstance(routes_field, str) and routes_field.startswith('['):
                try:
                    routes_data = json.loads(routes_field)
                    if isinstance(routes_data, list) and len(routes_data) > 0:
                        route_info = routes_data[0]
                        cost_type = route_info.get('cost_type', '')
                        if cost_type == 'truck_8':
                            tipo_camion = '🚛 Camión 8 Ton'
                        elif cost_type == 'truck_12_14':
                            tipo_camion = '🚚 Camión 12-14 Ton'
                        elif cost_type == 'short_rampla':
                            tipo_camion = '🚐 Rampla Corta'
                        elif cost_type == 'rampla':
                            tipo_camion = '🚛 Rampla'
                except:
                    pass
        
        datos_tabla.append({
            'Sel': False,
            'OC': oc['name'],
            'Fecha': oc['date_order'][:10] if oc.get('date_order') else 'N/A',
            'Transportista': transportista,
            'Ruta': nombre_ruta_logistica,
            'Número Ruta': numero_ruta if numero_ruta else 'N/A',
            'Kms': kms,
            'Kilos': kilos,
            'Costo': costo,
            '$/km': costo_por_km,
            'Tipo Camión': tipo_camion,
            '_oc_id': oc['id']
        })
    
    df = pd.DataFrame(datos_tabla)
    
    # Crear versión con formato chileno para visualización
    df_display = df.copy()
    df_display['Kms'] = df_display['Kms'].apply(lambda x: formato_numero_chileno(x, 0))
    df_display['Kilos'] = df_display['Kilos'].apply(lambda x: formato_numero_chileno(x, 1))
    df_display['Costo'] = df_display['Costo'].apply(lambda x: f"${formato_numero_chileno(x, 0)}")
    df_display['$/km'] = df_display['$/km'].apply(lambda x: f"${formato_numero_chileno(x, 0)}")
    
    # Detectar datos faltantes
    def detectar_datos_faltantes(df_data):
        """Detecta OCs con datos faltantes o incompletos"""
        problemas = []
        for idx, row in df_data.iterrows():
            issues = []
            if not row['Ruta'] or row['Ruta'] == 'Sin ruta':
                issues.append('Ruta')
            if row['Kms'] == 0 or pd.isna(row['Kms']):
                issues.append('Kms')
            if row['Kilos'] == 0 or pd.isna(row['Kilos']):
                issues.append('Kilos')
            if row['Costo'] == 0 or pd.isna(row['Costo']):
                issues.append('Costo')
            if not row['Tipo Camión'] or row['Tipo Camión'] == 'N/A':
                issues.append('Tipo Camión')
            
            if issues:
                problemas.append({
                    'indice': idx,
                    'oc': row['OC'],
                    'transportista': row['Transportista'],
                    'campos_faltantes': issues
                })
        return problemas
    
    # Analizar datos faltantes
    datos_faltantes = detectar_datos_faltantes(df)
    
    if datos_faltantes:
        st.warning(f"⚠️ Se detectaron {len(datos_faltantes)} OCs con datos incompletos")
        
        with st.expander(f"🔍 Ver detalles de {len(datos_faltantes)} OCs con datos faltantes"):
            for problema in datos_faltantes:
                st.markdown(f"**{problema['oc']}** ({problema['transportista']}): Faltan datos de **{', '.join(problema['campos_faltantes'])}**")
    
    # Tabla editable con opción de completar datos
    st.markdown("### 📋 Seleccione y complete las OCs para la proforma")
    
    # Inicializar session state para datos editados si no existe o si hay nuevos datos
    if 'df_proforma_editado' not in st.session_state or len(st.session_state.df_proforma_editado) != len(df):
        st.session_state.df_proforma_editado = df.copy()
    
    # Inicializar estado de selección si no existe
    if 'df_proforma_display' not in st.session_state:
        st.session_state.df_proforma_display = df_display.copy()
    
    # Inicializar DataFrame final si no existe
    if 'df_proforma_final' not in st.session_state:
        st.session_state.df_proforma_final = df.copy()
    
    # Tabs: Selección simple vs Editor completo
    tab_select, tab_editor = st.tabs(["✓ Selección Rápida", "✏️ Editor Completo (Completar Datos)"])
    
    # Variable para el df final editado (se asigna en cada tab)
    edited_df = df.copy()
    
    with tab_select:
        st.info("👉 Usa este modo para seleccionar OCs que ya tienen todos los datos completos")
        
        # Botón seleccionar todas
        col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 4])
        with col_btn1:
            if st.button("✅ Seleccionar Todas", use_container_width=True, key='btn_sel_all'):
                st.session_state.df_proforma_display['Sel'] = True
                df_display['Sel'] = True
        with col_btn2:
            if st.button("❌ Deseleccionar Todas", use_container_width=True, key='btn_desel_all'):
                st.session_state.df_proforma_display['Sel'] = False
                df_display['Sel'] = False
        
        edited_df_display = st.data_editor(
            df_display,
            column_config={
                'Sel': st.column_config.CheckboxColumn('Sel', default=False),
                'OC': st.column_config.TextColumn('OC', width='small'),
                'Fecha': st.column_config.TextColumn('Fecha', width='small'),
                'Transportista': st.column_config.TextColumn('Transportista', width='medium'),
                'Ruta': st.column_config.TextColumn('Ruta', width='medium'),
                'Número Ruta': st.column_config.TextColumn('Número Ruta', width='small'),
                'Kms': st.column_config.TextColumn('Kms', width='small'),
                'Kilos': st.column_config.TextColumn('Kilos', width='small'),
                'Costo': st.column_config.TextColumn('Costo', width='medium'),
                '$/km': st.column_config.TextColumn('$/km', width='small'),
                'Tipo Camión': st.column_config.TextColumn('Tipo Camión', width='medium'),
                '_oc_id': None  # Ocultar
            },
            disabled=['OC', 'Fecha', 'Transportista', 'Ruta', 'Número Ruta', 'Kms', 'Kilos', 'Costo', '$/km', 'Tipo Camión'],
            hide_index=True,
            key='editor_proforma_simple',
            use_container_width=True
        )
        
        # Guardar el estado actualizado y sincronizar selección con df original
        st.session_state.df_proforma_display = edited_df_display.copy()
        # BUGFIX: Asegurarse de preservar todos los datos numéricos originales
        edited_df = df.copy()
        edited_df['Sel'] = edited_df_display['Sel'].values
        # Guardar en session_state para uso posterior
        st.session_state.df_proforma_final = edited_df.copy()
    
    with tab_editor:
        st.info("✏️ Usa este editor para **completar datos faltantes** antes de generar la proforma")
        
        # Añadir columna de estado
        df_editor = st.session_state.df_proforma_editado.copy()
        
        # Marcar OCs con problemas
        df_editor['Estado'] = df_editor.apply(
            lambda row: '⚠️ Incompleto' if any(
                p['oc'] == row['OC'] for p in datos_faltantes
            ) else '✅ Completo',
            axis=1
        )
        
        # Editor completo con todos los campos editables
        edited_df_completo = st.data_editor(
            df_editor,
            column_config={
                'Sel': st.column_config.CheckboxColumn('☑️ Incluir', default=False),
                'Estado': st.column_config.TextColumn('Estado', width='small'),
                'OC': st.column_config.TextColumn('OC', width='small', disabled=True),
                'Fecha': st.column_config.TextColumn('Fecha', width='small', disabled=True),
                'Transportista': st.column_config.TextColumn('Transportista', width='medium', disabled=True),
                'Ruta': st.column_config.TextColumn('Ruta', width='medium', help='Editable: Nombre completo de la ruta'),
                'Número Ruta': st.column_config.TextColumn('Número Ruta', width='small', help='Editable: Número de ruta (ej: RT00259)'),
                'Kms': st.column_config.NumberColumn('Kms', format='%.0f', help='Editable: Kilómetros del viaje'),
                'Kilos': st.column_config.NumberColumn('Kilos', format='%.1f', help='Editable: Kilos transportados'),
                'Costo': st.column_config.NumberColumn('Costo', format='$%.0f', help='Editable: Costo total del flete'),
                '$/km': st.column_config.NumberColumn('$/km', format='$%.0f', help='Auto-calculado'),
                'Tipo Camión': st.column_config.SelectboxColumn(
                    'Tipo Camión',
                    options=['🚚 Camión 8 Ton', '🚛 Camión 12-14 Ton', '🚛 Camión 18 Ton', '🚛 Camión 24 Ton', 'N/A'],
                    help='Selecciona el tipo de camión'
                ),
                '_oc_id': None  # Ocultar
            },
            disabled=['OC', 'Fecha', 'Transportista', '$/km'],
            hide_index=True,
            key='editor_proforma_completo',
            use_container_width=True
        )
        
        # Recalcular $/km automáticamente
        edited_df_completo['$/km'] = edited_df_completo.apply(
            lambda row: (row['Costo'] / row['Kms']) if row['Kms'] > 0 else 0,
            axis=1
        )
        
        # Actualizar session state
        st.session_state.df_proforma_editado = edited_df_completo.copy()
        
        # Usar los datos editados como definitivos
        edited_df = edited_df_completo
        # Guardar en session_state para uso posterior
        st.session_state.df_proforma_final = edited_df.copy()
        
        # Botones de ayuda y resumen
        col_help1, col_help2 = st.columns(2)
        with col_help1:
            if st.button("🔄 Restaurar datos originales", help="Volver a los datos originales de Odoo", key="btn_restore"):
                st.session_state.df_proforma_editado = df.copy()
                st.rerun()
        
        with col_help2:
            nuevos_faltantes = detectar_datos_faltantes(edited_df_completo)
            if nuevos_faltantes:
                st.warning(f"⚠️ Aún quedan {len(nuevos_faltantes)} OCs incompletas")
            else:
                st.success("✅ Todas las OCs tienen datos completos")
    
    # Separador visual entre tabs y sección de generación
    st.divider()
    
    # Usar el DataFrame final guardado en session_state
    if 'df_proforma_final' in st.session_state and st.session_state.df_proforma_final is not None:
        edited_df = st.session_state.df_proforma_final
    
    # Resumen de seleccionados
    seleccionados = edited_df[edited_df['Sel'] == True]
    n_sel = len(seleccionados)
    
    if n_sel > 0:
        st.markdown(f"### ✅ {n_sel} OCs seleccionadas")
        
        # Detectar si hay datos faltantes en seleccionados
        problemas_seleccionados = detectar_datos_faltantes(seleccionados)
        if problemas_seleccionados:
            st.error(f"❌ {len(problemas_seleccionados)} OCs seleccionadas tienen datos incompletos. Ve al **Editor Completo** para corregirlas.")
            with st.expander("Ver OCs con problemas"):
                for p in problemas_seleccionados:
                    st.markdown(f"- **{p['oc']}**: Faltan {', '.join(p['campos_faltantes'])}")
        
        # Vista previa de datos consolidados
        with st.expander("👁️ Vista Previa - Cómo se verá en el PDF"):
            for transportista in seleccionados['Transportista'].unique():
                ocs_transp = seleccionados[seleccionados['Transportista'] == transportista]
                
                st.markdown(f"#### 🚛 {transportista}")
                st.markdown(f"**{len(ocs_transp)} OCs** | **{formato_numero_chileno(ocs_transp['Kms'].sum(), 0)} km** | **{formato_numero_chileno(ocs_transp['Kilos'].sum(), 1)} kg** | **${formato_numero_chileno(ocs_transp['Costo'].sum(), 0)}**")
                
                # Tabla preview
                preview_data = []
                for _, row in ocs_transp.iterrows():
                    preview_data.append({
                        'OC': row['OC'],
                        'Fecha': row['Fecha'],
                        'Ruta': row['Ruta'],
                        'Kms': formato_numero_chileno(row['Kms'], 0),
                        'Kilos': formato_numero_chileno(row['Kilos'], 1),
                        'Costo': f"${formato_numero_chileno(row['Costo'], 0)}",
                        '$/km': f"${formato_numero_chileno(row['$/km'], 0)}",
                        'Tipo': row['Tipo Camión']
                    })
                
                st.table(pd.DataFrame(preview_data))
                st.divider()
        
        # Agrupar por transportista
        transportistas = seleccionados.groupby('Transportista').agg({
            'Kms': 'sum',
            'Kilos': 'sum',
            'Costo': 'sum',
            'OC': 'count'
        }).reset_index()
        transportistas.rename(columns={'OC': 'Cant OCs'}, inplace=True)
        
        # Crear versión formateada para display
        transportistas_display = transportistas.copy()
        transportistas_display['Kms'] = transportistas_display['Kms'].apply(lambda x: formato_numero_chileno(x, 0))
        transportistas_display['Kilos'] = transportistas_display['Kilos'].apply(lambda x: formato_numero_chileno(x, 1))
        transportistas_display['Costo'] = transportistas_display['Costo'].apply(lambda x: f"${formato_numero_chileno(x, 0)}")
        
        st.dataframe(
            transportistas_display,
            column_config={
                'Transportista': st.column_config.TextColumn('Transportista'),
                'Cant OCs': st.column_config.NumberColumn('OCs', format='%d'),
                'Kms': st.column_config.TextColumn('Kms Totales'),
                'Kilos': st.column_config.TextColumn('Kilos Totales'),
                'Costo': st.column_config.TextColumn('Costo Total')
            },
            hide_index=True,
            use_container_width=True
        )
        
        # Advertencia final antes de generar
        if problemas_seleccionados:
            st.warning("⚠️ **ADVERTENCIA**: Algunas OCs tienen datos incompletos. El PDF se generará con los datos disponibles, pero puede verse incompleto.")
        
        # Botones para generar y enviar proforma
        st.markdown("### 🚀 Acciones")
        col_pdf, col_zip, col_excel, col_email = st.columns(4)
        
        with col_pdf:
            if st.button("📄 PDF Consolidado", use_container_width=True, help="Un solo PDF con todos los transportistas"):
                with st.spinner("Generando PDF consolidado..."):
                    # Agrupar datos por transportista
                    datos_consolidados = {}
                    
                    for _, row in seleccionados.iterrows():
                        transp = row['Transportista']
                        
                        if transp not in datos_consolidados:
                            datos_consolidados[transp] = {
                                'ocs': [],
                                'totales': {'kms': 0, 'kilos': 0, 'costo': 0}
                            }
                        
                        datos_consolidados[transp]['ocs'].append({
                            'oc_name': row['OC'],
                            'fecha': row['Fecha'],
                            'ruta': row['Ruta'],
                            'kms': row['Kms'],
                            'kilos': row['Kilos'],
                            'costo': row['Costo'],
                            'costo_por_km': row['$/km'],
                            'tipo_camion': row['Tipo Camión']
                        })
                        
                        datos_consolidados[transp]['totales']['kms'] += row['Kms']
                        datos_consolidados[transp]['totales']['kilos'] += row['Kilos']
                        datos_consolidados[transp]['totales']['costo'] += row['Costo']
                    
                    # Generar PDF consolidado
                    pdf_bytes = generar_pdf_proforma(datos_consolidados, fecha_desde_str, fecha_hasta_str)
                    
                    # Botón de descarga
                    st.download_button(
                        label="⬇️ Descargar PDF Consolidado",
                        data=pdf_bytes,
                        file_name=f"proforma_fletes_consolidado_{fecha_desde_str}_{fecha_hasta_str}.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
        
        with col_zip:
            if st.button("📦 ZIP por Transportista", type="primary", use_container_width=True, help="PDFs individuales organizados en carpetas"):
                with st.spinner("Generando PDFs individuales..."):
                    # Agrupar datos por transportista
                    datos_consolidados = {}
                    
                    for _, row in seleccionados.iterrows():
                        transp = row['Transportista']
                        
                        if transp not in datos_consolidados:
                            datos_consolidados[transp] = {
                                'ocs': [],
                                'totales': {'kms': 0, 'kilos': 0, 'costo': 0}
                            }
                        
                        datos_consolidados[transp]['ocs'].append({
                            'oc_name': row['OC'],
                            'fecha': row['Fecha'],
                            'ruta': row['Ruta'],
                            'kms': row['Kms'],
                            'kilos': row['Kilos'],
                            'costo': row['Costo'],
                            'costo_por_km': row['$/km'],
                            'tipo_camion': row['Tipo Camión']
                        })
                        
                        datos_consolidados[transp]['totales']['kms'] += row['Kms']
                        datos_consolidados[transp]['totales']['kilos'] += row['Kilos']
                        datos_consolidados[transp]['totales']['costo'] += row['Costo']
                    
                    # Generar ZIP con PDFs por transportista
                    zip_bytes = generar_zip_proformas_transportistas(datos_consolidados, fecha_desde_str, fecha_hasta_str)
                    
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    st.download_button(
                        label=f"⬇️ Descargar ZIP ({len(datos_consolidados)} transportistas)",
                        data=zip_bytes,
                        file_name=f"Proformas_Fletes_{timestamp}.zip",
                        mime="application/zip",
                        use_container_width=True
                    )
                    st.success(f"✅ {len(datos_consolidados)} PDFs generados")
        
        with col_excel:
            if st.button("📊 Generar Proforma Excel", use_container_width=True):
                with st.spinner("Generando Excel..."):
                    # Agrupar datos por transportista
                    datos_consolidados = {}
                    
                    for _, row in seleccionados.iterrows():
                        transp = row['Transportista']
                        
                        if transp not in datos_consolidados:
                            datos_consolidados[transp] = {
                                'ocs': [],
                                'totales': {'kms': 0, 'kilos': 0, 'costo': 0}
                            }
                        
                        datos_consolidados[transp]['ocs'].append({
                            'oc_name': row['OC'],
                            'fecha': row['Fecha'],
                            'ruta': row['Ruta'],
                            'kms': row['Kms'],
                            'kilos': row['Kilos'],
                            'costo': row['Costo'],
                            'costo_por_km': row['$/km'],
                            'tipo_camion': row['Tipo Camión']
                        })
                        
                        datos_consolidados[transp]['totales']['kms'] += row['Kms']
                        datos_consolidados[transp]['totales']['kilos'] += row['Kilos']
                        datos_consolidados[transp]['totales']['costo'] += row['Costo']
                    
                    # Generar Excel
                    excel_bytes = generar_excel_proforma(datos_consolidados, fecha_desde_str, fecha_hasta_str)
                    
                    # Botón de descarga
                    st.download_button(
                        label="⬇️ Descargar Proforma Excel",
                        data=excel_bytes,
                        file_name=f"proforma_fletes_{fecha_desde_str}_{fecha_hasta_str}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                    st.success("✅ Excel generado exitosamente")
        
        with col_email:
            if st.button("📧 Enviar por Email", use_container_width=True, help="Envío individual a cada transportista"):
                # Agrupar datos por transportista
                datos_consolidados = {}
                
                for _, row in seleccionados.iterrows():
                    transp = row['Transportista']
                    
                    if transp not in datos_consolidados:
                        datos_consolidados[transp] = {
                            'ocs': [],
                            'totales': {'kms': 0, 'kilos': 0, 'costo': 0}
                        }
                    
                    datos_consolidados[transp]['ocs'].append({
                        'oc_name': row['OC'],
                        'fecha': row['Fecha'],
                        'ruta': row['Ruta'],
                        'kms': row['Kms'],
                        'kilos': row['Kilos'],
                        'costo': row['Costo'],
                        'costo_por_km': row['$/km'],
                        'tipo_camion': row['Tipo Camión']
                    })
                    
                    datos_consolidados[transp]['totales']['kms'] += row['Kms']
                    datos_consolidados[transp]['totales']['kilos'] += row['Kilos']
                    datos_consolidados[transp]['totales']['costo'] += row['Costo']
                
                # Interfaz de envío mejorada
                st.markdown("---")
                st.markdown(f"### 📧 Enviando a {len(datos_consolidados)} transportista(s)")
                
                progress_bar = st.progress(0.0)
                status_container = st.empty()
                
                enviados = 0
                errores = []
                total = len(datos_consolidados)
                
                for idx, (transportista, data) in enumerate(datos_consolidados.items()):
                    with status_container:
                        st.info(f"📧 Enviando {idx + 1}/{total}: {transportista}")
                    
                    try:
                        # Generar PDF individual
                        pdf_bytes = generar_pdf_individual_transportista(transportista, data, fecha_desde_str, fecha_hasta_str)
                        
                        # Buscar email del transportista
                        transportista_info = models.execute_kw(
                            DB, uid, password,
                            'res.partner', 'search_read',
                            [[('name', '=', transportista)]],
                            {'fields': ['id', 'email'], 'limit': 1}
                        )
                        
                        if not transportista_info or not transportista_info[0].get('email'):
                            errores.append(f"{transportista}: Sin email configurado")
                            with status_container:
                                st.warning(f"⚠️ {transportista} no tiene email")
                            continue
                        
                        partner_id = transportista_info[0]['id']
                        email_destino = transportista_info[0]['email']
                        
                        # Codificar PDF
                        import base64
                        pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')
                        
                        # Crear adjunto
                        attachment_data = {
                            "name": f"Proforma_Fletes_{fecha_desde_str}_{fecha_hasta_str}.pdf",
                            "type": "binary",
                            "datas": pdf_base64,
                            "res_model": "res.partner",
                            "res_id": partner_id,
                            "mimetype": "application/pdf",
                            "description": f"Proforma de fletes enviada por correo"
                        }
                        
                        attachment_id = models.execute_kw(
                            DB, uid, password,
                            'ir.attachment', 'create',
                            [attachment_data]
                        )
                        
                        if isinstance(attachment_id, list):
                            attachment_id = attachment_id[0] if attachment_id else None
                        
                        if not attachment_id:
                            raise Exception("No se pudo crear el adjunto")
                        
                        # Generar email con template mejorado
                        email_data = get_email_template_transportista(transportista, data, fecha_desde_str, fecha_hasta_str)
                        
                        # Crear correo
                        mail_data = {
                            "subject": email_data['subject'],
                            "body_html": email_data['body_html'],
                            "email_to": email_destino,
                            "email_from": "notificaciones-rfp@riofuturo.cl",
                            "attachment_ids": [(6, 0, [attachment_id])],
                            "auto_delete": True
                        }
                        
                        mail_id = models.execute_kw(
                            DB, uid, password,
                            'mail.mail', 'create',
                            [mail_data]
                        )
                        
                        if isinstance(mail_id, list):
                            mail_id = mail_id[0] if mail_id else None
                        
                        if not mail_id:
                            raise Exception("No se pudo crear el correo")
                        
                        # Enviar
                        models.execute_kw(
                            DB, uid, password,
                            'mail.mail', 'send',
                            [[mail_id]]
                        )
                        
                        enviados += 1
                        with status_container:
                            st.success(f"✅ {transportista} enviada correctamente")
                    
                    except Exception as e:
                        errores.append(f"{transportista}: {str(e)}")
                        with status_container:
                            st.error(f"❌ Error enviando {transportista}: {str(e)}")
                    
                    progress_bar.progress((idx + 1) / total)
                
                # Resumen final
                st.markdown("---")
                st.markdown("### 📊 Resumen de Envío")
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("✅ Enviadas", enviados)
                with col2:
                    st.metric("❌ Errores", len(errores))
                with col3:
                    st.metric("📊 Total", total)
                
                if errores:
                    st.error("**Errores encontrados:**")
                    for error in errores:
                        st.caption(f"  • {error}")
                else:
                    st.success("🎉 ¡Todas las proformas fueron enviadas correctamente!")
                    st.balloons()
    else:
        st.info("👆 Seleccione las OCs que desea incluir en la proforma")
