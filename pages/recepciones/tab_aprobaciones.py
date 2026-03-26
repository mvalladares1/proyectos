"""
Tab: Aprobaciones MP
Aprobaciones de precios de materia prima y reportes.
"""
import streamlit as st
import pandas as pd
import requests
from datetime import datetime, timedelta
from pathlib import Path
from .shared import fmt_numero, fmt_dinero, fmt_fecha, API_URL
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from backend.services.aprobaciones_service import get_aprobaciones, save_aprobaciones, remove_aprobaciones

# Ruta del logo oficial
BASE_DIR = Path(__file__).resolve().parents[2]
LOGO_PATH = BASE_DIR / "data" / "RFP - LOGO OFICIAL.png"



@st.fragment
def _fragment_main_aprobaciones(username: str, password: str):
    """Fragment principal: Aprobaciones de precios."""
    st.markdown("### 📥 Aprobaciones de Precios MP")
    st.markdown("Valida masivamente los precios de recepción comparándolos con el presupuesto.")

    # Leyenda de semáforos
    st.markdown("""
    <div style="background: rgba(50,50,50,0.5); padding: 10px; border-radius: 8px; margin-bottom: 15px;">
        <b>🚦 Semáforo:</b> 
        <span style="color: #2ecc71;">🟢 OK (0-3%)</span> · 
        <span style="color: #f1c40f;">🟡 Alerta (3-8%)</span> · 
        <span style="color: #e74c3c;">🔴 Crítico (>8%)</span>
    </div>
    """, unsafe_allow_html=True)

    # Inicializar session_state
    if 'aprob_data' not in st.session_state:
        st.session_state.aprob_data = None
    if 'aprob_ppto' not in st.session_state:
        st.session_state.aprob_ppto = {}
    if 'aprob_ppto_detalle' not in st.session_state:
        st.session_state.aprob_ppto_detalle = {}
    if 'recep_aprob_productores_loading' not in st.session_state:
        st.session_state.recep_aprob_productores_loading = False

    # --- FILTROS DE FECHA ---
    col1, col2 = st.columns(2)
    with col1:
        fecha_inicio_aprob = st.date_input("Desde", datetime.now() - timedelta(days=7), key="fecha_inicio_aprob", format="DD/MM/YYYY")
    with col2:
        fecha_fin_aprob = st.date_input("Hasta", datetime.now() + timedelta(days=1), key="fecha_fin_aprob", format="DD/MM/YYYY")

    estado_filtro = st.radio("Estado", ["Pendientes", "Aprobadas", "Todas"], horizontal=True, key="estado_filtro_aprob")

    # Botón de carga
    if st.button("🔄 Cargar Recepciones", type="primary", use_container_width=True, disabled=st.session_state.recep_aprob_cargar_loading):
        st.session_state.recep_aprob_cargar_loading = True
        try:
            # Progress bar personalizado
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            try:
                # Fase 1: Conexión
                status_text.text("⏳ Fase 1/4: Conectando con Odoo...")
                progress_bar.progress(25)
                
                params = {
                    "username": username,
                    "password": password,
                    "fecha_inicio": fecha_inicio_aprob.strftime("%Y-%m-%d"),
                    "fecha_fin": fecha_fin_aprob.strftime("%Y-%m-%d")
                }
                
                # Construir URL con múltiples parámetros estados
                from urllib.parse import urlencode
                query_string_aprob = urlencode(params)
                for estado in ["assigned", "done"]:
                    query_string_aprob += f"&estados={estado}"
                
                url_aprob = f"{API_URL}/api/v1/recepciones-mp/?{query_string_aprob}"
                
                # Fase 2: Consulta recepciones
                status_text.text("⏳ Fase 2/4: Consultando recepciones...")
                progress_bar.progress(50)
                resp = requests.get(url_aprob, timeout=60)

                if resp.status_code == 200:
                    st.session_state.aprob_data = resp.json()
                    
                    # Fase 3: Carga presupuesto
                    status_text.text("⏳ Fase 3/4: Cargando presupuestos...")
                    progress_bar.progress(75)
                    
                    # Cargar PPTO General
                    try:
                        resp_ppto = requests.get(
                            f"{API_URL}/api/v1/recepciones-mp/abastecimiento/precios",
                            params={"planta": None, "especie": None},
                            timeout=30
                        )
                        if resp_ppto.status_code == 200:
                            st.session_state.aprob_ppto = {item.get('especie', ''): item.get('precio_proyectado', 0) for item in resp_ppto.json()}
                    except:
                        pass

                    # Cargar PPTO Detallado por Productor
                    try:
                        resp_ppto_det = requests.get(
                            f"{API_URL}/api/v1/recepciones-mp/abastecimiento/precios-detalle",
                            params={"planta": None, "especie": None},
                            timeout=30
                        )
                        if resp_ppto_det.status_code == 200:
                            # Clave compuesta: (productor, especie)
                            st.session_state.aprob_ppto_detalle = {
                                (item.get('productor', ''), item.get('especie', '')): item.get('precio_proyectado', 0) 
                                for item in resp_ppto_det.json()
                            }
                    except:
                        pass

                    # Fase 4: Completado
                    status_text.text("✅ Fase 4/4: Completado")
                    progress_bar.progress(100)
                    st.toast(f"✅ Cargadas {len(st.session_state.aprob_data)} recepciones", icon="✅")
                else:
                    progress_bar.empty()
                    status_text.empty()
                    st.error(f"Error: {resp.text}")
            except Exception as e:
                progress_bar.empty()
                status_text.empty()
                st.error(f"Error: {e}")
        finally:
            st.session_state.recep_aprob_cargar_loading = False

    # --- MOSTRAR DATOS SI EXISTEN ---
    if st.session_state.aprob_data:
        recepciones = st.session_state.aprob_data
        precios_ppto_dict = st.session_state.aprob_ppto
        precios_ppto_detalle = st.session_state.aprob_ppto_detalle
        aprobadas_ids = get_aprobaciones()

        # Procesar datos
        filas_aprobacion = []
        for rec in recepciones:
            # Filtro de QC: Solo mostrar recepciones con quality_state = 'pass'
            quality_state = rec.get('quality_state', '') or ''
            if quality_state != 'pass':
                continue

            recep_name = rec.get('albaran', '')
            picking_id = rec.get('id', 0)  # ID para link a Odoo
            fecha_recep = rec.get('fecha', '')
            productor = rec.get('productor', '')
            guia = rec.get('guia_despacho', '') or ''
            oc = rec.get('oc_asociada', '') or ''
            es_aprobada = recep_name in aprobadas_ids

            if estado_filtro == "Pendientes" and es_aprobada: continue
            if estado_filtro == "Aprobadas" and not es_aprobada: continue

            productos = rec.get('productos', []) or []
            for p in productos:
                cat = (p.get('Categoria') or '').strip().upper()
                if 'BANDEJ' in cat: continue

                kg = p.get('Kg Hechos', 0) or 0
                # Permitir kg = 0 para recepciones preparadas pero no terminadas
                # if kg <= 0: continue

                precio_real = float(p.get('Costo Unitario', 0) or p.get('precio', 0) or 0)

                # Obtener nombre del producto
                prod_name_raw = p.get('Producto') or p.get('producto') or ''
                prod_name = prod_name_raw.upper()

                # Usar TipoFruta del producto directamente (como hace KPIs)
                tipo_fruta = (p.get('TipoFruta') or rec.get('tipo_fruta') or '').strip()
                manejo = (p.get('Manejo') or '').strip()

                # QC Status visual (vacío si no hay datos)
                clasif_qc = rec.get('calific_final', '') or ""

                # Si hay TipoFruta directo, usarlo
                if tipo_fruta:
                    esp_base = tipo_fruta
                    man_norm = manejo if manejo else 'Convencional'
                else:
                    # Fallback: detectar desde nombre del producto
                    man_norm = 'Orgánico' if ('ORG' in prod_name or 'ORGAN' in manejo.upper()) else 'Convencional'
                    esp_base = 'Otro'
                    # Detectar especie desde código/nombre
                    if 'AR ' in prod_name or 'AR HB' in prod_name or 'ARAND' in prod_name or 'BLUEBERRY' in prod_name:
                        esp_base = 'Arándano'
                    elif 'FR ' in prod_name or 'FRAM' in prod_name or 'MEEKER' in prod_name or 'HERITAGE' in prod_name or 'RASPBERRY' in prod_name:
                        esp_base = 'Frambuesa'
                    elif 'FT ' in prod_name or 'FRUTI' in prod_name or 'STRAW' in prod_name:
                        esp_base = 'Frutilla'
                    elif 'MO ' in prod_name or 'MORA' in prod_name or 'BLACKBERRY' in prod_name:
                        esp_base = 'Mora'
                    elif 'CE ' in prod_name or 'CEREZA' in prod_name or 'CHERRY' in prod_name:
                        esp_base = 'Cereza'

                especie_manejo = f"{esp_base} {man_norm}"

                # Obtener PPTO (Prioridad: Productor+Especie -> Solo Especie)
                ppto_val = 0
                if precios_ppto_detalle:
                     ppto_val = precios_ppto_detalle.get((productor, especie_manejo), 0)

                if ppto_val == 0:
                     ppto_val = precios_ppto_dict.get(especie_manejo, 0)

                desv = 0
                if ppto_val > 0:
                    desv = ((precio_real - ppto_val) / ppto_val)

                sema = "🟢"
                if desv > 0.08: sema = "🔴"
                elif desv > 0.03: sema = "🟡"

                calificacion = rec.get('calific_final', '') or ""

                # Desviación con color combinado
                desv_pct = f"{desv*100:.1f}% {sema}"

                filas_aprobacion.append({
                    "Sel": es_aprobada,
                    "Recepción": recep_name,
                    "Fecha": fmt_fecha(fecha_recep),
                    "Productor": productor,
                    "OC": oc,
                    "Producto": prod_name_raw[:40] if len(prod_name_raw) > 40 else prod_name_raw,
                    "Kg": fmt_numero(kg, 2),
                    "$/Kg": fmt_dinero(precio_real),
                    "PPTO": fmt_dinero(ppto_val),
                    "Desv": desv_pct,
                    "Calidad": calificacion,
                    "_id": recep_name,
                    "_kg_raw": kg,
                    "_picking_id": picking_id,
                    "Especie": especie_manejo,
                    "Estado": "✅ Hecho" if rec.get('state') == 'done' else "📦 Preparado"
                })

        if filas_aprobacion:
            df_full = pd.DataFrame(filas_aprobacion)

            # --- FILTROS ADICIONALES ---
            with st.expander("🔍 Filtros adicionales", expanded=False):
                col_f1, col_f2, col_f3, col_f4 = st.columns(4)
                with col_f1:
                    filtro_recep = st.text_input("Recepción", "", key="filtro_recep", placeholder="Buscar...")
                with col_f2:
                    filtro_prod = st.selectbox("Productor", ["Todos"] + sorted(df_full["Productor"].unique().tolist()), key="filtro_prod")
                with col_f3:
                    filtro_esp = st.selectbox("Especie", ["Todos"] + sorted(df_full["Especie"].unique().tolist()), key="filtro_esp")
                with col_f4:
                    filtro_cal = st.selectbox("Calidad", ["Todos"] + sorted(df_full["Calidad"].unique().tolist()), key="filtro_cal")

            with st.expander("🔍 Más filtros", expanded=False):
                col_f5, col_f6 = st.columns(2)
                with col_f5:
                    filtro_oc = st.text_input("OC", "", key="filtro_oc", placeholder="Buscar OC...")
                with col_f6:
                    opciones_estado = ["Todos"] + sorted(df_full["Estado"].unique().tolist())
                    # Intentar seleccionar "Preparado" por defecto si existe
                    idx_def = 0
                    try:
                        idx_def = opciones_estado.index("📦 Preparado")
                    except ValueError:
                        idx_def = 0

                    filtro_est_odoo = st.selectbox("Estado Odoo", opciones_estado, index=idx_def, key="filtro_est_odoo")

            # Aplicar filtros
            df_filtered = df_full.copy()
            if filtro_recep:
                df_filtered = df_filtered[df_filtered["Recepción"].str.contains(filtro_recep, case=False, na=False)]
            if filtro_prod != "Todos":
                df_filtered = df_filtered[df_filtered["Productor"] == filtro_prod]
            if filtro_esp != "Todos":
                df_filtered = df_filtered[df_filtered["Especie"] == filtro_esp]
            if filtro_oc:
                df_filtered = df_filtered[df_filtered["OC"].str.contains(filtro_oc, case=False, na=False)]
            if filtro_cal != "Todos":
                df_filtered = df_filtered[df_filtered["Calidad"] == filtro_cal]
            if filtro_est_odoo != "Todos":
                df_filtered = df_filtered[df_filtered["Estado"] == filtro_est_odoo]

            # --- GENERAR LINK A ODOO (reemplazar Recepción con URL) ---
            ODOO_BASE = "https://riofuturo.server98c6e.oerpondemand.net/web#"
            df_filtered["Recepción"] = df_filtered.apply(
                lambda row: f"{ODOO_BASE}id={row['_picking_id']}&menu_id=350&cids=1&action=540&model=stock.picking&view_type=form&display_name={row['Recepción']}" if row['_picking_id'] else row['Recepción'],
                axis=1
            )

            # --- CHECKBOX SELECCIONAR TODO INTELIGENTE ---
            # Guardar estado previo para detectar cambios
            if 'sel_all_state' not in st.session_state:
                st.session_state.sel_all_state = False
            if 'editor_key' not in st.session_state:
                st.session_state.editor_key = 0

            col_sel_all, col_info = st.columns([1, 3])
            with col_sel_all:
                sel_all_input = st.checkbox("☑️ Seleccionar todo", value=st.session_state.sel_all_state, key="chk_sel_all")

            # Detectar si cambió
            if sel_all_input != st.session_state.sel_all_state:
                st.session_state.sel_all_state = sel_all_input
                st.session_state.editor_key += 1 # Forzar reinicio del editor
                df_filtered["Sel"] = sel_all_input

            with col_info:
                st.caption(f"📋 {len(df_filtered)} líneas filtradas")

            # Mostrar tabla
            edited_df = st.data_editor(
                df_filtered,
                column_config={
                    "Sel": st.column_config.CheckboxColumn("✓", default=False, width="small"),
                    "Recepción": st.column_config.LinkColumn("Recepción", display_text=r"display_name=(.*?)(&|$)", width="medium"),
                    "Producto": st.column_config.TextColumn("Producto", width="medium"),
                    "Desv": st.column_config.TextColumn("Desv", width="small"),
                    "Calidad": st.column_config.TextColumn("Calidad", width="small"),
                    "Estado": st.column_config.TextColumn("Estado", width="small"),
                    "$/Kg": st.column_config.TextColumn("$/Kg"),
                    "PPTO": st.column_config.TextColumn("PPTO"),
                    "Kg": st.column_config.TextColumn("Kg"),
                    "Especie": None,
                    "_id": None,
                    "_kg_raw": None,
                    "_picking_id": None,
                },
                column_order=["Sel", "Recepción", "Fecha", "Estado", "Productor", "OC", "Producto", "Kg", "$/Kg", "PPTO", "Desv", "Calidad"],
                disabled=["Recepción", "Fecha", "Estado", "Productor", "OC", "Producto", "Kg", "$/Kg", "PPTO", "Desv", "Calidad"],
                hide_index=True,
                key=f"editor_aprob_{st.session_state.editor_key}",
                height=500,
                use_container_width=True
            )

            # --- BARRA DE ESTADO ---
            seleccionados = edited_df[edited_df["Sel"] == True]
            n_sel = len(seleccionados)
            kg_sel = seleccionados["_kg_raw"].sum() if n_sel > 0 else 0
            receps_sel = seleccionados["_id"].nunique() if n_sel > 0 else 0

            st.markdown(f"""
            <div style="background: rgba(50,100,150,0.3); padding: 10px; border-radius: 8px; margin: 10px 0;">
                <b>📊 Selección:</b> {n_sel} líneas · {receps_sel} recepciones · {fmt_numero(kg_sel, 2)} Kg
            </div>
            """, unsafe_allow_html=True)

            # --- BOTONES DE ACCIÓN --- (Habilitado para todos los usuarios autenticados)
            col_a, col_b = st.columns([1, 1])
            with col_a:
                if st.button("✅ Aprobar Seleccionadas", type="primary", use_container_width=True, disabled=st.session_state.recep_aprob_aprobar_loading):
                    st.session_state.recep_aprob_aprobar_loading = True
                    try:
                        ids_names = seleccionados["_id"].unique().tolist()
                        picking_ids = [int(pid) for pid in seleccionados["_picking_id"].unique().tolist() if pid]

                        if picking_ids:
                            with st.spinner("Validando en Odoo..."):
                                try:
                                    resp_val = requests.post(
                                        f"{API_URL}/api/v1/recepciones-mp/validate",
                                        params={"username": username, "password": password},
                                        json=picking_ids,
                                        timeout=60
                                    )
                                    if resp_val.status_code == 200:
                                        res_json = resp_val.json()
                                        if res_json.get("success"):
                                            save_aprobaciones(ids_names)
                                            st.toast(f"✅ {len(picking_ids)} recepciones validadas correctamente")
                                        else:
                                            # Mostrar lista de errores si existen
                                            errores_msg = "\n".join(res_json.get("errores", []))
                                            st.error(f"Error al validar algunas recepciones:\n{errores_msg}")
                                    else:
                                        st.error(f"Error en el servidor: {resp_val.text}")
                                except Exception as e:
                                    st.error(f"Error al conectar con la API: {e}")
                        else:
                            st.warning("Selecciona al menos una recepción con ID válido.")
                    finally:
                        st.session_state.recep_aprob_aprobar_loading = False
            with col_b:
                if estado_filtro != "Pendientes":
                    if st.button("↩️ Quitar Aprobación", use_container_width=True):
                        ids_del = seleccionados["_id"].unique().tolist()
                        if ids_del:
                            if remove_aprobaciones(ids_del):
                                st.warning(f"Se quitó aprobación a {len(ids_del)} recepciones.")
        else:
            st.info("No hay datos con los filtros seleccionados.")
    else:
        st.info("👆 Selecciona un rango de fechas y presiona **Cargar Recepciones**")


@st.fragment
def _fragment_pdf_reports(username: str, password: str):
    """Fragment secundario: Reportes PDF."""

    # ========== SECCIÓN INDEPENDIENTE: REPORTE PDF PARA PRODUCTORES ==========
    st.markdown("---")
    st.markdown("### 📄 Reportes PDF para Productores")
    st.caption("Genera reportes con datos de recepción para entregar a productores.")

    with st.expander("📊 Generar Reporte de Recepciones", expanded=False):
        # Filtros propios independientes
        col_r1, col_r2 = st.columns(2)
        with col_r1:
            fecha_ini_rep = st.date_input("Desde", datetime.now() - timedelta(days=30), key="fecha_ini_rep", format="DD/MM/YYYY")
        with col_r2:
            fecha_fin_rep = st.date_input("Hasta", datetime.now(), key="fecha_fin_rep", format="DD/MM/YYYY")

        estado_rep = st.radio("Estado de Recepción", ["Hechas", "Todas"], horizontal=True, key="estado_rep")

        if st.button("🔄 Cargar Productores", type="secondary", use_container_width=True, disabled=st.session_state.recep_aprob_productores_loading):
            st.session_state.recep_aprob_productores_loading = True
            try:
                # Progress bar personalizado
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                status_text.text("⏳ Fase 1/4: Conectando con API...")
                progress_bar.progress(25)
                
                try:
                    params_rep = {
                        "username": username,
                        "password": password,
                        "fecha_inicio": fecha_ini_rep.strftime("%Y-%m-%d"),
                        "fecha_fin": fecha_fin_rep.strftime("%Y-%m-%d")
                    }
                    
                    status_text.text("⏳ Fase 2/4: Consultando recepciones...")
                    progress_bar.progress(50)
                    
                    resp_rep = requests.get(f"{API_URL}/api/v1/recepciones-mp/", params=params_rep, timeout=60)

                    if resp_rep.status_code == 200:
                        status_text.text("⏳ Fase 3/4: Procesando datos...")
                        progress_bar.progress(75)
                        
                        recepciones_rep = resp_rep.json()

                        # Filtrar por estado si es necesario
                        if estado_rep == "Hechas":
                            recepciones_rep = [r for r in recepciones_rep if r.get('state') == 'done']

                        st.session_state.reporte_recepciones = recepciones_rep
                        
                        status_text.text("✅ Fase 4/4: Completado")
                        progress_bar.progress(100)
                        st.toast(f"✅ Cargadas {len(recepciones_rep)} recepciones")
                    else:
                        progress_bar.empty()
                        status_text.empty()
                        st.error(f"Error: {resp_rep.text}")
                except Exception as e:
                    progress_bar.empty()
                    status_text.empty()
                    st.error(f"Error: {e}")
            finally:
                st.session_state.recep_aprob_productores_loading = False

        # Si hay datos cargados para reporte
        if 'reporte_recepciones' in st.session_state and st.session_state.reporte_recepciones:
            recepciones_rep = st.session_state.reporte_recepciones

            # Obtener productores únicos
            productores_rep = sorted(list(set(r.get('productor', '') for r in recepciones_rep if r.get('productor'))))

            if productores_rep:
                productor_sel = st.selectbox("Seleccionar Productor", productores_rep, key="productor_rep_sel")

                col_btn1, col_btn2 = st.columns(2)
                with col_btn1:
                    gen_pdf = st.button("📥 Generar PDF", type="primary", use_container_width=True)
                with col_btn2:
                    gen_excel = st.button("📊 Generar Excel", type="secondary", use_container_width=True)

                # Filtrar recepciones del productor
                recs_prod = [r for r in recepciones_rep if r.get('productor') == productor_sel]

                if gen_pdf:
                    if recs_prod:
                        try:
                            from io import BytesIO
                            from reportlab.lib import colors
                            from reportlab.lib.pagesizes import letter, landscape
                            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                            from reportlab.lib.units import inch
                            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
                            from reportlab.lib.enums import TA_CENTER, TA_LEFT

                            buffer = BytesIO()
                            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                                                    leftMargin=0.4*inch, rightMargin=0.4*inch,
                                                    topMargin=0.4*inch, bottomMargin=0.4*inch)

                            styles = getSampleStyleSheet()
                            titulo_style = ParagraphStyle('Titulo', parent=styles['Heading1'],
                                                          fontSize=14, alignment=TA_CENTER, spaceAfter=15)
                            subtitulo_style = ParagraphStyle('Subtitulo', parent=styles['Normal'],
                                                             fontSize=9, alignment=TA_CENTER, spaceAfter=8)

                            elements = []
                            
                            # Agregar logo si existe
                            if LOGO_PATH.exists():
                                try:
                                    logo = Image(str(LOGO_PATH), width=1.5*inch, height=0.75*inch)
                                    logo.hAlign = 'LEFT'
                                    elements.append(logo)
                                    elements.append(Spacer(1, 10))
                                except:
                                    pass
                            
                            elements.append(Paragraph("Reporte de Recepciones de Materia Prima", titulo_style))
                            elements.append(Paragraph(f"Productor: {productor_sel}", subtitulo_style))
                            elements.append(Paragraph(f"Período: {fecha_ini_rep.strftime('%d/%m/%Y')} al {fecha_fin_rep.strftime('%d/%m/%Y')}", subtitulo_style))
                            elements.append(Spacer(1, 15))

                            # Encabezados de tabla
                            data = [["Fecha", "Guía", "Recepción", "Producto", "Kg", "$/Kg", "IQF%", "Block%", "Calif."]]

                            total_kg = 0
                            for rec in recs_prod:
                                fecha_str = fmt_fecha(rec.get('fecha', ''))
                                guia = rec.get('guia_despacho', '') or rec.get('x_studio_gua_de_despacho', '') or ''
                                recepcion = rec.get('albaran', '')
                                iqf_pct = rec.get('total_iqf', 0) or 0
                                block_pct = rec.get('total_block', 0) or 0
                                calific = rec.get('calific_final', '') or ''

                                productos = rec.get('productos', []) or []
                                for p in productos:
                                    cat = (p.get('Categoria') or '').upper()
                                    if 'BANDEJ' in cat:
                                        continue
                                    kg = p.get('Kg Hechos', 0) or 0
                                    if kg <= 0:
                                        continue

                                    prod_name = (p.get('Producto') or '')[:45]
                                    precio = p.get('Costo Unitario', 0) or 0

                                    data.append([
                                        fecha_str,
                                        str(guia)[:10],
                                        recepcion[-8:] if len(recepcion) > 8 else recepcion,
                                        prod_name,
                                        fmt_numero(kg, 2),
                                        fmt_dinero(precio),
                                        f"{iqf_pct:.1f}%" if iqf_pct else "",
                                        f"{block_pct:.1f}%" if block_pct else "",
                                        calific[:3] if calific else ""
                                    ])
                                    total_kg += kg

                            # Fila totales
                            data.append(["", "", "", "TOTAL", fmt_numero(total_kg, 2), "", "", "", ""])

                            # Crear tabla
                            col_widths = [0.75*inch, 0.55*inch, 0.85*inch, 3*inch, 0.85*inch, 0.8*inch, 0.55*inch, 0.55*inch, 0.45*inch]
                            table = Table(data, colWidths=col_widths)
                            table.setStyle(TableStyle([
                                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5276')),
                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                ('FONTSIZE', (0, 0), (-1, 0), 8),
                                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                                ('TOPPADDING', (0, 0), (-1, 0), 6),
                                ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#f8f9fa')),
                                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#2980b9')),
                                ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
                                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
                                ('FONTSIZE', (0, 1), (-1, -1), 7),
                                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                            ]))

                            elements.append(table)
                            doc.build(elements)

                            pdf_data = buffer.getvalue()
                            buffer.close()

                            # Nombre archivo
                            prod_clean = "".join(c for c in productor_sel if c.isalnum() or c in " _-")[:25]
                            filename = f"Recepciones_{prod_clean}_{fecha_ini_rep.strftime('%Y%m%d')}.pdf"

                            st.download_button(
                                label="⬇️ Descargar PDF",
                                data=pdf_data,
                                file_name=filename,
                                mime="application/pdf",
                                use_container_width=True
                            )
                            st.success(f"✅ PDF generado: {len(data)-2} líneas, {fmt_numero(total_kg, 2)} Kg")

                        except ImportError:
                            st.error("⚠️ Instalar reportlab: pip install reportlab")
                        except Exception as e:
                            st.error(f"Error generando PDF: {e}")
                    else:
                        st.warning("No hay recepciones para este productor.")

                if gen_excel:
                    if recs_prod:
                        try:
                            from io import BytesIO
                            from openpyxl import Workbook
                            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                            from openpyxl.drawing.image import Image as XLImage
                            from openpyxl.utils.dataframe import dataframe_to_rows

                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Recepciones"

                            # Agregar logo si existe
                            row_start = 1
                            if LOGO_PATH.exists():
                                try:
                                    img = XLImage(str(LOGO_PATH))
                                    img.width = 150
                                    img.height = 75
                                    ws.add_image(img, 'A1')
                                    row_start = 6  # Dejar espacio para el logo
                                except:
                                    pass

                            # Título
                            ws.merge_cells(start_row=row_start, start_column=1, end_row=row_start, end_column=9)
                            ws.cell(row=row_start, column=1, value="Reporte de Recepciones de Materia Prima")
                            ws.cell(row=row_start, column=1).font = Font(bold=True, size=14)
                            ws.cell(row=row_start, column=1).alignment = Alignment(horizontal='center')

                            ws.merge_cells(start_row=row_start+1, start_column=1, end_row=row_start+1, end_column=9)
                            ws.cell(row=row_start+1, column=1, value=f"Productor: {productor_sel}")
                            ws.cell(row=row_start+1, column=1).alignment = Alignment(horizontal='center')

                            ws.merge_cells(start_row=row_start+2, start_column=1, end_row=row_start+2, end_column=9)
                            ws.cell(row=row_start+2, column=1, value=f"Período: {fecha_ini_rep.strftime('%d/%m/%Y')} al {fecha_fin_rep.strftime('%d/%m/%Y')}")
                            ws.cell(row=row_start+2, column=1).alignment = Alignment(horizontal='center')

                            # Encabezados
                            header_row = row_start + 4
                            headers = ["Fecha", "Guía", "Recepción", "Producto", "Kg", "$/Kg", "IQF%", "Block%", "Calif."]
                            header_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
                            header_font = Font(bold=True, color="FFFFFF")
                            thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )

                            for col, header in enumerate(headers, 1):
                                cell = ws.cell(row=header_row, column=col, value=header)
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = Alignment(horizontal='center')
                                cell.border = thin_border

                            # Datos
                            data_row = header_row + 1
                            total_kg = 0
                            for rec in recs_prod:
                                fecha_str = fmt_fecha(rec.get('fecha', ''))
                                guia = rec.get('guia_despacho', '') or rec.get('x_studio_gua_de_despacho', '') or ''
                                recepcion = rec.get('albaran', '')
                                iqf_pct = rec.get('total_iqf', 0) or 0
                                block_pct = rec.get('total_block', 0) or 0
                                calific = rec.get('calific_final', '') or ''

                                productos = rec.get('productos', []) or []
                                for p in productos:
                                    cat = (p.get('Categoria') or '').upper()
                                    if 'BANDEJ' in cat:
                                        continue
                                    kg = p.get('Kg Hechos', 0) or 0
                                    if kg <= 0:
                                        continue

                                    prod_name = (p.get('Producto') or '')[:45]
                                    precio = p.get('Costo Unitario', 0) or 0

                                    ws.cell(row=data_row, column=1, value=fecha_str).border = thin_border
                                    ws.cell(row=data_row, column=2, value=str(guia)[:10]).border = thin_border
                                    ws.cell(row=data_row, column=3, value=recepcion[-8:] if len(recepcion) > 8 else recepcion).border = thin_border
                                    ws.cell(row=data_row, column=4, value=prod_name).border = thin_border
                                    ws.cell(row=data_row, column=5, value=kg).border = thin_border
                                    ws.cell(row=data_row, column=5).number_format = '#,##0.00'
                                    ws.cell(row=data_row, column=6, value=precio).border = thin_border
                                    ws.cell(row=data_row, column=6).number_format = '$#,##0'
                                    ws.cell(row=data_row, column=7, value=f"{iqf_pct:.1f}%" if iqf_pct else "").border = thin_border
                                    ws.cell(row=data_row, column=8, value=f"{block_pct:.1f}%" if block_pct else "").border = thin_border
                                    ws.cell(row=data_row, column=9, value=calific[:3] if calific else "").border = thin_border
                                    total_kg += kg
                                    data_row += 1

                            # Fila de totales
                            total_fill = PatternFill(start_color="2980b9", end_color="2980b9", fill_type="solid")
                            total_font = Font(bold=True, color="FFFFFF")
                            for col in range(1, 10):
                                ws.cell(row=data_row, column=col).fill = total_fill
                                ws.cell(row=data_row, column=col).font = total_font
                                ws.cell(row=data_row, column=col).border = thin_border
                            ws.cell(row=data_row, column=4, value="TOTAL")
                            ws.cell(row=data_row, column=5, value=total_kg)
                            ws.cell(row=data_row, column=5).number_format = '#,##0.00'

                            # Ajustar anchos de columna
                            ws.column_dimensions['A'].width = 12
                            ws.column_dimensions['B'].width = 10
                            ws.column_dimensions['C'].width = 12
                            ws.column_dimensions['D'].width = 45
                            ws.column_dimensions['E'].width = 12
                            ws.column_dimensions['F'].width = 12
                            ws.column_dimensions['G'].width = 8
                            ws.column_dimensions['H'].width = 8
                            ws.column_dimensions['I'].width = 8

                            # Guardar a buffer
                            excel_buffer = BytesIO()
                            wb.save(excel_buffer)
                            excel_data = excel_buffer.getvalue()
                            excel_buffer.close()

                            # Nombre archivo
                            prod_clean = "".join(c for c in productor_sel if c.isalnum() or c in " _-")[:25]
                            filename_xl = f"Recepciones_{prod_clean}_{fecha_ini_rep.strftime('%Y%m%d')}.xlsx"

                            st.download_button(
                                label="⬇️ Descargar Excel",
                                data=excel_data,
                                file_name=filename_xl,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                            st.success(f"✅ Excel generado: {data_row - header_row - 1} líneas, {fmt_numero(total_kg, 2)} Kg")

                        except ImportError as ie:
                            st.error(f"⚠️ Instalar openpyxl: pip install openpyxl - {ie}")
                        except Exception as e:
                            st.error(f"Error generando Excel: {e}")
                    else:
                        st.warning("No hay recepciones para este productor.")
            else:
                st.info("No hay productores en el rango de fechas seleccionado.")


def render(username: str, password: str):
    """Renderiza el contenido del tab Aprobaciones MP (Orquestador de fragments)."""
    _fragment_main_aprobaciones(username, password)
    _fragment_pdf_reports(username, password)
