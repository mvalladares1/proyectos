"""
Reporte de Bandejas por Productor
Genera Excel y PDF con detalle de bandejas despachadas (limpias) y recepcionadas (sucias)
por cada productor, incluyendo cuadro resumen con diferencias y pallets.
"""
import xmlrpc.client
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime
import os
import sys
import re

# Instalar reportlab si no existe
try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    HAS_REPORTLAB = True
except ImportError:
    print("Instalando reportlab...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "reportlab"])
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    HAS_REPORTLAB = True

# ── Conexión a Odoo ──────────────────────────────────────────────────────────
url = "https://riofuturo.server98c6e.oerpondemand.net"
db = "riofuturo-master"
username = "mvalladares@riofuturo.cl"
password = "c0766224bec30cac071ffe43a858c9ccbd521ddd"

print("Conectando a Odoo...")
common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common")
uid = common.authenticate(db, username, password, {})
models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object")


def search_read(model, domain, fields, limit=0, order=""):
    kwargs = {"fields": fields, "limit": limit}
    if order:
        kwargs["order"] = order
    return models.execute_kw(
        db, uid, password, model, "search_read", [domain], kwargs
    )


# ── Bandejas por pallet (calculado del Excel del usuario) ────────────────────
BANDEJAS_POR_PALLET = {
    "Bandeja Blanca 1/8 50x30": 200,
    "Bandejón IQF 60x40": 75,
    "Bandeja Verde 45x34": 120,
    "Bandeja Negra 50x30": 200,
    "Bandejón 3/4 Esparraguera": 60,
    "Bandeja azul 50x30": 200,
    "Bandeja Cosecha Mecanica": 120,
    "Bandeja baja IQF Gris 56x44": 75,
}

# ── Filtro de fechas ──────────────────────────────────────────────────────────
FECHA_DESDE = "2025-11-01"  # Desde noviembre 2025
FECHA_HASTA = "2026-03-11"  # Hasta hoy (inclusive)

# ── Picking types ────────────────────────────────────────────────────────────
# Recepciones MP (incoming fruit + sucia trays)
RECEPCION_TYPE_IDS = [1, 151, 164, 217]  # RF, RF(dup), San Jose, VLK
# Delivery Orders (outgoing clean trays)
DESPACHO_TYPE_IDS = [2, 194]  # RF Delivery Orders, VLK Expediciones


def extraer_tipo_bandeja(product_name):
    """Extrae el tipo de bandeja limpio del nombre del producto."""
    name = product_name
    # Quitar código entre corchetes al inicio
    name = re.sub(r"^\[.*?\]\s*", "", name)
    # Quitar "(A PRODUCTOR)" y variantes
    name = re.sub(r"\s*\(A PRODUCTOR\)", "", name, flags=re.IGNORECASE)
    # Quitar "- Sucia", "- Limpia", "(Productor)"
    name = re.sub(r"\s*-\s*(Sucia|Limpia|copia).*$", "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s*\(Productor\).*$", "", name, flags=re.IGNORECASE)
    return name.strip()


def es_limpia(product):
    """Determina si un producto es bandeja limpia."""
    code = product.get("default_code", "") or ""
    name = product.get("name", "")
    return code.endswith("L") or "Limpia" in name


# ── 1. Buscar productos de bandeja "A PRODUCTOR" ────────────────────────────
print("\nBuscando productos de bandeja 'A PRODUCTOR'...")
all_bandeja_products = search_read(
    "product.product",
    [
        "|",
        ("name", "ilike", "A PRODUCTOR"),
        ("name", "ilike", "Productor"),
    ],
    ["id", "name", "default_code"],
)
# Filtrar solo los que son realmente bandejas (no otros productos con "productor")
bandeja_products = [
    p
    for p in all_bandeja_products
    if any(
        kw in p["name"].lower()
        for kw in ["bandeja", "bandejón", "bandejon"]
    )
]

print(f"Productos de bandeja encontrados: {len(bandeja_products)}")

# Separar en sucias y limpias
sucia_products = [p for p in bandeja_products if not es_limpia(p)]
limpia_products = [p for p in bandeja_products if es_limpia(p)]

sucia_ids = [p["id"] for p in sucia_products]
limpia_ids = [p["id"] for p in limpia_products]

# Mapeo product_id → tipo bandeja
product_tipo = {}
for p in bandeja_products:
    product_tipo[p["id"]] = extraer_tipo_bandeja(p["name"])

print(f"  Sucias: {len(sucia_products)} productos")
for p in sucia_products:
    print(f"    [{p.get('default_code','')}] {p['name']} → {product_tipo[p['id']]}")
print(f"  Limpias: {len(limpia_products)} productos")
for p in limpia_products:
    print(f"    [{p.get('default_code','')}] {p['name']} → {product_tipo[p['id']]}")

all_bandeja_ids = sucia_ids + limpia_ids

# ── 2. Obtener todos los stock.move de bandejas en pickings completados ──────
print("\nBuscando movimientos de bandejas...")

all_moves = []
batch_size = 500

# Buscar moves de SUCIAS en recepciones
print("  Buscando sucias en recepciones...")
sucia_move_ids = models.execute_kw(
    db, uid, password, "stock.move", "search",
    [[
        ("product_id", "in", sucia_ids),
        ("state", "=", "done"),
        ("picking_id", "!=", False),
        ("picking_type_id", "in", RECEPCION_TYPE_IDS),
        ("date", ">=", FECHA_DESDE),
        ("date", "<", FECHA_HASTA),
    ]],
)
print(f"    {len(sucia_move_ids)} movimientos de sucias encontrados")

for i in range(0, len(sucia_move_ids), batch_size):
    batch = sucia_move_ids[i : i + batch_size]
    moves = search_read(
        "stock.move",
        [("id", "in", batch)],
        ["id", "picking_id", "product_id", "quantity_done"],
    )
    for m in moves:
        m["tipo"] = "recepcionada"
    all_moves.extend(moves)
    print(f"    Leídos {min(i + batch_size, len(sucia_move_ids))}/{len(sucia_move_ids)}")

# Buscar moves de LIMPIAS en despachos
print("  Buscando limpias en despachos...")
limpia_move_ids = models.execute_kw(
    db, uid, password, "stock.move", "search",
    [[
        ("product_id", "in", limpia_ids),
        ("state", "=", "done"),
        ("picking_id", "!=", False),
        ("picking_type_id", "in", DESPACHO_TYPE_IDS),
        ("date", ">=", FECHA_DESDE),
        ("date", "<", FECHA_HASTA),
    ]],
)
print(f"    {len(limpia_move_ids)} movimientos de limpias encontrados")

for i in range(0, len(limpia_move_ids), batch_size):
    batch = limpia_move_ids[i : i + batch_size]
    moves = search_read(
        "stock.move",
        [("id", "in", batch)],
        ["id", "picking_id", "product_id", "quantity_done"],
    )
    for m in moves:
        m["tipo"] = "despachada"
    all_moves.extend(moves)
    print(f"    Leídos {min(i + batch_size, len(limpia_move_ids))}/{len(limpia_move_ids)}")

print(f"\nTotal movimientos: {len(all_moves)}")

# ── 3. Obtener datos de los pickings ─────────────────────────────────────────
print("\nObteniendo datos de las transferencias...")
picking_ids = list({m["picking_id"][0] for m in all_moves if m["picking_id"]})
print(f"  {len(picking_ids)} transferencias únicas")

picking_data = {}
picking_fields = [
    "id", "name", "partner_id", "date_done", "scheduled_date",
    "x_studio_gua_de_despacho", "x_studio_fecha_de_emisin_gdd",
    "picking_type_id",
]

for i in range(0, len(picking_ids), batch_size):
    batch = picking_ids[i : i + batch_size]
    picks = search_read("stock.picking", [("id", "in", batch)], picking_fields)
    for p in picks:
        picking_data[p["id"]] = p
    print(f"  Leídos {min(i + batch_size, len(picking_ids))}/{len(picking_ids)}")

# ── 4. Agrupar datos por productor ────────────────────────────────────────────
print("\nAgrupando datos por productor...")

# Estructura: {partner_name: {despachadas: [...], recepcionadas: [...]}}
por_productor = defaultdict(lambda: {"despachadas": [], "recepcionadas": []})

for move in all_moves:
    picking_id = move["picking_id"][0]
    pick = picking_data.get(picking_id)
    if not pick or not pick.get("partner_id"):
        continue

    partner_name = pick["partner_id"][1]
    product_id = move["product_id"][0]
    tipo_bandeja = product_tipo.get(product_id, move["product_id"][1])
    qty = move["quantity_done"] or 0
    if qty <= 0:
        continue

    # Fecha: preferir fecha de emisión GdD, luego date_done
    fecha_str = pick.get("x_studio_fecha_de_emisin_gdd") or ""
    if fecha_str:
        try:
            fecha = datetime.strptime(str(fecha_str), "%Y-%m-%d")
        except ValueError:
            fecha = None
    else:
        fecha = None
    if not fecha and pick.get("date_done"):
        try:
            fecha = datetime.strptime(str(pick["date_done"])[:10], "%Y-%m-%d")
        except ValueError:
            fecha = None

    guia = pick.get("x_studio_gua_de_despacho") or ""
    picking_name = pick.get("name", "")

    entry = {
        "fecha": fecha,
        "guia": guia,
        "tipo_bandeja": tipo_bandeja,
        "cantidad": qty,
        "picking": picking_name,
    }

    if move["tipo"] == "despachada":
        por_productor[partner_name]["despachadas"].append(entry)
    else:
        por_productor[partner_name]["recepcionadas"].append(entry)

# Ordenar cada lista por fecha
for partner, data in por_productor.items():
    data["despachadas"].sort(key=lambda x: x["fecha"] or datetime.min)
    data["recepcionadas"].sort(key=lambda x: x["fecha"] or datetime.min)

print(f"Productores encontrados: {len(por_productor)}")
for name, data in sorted(por_productor.items()):
    print(f"  {name}: {len(data['despachadas'])} desp, {len(data['recepcionadas'])} recep")

# ── 5. Crear directorio de salida ─────────────────────────────────────────────
output_dir = os.path.join(os.path.dirname(__file__), "output", "bandejas_productores")
os.makedirs(output_dir, exist_ok=True)

# ── Estilos comunes Excel ─────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
DATA_FONT = Font(size=10, name="Calibri")
TOTAL_FONT = Font(bold=True, size=11, name="Calibri")
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUMMARY_HEADER_FILL = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
POSITIVE_FONT = Font(bold=True, size=10, name="Calibri", color="006100")
NEGATIVE_FONT = Font(bold=True, size=10, name="Calibri", color="9C0006")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)


def safe_filename(name):
    """Genera nombre de archivo seguro."""
    return re.sub(r'[<>:"/\\|?*]', "_", name).strip()[:80]


def apply_cell_style(cell, font=DATA_FONT, fill=None, align=None, border=THIN_BORDER):
    cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if border:
        cell.border = border


# ── 6. Generar Excel por productor ────────────────────────────────────────────
print("\nGenerando archivos Excel y PDF...")


def generar_excel(partner_name, data, filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bandejas"

    row = 1

    # ── Título ────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    title_cell = ws.cell(row=row, column=1, value=f"Control de Bandejas - {partner_name}")
    title_cell.font = Font(bold=True, size=14, name="Calibri", color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    row += 1

    # Fecha de generación
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    date_cell = ws.cell(
        row=row, column=1,
        value=f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    )
    date_cell.font = Font(size=9, italic=True, name="Calibri", color="666666")
    date_cell.alignment = Alignment(horizontal="center")
    row += 2

    headers = ["Fecha", "Guía Despacho", "Tipo Bandeja", "Cantidad", "Referencia"]

    def escribir_seccion(titulo, entries, start_row):
        r = start_row
        # Título sección
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        sec_cell = ws.cell(row=r, column=1, value=titulo)
        sec_cell.font = Font(bold=True, size=12, name="Calibri", color="FFFFFF")
        sec_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        sec_cell.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(2, 6):
            ws.cell(row=r, column=c).fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        ws.row_dimensions[r].height = 25
        r += 1

        # Headers
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=col, value=h)
            apply_cell_style(cell, SUBHEADER_FONT, SUBHEADER_FILL,
                             Alignment(horizontal="center", vertical="center"))
        ws.row_dimensions[r].height = 25
        r += 1

        # Datos
        total_qty = 0
        if not entries:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            ws.cell(row=r, column=1, value="Sin registros").font = Font(
                size=10, italic=True, name="Calibri", color="999999"
            )
            r += 1
        else:
            for idx, entry in enumerate(entries):
                fill = ALT_FILL if idx % 2 == 0 else None
                fecha_val = entry["fecha"].strftime("%d-%m-%Y") if entry["fecha"] else ""
                values = [fecha_val, entry["guia"], entry["tipo_bandeja"],
                          entry["cantidad"], entry["picking"]]
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=r, column=col, value=val)
                    align = Alignment(horizontal="right") if col == 4 else Alignment(horizontal="left")
                    if col == 4:
                        cell.number_format = "#,##0"
                    apply_cell_style(cell, DATA_FONT, fill, align)
                total_qty += entry["cantidad"]
                r += 1

        # Fila total
        for col in range(1, 6):
            cell = ws.cell(row=r, column=col)
            apply_cell_style(cell, TOTAL_FONT, TOTAL_FILL)
        ws.cell(row=r, column=1, value="TOTAL").alignment = Alignment(horizontal="right")
        total_cell = ws.cell(row=r, column=4, value=total_qty)
        total_cell.number_format = "#,##0"
        total_cell.alignment = Alignment(horizontal="right")
        apply_cell_style(total_cell, TOTAL_FONT, TOTAL_FILL, Alignment(horizontal="right"))
        r += 1

        return r, total_qty

    # ── Sección Despachadas (Limpias) ─────────────────────────────────────
    row, total_desp = escribir_seccion(
        "BANDEJAS DESPACHADAS (LIMPIAS)", data["despachadas"], row
    )
    row += 1

    # ── Sección Recepcionadas (Sucias) ────────────────────────────────────
    row, total_recep = escribir_seccion(
        "BANDEJAS RECEPCIONADAS (SUCIAS)", data["recepcionadas"], row
    )
    row += 2

    # ── Cuadro Resumen ────────────────────────────────────────────────────
    # Calcular totales por tipo
    tipos_recep = defaultdict(float)
    tipos_desp = defaultdict(float)
    for e in data["recepcionadas"]:
        tipos_recep[e["tipo_bandeja"]] += e["cantidad"]
    for e in data["despachadas"]:
        tipos_desp[e["tipo_bandeja"]] += e["cantidad"]

    all_tipos = sorted(set(list(tipos_recep.keys()) + list(tipos_desp.keys())))

    # Título resumen
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value="RESUMEN POR TIPO DE BANDEJA").font = Font(
        bold=True, size=12, name="Calibri", color="FFFFFF"
    )
    ws.cell(row=row, column=1).fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    for c in range(2, 5):
        ws.cell(row=row, column=c).fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
    row += 1

    # Headers resumen
    resumen_headers = ["TIPO DE BANDEJA", "RECEPCIONADAS", "DESPACHADAS", "DIFERENCIA"]
    for col, h in enumerate(resumen_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        apply_cell_style(cell, SUBHEADER_FONT, SUMMARY_HEADER_FILL,
                         Alignment(horizontal="center", vertical="center", wrap_text=True))
    row += 1

    # Datos resumen
    total_recep_all = 0
    total_desp_all = 0
    for idx, tipo in enumerate(all_tipos):
        fill = ALT_FILL if idx % 2 == 0 else None
        recep = tipos_recep.get(tipo, 0)
        desp = tipos_desp.get(tipo, 0)
        diff = recep - desp
        total_recep_all += recep
        total_desp_all += desp

        ws.cell(row=row, column=1, value=tipo)
        apply_cell_style(ws.cell(row=row, column=1), DATA_FONT, fill, Alignment(horizontal="left"))

        for col, val in [(2, recep), (3, desp)]:
            cell = ws.cell(row=row, column=col, value=int(val))
            cell.number_format = "#,##0"
            apply_cell_style(cell, DATA_FONT, fill, Alignment(horizontal="right"))

        diff_cell = ws.cell(row=row, column=4, value=int(diff))
        diff_cell.number_format = "#,##0"
        diff_font = POSITIVE_FONT if diff >= 0 else NEGATIVE_FONT
        apply_cell_style(diff_cell, diff_font, fill, Alignment(horizontal="right"))
        row += 1

    # Total resumen
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = TOTAL_FILL
        ws.cell(row=row, column=col).border = THIN_BORDER
    ws.cell(row=row, column=1, value="TOTAL GENERAL").font = TOTAL_FONT
    ws.cell(row=row, column=2, value=int(total_recep_all)).font = TOTAL_FONT
    ws.cell(row=row, column=2).number_format = "#,##0"
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=3, value=int(total_desp_all)).font = TOTAL_FONT
    ws.cell(row=row, column=3).number_format = "#,##0"
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="right")
    total_diff = int(total_recep_all - total_desp_all)
    diff_cell = ws.cell(row=row, column=4, value=total_diff)
    diff_cell.number_format = "#,##0"
    diff_cell.font = POSITIVE_FONT if total_diff >= 0 else NEGATIVE_FONT
    diff_cell.alignment = Alignment(horizontal="right")
    row += 2

    # ── Cuadro Cantidad de Pallets ────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value="CANTIDAD DE PALLETS (Diferencia)").font = Font(
        bold=True, size=11, name="Calibri", color="FFFFFF"
    )
    ws.cell(row=row, column=1).fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    ws.cell(row=row, column=2).fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    row += 1

    pallet_headers = ["TIPO DE BANDEJA", "CANTIDAD DE PALLET"]
    for col, h in enumerate(pallet_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        apply_cell_style(cell, SUBHEADER_FONT, SUMMARY_HEADER_FILL,
                         Alignment(horizontal="center", vertical="center"))
    row += 1

    for idx, tipo in enumerate(all_tipos):
        fill = ALT_FILL if idx % 2 == 0 else None
        diff = tipos_recep.get(tipo, 0) - tipos_desp.get(tipo, 0)
        bpp = BANDEJAS_POR_PALLET.get(tipo, 0)
        pallet_qty = round(diff / bpp, 2) if bpp > 0 else 0

        ws.cell(row=row, column=1, value=tipo)
        apply_cell_style(ws.cell(row=row, column=1), DATA_FONT, fill,
                         Alignment(horizontal="left"))

        pallet_cell = ws.cell(row=row, column=2, value=pallet_qty)
        pallet_cell.number_format = "#,##0.00"
        pf = POSITIVE_FONT if pallet_qty >= 0 else NEGATIVE_FONT
        apply_cell_style(pallet_cell, pf, fill, Alignment(horizontal="right"))
        row += 1

    # Ajustar anchos
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 25

    ws.freeze_panes = "A4"
    wb.save(filepath)
    return filepath


# ── 7. Generar PDF por productor ──────────────────────────────────────────────

def generar_pdf(partner_name, data, filepath):
    doc = SimpleDocTemplate(
        filepath,
        pagesize=landscape(A4),
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleCustom", parent=styles["Title"],
        fontSize=16, textColor=colors.HexColor("#1F4E79"),
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "SubtitleCustom", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#666666"),
        spaceAfter=12,
    )
    section_style = ParagraphStyle(
        "SectionCustom", parent=styles["Heading2"],
        fontSize=12, textColor=colors.white,
        backColor=colors.HexColor("#1F4E79"),
        spaceAfter=4, spaceBefore=12,
        leftIndent=4, rightIndent=4,
    )

    elements = []

    # Logo + Título en una tabla lado a lado
    logo_path = os.path.join(os.path.dirname(__file__), "data", "RFP - LOGO OFICIAL.png")
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=3 * cm, height=3 * cm)
        logo.hAlign = "LEFT"
        title_para = Paragraph(f"Control de Bandejas<br/><font size=10>{partner_name}</font>", title_style)
        date_para = Paragraph(
            f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}", subtitle_style
        )
        from reportlab.platypus import KeepTogether
        header_table = Table(
            [[logo, [title_para, date_para]]],
            colWidths=[3.5 * cm, None],
        )
        header_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 8))
    else:
        elements.append(Paragraph(f"Control de Bandejas - {partner_name}", title_style))
        elements.append(Paragraph(
            f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}", subtitle_style
        ))

    # Colores para tablas
    header_bg = colors.HexColor("#2E75B6")
    alt_bg = colors.HexColor("#F2F7FB")
    total_bg = colors.HexColor("#D6E4F0")

    def crear_tabla_detalle(titulo, entries):
        elements.append(Paragraph(titulo, section_style))
        elements.append(Spacer(1, 4))

        table_data = [["Fecha", "Guía Despacho", "Tipo Bandeja", "Cantidad", "Referencia"]]

        total_qty = 0
        if not entries:
            table_data.append(["Sin registros", "", "", "", ""])
        else:
            for entry in entries:
                fecha_val = entry["fecha"].strftime("%d-%m-%Y") if entry["fecha"] else ""
                table_data.append([
                    fecha_val,
                    str(entry["guia"]),
                    entry["tipo_bandeja"],
                    f"{int(entry['cantidad']):,}",
                    entry["picking"],
                ])
                total_qty += entry["cantidad"]

        table_data.append(["", "", "TOTAL", f"{int(total_qty):,}", ""])

        col_widths = [3 * cm, 3.5 * cm, 7 * cm, 3 * cm, 5 * cm]
        t = Table(table_data, colWidths=col_widths, repeatRows=1)

        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (3, 0), (3, -1), "RIGHT"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B4C6E7")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, alt_bg]),
            ("BACKGROUND", (0, -1), (-1, -1), total_bg),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        t.setStyle(TableStyle(style_cmds))
        elements.append(t)
        elements.append(Spacer(1, 8))

    # Sección Despachadas
    crear_tabla_detalle("BANDEJAS DESPACHADAS (LIMPIAS)", data["despachadas"])

    # Sección Recepcionadas
    crear_tabla_detalle("BANDEJAS RECEPCIONADAS (SUCIAS)", data["recepcionadas"])

    # ── Cuadro Resumen ────────────────────────────────────────────────────
    tipos_recep = defaultdict(float)
    tipos_desp = defaultdict(float)
    for e in data["recepcionadas"]:
        tipos_recep[e["tipo_bandeja"]] += e["cantidad"]
    for e in data["despachadas"]:
        tipos_desp[e["tipo_bandeja"]] += e["cantidad"]
    all_tipos = sorted(set(list(tipos_recep.keys()) + list(tipos_desp.keys())))

    elements.append(Spacer(1, 12))
    elements.append(Paragraph("RESUMEN POR TIPO DE BANDEJA", section_style))
    elements.append(Spacer(1, 4))

    resumen_data = [["TIPO DE BANDEJA", "RECEPCIONADAS", "DESPACHADAS", "DIFERENCIA"]]
    total_r = total_d = 0
    for tipo in all_tipos:
        r = int(tipos_recep.get(tipo, 0))
        d = int(tipos_desp.get(tipo, 0))
        diff = r - d
        total_r += r
        total_d += d
        resumen_data.append([tipo, f"{r:,}", f"{d:,}", f"{diff:,}"])
    resumen_data.append(["TOTAL GENERAL", f"{total_r:,}", f"{total_d:,}", f"{total_r - total_d:,}"])

    col_widths_r = [7 * cm, 4 * cm, 4 * cm, 4 * cm]
    rt = Table(resumen_data, colWidths=col_widths_r, repeatRows=1)
    rt_style = [
        ("BACKGROUND", (0, 0), (-1, 0), header_bg),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B4C6E7")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, alt_bg]),
        ("BACKGROUND", (0, -1), (-1, -1), total_bg),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    rt.setStyle(TableStyle(rt_style))
    elements.append(rt)

    # ── Cuadro Pallets ────────────────────────────────────────────────────
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("CANTIDAD DE PALLETS (Diferencia)", section_style))
    elements.append(Spacer(1, 4))

    pallet_data = [["TIPO DE BANDEJA", "CANTIDAD DE PALLET"]]
    for tipo in all_tipos:
        diff = tipos_recep.get(tipo, 0) - tipos_desp.get(tipo, 0)
        bpp = BANDEJAS_POR_PALLET.get(tipo, 0)
        pq = round(diff / bpp, 2) if bpp > 0 else 0
        pallet_data.append([tipo, f"{pq:,.2f}"])

    col_widths_p = [7 * cm, 5 * cm]
    pt = Table(pallet_data, colWidths=col_widths_p, repeatRows=1)
    pt_style = [
        ("BACKGROUND", (0, 0), (-1, 0), header_bg),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B4C6E7")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, alt_bg]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    pt.setStyle(TableStyle(pt_style))
    elements.append(pt)

    doc.build(elements)
    return filepath


# ── 8. Generar archivos ──────────────────────────────────────────────────────
total_excel = 0
total_pdf = 0

for partner_name, data in sorted(por_productor.items()):
    safe_name = safe_filename(partner_name)

    # Excel
    excel_path = os.path.join(output_dir, f"Bandejas_{safe_name}.xlsx")
    try:
        generar_excel(partner_name, data, excel_path)
        total_excel += 1
    except Exception as e:
        print(f"  ERROR Excel {partner_name}: {e}")

    # PDF
    pdf_path = os.path.join(output_dir, f"Bandejas_{safe_name}.pdf")
    try:
        generar_pdf(partner_name, data, pdf_path)
        total_pdf += 1
    except Exception as e:
        print(f"  ERROR PDF {partner_name}: {e}")

print(f"\n{'=' * 60}")
print(f"  Archivos generados en: {output_dir}")
print(f"  Productores: {len(por_productor)}")
print(f"  Excel: {total_excel} archivos")
print(f"  PDF:   {total_pdf} archivos")
print(f"{'=' * 60}")

# Limpiar archivo de descubrimiento si existe
discovery_file = os.path.join(os.path.dirname(__file__), "_discover_bandejas.py")
if os.path.exists(discovery_file):
    os.remove(discovery_file)
    print(f"\nLimpiado: {discovery_file}")
