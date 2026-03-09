import xmlrpc.client
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from collections import defaultdict
from datetime import datetime, timedelta

# Conexión a Odoo
url = "https://riofuturo.server98c6e.oerpondemand.net"
db = "riofuturo-master"
username = "mvalladares@riofuturo.cl"
password = "c0766224bec30cac071ffe43a858c9ccbd521ddd"

MESES = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
          7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}

# Mapeo de prefijo de producto a familia de fruta
FAMILIAS_FRUTA = {
    'AR': 'Arándano',
    'FB': 'Frambuesa',
    'FT': 'Frutilla',
    'MO': 'Mora',
}

def get_familia(product_name):
    """Extrae la familia de fruta del nombre del producto (AR, FB, FT, MO)."""
    name = product_name.strip()
    # Saltar prefijo [código] si existe
    if name.startswith('['):
        idx = name.find(']')
        if idx >= 0:
            name = name[idx+1:].strip()
    prefix = name[:2].upper()
    return FAMILIAS_FRUTA.get(prefix, 'Otro')

# Tag de producto: MP (ID 18) para todos los procesos
MP_TAG_ID = 18
# Categoría de producto: solo PRODUCTOS / MP (ID 4) - para San José
MP_CATEG_ID = 4

print("Conectando a Odoo...")
common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common", context=None)
uid = common.authenticate(db, username, password, {})
models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object", context=None)

def search_read(model, domain, fields, limit=0):
    return models.execute_kw(db, uid, password, model, 'search_read', [domain], {'fields': fields, 'limit': limit})

registros = []
batch_size = 200

# Pre-cargar productos con tag MP
print("Cargando productos con tag MP...")
mp_tag_products = search_read('product.product', [('product_tag_ids', 'in', [MP_TAG_ID])], ['id', 'name'])
mp_tag_product_ids = set(p['id'] for p in mp_tag_products)
print(f"Productos con tag MP: {len(mp_tag_product_ids)}")

# Para San José: categoría MP
mp_categ_products = search_read('product.product', [('categ_id', '=', MP_CATEG_ID)], ['id', 'name'])
mp_categ_product_ids = set(p['id'] for p in mp_categ_products)
print(f"Productos categoría MP (San José): {len(mp_categ_product_ids)}")

# Para túneles: unión de tag MP + categoría MP (incluye FB MK Conv. IQF en Bandeja)
mp_all_product_ids = mp_tag_product_ids | mp_categ_product_ids
print(f"Productos MP combinados (túneles): {len(mp_all_product_ids)}")

# ==========================================
# PARTE 1: Túneles Estáticos 1,2,3 + VLK + Continuo
# ==========================================
print("\n=== PROCESOS DE CONGELADO ===")

# Buscar todos los productos de proceso congelado túnel
tunnel_products = search_read('product.product', [
    ('name', 'ilike', 'PROCESO CONGELADO TÚNEL'),
], ['id', 'name'])
if not tunnel_products:
    tunnel_products = search_read('product.product', [
        ('name', 'ilike', 'PROCESO CONGELADO TUNEL'),
    ], ['id', 'name'])

print(f"Procesos encontrados: {len(tunnel_products)}")
for p in tunnel_products:
    print(f"  - {p['name']} (ID: {p['id']})")

product_ids = [p['id'] for p in tunnel_products]

short_names = {}
for p in tunnel_products:
    name = p['name'].upper()
    if 'VLK' in name:
        short_names[p['id']] = '[1.1.1] Túnel Estático VLK'
    elif 'ESTÁTICO 1' in name or 'ESTATICO 1' in name:
        short_names[p['id']] = '[1.1] Túnel Estático 1'
    elif 'ESTÁTICO 2' in name or 'ESTATICO 2' in name:
        short_names[p['id']] = '[1.2] Túnel Estático 2'
    elif 'ESTÁTICO 3' in name or 'ESTATICO 3' in name:
        short_names[p['id']] = '[1.3] Túnel Estático 3'
    elif 'CONTÍNUO' in name or 'CONTINUO' in name:
        short_names[p['id']] = '[1.4] Túnel Continuo'
    else:
        short_names[p['id']] = p['name']

def get_planta(tunnel_short):
    if 'VLK' in tunnel_short:
        return 'Vilkun'
    elif 'San José' in tunnel_short:
        return 'San José'
    return 'Río Futuro Procesos'

def utc_to_chile(date_raw):
    """Convierte fecha UTC de Odoo a hora Chile (UTC-3)."""
    s = str(date_raw or '')
    try:
        dt = datetime.strptime(s[:19], '%Y-%m-%d %H:%M:%S')
        dt = dt - timedelta(hours=3)
        return dt.strftime('%Y-%m-%d')
    except:
        return s[:10]

# Buscar producciones terminadas
print("Buscando producciones terminadas...")
productions = search_read('mrp.production', [
    ('product_id', 'in', product_ids),
    ('state', '=', 'done'),
], ['id', 'name', 'product_id', 'move_raw_ids', 'qty_produced', 'date_finished'])
print(f"Producciones encontradas: {len(productions)}")

# Mapear move_id -> info
all_raw_ids = []
move_to_info = {}
for prod in productions:
    tunnel_id = prod['product_id'][0]
    tunnel_short = short_names.get(tunnel_id, prod['product_id'][1])
    date_str = utc_to_chile(prod.get('date_finished', ''))
    year = int(date_str[:4]) if len(date_str) >= 4 and date_str[:4].isdigit() else 0
    month = int(date_str[5:7]) if len(date_str) >= 7 and date_str[5:7].isdigit() else 0
    mes_nombre = MESES.get(month, '?')
    
    for mid in prod['move_raw_ids']:
        all_raw_ids.append(mid)
        move_to_info[mid] = {
            'tunnel': tunnel_short,
            'date': date_str,
            'month': month,
            'year': year,
            'mes': f"{year}-{month:02d} {mes_nombre}",
            'planta': get_planta(tunnel_short),
        }

print(f"Leyendo {len(all_raw_ids)} movimientos de materia prima...")

for i in range(0, len(all_raw_ids), batch_size):
    batch = all_raw_ids[i:i+batch_size]
    moves = search_read('stock.move', [('id', 'in', batch)], [
        'id', 'product_id', 'product_uom_qty', 'quantity_done', 'product_uom'
    ])
    for move in moves:
        product_id = move['product_id'][0] if move['product_id'] else 0
        product_name = move['product_id'][1] if move['product_id'] else 'Desconocido'
        info = move_to_info.get(move['id'], {})
        # Solo productos MP (tag MP o categoría MP)
        if product_id not in mp_all_product_ids:
            continue
        # VLK: usar product_uom_qty como fallback si quantity_done=0
        qty = move.get('quantity_done', 0) or 0
        if qty <= 0 and 'VLK' in info.get('tunnel', ''):
            qty = move.get('product_uom_qty', 0) or 0
        if qty <= 0:
            continue
        registros.append({
            'mes': info.get('mes', ''),
            'fecha': info.get('date', ''),
            'planta': info.get('planta', ''),
            'tunel': info.get('tunnel', ''),
            'tipo': 'MP',
            'producto': product_name,
            'familia': get_familia(product_name),
            'kg': qty,
        })
    print(f"  Procesos: {min(i+batch_size, len(all_raw_ids))}/{len(all_raw_ids)}")

total_procesos = sum(r['kg'] for r in registros)
print(f"Total KG Procesos Congelado: {total_procesos:,.2f}")

# ==========================================
# PARTE 2: San José = Recepciones tipo 164 + RF2026M + pallets PACK específicos
# ==========================================
print("\n=== RECEPCIONES SAN JOSÉ ===")

# 1) Todas las recepciones San José (picking_type_id=164)
SJ_PICKING_TYPE = 164
sj_picks = search_read('stock.picking', [
    ('picking_type_id', '=', SJ_PICKING_TYPE),
    ('state', '=', 'done'),
], ['id'])
sj_pick_ids_164 = set(p['id'] for p in sj_picks)
print(f"Recepciones San José (tipo 164): {len(sj_pick_ids_164)}")

# 2) Recepciones de OTROS tipos que tengan paquetes RF2026M
all_packages = search_read('stock.quant.package', [('name', '=like', 'RF2026M%')], ['id', 'name'])
print(f"Paquetes RF2026M encontrados: {len(all_packages)}")
pkg_ids = [p['id'] for p in all_packages]

print("Buscando move lines con paquetes RF2026M...")
all_ml_rf = []
for i in range(0, len(pkg_ids), batch_size):
    batch = pkg_ids[i:i+batch_size]
    mls = search_read('stock.move.line', [
        ('result_package_id', 'in', batch),
        ('state', '=', 'done'),
    ], ['id', 'picking_id'])
    all_ml_rf.extend(mls)
    print(f"  RF2026M: {min(i+batch_size, len(pkg_ids))}/{len(pkg_ids)} -> {len(mls)} lines")

rf_pick_ids = set(ml['picking_id'][0] for ml in all_ml_rf if ml.get('picking_id'))
extra_rf_picks = rf_pick_ids - sj_pick_ids_164
print(f"Recepciones con RF2026M en otros tipos de operación: {len(extra_rf_picks)}")

# 3) Pallets PACK específicos que pertenecen a San José
SJ_PALLETS = [
    'PACK0011254','PACK0011354','PACK0011363','PACK0011333',
    'PACK0011334','PACK0011277','PACK0011362','PACK0011278',
    'PACK0013332','PACK0013297','PACK0013316','PACK0010981',
    'PACK0012532','PACK0013203','PACK0013329','PACK0010723',
    'PACK0011352','PACK0013283','PACK0013204','PACK0013207',
    'PACK0013348','PACK0011357','PACK0013346','PACK0011359',
    'PACK0014040','PACK0014039','PACK0013229','PACK0012597',
    'PACK0013198','PACK0012598','PACK0013254','PACK0013363',
    'PACK0012635','PACK0012647','PACK0013193','PACK0012626',
    'PACK0013197','PACK0012664','PACK0013282','PACK0013192',
    'PACK0011365','PACK0012652','PACK0013208',
    'PACK0011364','PACK0013228','PACK0012639','PACK0012653',
]
SJ_PALLETS = list(dict.fromkeys(SJ_PALLETS))  # únicos
sj_pack_pkgs = search_read('stock.quant.package', [('name', 'in', SJ_PALLETS)], ['id', 'name'])
sj_pack_ids = [p['id'] for p in sj_pack_pkgs]
print(f"Pallets PACK San José encontrados: {len(sj_pack_pkgs)}")

# Buscar move lines de esos pallets (solo las específicas, no toda la recepción)
sj_pack_mls = search_read('stock.move.line', [
    ('result_package_id', 'in', sj_pack_ids),
    ('state', '=', 'done'),
], ['id', 'picking_id', 'product_id', 'qty_done'])
sj_pack_pick_ids = set(ml['picking_id'][0] for ml in sj_pack_mls if ml.get('picking_id'))
# Recepciones que SOLO están en los PACK (no en 164 ni RF2026M)
extra_pack_picks = sj_pack_pick_ids - sj_pick_ids_164 - rf_pick_ids
print(f"Recepciones adicionales por pallets PACK: {len(extra_pack_picks)}")
# IDs de move lines de los pallets PACK (para evitar duplicar si la recepción ya está en SJ)
sj_pack_ml_ids = set(ml['id'] for ml in sj_pack_mls)
print(f"Move lines de pallets PACK: {len(sj_pack_ml_ids)}")

# 4) Para recepciones: solo tipo 164 + RF2026M (completas), NO las de PACK
sj_pick_ids = list(sj_pick_ids_164 | rf_pick_ids)
print(f"Recepciones San José (tipo 164 + RF2026M): {len(sj_pick_ids)}")

# Obtener TODAS las move lines de esos pickings
print(f"Obteniendo move lines de {len(sj_pick_ids)} recepciones...")
all_ml_sj = []
for i in range(0, len(sj_pick_ids), batch_size):
    batch = sj_pick_ids[i:i+batch_size]
    mls = search_read('stock.move.line', [
        ('picking_id', 'in', batch),
        ('state', '=', 'done'),
    ], ['id', 'product_id', 'qty_done', 'picking_id', 'result_package_id'])
    all_ml_sj.extend(mls)

# Obtener info de pickings
pick_info_sj = {}
for i in range(0, len(sj_pick_ids), batch_size):
    batch = sj_pick_ids[i:i+batch_size]
    picks = search_read('stock.picking', [('id', 'in', batch)], ['id', 'name', 'date_done', 'partner_id'])
    for p in picks:
        pick_info_sj[p['id']] = p

print(f"Total move lines en recepciones San José: {len(all_ml_sj)}")

# Track IDs de move lines ya procesados
processed_ml_ids = set()

for ml in all_ml_sj:
    processed_ml_ids.add(ml['id'])
    product_id = ml['product_id'][0] if ml['product_id'] else 0
    product = ml['product_id'][1] if ml['product_id'] else 'Desconocido'
    # Solo productos categoría MP (San José)
    if product_id not in mp_categ_product_ids:
        continue
    qty = ml['qty_done'] or 0
    if qty <= 0:
        continue
    pick_id = ml['picking_id'][0] if ml['picking_id'] else None
    pick = pick_info_sj.get(pick_id, {})
    date_str = utc_to_chile(pick.get('date_done', ''))
    year = int(date_str[:4]) if len(date_str) >= 4 and date_str[:4].isdigit() else 0
    month = int(date_str[5:7]) if len(date_str) >= 7 and date_str[5:7].isdigit() else 0
    mes_nombre = MESES.get(month, '?')
    partner = pick.get('partner_id', [0, ''])[1] if pick.get('partner_id') else ''
    
    registros.append({
        'mes': f"{year}-{month:02d} {mes_nombre}" if year else '',
        'fecha': date_str,
        'planta': 'San José',
        'tunel': 'Congelados SJ',
        'tipo': 'MP',
        'producto': product,
        'familia': get_familia(product),
        'kg': qty,
        'proveedor': partner,
    })

# 5) Agregar move lines de pallets PACK que NO están en las recepciones ya procesadas
pack_extra_count = 0
# Cargar info de picking para las recepciones PACK que no están en SJ
extra_pack_pick_list = list(sj_pack_pick_ids - set(sj_pick_ids))
if extra_pack_pick_list:
    for i in range(0, len(extra_pack_pick_list), batch_size):
        batch = extra_pack_pick_list[i:i+batch_size]
        picks = search_read('stock.picking', [('id', 'in', batch)], ['id', 'name', 'date_done', 'partner_id'])
        for p in picks:
            pick_info_sj[p['id']] = p

# Re-leer move lines PACK con más campos
sj_pack_mls_full = search_read('stock.move.line', [
    ('result_package_id', 'in', sj_pack_ids),
    ('state', '=', 'done'),
], ['id', 'picking_id', 'product_id', 'qty_done'])

for ml in sj_pack_mls_full:
    if ml['id'] in processed_ml_ids:
        continue  # ya incluida desde recepciones 164/RF2026M
    product_id = ml['product_id'][0] if ml['product_id'] else 0
    product = ml['product_id'][1] if ml['product_id'] else 'Desconocido'
    if product_id not in mp_categ_product_ids:
        continue
    qty = ml['qty_done'] or 0
    if qty <= 0:
        continue
    pick_id = ml['picking_id'][0] if ml['picking_id'] else None
    pick = pick_info_sj.get(pick_id, {})
    date_str = utc_to_chile(pick.get('date_done', ''))
    year = int(date_str[:4]) if len(date_str) >= 4 and date_str[:4].isdigit() else 0
    month = int(date_str[5:7]) if len(date_str) >= 7 and date_str[5:7].isdigit() else 0
    mes_nombre = MESES.get(month, '?')
    partner = pick.get('partner_id', [0, ''])[1] if pick.get('partner_id') else ''
    
    registros.append({
        'mes': f"{year}-{month:02d} {mes_nombre}" if year else '',
        'fecha': date_str,
        'planta': 'San José',
        'tunel': 'Congelados SJ',
        'tipo': 'MP',
        'producto': product,
        'familia': get_familia(product),
        'kg': qty,
        'proveedor': partner,
    })
    pack_extra_count += 1

print(f"Move lines PACK extra (no en recepciones 164/RF2026M): {pack_extra_count}")

total_general = sum(r['kg'] for r in registros)
total_sj = sum(r['kg'] for r in registros if r['planta'] == 'San José')
print(f"Total KG MP San José: {total_sj:,.2f}")

# ==========================================
# GENERAR EXCEL
# ==========================================
print("\nGenerando Excel...")
wb = openpyxl.Workbook()

# --- Estilos ---
AZUL_OSCURO = "1F4E79"
AZUL_MEDIO = "2E75B6"
AZUL_CLARO = "D6E4F0"
GRIS_CLARO = "F2F7FB"
VERDE = "548235"
NARANJA = "ED7D31"

hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
hdr_fill = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type="solid")
hdr_fill2 = PatternFill(start_color=AZUL_MEDIO, end_color=AZUL_MEDIO, fill_type="solid")
data_font = Font(size=10, name="Calibri")
bold_font = Font(bold=True, size=11, name="Calibri")
bold_font_white = Font(bold=True, size=11, name="Calibri", color="FFFFFF")
title_font = Font(bold=True, size=16, name="Calibri", color=AZUL_OSCURO)
subtitle_font = Font(bold=True, size=12, name="Calibri", color=AZUL_MEDIO)
total_fill = PatternFill(start_color=AZUL_CLARO, end_color=AZUL_CLARO, fill_type="solid")
grand_fill = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type="solid")
grand_font = Font(bold=True, size=13, name="Calibri", color="FFFFFF")
alt_fill = PatternFill(start_color=GRIS_CLARO, end_color=GRIS_CLARO, fill_type="solid")
verde_fill = PatternFill(start_color=VERDE, end_color=VERDE, fill_type="solid")
naranja_fill = PatternFill(start_color=NARANJA, end_color=NARANJA, fill_type="solid")
border = Border(
    left=Side(style='thin', color='B4C6E7'), right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'), bottom=Side(style='thin', color='B4C6E7')
)
border_thick = Border(
    left=Side(style='thin', color='B4C6E7'), right=Side(style='thin', color='B4C6E7'),
    top=Side(style='medium', color=AZUL_OSCURO), bottom=Side(style='medium', color=AZUL_OSCURO)
)
num_fmt = '#,##0.00'
pct_fmt = '0.0%'

def style_header_row(ws, row, col_start, col_end, fill=None):
    f = fill or hdr_fill
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font; cell.fill = f
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

def style_data_cell(ws, row, col, value, is_number=False, bold=False, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = bold_font if bold else data_font
    cell.border = border
    if fill:
        cell.fill = fill
    if is_number:
        cell.number_format = num_fmt
        cell.alignment = Alignment(horizontal='right')
    return cell

# ============================================================
#  HOJA 1: RESUMEN EJECUTIVO
# ============================================================
ws1 = wb.active
ws1.title = "Resumen"
ws1.sheet_properties.tabColor = AZUL_OSCURO

# Título
ws1.merge_cells('A1:F1')
c = ws1.cell(row=1, column=1, value="REPORTE DE KILOGRAMOS CONGELADOS")
c.font = title_font; c.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 35

ws1.merge_cells('A2:F2')
c = ws1.cell(row=2, column=1, value="Río Futuro Procesos | Vilkun | San José — Solo Materia Prima (MP)")
c.font = subtitle_font; c.alignment = Alignment(horizontal='center')
ws1.row_dimensions[2].height = 22

# --- Tabla resumen por planta ---
row = 4
ws1.merge_cells(f'A{row}:D{row}')
c = ws1.cell(row=row, column=1, value="RESUMEN POR PROCESO / PLANTA")
c.font = bold_font_white; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center')
for cc in range(2, 5):
    ws1.cell(row=row, column=cc).fill = hdr_fill; ws1.cell(row=row, column=cc).border = border

row += 1
headers_resumen = ["Nº", "Proceso / Planta", "KG Congelados", "% del Total"]
for ci, h in enumerate(headers_resumen, 1):
    c = ws1.cell(row=row, column=ci, value=h)
    c.font = hdr_font; c.fill = hdr_fill2; c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

# Orden lógico de túneles
orden_tuneles = [
    '[1.1] Túnel Estático 1',
    '[1.2] Túnel Estático 2',
    '[1.3] Túnel Estático 3',
    '[1.4] Túnel Continuo',
    '[1.1.1] Túnel Estático VLK',
    'Congelados SJ',
]

# Calcular totales por túnel
totales_tunel = defaultdict(float)
for r in registros:
    totales_tunel[r['tunel']] += r['kg']

row += 1
for idx, tunel in enumerate(orden_tuneles):
    kg = totales_tunel.get(tunel, 0)
    pct = kg / total_general if total_general > 0 else 0
    fill = alt_fill if idx % 2 == 0 else PatternFill(fill_type=None)
    style_data_cell(ws1, row, 1, idx + 1, fill=fill)
    ws1.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    style_data_cell(ws1, row, 2, tunel, fill=fill)
    style_data_cell(ws1, row, 3, round(kg, 2), is_number=True, fill=fill)
    c = style_data_cell(ws1, row, 4, pct, fill=fill)
    c.number_format = pct_fmt; c.alignment = Alignment(horizontal='center')
    row += 1

# Fila TOTAL
for cc in range(1, 5):
    ws1.cell(row=row, column=cc).fill = grand_fill; ws1.cell(row=row, column=cc).border = border_thick
ws1.cell(row=row, column=1, value="").fill = grand_fill
ws1.cell(row=row, column=2, value="TOTAL GENERAL").font = grand_font
ws1.cell(row=row, column=2).fill = grand_fill
c = ws1.cell(row=row, column=3, value=round(total_general, 2))
c.font = grand_font; c.fill = grand_fill; c.number_format = num_fmt; c.alignment = Alignment(horizontal='right')
c = ws1.cell(row=row, column=4, value=1.0)
c.font = grand_font; c.fill = grand_fill; c.number_format = pct_fmt; c.alignment = Alignment(horizontal='center')

# --- Subtotales por grupo ---
row += 2
ws1.merge_cells(f'A{row}:D{row}')
c = ws1.cell(row=row, column=1, value="DESGLOSE POR GRUPO")
c.font = bold_font_white; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center')
for cc in range(2, 5):
    ws1.cell(row=row, column=cc).fill = hdr_fill

row += 1
kg_rf = sum(totales_tunel.get(t, 0) for t in orden_tuneles[:4])  # Est 1,2,3 + Continuo
kg_vlk = totales_tunel.get('[1.1.1] Túnel Estático VLK', 0)
kg_sj = totales_tunel.get('Congelados SJ', 0)

grupos = [
    ("Río Futuro Procesos (Est. 1+2+3 + Continuo)", kg_rf),
    ("Vilkun (Túnel Estático VLK)", kg_vlk),
    ("San José (Congelados SJ)", kg_sj),
]
for idx, (nombre, kg) in enumerate(grupos):
    fill = alt_fill if idx % 2 == 0 else PatternFill(fill_type=None)
    style_data_cell(ws1, row, 1, idx + 1, fill=fill)
    ws1.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    style_data_cell(ws1, row, 2, nombre, fill=fill)
    style_data_cell(ws1, row, 3, round(kg, 2), is_number=True, bold=True, fill=fill)
    pct = kg / total_general if total_general > 0 else 0
    c = style_data_cell(ws1, row, 4, pct, fill=fill)
    c.number_format = pct_fmt; c.alignment = Alignment(horizontal='center')
    row += 1

ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 36
ws1.column_dimensions['C'].width = 22
ws1.column_dimensions['D'].width = 16

# ============================================================
#  HOJA 2: KG POR MES Y PLANTA (Tabla cruzada)
# ============================================================
ws2 = wb.create_sheet("KG por Mes y Planta")
ws2.sheet_properties.tabColor = AZUL_MEDIO

pivot_planta = defaultdict(lambda: defaultdict(float))
for r in registros:
    pivot_planta[r['mes']][r['tunel']] += r['kg']

all_meses = sorted(pivot_planta.keys())

# Título
ws2.merge_cells(f'A1:{openpyxl.utils.get_column_letter(len(orden_tuneles)+2)}1')
c = ws2.cell(row=1, column=1, value="KG CONGELADOS POR MES Y PROCESO")
c.font = title_font; c.alignment = Alignment(horizontal='center')
ws2.row_dimensions[1].height = 30

headers3 = ["Mes"] + orden_tuneles + ["TOTAL MES"]
row = 3
style_header_row(ws2, row, 1, len(headers3))
for ci, h in enumerate(headers3, 1):
    ws2.cell(row=row, column=ci, value=h)

row = 4
for idx, mes in enumerate(all_meses):
    fill = alt_fill if idx % 2 == 0 else PatternFill(fill_type=None)
    style_data_cell(ws2, row, 1, mes, fill=fill)
    total_mes = 0
    for ci, tunel in enumerate(orden_tuneles, 2):
        val = pivot_planta[mes].get(tunel, 0)
        style_data_cell(ws2, row, ci, round(val, 2) if val > 0 else "", is_number=val > 0, fill=fill)
        total_mes += val
    style_data_cell(ws2, row, len(headers3), round(total_mes, 2), is_number=True, bold=True, fill=fill)
    row += 1

# Fila TOTAL
for cc in range(1, len(headers3) + 1):
    ws2.cell(row=row, column=cc).fill = grand_fill; ws2.cell(row=row, column=cc).border = border_thick
c = ws2.cell(row=row, column=1, value="TOTAL"); c.font = grand_font; c.fill = grand_fill
for ci, tunel in enumerate(orden_tuneles, 2):
    total_t = sum(pivot_planta[m].get(tunel, 0) for m in all_meses)
    c = ws2.cell(row=row, column=ci, value=round(total_t, 2))
    c.font = grand_font; c.fill = grand_fill; c.number_format = num_fmt; c.alignment = Alignment(horizontal='right')
c = ws2.cell(row=row, column=len(headers3), value=round(total_general, 2))
c.font = grand_font; c.fill = grand_fill; c.number_format = num_fmt; c.alignment = Alignment(horizontal='right')

ws2.column_dimensions['A'].width = 22
for ci in range(2, len(headers3) + 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 22
ws2.freeze_panes = 'B4'

# ============================================================
#  HOJA 3: KG POR MES Y TIPO DE FRUTA
# ============================================================
ws3 = wb.create_sheet("KG por Mes y Fruta")
ws3.sheet_properties.tabColor = VERDE

pivot_fruta = defaultdict(lambda: defaultdict(float))
for r in registros:
    pivot_fruta[r['mes']][r['familia']] += r['kg']

all_frutas_set = set()
for m_data in pivot_fruta.values():
    all_frutas_set.update(m_data.keys())
all_frutas_sorted = sorted(all_frutas_set)

# Título
ws3.merge_cells(f'A1:{openpyxl.utils.get_column_letter(len(all_meses)+2)}1')
c = ws3.cell(row=1, column=1, value="KG CONGELADOS POR MES Y TIPO DE FRUTA")
c.font = title_font; c.alignment = Alignment(horizontal='center')
ws3.row_dimensions[1].height = 30

headers_fruta = ["Familia de Fruta"] + all_meses + ["TOTAL"]
row = 3
style_header_row(ws3, row, 1, len(headers_fruta))
for ci, h in enumerate(headers_fruta, 1):
    ws3.cell(row=row, column=ci, value=h)

row = 4
for idx, fruta in enumerate(all_frutas_sorted):
    fill = alt_fill if idx % 2 == 0 else PatternFill(fill_type=None)
    style_data_cell(ws3, row, 1, fruta, fill=fill)
    total_fruta = 0
    for ci, mes in enumerate(all_meses, 2):
        val = pivot_fruta[mes].get(fruta, 0)
        style_data_cell(ws3, row, ci, round(val, 2) if val > 0 else "", is_number=val > 0, fill=fill)
        total_fruta += val
    style_data_cell(ws3, row, len(headers_fruta), round(total_fruta, 2), is_number=True, bold=True, fill=fill)
    row += 1

# Fila TOTAL
for cc in range(1, len(headers_fruta) + 1):
    ws3.cell(row=row, column=cc).fill = grand_fill; ws3.cell(row=row, column=cc).border = border_thick
c = ws3.cell(row=row, column=1, value="TOTAL"); c.font = grand_font; c.fill = grand_fill
for ci, mes in enumerate(all_meses, 2):
    total_mes = sum(pivot_fruta[mes].values())
    c = ws3.cell(row=row, column=ci, value=round(total_mes, 2))
    c.font = grand_font; c.fill = grand_fill; c.number_format = num_fmt; c.alignment = Alignment(horizontal='right')
c = ws3.cell(row=row, column=len(headers_fruta), value=round(total_general, 2))
c.font = grand_font; c.fill = grand_fill; c.number_format = num_fmt; c.alignment = Alignment(horizontal='right')

ws3.column_dimensions['A'].width = 18
for ci in range(2, len(headers_fruta) + 1):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 18
ws3.freeze_panes = 'B4'

# ============================================================
#  HOJA 4: KG POR FRUTA Y PLANTA (qué fruta se congela dónde)
# ============================================================
ws4 = wb.create_sheet("KG por Fruta y Planta")
ws4.sheet_properties.tabColor = NARANJA

pivot_fp = defaultdict(lambda: defaultdict(float))
for r in registros:
    pivot_fp[r['familia']][r['tunel']] += r['kg']

ws4.merge_cells(f'A1:{openpyxl.utils.get_column_letter(len(orden_tuneles)+2)}1')
c = ws4.cell(row=1, column=1, value="KG CONGELADOS POR FAMILIA DE FRUTA Y PROCESO")
c.font = title_font; c.alignment = Alignment(horizontal='center')
ws4.row_dimensions[1].height = 30

headers_fp = ["Familia de Fruta"] + orden_tuneles + ["TOTAL"]
row = 3
style_header_row(ws4, row, 1, len(headers_fp))
for ci, h in enumerate(headers_fp, 1):
    ws4.cell(row=row, column=ci, value=h)

row = 4
frutas_sorted = sorted(pivot_fp.keys())
for idx, fruta in enumerate(frutas_sorted):
    fill = alt_fill if idx % 2 == 0 else PatternFill(fill_type=None)
    style_data_cell(ws4, row, 1, fruta, fill=fill)
    total_fruta = 0
    for ci, tunel in enumerate(orden_tuneles, 2):
        val = pivot_fp[fruta].get(tunel, 0)
        style_data_cell(ws4, row, ci, round(val, 2) if val > 0 else "", is_number=val > 0, fill=fill)
        total_fruta += val
    style_data_cell(ws4, row, len(headers_fp), round(total_fruta, 2), is_number=True, bold=True, fill=fill)
    row += 1

# Fila TOTAL
for cc in range(1, len(headers_fp) + 1):
    ws4.cell(row=row, column=cc).fill = grand_fill; ws4.cell(row=row, column=cc).border = border_thick
c = ws4.cell(row=row, column=1, value="TOTAL"); c.font = grand_font; c.fill = grand_fill
for ci, tunel in enumerate(orden_tuneles, 2):
    total_t = sum(pivot_fp[f].get(tunel, 0) for f in frutas_sorted)
    c = ws4.cell(row=row, column=ci, value=round(total_t, 2))
    c.font = grand_font; c.fill = grand_fill; c.number_format = num_fmt; c.alignment = Alignment(horizontal='right')
c = ws4.cell(row=row, column=len(headers_fp), value=round(total_general, 2))
c.font = grand_font; c.fill = grand_fill; c.number_format = num_fmt; c.alignment = Alignment(horizontal='right')

ws4.column_dimensions['A'].width = 18
for ci in range(2, len(headers_fp) + 1):
    ws4.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 22
ws4.freeze_panes = 'B4'

# ============================================================
#  HOJA 5: DATOS DETALLADOS (para filtros)
# ============================================================
ws5 = wb.create_sheet("Detalle")
ws5.sheet_properties.tabColor = "808080"

ws5.merge_cells('A1:G1')
c = ws5.cell(row=1, column=1, value="DETALLE DE TODOS LOS REGISTROS (usar filtros para explorar)")
c.font = subtitle_font; c.alignment = Alignment(horizontal='center')
ws5.row_dimensions[1].height = 25

headers_det = ["Mes", "Fecha", "Planta", "Proceso / Origen", "Familia Fruta", "Producto", "KG"]
row = 3
style_header_row(ws5, row, 1, len(headers_det))
for ci, h in enumerate(headers_det, 1):
    ws5.cell(row=row, column=ci, value=h)

registros_sorted = sorted(registros, key=lambda x: (x['tunel'], x['mes'], x['producto']))
row = 4
for idx, r in enumerate(registros_sorted):
    fill = alt_fill if idx % 2 == 0 else PatternFill(fill_type=None)
    style_data_cell(ws5, row, 1, r['mes'], fill=fill)
    style_data_cell(ws5, row, 2, r['fecha'], fill=fill)
    style_data_cell(ws5, row, 3, r['planta'], fill=fill)
    style_data_cell(ws5, row, 4, r['tunel'], fill=fill)
    style_data_cell(ws5, row, 5, r['familia'], fill=fill)
    style_data_cell(ws5, row, 6, r['producto'], fill=fill)
    style_data_cell(ws5, row, 7, round(r['kg'], 2), is_number=True, fill=fill)
    row += 1

# Fila total
for cc in range(1, 8):
    ws5.cell(row=row, column=cc).fill = grand_fill; ws5.cell(row=row, column=cc).border = border_thick
c = ws5.cell(row=row, column=1, value="TOTAL"); c.font = grand_font; c.fill = grand_fill
c = ws5.cell(row=row, column=7, value=round(total_general, 2))
c.font = grand_font; c.fill = grand_fill; c.number_format = num_fmt; c.alignment = Alignment(horizontal='right')

ws5.column_dimensions['A'].width = 22; ws5.column_dimensions['B'].width = 14; ws5.column_dimensions['C'].width = 18
ws5.column_dimensions['D'].width = 30; ws5.column_dimensions['E'].width = 16; ws5.column_dimensions['F'].width = 40
ws5.column_dimensions['G'].width = 16
ws5.freeze_panes = 'A4'
ws5.auto_filter.ref = f"A3:G{row - 1}"

# Guardar
output_file = "KG_Congelados_Todas_Plantas.xlsx"
wb.save(output_file)

print(f"\n{'='*65}")
print(f"  Excel generado: {output_file}")
print(f"  Registros totales: {len(registros)}")
print(f"  Hojas:")
print(f"    1. Resumen         — Totales por proceso y grupo")
print(f"    2. KG por Mes y Planta — Tabla cruzada mes × proceso")
print(f"    3. KG por Mes y Fruta  — Tabla cruzada mes × fruta")
print(f"    4. KG por Fruta y Planta — Qué fruta va a cada proceso")
print(f"    5. Detalle         — Todos los registros con filtros")
print()
for t in orden_tuneles:
    kg = totales_tunel.get(t, 0)
    pct = (kg / total_general * 100) if total_general > 0 else 0
    print(f"  {t}: {kg:>14,.2f} kg  ({pct:.1f}%)")
print(f"  {'─'*55}")
print(f"  TOTAL GENERAL: {total_general:>14,.2f} kg")
print(f"{'='*65}")
