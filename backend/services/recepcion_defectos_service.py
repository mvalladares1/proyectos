"""
Servicio para generar reporte Excel de recepciones con detalles de defectos.
Basado en el script generar_reporte_recepciones_excel.py
"""
from shared.odoo_client import OdooClient
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from typing import List, Optional


def generar_reporte_defectos_excel(
    username: str,
    password: str,
    fecha_inicio: str,
    fecha_fin: str,
    origenes: Optional[List[str]] = None,
    solo_hechas: bool = True
) -> bytes:
    """
    Genera un reporte Excel detallado de recepciones con información de defectos.
    
    Args:
        username: Usuario de Odoo
        password: Password de Odoo
        fecha_inicio: Fecha inicio en formato YYYY-MM-DD
        fecha_fin: Fecha fin en formato YYYY-MM-DD
        origenes: Lista de orígenes ['RFP', 'VILKUN', 'SAN JOSE']. Si None, incluye todos.
        solo_hechas: Si True, solo recepciones en estado 'done'
        
    Returns:
        bytes del archivo Excel generado
    """
    
    # Conectar a Odoo
    odoo = OdooClient(username=username, password=password)
    
    # Analizar campos disponibles
    campos_picking = odoo.execute('stock.picking', 'fields_get', [], {'attributes': ['string', 'type']})
    campos_template = odoo.execute('product.template', 'fields_get', [], {'attributes': ['string', 'type']})
    campos_quality = odoo.execute('quality.check', 'fields_get', [], {'attributes': ['string', 'type']})
    
    # Determinar campos a usar
    campo_guia = _determinar_campo(campos_picking, [
        'x_studio_gua_de_despacho', 'x_studio_gua_despacho', 'x_studio_guia_despacho'
    ])
    
    campo_categoria_prod = _determinar_campo(campos_picking, [
        'x_studio_categora_de_producto', 'x_studio_categoria_de_producto'
    ])
    
    campos_template_usar = {
        'variedad': _determinar_campo(campos_template, ['x_studio_categora_variedad', 'x_studio_variedad']),
        'manejo': _determinar_campo(campos_template, ['x_studio_categora_tipo_de_manejo', 'x_studio_manejo_del_producto']),
        'categoria': 'categ_id',  # Campo estándar de Odoo para categoría de producto
    }
    
    campos_quality_usar = {
        'tipo_fruta_qc': _determinar_campo(campos_quality, ['x_studio_tipo_de_fruta']),
        'pallet': _determinar_campo(campos_quality, ['x_studio_n_de_palet_o_paquete', 'x_studio_n_palet']),
        'clasificacion': _determinar_campo(campos_quality, ['x_studio_calific_final', 'x_studio_calificacin_final']),
        'total_defectos': _determinar_campo(campos_quality, ['x_studio_totdefectoarandano', 'x_studio_total_def_calidad', 'x_studio_total_def_calidad_1', 'x_studio_total_de_defectos_']),
        'temperatura': _determinar_campo(campos_quality, ['x_studio_temperatura']),
        'pct_iqf': _determinar_campo(campos_quality, ['x_studio_total_iqf_', 'x_studio_total_iqf']),
        'pct_block': _determinar_campo(campos_quality, ['x_studio_total_block_', 'x_studio_total_block']),
        'hongos': _determinar_campo(campos_quality, ['x_studio_tothongos', 'x_studio_hongos', 'x_studio_hongos_1']),
        'inmadura': _determinar_campo(campos_quality, ['x_studio_totinmadura', 'x_studio_inmadura', 'x_studio_inmadura_1']),
        'sobremadura': _determinar_campo(campos_quality, ['x_studio_totsobremadurez', 'x_studio_sobremadura', 'x_studio_sobremadurez_1', 'x_studio_sobre_madura']),
        'deshidratado': _determinar_campo(campos_quality, ['x_studio_totdeshidratado', 'x_studio_deshidratado', 'x_studio_deshidratado_1']),
        'crumble': _determinar_campo(campos_quality, ['x_studio_crumble', 'x_studio_crumble_1']),
        'dano_mecanico': _determinar_campo(campos_quality, ['x_studio_totdaomecanico', 'x_studio_dao_mecanico', 'x_studio_dano_mecanico']),
        'dano_insecto': _determinar_campo(campos_quality, ['x_studio_totpresenciainsecto', 'x_studio_presencia_de_insectos', 'x_studio_totdaoinsecto', 'x_studio_presencia_de_insectos_1']),
        'deformes': _determinar_campo(campos_quality, ['x_studio_totfrutosdeformes', 'x_studio_frutos_deformes', 'x_studio_deformes', 'x_studio_frutos_deformes_1']),
        'fruta_verde': _determinar_campo(campos_quality, ['x_studio_totfrutaverde', 'x_studio_fruta_verde', 'x_studio_fruta_verde_1']),
        'herida_partida': _determinar_campo(campos_quality, ['x_studio_totheridapartidamolida', 'x_studio_heridapartidamolida', 'x_studio_heridapartiduramolida', 'x_studio_heridapartidamolida_1']),
        'materias_extranas': _determinar_campo(campos_quality, ['x_studio_totmateriaextraa', 'x_studio_materias_extraas', 'x_studio_materias_extranas', 'x_studio_materias_extraas_1']),
    }
    
    # Mapeo de campos en líneas de calidad (one2many)
    campos_linea_defectos = {
        'total_defectos': ['x_studio_total_defectos'],
        'hongos': ['x_studio_pudricin_hongos', 'x_studio_hongos'],
        'inmadura': ['x_studio_inmadura', 'x_studio_frutos_inmaduros'],
        'sobremadura': ['x_studio_sobremadura', 'x_studio_frutos_sobre_maduros'],
        'dano_insecto': ['x_studio_dao_por_insecto', 'x_studio_daos_por_insectos'],
        'dano_mecanico': ['x_studio_dao_mecanico'],
        'golpe_sol': ['x_studio_golpe_de_sol'],
        'pallet': ['x_studio_n_palet'],
        'temperatura': ['x_studio_temperatura'],
    }
    
    # Modelos de líneas de calidad
    modelos_lineas_qc = {
        'x_studio_one2many_field_3jSXq': 'x_quality_check_line_2a594',
        'x_studio_one2many_field_eNeCg': 'x_quality_check_line_0d011',
        'x_studio_one2many_field_ipdDS': 'x_quality_check_line_35406',
        'x_studio_one2many_field_mZmK2': 'x_quality_check_line_46726',
        'x_studio_one2many_field_nsxt0': 'x_quality_check_line_17bfb',
        'x_studio_one2many_field_RdQtm': 'x_quality_check_line_2efd1',
        'x_studio_one2many_field_rgA7I': 'x_quality_check_line_1d183',
        'x_studio_one2many_field_vloaS': 'x_quality_check_line_f0f7b',
    }
    
    # Mapeo de picking types
    ORIGEN_PICKING_MAP = {
        "RFP": 1,
        "VILKUN": 217,
        "SAN JOSE": 164
    }
    
    # Si no se especifican orígenes, usar todos
    if not origenes:
        origenes = list(ORIGEN_PICKING_MAP.keys())
    
    # Obtener recepciones
    todas_recepciones = []
    for origen in origenes:
        if origen not in ORIGEN_PICKING_MAP:
            continue
            
        picking_type_id = ORIGEN_PICKING_MAP[origen]
        
        domain = [
            ('picking_type_id', '=', picking_type_id),
            ('scheduled_date', '>=', fecha_inicio),
            ('scheduled_date', '<=', fecha_fin),
        ]
        
        if solo_hechas:
            domain.append(('state', '=', 'done'))
        
        if campo_categoria_prod:
            domain.append((campo_categoria_prod, '=', 'MP'))
        
        campos_leer = ['id', 'name', 'scheduled_date', 'partner_id']
        if campo_guia:
            campos_leer.append(campo_guia)
        
        recepciones = odoo.search_read('stock.picking', domain, campos_leer, limit=5000)
        
        for rec in recepciones:
            rec['origen'] = origen
        
        todas_recepciones.extend(recepciones)
    
    if not todas_recepciones:
        # Generar Excel vacío
        return _generar_excel_vacio()
    
    # Obtener movimientos
    picking_ids = [r['id'] for r in todas_recepciones]
    movimientos = odoo.search_read(
        'stock.move',
        [('picking_id', 'in', picking_ids), ('state', '=', 'done')],
        ['id', 'product_id', 'quantity_done', 'picking_id'],
        limit=50000
    )
    
    # Obtener productos y templates
    product_ids = list(set([m['product_id'][0] for m in movimientos if m.get('product_id')]))
    productos = odoo.search_read(
        'product.product',
        [('id', 'in', product_ids)],
        ['id', 'name', 'product_tmpl_id'],
        limit=10000
    )
    productos_map = {p['id']: p for p in productos}
    
    template_ids = list(set([p['product_tmpl_id'][0] for p in productos if p.get('product_tmpl_id')]))
    template_campos_leer = ['id']
    for campo in campos_template_usar.values():
        if campo:
            template_campos_leer.append(campo)
    
    templates = odoo.search_read(
        'product.template',
        [('id', 'in', template_ids)],
        template_campos_leer,
        limit=10000
    )
    templates_map = {t['id']: t for t in templates}
    
    # Obtener quality checks
    quality_campos_leer = ['id', 'picking_id', 'create_date']
    for campo in campos_quality_usar.values():
        if campo:
            quality_campos_leer.append(campo)
    
    quality_checks = odoo.search_read(
        'quality.check',
        [('picking_id', 'in', picking_ids)],
        quality_campos_leer,
        limit=50000
    )
    
    # Organizar quality checks por picking
    quality_by_picking = {}
    for qc in quality_checks:
        picking_id = qc.get('picking_id')
        if picking_id:
            picking_id = picking_id[0] if isinstance(picking_id, (list, tuple)) else picking_id
            if picking_id not in quality_by_picking:
                quality_by_picking[picking_id] = []
            quality_by_picking[picking_id].append(qc)
    
    # Construir datos para Excel
    datos_excel = []
    
    for recepcion in todas_recepciones:
        picking_id = recepcion['id']
        albaran = recepcion.get('name', '')
        fecha = recepcion.get('scheduled_date', '')
        if fecha:
            fecha = datetime.strptime(fecha, '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y')
        
        proveedor = recepcion.get('partner_id')
        proveedor = proveedor[1] if isinstance(proveedor, (list, tuple)) else str(proveedor) if proveedor else ''
        
        guia_despacho = recepcion.get(campo_guia, '') if campo_guia else ''
        origen = recepcion.get('origen', '')
        
        movs_recepcion = [m for m in movimientos if m.get('picking_id') and m['picking_id'][0] == picking_id]
        qcs_recepcion = quality_by_picking.get(picking_id, [])
        
        # Si NO hay quality checks, saltar esta recepción (es IQF/BLOCK sin control de calidad)
        if not qcs_recepcion:
            continue
            
        # Para cada quality check (pallet)
        for qc in qcs_recepcion:
            n_pallet = _get_field(qc, campos_quality_usar.get('pallet'))
            calificacion = _get_field(qc, campos_quality_usar.get('clasificacion'))
            temperatura = _get_field(qc, campos_quality_usar.get('temperatura'), 0)
            tipo_fruta_from_qc = _get_field(qc, campos_quality_usar.get('tipo_fruta_qc'))
            pct_iqf = _get_field(qc, campos_quality_usar.get('pct_iqf'), 0)
            pct_block = _get_field(qc, campos_quality_usar.get('pct_block'), 0)
            
            # Intentar leer defectos de las líneas (one2many) primero
            lineas_defectos = _leer_lineas_calidad(odoo, qc, modelos_lineas_qc, campos_linea_defectos)
            
            # Si hay líneas, usar los defectos de las líneas; si no, usar los del QC principal
            if lineas_defectos:
                # Procesar cada línea como un registro separado
                for linea in lineas_defectos:
                    # Defectos de la línea
                    hongos = linea.get('hongos', 0)
                    inmadura = linea.get('inmadura', 0)
                    sobremadura = linea.get('sobremadura', 0)
                    dano_mecanico = linea.get('dano_mecanico', 0)
                    dano_insecto = linea.get('dano_insecto', 0)
                    total_defectos_gramos = linea.get('total_defectos', 0)
                    
                    # Sobrescribir con datos de línea si existen
                    if linea.get('pallet'):
                        n_pallet = linea.get('pallet')
                    if linea.get('temperatura'):
                        temperatura = linea.get('temperatura')
                    
                    # Valores fijos de la línea (no aplican estos en líneas actualmente)
                    deshidratado = linea.get('deshidratado', 0)
                    crumble = 0
                    deformes = 0
                    fruta_verde = 0
                    herida_partida = 0
                    materias_extranas = 0
                    golpe_sol = linea.get('golpe_sol', 0)
                    
                    # Calcular % defectos
                    total_defectos_pct = 0
                    if total_defectos_gramos > 0:
                        total_defectos_pct = round((total_defectos_gramos / 1000) * 100, 2)
                    
                    # Agregar registro con defectos de la línea
                    _agregar_registros_excel(
                        datos_excel, movs_recepcion, productos_map, templates_map,
                        albaran, fecha, proveedor, guia_despacho, origen, n_pallet,
                        calificacion, temperatura, pct_iqf, pct_block, tipo_fruta_from_qc,
                        total_defectos_gramos, total_defectos_pct,
                        hongos, inmadura, sobremadura, deshidratado, crumble, dano_mecanico,
                        dano_insecto, deformes, fruta_verde, herida_partida, materias_extranas,
                        campos_template_usar
                    )
            else:
                # No hay líneas, usar defectos del QC principal (comportamiento antiguo)
                total_defectos_gramos = _get_field(qc, campos_quality_usar.get('total_defectos'), 0)
                
                # Defectos en gramos del QC principal
                hongos = _get_field(qc, campos_quality_usar.get('hongos'), 0)
                inmadura = _get_field(qc, campos_quality_usar.get('inmadura'), 0)
                sobremadura = _get_field(qc, campos_quality_usar.get('sobremadura'), 0)
                deshidratado = _get_field(qc, campos_quality_usar.get('deshidratado'), 0)
                crumble = _get_field(qc, campos_quality_usar.get('crumble'), 0)
                dano_mecanico = _get_field(qc, campos_quality_usar.get('dano_mecanico'), 0)
                dano_insecto = _get_field(qc, campos_quality_usar.get('dano_insecto'), 0)
                deformes = _get_field(qc, campos_quality_usar.get('deformes'), 0)
                fruta_verde = _get_field(qc, campos_quality_usar.get('fruta_verde'), 0)
                herida_partida = _get_field(qc, campos_quality_usar.get('herida_partida'), 0)
                materias_extranas = _get_field(qc, campos_quality_usar.get('materias_extranas'), 0)
                golpe_sol = 0
                
                # Calcular % defectos
                total_defectos_pct = 0
                if total_defectos_gramos > 0:
                    total_defectos_pct = round((total_defectos_gramos / 1000) * 100, 2)
                
                # Agregar registros con defectos del QC principal
                _agregar_registros_excel(
                    datos_excel, movs_recepcion, productos_map, templates_map,
                    albaran, fecha, proveedor, guia_despacho, origen, n_pallet,
                    calificacion, temperatura, pct_iqf, pct_block, tipo_fruta_from_qc,
                    total_defectos_gramos, total_defectos_pct,
                    hongos, inmadura, sobremadura, deshidratado, crumble, dano_mecanico,
                    dano_insecto, deformes, fruta_verde, herida_partida, materias_extranas,
                    campos_template_usar
                )


def _agregar_registros_excel(
    datos_excel, movs_recepcion, productos_map, templates_map,
    albaran, fecha, proveedor, guia_despacho, origen, n_pallet,
    calificacion, temperatura, pct_iqf, pct_block, tipo_fruta_from_qc,
    total_defectos_gramos, total_defectos_pct,
    hongos, inmadura, sobremadura, deshidratado, crumble, dano_mecanico,
    dano_insecto, deformes, fruta_verde, herida_partida, materias_extranas,
    campos_template_usar
):
    """Función auxiliar para agregar registros al Excel (evitar duplicación de código)."""
    for mov in movs_recepcion:
        qty = mov.get('quantity_done', 0) or 0
        if qty == 0:
            continue
        
        prod_id = mov.get('product_id')
        if not prod_id:
            continue
        prod_id = prod_id[0] if isinstance(prod_id, (list, tuple)) else prod_id
        
        prod_info = productos_map.get(prod_id, {})
        producto_nombre = prod_info.get('name', '')
        
        tmpl_id = prod_info.get('product_tmpl_id')
        tmpl_id = tmpl_id[0] if isinstance(tmpl_id, (list, tuple)) else tmpl_id
        tmpl_info = templates_map.get(tmpl_id, {})
        
        # FILTRO: Excluir productos de categoría BANDEJAS
        categoria = tmpl_info.get('categ_id', '')
        categoria_nombre = categoria[1] if isinstance(categoria, (list, tuple)) and len(categoria) > 1 else ''
        if 'BANDEJA' in categoria_nombre.upper():
            continue
        
        variedad = _extract_many2one(tmpl_info.get(campos_template_usar.get('variedad'), ''))
        manejo = _extract_many2one(tmpl_info.get(campos_template_usar.get('manejo'), ''))
        
        # Tipo de fruta: priorizar desde quality.check
        tipo_fruta = tipo_fruta_from_qc or ''
        
        datos_excel.append({
            'Fecha': fecha,
            'Proveedor': proveedor,
            'Guía Despacho': guia_despacho,
            'Origen': origen,
            'Albarán': albaran,
            'Producto': producto_nombre,
            'Variedad': variedad,
            'Tipo Fruta': tipo_fruta,
            'Manejo': manejo,
            'Kg': qty,
            'N° Pallet': n_pallet,
            'Calificación': calificacion,
            '% IQF': pct_iqf,
            '% BLOCK': pct_block,
            'Temperatura °C': temperatura,
            '% Defectos': total_defectos_pct,
            'Hongos (g)': hongos,
            'Inmadura (g)': inmadura,
            'Sobremadura (g)': sobremadura,
            'Deshidratado (g)': deshidratado,
            'Crumble (g)': crumble,
            'Daño Mecánico (g)': dano_mecanico,
            'Daño Insecto (g)': dano_insecto,
            'Deformes (g)': deformes,
            'Fruta Verde (g)': fruta_verde,
            'Herida/Partida (g)': herida_partida,
            'Materias Extrañas (g)': materias_extranas,
        })
    

def _generar_excel_desde_datos(datos_excel, fecha_desde, fecha_hasta):
    """Genera un archivo Excel desde los datos procesados."""
    # Crear DataFrame
    df = pd.DataFrame(datos_excel)
    
    if df.empty:
        df = pd.DataFrame(columns=[
            'Fecha', 'Proveedor', 'Guía Despacho', 'Origen', 'Albarán', 
            'Producto', 'Variedad', 'Tipo Fruta', 'Manejo', 'Kg', 'N° Pallet',
            'Calificación', '% IQF', '% BLOCK', 'Temperatura °C', '% Defectos',
            'Hongos (g)', 'Inmadura (g)', 'Sobremadura (g)', 'Deshidratado (g)',
            'Crumble (g)', 'Daño Mecánico (g)', 'Daño Insecto (g)', 'Deformes (g)',
            'Fruta Verde (g)', 'Herida/Partida (g)', 'Materias Extrañas (g)'
        ])
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Recepciones Defectos"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = df.columns.tolist()
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Datos
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            
            if col_num > 10:  # Columnas numéricas
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
    
    # Ajustar anchos
    for col_num, header in enumerate(headers, 1):
        column_letter = ws.cell(row=1, column=col_num).column_letter
        
        if 'Fecha' in header:
            ws.column_dimensions[column_letter].width = 12
        elif 'Proveedor' in header or 'Producto' in header:
            ws.column_dimensions[column_letter].width = 30
        elif '(g)' in header or '°C' in header or '%' in header:
            ws.column_dimensions[column_letter].width = 12
        else:
            ws.column_dimensions[column_letter].width = 15
    
    ws.freeze_panes = "A2"
    
    # Guardar a bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


def _determinar_campo(campos_dict, opciones):
    """Determina qué campo usar de una lista de opciones."""
    for opcion in opciones:
        if opcion in campos_dict:
            return opcion
    return None


def _get_field(data_dict, field_name, default=''):
    """Obtiene un campo del diccionario de forma segura."""
    if not field_name:
        return default
    value = data_dict.get(field_name, default)
    # No usar evaluación booleana para permitir 0, False, etc.
    if value is None or value == '':
        return default
    return value


def _leer_lineas_calidad(odoo, qc_data, modelos_lineas_qc, campos_linea_defectos):
    """
    Lee las líneas de calidad (one2many) que contienen los defectos reales.
    Retorna lista de líneas con sus defectos.
    """
    lineas_defectos = []
    
    # Buscar campos one2many con registros
    for campo_one2many, modelo_linea in modelos_lineas_qc.items():
        ids_lineas = qc_data.get(campo_one2many, [])
        if not ids_lineas or not isinstance(ids_lineas, list) or len(ids_lineas) == 0:
            continue
        
        # Leer las líneas del modelo
        try:
            lineas = odoo.execute(
                modelo_linea, 'read',
                ids_lineas,
                list(set([campo for campos in campos_linea_defectos.values() for campo in campos]))
            )
            
            for linea in lineas:
                # Extraer defectos de la línea usando el mapeo
                defectos_linea = {}
                for key, posibles_campos in campos_linea_defectos.items():
                    campo_encontrado = _determinar_campo({k: True for k in linea.keys()}, posibles_campos)
                    if campo_encontrado:
                        defectos_linea[key] = _get_field(linea, campo_encontrado, 0)
                
                lineas_defectos.append(defectos_linea)
        except Exception as e:
            # Si falla la lectura de un modelo, continuar con el siguiente
            continue
    
    return lineas_defectos


def _extract_many2one(value):
    """Extrae el valor de un campo many2one."""
    if isinstance(value, (list, tuple)) and len(value) > 1:
        return value[1]
    elif value:
        return str(value)
    return ''


def _generar_excel_vacio():
    """Genera un Excel vacío cuando no hay datos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sin Datos"
    ws['A1'] = "No se encontraron recepciones con los filtros especificados"
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
