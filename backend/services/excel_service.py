from io import BytesIO
from typing import List, Dict, Any, Optional
from datetime import datetime, date, timedelta
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter

from backend.services.recepcion_service import get_recepciones_mp


# --- Funciones de formateo chileno ---
def fmt_fecha(fecha_str):
    """Convierte fecha ISO a formato DD/MM/AAAA"""
    if not fecha_str:
        return ""
    try:
        if isinstance(fecha_str, (date, datetime)):
            return fecha_str.strftime("%d/%m/%Y")
        if isinstance(fecha_str, str):
            if " " in fecha_str:
                fecha_str = fecha_str.split(" ")[0]
            elif "T" in fecha_str:
                fecha_str = fecha_str.split("T")[0]
            dt = datetime.strptime(fecha_str, "%Y-%m-%d")
            return dt.strftime("%d/%m/%Y")
    except:
        pass
    return str(fecha_str)

def fmt_numero(valor, decimales=0):
    """Formatea número con punto como miles y coma como decimal"""
    if valor is None:
        return "0"
    try:
        if decimales > 0:
            formatted = f"{valor:,.{decimales}f}"
        else:
            formatted = f"{valor:,.0f}"
        formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
        return formatted
    except:
        return str(valor)

def fmt_dinero(valor, decimales=0):
    """Formatea valor monetario con símbolo $"""
    return f"${fmt_numero(valor, decimales)}"


MAX_DAYS_FETCH = 90


def _auto_fit_columns(ws):
    """Adjust column widths based on max length in each column."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ""
            except Exception:
                v = ""
            if len(v) > max_length:
                max_length = len(v)
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width


def generate_recepciones_excel(username: str, password: str, fecha_inicio: str, fecha_fin: str,
                               include_prev_week: bool = False, include_month_accum: bool = False,
                               solo_hechas: bool = True,
                               filter_tipo_fruta: List[str] = None,
                               filter_clasificacion: List[str] = None,
                               filter_manejo: List[str] = None,
                               filter_productor: List[str] = None) -> bytes:
    """Genera un Excel con detalle de recepciones y productos desglosados.

    Retorna bytes del archivo .xlsx listo para enviar por StreamingResponse.
    """
    # parse dates
    f_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    f_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()

    prev_end = f_inicio - timedelta(days=1)
    delta = (f_fin - f_inicio) + timedelta(days=1)
    prev_start = prev_end - (delta - timedelta(days=1))
    month_start = date(f_fin.year, f_fin.month, 1)
    fetch_start = min(month_start, prev_start)

    total_days = (f_fin - fetch_start).days + 1
    if total_days > MAX_DAYS_FETCH:
        raise Exception(f"Rango demasiado grande: {total_days} días. Límite = {MAX_DAYS_FETCH} días.")

    recepciones_all = get_recepciones_mp(username, password, fetch_start.isoformat(), f_fin.isoformat(), solo_hechas=solo_hechas)

    # Helper to filter by date portion
    def _in_range(r, start_date: date, end_date: date) -> bool:
        try:
            rd = datetime.fromisoformat(r.get('fecha')).date()
        except Exception:
            try:
                rd = datetime.strptime(r.get('fecha', '')[:10], '%Y-%m-%d').date()
            except Exception:
                return False
        return start_date <= rd <= end_date

    # Helper para verificar manejo en lista de productos (para filtado nivel recepción si aplica)
    # Pero los filtros de manejo y tipo fruta aplican más estrictamente a nivel de fila de producto
    
    # 1. Filtrar RECEPCIONES por Productor y Clasificación
    recepciones_main = []
    for r in recepciones_all:
        if not _in_range(r, f_inicio, f_fin):
            continue
            
        # Filtro Productor
        if filter_productor and r.get('productor') not in filter_productor:
            continue
            
        # Filtro Clasificación
        if filter_clasificacion and r.get('calific_final') not in filter_clasificacion:
            continue
            
        # Si hay filtros de producto, verificamos si la recepción contiene AL MENOS UN producto válido
        # para decidir si incluirla (aunque el filtrado fino se hace al escribir filas)
        if filter_tipo_fruta or filter_manejo:
            has_valid_prod = False
            for p in r.get('productos', []) or []:
                tf = (p.get('TipoFruta') or r.get('tipo_fruta') or '').strip()
                if filter_tipo_fruta and tf not in filter_tipo_fruta:
                    continue
                
                man = (p.get('Manejo') or '').strip()
                if filter_manejo and man not in filter_manejo:
                    continue
                
                has_valid_prod = True
                break
            
            if not has_valid_prod:
                continue
        
        recepciones_main.append(r)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Detalle'

    headers = [
        'Albarán', 'Fecha', 'Productor', 'Tipo Fruta', 'OC Asociada', 'Guía Despacho',
        'Producto', 'Categoría', 'Manejo', 'Bandejas (unidades)', 'Kg Hechos', 'UOM',
        'Costo Unitario', 'Costo Total', 'Clasificación', '% IQF', '% Block'
    ]
    ws.append(headers)
    header_font = Font(bold=True)
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Fill detalle rows
    for r in recepciones_main:
        albaran = r.get('albaran', '')
        fecha = fmt_fecha(r.get('fecha', ''))
        productor = r.get('productor', '')
        tipo_fruta = r.get('tipo_fruta', '')
        guia = r.get('guia_despacho', '')
        calific = r.get('calific_final', '')
        pct_iqf = r.get('total_iqf', 0) or 0
        pct_block = r.get('total_block', 0) or 0

        productos = r.get('productos', []) or []
        oc_asociada = r.get('oc_asociada', '')
        if not productos:
            # Solo agregar si hay kg recepcionados
            kg_rec = r.get('kg_recepcionados', 0) or 0
            # Solo si pasa filtros de producto vacíos (o se asume válido si no hay productos pero pasó el filtro de recepción)
            # Como filtramos recepciones_main antes, si estamos aquí es porque la recepción es válida
            # Pero si hay filtro estricto de tipo fruta y la recepción no tiene productos, ¿qué hacemos?
            # Asumimos que si filter_tipo_fruta está activo, requerimos match explicito.
            if kg_rec > 0 and not filter_tipo_fruta and not filter_manejo:
                 ws.append([albaran, fecha, productor, tipo_fruta, oc_asociada, guia, '', '', '', '', kg_rec, '', '', '', calific, pct_iqf, pct_block])
        else:
            for p in productos:
                kg_hechos = p.get('Kg Hechos', 0) or 0
                # Ignorar productos con 0 kg
                if kg_hechos <= 0:
                    continue
                
                # Usar TipoFruta del producto, fallback a tipo_fruta de la recepción
                tipo_fruta_prod = (p.get('TipoFruta') or tipo_fruta or '').strip()
                manejo = (p.get('Manejo') or '').strip()
                
                # --- FILTROS DE PRODUCTO ---
                if filter_tipo_fruta and tipo_fruta_prod not in filter_tipo_fruta:
                    continue
                if filter_manejo and manejo not in filter_manejo:
                    continue
                # ---------------------------
                    
                prod_name = p.get('Producto', '')
                categoria = p.get('Categoria', '')
                uom = p.get('UOM', '')
                costo_unit = p.get('Costo Unitario', 0) or 0
                costo_total = p.get('Costo Total', 0) or 0
                bandejas_units = kg_hechos if (categoria or '').upper() == 'BANDEJAS' else ''
                
                ws.append([
                    albaran, fecha, productor, tipo_fruta_prod, oc_asociada, guia,
                    prod_name, categoria, manejo, bandejas_units, kg_hechos, uom,
                    costo_unit, costo_total, calific, pct_iqf, pct_block
                ])

    # Apply number formats for numeric columns
    # Find indices
    col_map = {h: i + 1 for i, h in enumerate(headers)}
    kg_col = col_map['Kg Hechos']
    costo_unit_col = col_map['Costo Unitario']
    costo_total_col = col_map['Costo Total']
    bandejas_col = col_map['Bandejas (unidades)']

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=len(headers)):
        # kg
        row[kg_col - 1].number_format = '#,##0.00'
        # cost
        row[costo_unit_col - 1].number_format = '[$$-en-US]#,##0.00'
        row[costo_total_col - 1].number_format = '[$$-en-US]#,##0'
        # bandejas as integer
        if row[bandejas_col - 1].value != '':
            try:
                row[bandejas_col - 1].number_format = '#,##0'
            except Exception:
                pass

    # Freeze header and enable filter
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'
    _auto_fit_columns(ws)

    # Summary sheet
    summary = wb.create_sheet('Resumen')
    summary_headers = ['Tipo Fruta', 'Kg', 'Costo Total', 'Costo Promedio/kg', '# Recepciones']
    summary.append(summary_headers)
    for c_idx, h in enumerate(summary_headers, start=1):
        cell = summary.cell(row=1, column=c_idx)
        cell.font = header_font

    # Aggregate by tipo_fruta del PRODUCTO (no de la recepción)
    agrup = {}
    for r in recepciones_main:
        # Nota: recepciones_main ya tiene filtros de Productor y Clasificación aplicados
        tipo_fruta_rec = (r.get('tipo_fruta') or '').strip()
        
        # Verificar si la recepción tiene productos válidos después de filtros (para conteo de recepciones)
        # Esto es complejo porque una recepción puede tener mix de productos validos e invalidos.
        # Contaremos recepción si al menos un producto contribuyó al resumen de ese tipo.
        
        products_contrib = set() # Tipos de fruta a los que esta recepción contribuyó
        
        for p in r.get('productos', []) or []:
            cat = (p.get('Categoria') or '').upper()
            if cat == 'BANDEJAS':
                # skip bandejas from kg/costo of fruta
                continue
            
            kg = p.get('Kg Hechos', 0) or 0
            if kg <= 0:
                continue
                
            # Usar TipoFruta del producto, fallback a tipo_fruta de la recepción
            tipo = (p.get('TipoFruta') or tipo_fruta_rec or '').strip()
            manejo = (p.get('Manejo') or '').strip()
            
            # --- FILTROS DE PRODUCTO ---
            if filter_tipo_fruta and tipo not in filter_tipo_fruta:
                continue
            if filter_manejo and manejo not in filter_manejo:
                continue
            # ---------------------------
            
            if not tipo:
                tipo = 'SIN_TIPO'
            
            if tipo not in agrup:
                agrup[tipo] = {'kg': 0.0, 'costo': 0.0, 'recepciones_ids': set()}
            
            agrup[tipo]['kg'] += kg
            agrup[tipo]['costo'] += p.get('Costo Total', 0) or 0
            agrup[tipo]['recepciones_ids'].add(r.get('name')) # Usar nombre o ID único para contar

    for tipo, vals in sorted(agrup.items(), key=lambda x: x[0]):
        kg = vals['kg']
        costo = vals['costo']
        n_recep = len(vals['recepciones_ids'])
        costo_prom = (costo / kg) if kg > 0 else None
        summary.append([tipo, kg, costo, costo_prom if costo_prom is not None else '', n_recep])

    # Format summary numeric columns
    for row in summary.iter_rows(min_row=2, min_col=1, max_col=5):
        row[1].number_format = '#,##0.00'  # Kg
        row[2].number_format = '[$$-en-US]#,##0'  # costo total
        try:
            row[3].number_format = '[$$-en-US]#,##0.00'
        except Exception:
            pass
        row[4].number_format = '#,##0' # N recepciones

    _auto_fit_columns(summary)

    # Save to bytes
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    data = buf.getvalue()
    buf.close()
    return data


def generate_pallets_excel(username: str, password: str, fecha_inicio: str, fecha_fin: str,
                            manejo_filtros: Optional[List[str]] = None,
                            tipo_fruta_filtros: Optional[List[str]] = None,
                            origen_filtros: Optional[List[str]] = None) -> bytes:
    """
    Genera un Excel con el detalle de cada pallet (uno por fila).
    
    Args:
        username: Usuario Odoo
        password: Contraseña Odoo
        fecha_inicio: Fecha inicio en formato YYYY-MM-DD
        fecha_fin: Fecha fin en formato YYYY-MM-DD
        manejo_filtros: Lista de manejos a filtrar (opcional)
        tipo_fruta_filtros: Lista de tipos de fruta a filtrar (opcional)
        origen_filtros: Lista de orígenes a filtrar (RFP, VILKUN, SAN JOSE) (opcional)
    
    Returns:
        bytes: Archivo Excel en formato bytes
    """
    from backend.services.recepcion_service import get_recepciones_pallets_detailed
    
    # Obtener datos detallados (un pallet por fila)
    pallets_detail = get_recepciones_pallets_detailed(
        username, password, fecha_inicio, fecha_fin, 
        manejo_filtros, tipo_fruta_filtros, origen_filtros
    )
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle de Pallets"
    
    # Encabezados
    headers = [
        "Fecha", "Planta/Origen", "Albarán", "Productor", "Guía Despacho", 
        "Pallet ID", "Producto", "Manejo", "Tipo Fruta", "Kg"
    ]
    ws.append(headers)
    
    # Formato de encabezados
    header_font = Font(bold=True)
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Agregar datos
    for p in pallets_detail:
        ws.append([
            p.get("fecha", ""),
            p.get("origen", ""),
            p.get("albaran", ""),
            p.get("productor", ""),
            p.get("guia_despacho", ""),
            p.get("pallet_name", ""),
            p.get("producto_name", ""),
            p.get("manejo", ""),
            p.get("tipo_fruta", ""),
            p.get("kg", 0)
        ])
    
    # Formatear columna de Kg
    kg_col_idx = headers.index("Kg") + 1
    for row in ws.iter_rows(min_row=2, min_col=kg_col_idx, max_col=kg_col_idx):
        row[0].number_format = '#,##0.00'
    
    # Aplicar autofiltro y congelar paneles
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'
    _auto_fit_columns(ws)
    
    # Guardar a bytes
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    data = buf.getvalue()
    buf.close()
    return data
