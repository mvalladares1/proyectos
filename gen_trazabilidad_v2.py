import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# Estilos
header_font = Font(bold=True, color="FFFFFF", size=10)
header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
pack_font = Font(bold=True, size=11, color="FFFFFF")
pack_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
level_font = Font(size=10)
gd_font = Font(bold=True, size=10, color="006100")
gd_fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")
prod_font = Font(bold=True, size=10)
no_traz_font = Font(italic=True, color="c0392b", size=10, bold=True)
no_traz_fill = PatternFill(start_color="fce4ec", end_color="fce4ec", fill_type="solid")
thin_border = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)
center = Alignment(horizontal="center", vertical="center")

# Cada fila: (pallet_consumido, nivel2, nivel3, nivel4, guia, productor)
packs_data = [
    {
        "name": "PACK 769 (A)",
        "pallets_consumidos": "048229 - 046485 - 106069 - 044126 - 047453",
        "rows": [
            ("048229", "42534", "104811-C → 104811", "", "GD2054", "VICTOR OSVALDO FROHLICH MOHR"),
            ("048229", "041291", "025107", "008274", "GD158", "AGRÍCOLA SANTA ELENA LIMITADA"),
            ("046485", "SIN TRAZABILIDAD", "", "", "", ""),
            ("106069", "", "", "", "GD318", "LVR SOLUCIONES SPA"),
            ("044126", "042950", "105255-C → 105255", "", "GD4300", "SOC. AGRÍCOLA DEL HUERTO"),
            ("044126", "042950", "105256-C → 105256", "", "GD4300", "SOC. AGRÍCOLA DEL HUERTO"),
            ("044126", "042954", "105175-C → 105175", "", "GD4294", "SOC. AGRÍCOLA DEL HUERTO"),
            ("044126", "042954", "105152-C → 105152", "", "GD2382", "COMERCIAL ZENIZ"),
            ("044126", "042954", "105149-C → 105149", "", "GD2382", "COMERCIAL ZENIZ"),
            ("047453", "106334", "", "", "GD358", "LVR SOLUCIONES SPA"),
            ("047453", "106336", "", "", "GD358", "LVR SOLUCIONES SPA"),
            ("047453", "106335", "", "", "GD358", "LVR SOLUCIONES SPA"),
        ]
    },
    {
        "name": "PACK 769 (B)",
        "pallets_consumidos": "048229 - 046485 - 106069 - 044126 - 048709 - 047453",
        "rows": [
            ("048229", "42534", "104810-C → 104810", "", "GD2054", "VICTOR OSVALDO FROHLICH MOHR"),
            ("048229", "041291", "025107", "008274", "GD158", "AGRÍCOLA SANTA ELENA LIMITADA"),
            ("046485", "SIN TRAZABILIDAD", "", "", "", ""),
            ("106069", "", "", "", "GD318", "LVR SOLUCIONES SPA"),
            ("044126", "042950", "105255-C → 105255", "", "GD4300", "SOC. AGRÍCOLA DEL HUERTO"),
            ("044126", "042950", "105256-C → 105256", "", "GD4300", "SOC. AGRÍCOLA DEL HUERTO"),
            ("044126", "042954", "105191-C → 105175", "", "GD4294", "SOC. AGRÍCOLA DEL HUERTO"),
            ("044126", "042954", "105186-C → 105152", "", "GD2382", "COMERCIAL ZENIZ"),
            ("044126", "042954", "105184-C → 105149", "", "GD2382", "COMERCIAL ZENIZ"),
            ("047453", "106330", "", "", "GD358", "LVR SOLUCIONES SPA"),
            ("047453", "106335", "", "", "GD358", "LVR SOLUCIONES SPA"),
        ]
    },
    {
        "name": "PACK 048783",
        "pallets_consumidos": "048229 - 046485 - 106069 - 044126 - 048709 - 047453",
        "rows": [
            ("048229", "42534", "104810-C → 104810", "", "GD2054", "VICTOR OSVALDO FROHLICH MOHR"),
            ("048229", "041291", "025107", "008274", "GD158", "AGRÍCOLA SANTA ELENA LIMITADA"),
            ("046485", "SIN TRAZABILIDAD", "", "", "", ""),
            ("106069", "", "", "", "GD318", "LVR SOLUCIONES SPA"),
            ("044126", "042950", "105266-C → 105266", "", "GD4300", "SOC. AGRÍCOLA DEL HUERTO"),
            ("044126", "042950", "105267-C → 105267", "", "GD4300", "SOC. AGRÍCOLA DEL HUERTO"),
            ("044126", "042954", "105191-C → 105191", "", "GD4294", "SOC. AGRÍCOLA DEL HUERTO"),
            ("044126", "042954", "105186-C → 105186", "", "GD2382", "COMERCIAL ZENIZ"),
            ("044126", "042954", "105184-C → 105184", "", "GD2382", "COMERCIAL ZENIZ"),
            ("047453", "106330", "", "", "GD358", "LVR SOLUCIONES SPA"),
            ("047453", "106335", "", "", "GD358", "LVR SOLUCIONES SPA"),
        ]
    },
]

ws = wb.active
ws.title = "Trazabilidad Packs"

# Headers
headers = [
    "Pack",
    "Pallets Consumidos",
    "Pallet Consumido\n(Nivel 1)",
    "Nivel 2\n(Sub-pallet)",
    "Nivel 3\n(Lote / Componente)",
    "Nivel 4\n(Lote origen)",
    "Guía Despacho",
    "Productor"
]
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
ws.row_dimensions[1].height = 40

row = 2
for pack in packs_data:
    pack_start = row

    for r in pack["rows"]:
        pallet, n2, n3, n4, guia, productor = r

        # Col A - Pack
        c = ws.cell(row=row, column=1, value=pack["name"])
        c.font = pack_font
        c.fill = pack_fill
        c.alignment = center
        c.border = thin_border

        # Col B - Pallets consumidos
        c = ws.cell(row=row, column=2, value=pack["pallets_consumidos"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.font = Font(size=9)
        c.border = thin_border

        # Col C - Pallet consumido (nivel 1)
        c = ws.cell(row=row, column=3, value=pallet)
        c.alignment = center
        c.font = Font(bold=True, size=10, color="1e3a5f")
        c.border = thin_border

        # Col D - Nivel 2
        c = ws.cell(row=row, column=4, value=n2)
        c.alignment = center
        c.border = thin_border
        if n2 == "SIN TRAZABILIDAD":
            c.font = no_traz_font
            c.fill = no_traz_fill
            for cc in range(5, 9):
                cell2 = ws.cell(row=row, column=cc)
                cell2.fill = no_traz_fill
                cell2.border = thin_border
        else:
            c.font = level_font

        # Col E - Nivel 3
        c = ws.cell(row=row, column=5, value=n3)
        c.alignment = center
        c.font = level_font
        c.border = thin_border

        # Col F - Nivel 4
        c = ws.cell(row=row, column=6, value=n4)
        c.alignment = center
        c.font = level_font
        c.border = thin_border

        # Col G - Guía Despacho
        c = ws.cell(row=row, column=7, value=guia)
        c.alignment = center
        c.border = thin_border
        if guia:
            c.font = gd_font
            c.fill = gd_fill

        # Col H - Productor
        c = ws.cell(row=row, column=8, value=productor)
        c.border = thin_border
        if productor:
            c.font = prod_font

        row += 1

    # Merge Pack y Pallets Consumidos
    if row - pack_start > 1:
        ws.merge_cells(start_row=pack_start, start_column=1, end_row=row - 1, end_column=1)
        ws.cell(row=pack_start, column=1).alignment = center
        ws.merge_cells(start_row=pack_start, start_column=2, end_row=row - 1, end_column=2)
        ws.cell(row=pack_start, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Separador
    for c in range(1, 9):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        cell.border = thin_border
    ws.row_dimensions[row].height = 6
    row += 1

# Anchos
ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 24
ws.column_dimensions['F'].width = 16
ws.column_dimensions['G'].width = 18
ws.column_dimensions['H'].width = 42

ws.freeze_panes = "A2"

output = r"c:\Users\HP\Desktop\Trazabilidad_Pack769_048783.xlsx"
wb.save(output)
print(f"Excel generado: {output}")
